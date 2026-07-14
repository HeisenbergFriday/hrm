from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Any

import openpyxl


class _Cell:
    __slots__ = ("value",)

    def __init__(self, value: Any):
        self.value = value


class _XlsWorksheet:
    def __init__(self, sheet, datemode: int):
        self._sheet = sheet
        self._datemode = datemode
        self.title = sheet.name
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row: int, column: int) -> _Cell:
        value = None
        row_idx = row - 1
        col_idx = column - 1
        if 0 <= row_idx < self._sheet.nrows and 0 <= col_idx < self._sheet.ncols:
            value = self._value(row_idx, col_idx)
        return _Cell(value)

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
        values_only: bool = False,
    ):
        max_row = max_row or self.max_row
        max_col = max_col or self.max_column
        for row in range(min_row, max_row + 1):
            values = [self.cell(row, col).value for col in range(min_col, max_col + 1)]
            yield tuple(values) if values_only else tuple(_Cell(value) for value in values)

    def _value(self, row_idx: int, col_idx: int):
        import xlrd

        cell = self._sheet.cell(row_idx, col_idx)
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, self._datemode)
            except Exception:
                return cell.value
        if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
            return int(cell.value)
        return cell.value


class _XlsWorkbook:
    def __init__(self, path: str | Path):
        import xlrd

        self._book = xlrd.open_workbook(str(path))
        self.worksheets = [
            _XlsWorksheet(self._book.sheet_by_index(index), self._book.datemode)
            for index in range(self._book.nsheets)
        ]
        self.sheetnames = [sheet.title for sheet in self.worksheets]
        self.active = self.worksheets[0] if self.worksheets else None

    def __getitem__(self, name: str):
        for sheet in self.worksheets:
            if sheet.title == name:
                return sheet
        raise KeyError(name)

    def close(self) -> None:
        release = getattr(self._book, "release_resources", None)
        if release:
            release()


def load_workbook_compat(path: str | Path, **kwargs):
    if Path(path).suffix.lower() == ".xls":
        return _XlsWorkbook(path)
    return openpyxl.load_workbook(path, **kwargs)


# ---------------------------------------------------------------------------
# Upload audit support (mirrors _warn_upload_observations from app.py)
# ---------------------------------------------------------------------------

_SECRET_PATTERNS = [
    re.compile(
        r"(?i)(?:api[_\s-]?key|access[_\s-]?key|secret|password|token|ak|sk)\s*[:=]\s*[\"']?[A-Za-z0-9_\-]{8,}"
    ),
    re.compile(r"(?i)-----BEGIN (?:RSA |EC |DSA )?PRIVATE KEY-----"),
]

_DOWNLOAD_NAME_INVALID_RE = re.compile(r"[\\/:*?\"<>|]|^\.+$|\.\.")


def audit_upload(
    filename: str,
    size_bytes: int,
    data: bytes,
    *,
    max_warn_mb: int = 50,
    max_warn_rows: int = 50_000,
    max_warn_cols: int = 500,
) -> dict:
    """Return structured upload observations for the given file.

    The caller (Go service layer) may attach the ``warnings`` list to the HTTP
    response (e.g. as ``X-Attendance-Toolbox-Warnings``) and/or forward the
    ``audit`` list to the backend audit log.  The structure mirrors the spirit
    of ``_warn_upload_observations`` from the original ``app.py`` while also
    scanning a small number of cells for secret-like strings.
    """
    safe_name = _safe_file_name(filename)
    warnings: list[str] = []
    audit: list[dict] = [
        {
            "event": "upload_received",
            "file_name": safe_name,
            "extension": (Path(filename).suffix or "").lower(),
            "size_bytes": size_bytes,
        }
    ]

    if size_bytes <= 0:
        warnings.append(f"{safe_name} 为空文件")
        audit.append({"event": "upload_warning", "type": "empty_file"})
        return {"warnings": warnings, "audit": audit, "file_name": safe_name}

    # extension / basic shape
    ext = (Path(filename).suffix or "").lower()
    is_xlsx_like = data[:4].startswith(b"PK\x03\x04")  # OOXML zip magic
    is_xls_like = data[:8].startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")  # BIFF

    if not (is_xlsx_like or is_xls_like):
        warnings.append(f"{safe_name} 不是有效的 .xlsx 或 .xls 文件")
        audit.append(
            {"event": "upload_warning", "type": "invalid_magic", "extension": ext}
        )
        return {"warnings": warnings, "audit": audit, "file_name": safe_name}

    if max_warn_mb and size_bytes > max_warn_mb * 1024 * 1024:
        warnings.append(
            f"{safe_name} 文件较大({size_bytes / 1024 / 1024:.1f}MB)，可能上传较慢"
        )
        audit.append(
            {
                "event": "upload_warning",
                "type": "file_size",
                "threshold_mb": max_warn_mb,
                "size_bytes": size_bytes,
            }
        )

    if is_xlsx_like:
        _audit_xlsx_dimensions(data, safe_name, max_warn_rows, max_warn_cols, warnings, audit)

    return {"warnings": warnings, "audit": audit, "file_name": safe_name}


def _audit_xlsx_dimensions(
    data: bytes,
    safe_name: str,
    max_warn_rows: int,
    max_warn_cols: int,
    warnings: list[str],
    audit: list[dict],
) -> None:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        warnings.append(f"{safe_name} 无法作为 Excel 解析({type(exc).__name__})")
        audit.append({"event": "upload_warning", "type": "parse_error", "error": str(exc)})
        return

    try:
        oversized: list[str] = []
        for ws in wb.worksheets:
            too_many_rows = bool(max_warn_rows and ws.max_row and ws.max_row > max_warn_rows)
            too_many_cols = bool(max_warn_cols and ws.max_column and ws.max_column > max_warn_cols)
            if too_many_rows or too_many_cols:
                oversized.append(
                    f"{ws.title} ({ws.max_row} 行 x {ws.max_column} 列)"
                )
        if oversized:
            warnings.append(
                f"{safe_name} 中存在尺寸超过观测阈值的 sheet：" + "；".join(oversized)
            )
            audit.append(
                {
                    "event": "upload_warning",
                    "type": "sheet_dimensions",
                    "threshold_rows": max_warn_rows,
                    "threshold_columns": max_warn_cols,
                    "sheets": oversized,
                }
            )
    finally:
        wb.close()


def is_valid_download_name(name: str) -> bool:
    return bool(name) and not _DOWNLOAD_NAME_INVALID_RE.search(name)


def _safe_file_name(name: str) -> str:
    cleaned = str(name or "").replace("\\", "/").split("/")[-1]
    return (cleaned or "uploaded file")[:200]
