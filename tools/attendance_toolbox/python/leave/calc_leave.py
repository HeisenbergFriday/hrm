"""
calc_leave.py
──────────────────────────────────────────────────────────────────────────────
功能：
  读取 请假系统导出.xlsx（原始源表），经过清洗与计算后直接写出 请假明细表.xlsx。

  步骤一（清洗，内存处理，不生成中间文件）：
    过滤审批状态/结果，拆分部门层级，规范化时长字段。
    过滤规则：
      - 只保留审批状态为"完成"、"审批中"、"已修改"的行
      - 审批状态为"完成"但审批结果为"拒绝"的行也剔除

  步骤二（计算）：
    根据 作息表.xlsx 工作日补充三个字段：
      最终请假时长 / 最终请假天数 / 备注
    将结果写入 请假明细表.xlsx

计算规则（由行政沟通确认）：
  - 请假在目标月内、sys为数值 → final = sys（信任HR系统）
  - 请假跨越上月→目标月、或 sys 为文字天数 → 重新按目标月工作日计算
  - 请假目标月开始→下月结束   → 只计算目标月内工作日部分
  - 请假纯在下月及以后        → 标记为「X月考勤」
  - 产假按作息表应出勤日计算，不按自然日计算
  - 婚/陪产等文字sys且纯在目标月内 → N×8（N 为天数）
  - 二/三级部门为售后运营组、电话客服组、成都运营、成都客服且在目标月内 → 直接使用系统时长
  - 成都员工（部门字段含"成都"）使用成都工作日历
──────────────────────────────────────────────────────────────────────────────
"""

import argparse
import decimal
import os
import re
from datetime import date, datetime, timedelta
from functools import lru_cache

import openpyxl
from openpyxl import load_workbook

# ── 路径 ─────────────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
EXPORT_FILE   = os.path.join(BASE_DIR, "请假系统导出.xlsx")
OUT_FILE      = os.path.join(BASE_DIR, "请假明细表.xlsx")
SCHEDULE_FILE = os.path.join(BASE_DIR, "作息表.xlsx")

# ── 清洗：允许通过的审批状态 ──────────────────────────────────────────────────
KEEP_STATUSES = {"完成", "审批中", "已修改"}

# ── 标准工作时间 ──────────────────────────────────────────────────────────────
WORK_START  = (9,  0)
LUNCH_START = (12, 30)
LUNCH_END   = (14, 0)
WORK_END    = (18, 30)
WORKING_DAY_COLOR = "FFFFFF00"
STATUTORY_HOLIDAY_COLOR = "FFFF0000"
COMPANY_WELFARE_COLOR = "FF0070C0"
_SCHEDULE_COLOR_NORMALIZE_DISTANCE = 32
_SCHEDULE_COLOR_TARGETS = (
    WORKING_DAY_COLOR,
    STATUTORY_HOLIDAY_COLOR,
    COMPANY_WELFARE_COLOR,
)
MONTH_LABELS = {
    1: "一",
    2: "二",
    3: "三",
    4: "四",
    5: "五",
    6: "六",
    7: "七",
    8: "八",
    9: "九",
    10: "十",
    11: "十一",
    12: "十二",
}

ROW_EMP_ID = 0
ROW_EMP_NAME = 1
ROW_DEPT_1 = 2
ROW_DEPT_2 = 3
ROW_DEPT_3 = 4
ROW_LEAVE_TYPE = 5
ROW_START = 6
ROW_END = 7
ROW_SYS_DURATION = 8
ROW_LAUNCH_TIME = 9
ROW_FINISH_TIME = 10
ROW_APPROVAL_ID = 11
ROW_APPROVAL_STATUS = 12
ROW_APPROVAL_RESULT = 13
ROW_IS_INTERN = 14
ROW_SOURCE_ROW = 15

SYSTEM_DURATION_DEPTS = {"售后运营组", "电话客服组", "成都运营", "成都客服"}

INTERN_HINT_COLUMNS = (
    "职位",
    "岗位",
    "岗位名称",
    "职务",
    "编制类型",
    "员工类型",
    "用工形式",
    "聘用形式",
)
EMPLOYEE_TYPE_ID_COLUMNS = (
    "工号",
    "员工工号",
    "人员工号",
    "发起人工号",
    "员工编号",
    "职工号",
)
EMPLOYEE_TYPE_COLUMNS = (
    "员工类型",
    "人员类型",
    "用工类型",
    "编制类型",
    "用工形式",
    "聘用形式",
)
EMPLOYEE_TYPE_TARGET_KEYWORDS = ("实习", "外包")
SPECIAL_EMPLOYEE_SPLIT_RE = re.compile(r"[\s,，;；、]+")
DEFAULT_SPECIAL_EMPLOYEE_NAMES: tuple[str, ...] = (
    # 固定特殊人员姓名写在这里，页面会默认带出，计算前也可临时追加。
    "梁伯林",
    "陈秋宇",
    "蔡依诺",
    "王天钦",
    "周灵京",
    "秦洋",
    "翁钰雪",
    "杨冉",
    "林彤",
    "李玉琳",
    "杨静淇",
    "余固",
    "郑红军",
)
DEFAULT_CHENGDU_WORK_LOCATION_NAMES: tuple[str, ...] = (
    # 部门名称不含"成都"，但实际按成都办公作息计算的人员。
    "费婷玉",
    "张莹",
    "陈星雨",
    "伏鸣",
)

OFFSITE_DURATION_ID_COLUMNS = EMPLOYEE_TYPE_ID_COLUMNS
OFFSITE_DURATION_NAME_COLUMNS = (
    "姓名",
    "员工姓名",
    "人员姓名",
    "发起人姓名",
)
OFFSITE_DURATION_COLUMNS = (
    "请假时长",
    "申请系统时长",
    "系统时长",
    "最终请假时长",
    "申请时长",
    "时长",
)
OFFSITE_DURATION_LEAVE_TYPE_COLUMNS = (
    "请假类型",
    "假期类型",
    "假别",
)
OFFSITE_DURATION_START_COLUMNS = (
    "开始时间",
    "请假开始时间",
    "起始时间",
)
OFFSITE_DURATION_END_COLUMNS = (
    "结束时间",
    "请假结束时间",
    "截止时间",
)
OFFSITE_DURATION_APPROVAL_COLUMNS = (
    "审批编号",
    "审批单号",
    "申请编号",
)
OFFSITE_ATTENDANCE_GROUP_COLUMNS = (
    "考勤组",
    "考勤分组",
    "考勤组名称",
)
HEADER_SCAN_LIMIT = None
_HEADER_NOISE_RE = re.compile(r"[\s\u00a0\u2000-\u200f\u2028-\u202f\u205f\u3000\ufeff:：,，;；/\\|、._\-—–·*（）()【】\[\]{}<>《》]+")


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  工具函数                                                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def to_datetime(val):
    """将单元格值（str/datetime/date/int）统一转换为 datetime 对象。"""
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, (int, float)):
        # Excel date serial number（openpyxl 有时不自动转换）
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(val)
        except Exception:
            return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                pass
    return None


def has_time_component(val):
    """判断原始值是否含有时间部分（不只是日期）。"""
    if isinstance(val, datetime):
        return val.hour != 0 or val.minute != 0
    if isinstance(val, str):
        return ' ' in val.strip()
    return False


def parse_day_text(sys_val):
    """从 '4天'、'0.5天' 等文本中提取天数，失败返回 None。"""
    if isinstance(sys_val, str):
        m = re.match(r'([\d.]+)\s*天$', sys_val.strip())
        if m:
            return float(m.group(1))
    return None


def t2min(t):
    return t[0] * 60 + t[1]


def _cell_fg_rgb(cell) -> str:
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb":
            return normalize_schedule_color(fg.rgb)
    except Exception:
        pass
    return ""


def _rgb_parts(argb: str) -> tuple[int, int, int] | None:
    text = str(argb or "").strip().upper()
    if len(text) == 8:
        text = text[2:]
    if len(text) != 6 or not re.fullmatch(r"[0-9A-F]{6}", text):
        return None
    return int(text[0:2], 16), int(text[2:4], 16), int(text[4:6], 16)


def normalize_schedule_color(argb: str) -> str:
    """Normalize legacy .xls palette approximations to the app's schedule colors."""
    text = str(argb or "").strip().upper()
    if len(text) == 6:
        text = "FF" + text
    if text in _SCHEDULE_COLOR_TARGETS:
        return text

    source = _rgb_parts(text)
    if source is None:
        return text

    for target in _SCHEDULE_COLOR_TARGETS:
        target_rgb = _rgb_parts(target)
        if target_rgb is None:
            continue
        distance = sum((source[idx] - target_rgb[idx]) ** 2 for idx in range(3)) ** 0.5
        if distance <= _SCHEDULE_COLOR_NORMALIZE_DISTANCE:
            return target
    return text


def _parse_title(title_str: str):
    match = re.search(r"(\d{4})\D*(\d{1,2})\D*月", title_str or "")
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _next_month_start(year: int, month: int) -> date:
    if month == 12:
        return date(year + 1, 1, 1)
    return date(year, month + 1, 1)


def _qingming_date(year: int) -> date:
    """
    计算清明节公历日期。
    公式适用于常规考勤处理年份范围；如未来年度规则变更，应优先接入官方节假日数据源。
    """
    if 1900 <= year <= 1999:
        century_const = 5.59
    elif 2000 <= year <= 2099:
        century_const = 4.81
    else:
        raise ValueError(f"暂未内置 {year} 年清明节计算规则")

    yy = year % 100
    day = int(yy * 0.2422 + century_const) - int(yy / 4)
    return date(year, 4, day)


@lru_cache(maxsize=None)
def _statutory_legal_holidays_for_year(year: int) -> frozenset[date]:
    """返回国家法定节假日当天，不包含调休形成的连休日。"""
    try:
        from lunardate import LunarDate
    except ImportError as exc:
        raise RuntimeError("缺少 lunardate，无法计算春节/端午/中秋法定节假日") from exc

    lunar_new_year = LunarDate(year, 1, 1).toSolarDate()

    if year >= 2025:
        spring_festival = {lunar_new_year - timedelta(days=1)}
        spring_festival.update(lunar_new_year + timedelta(days=offset) for offset in range(3))
        labour_day = {date(year, 5, 1), date(year, 5, 2)}
    else:
        spring_festival = {lunar_new_year + timedelta(days=offset) for offset in range(3)}
        labour_day = {date(year, 5, 1)}

    legal_days = {
        date(year, 1, 1),
        _qingming_date(year),
        LunarDate(year, 5, 5).toSolarDate(),
        LunarDate(year, 8, 15).toSolarDate(),
        date(year, 10, 1),
        date(year, 10, 2),
        date(year, 10, 3),
    }
    legal_days.update(spring_festival)
    legal_days.update(labour_day)
    return frozenset(legal_days)


def _normalize_paid_statutory_holidays(days: set[date]) -> set[date]:
    """
    将作息表红底假期日期收敛为计入应出勤的法定假日。
    红底可能包含调休连休日，因此这里按国家法定节假日当天取交集。
    """
    legal_days: set[date] = set()
    for year in {d.year for d in days}:
        legal_days.update(_statutory_legal_holidays_for_year(year))
    return {d for d in days if d in legal_days}


# 补班周六（按需维护，跨年时可扩展为从作息表读取）
_EXTRA_WORKDAYS: set[date] = {
    date(2026, 2, 7),
}


def count_standard_workdays_in_month(d_start: date, d_end: date,
                                      month_start: date, month_end: date) -> int:
    """
    计算陪产/婚假等文字天数假期在目标月内的标准工作日数。
    标准工作日 = 周一~周五 + _EXTRA_WORKDAYS 中的补班日。
    口径：闭区间 [max(d_start, month_start), min(d_end+1天, month_end)]。
    d_end+1 原因：结束日当天为最后一天，下一个工作日才是实际复工日，
    行政将该复工日也计入本月（与实际观测值吻合）。
    """
    clip_start = max(d_start, month_start)
    clip_end   = min(d_end + timedelta(days=1), month_end)  # 闭区间上限
    if clip_start > clip_end:
        return 0
    count = 0
    cur = clip_start
    while cur <= clip_end:
        if cur.weekday() < 5 or cur in _EXTRA_WORKDAYS:
            count += 1
        cur += timedelta(days=1)
    return count


def _fix_date(cell_val, target_year: int, target_month: int):
    if isinstance(cell_val, datetime):
        try:
            return date(target_year, target_month, cell_val.day)
        except ValueError:
            return None
    if isinstance(cell_val, str):
        # 从字符串中提取日（兼容 "2026/2/29"、"29" 等格式）
        parts = re.split(r'[/\-.]', cell_val.strip())
        day = None
        for part in reversed(parts):
            try:
                day = int(part)
                break
            except ValueError:
                continue
        if day:
            try:
                return date(target_year, target_month, day)
            except ValueError:
                return None
    if isinstance(cell_val, (int, float)):
        try:
            return date(target_year, target_month, int(cell_val))
        except ValueError:
            return None
    return None


def _parse_schedule_block(ws, title_row_idx: int) -> dict | None:
    title = str(ws.cell(title_row_idx, 1).value or "").strip()
    year, month = _parse_title(title)
    if not year or not month:
        return None

    header_row_idx = None
    week_key = normalize_header_name("周数")
    for row_idx in range(title_row_idx + 1, min(ws.max_row, title_row_idx + 8) + 1):
        if normalize_header_name(ws.cell(row_idx, 1).value) == week_key:
            header_row_idx = row_idx
            break
    if header_row_idx is None:
        return None

    working_days = set()
    statutory_holiday_days = set()
    company_welfare_days = set()
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        first_val = ws.cell(row_idx, 1).value
        if isinstance(first_val, (int, float)):
            for col_idx in range(2, min(8, ws.max_column) + 1):
                cell = ws.cell(row_idx, col_idx)
                current = _fix_date(cell.value, year, month)
                if not current:
                    continue
                color = _cell_fg_rgb(cell)
                if color == WORKING_DAY_COLOR:
                    working_days.add(current)
                elif color == STATUTORY_HOLIDAY_COLOR:
                    statutory_holiday_days.add(current)
                elif color == COMPANY_WELFARE_COLOR:
                    company_welfare_days.add(current)
            continue
        if first_val is not None:
            break

    paid_statutory_days = _normalize_paid_statutory_holidays(statutory_holiday_days)
    expected_attendance_days = working_days | paid_statutory_days | company_welfare_days

    return {
        "title": title,
        "year": year,
        "month": month,
        "is_chengdu": "成都" in title or "成都" in ws.title,
        "working_days": working_days,
        "statutory_holiday_days": statutory_holiday_days,
        "paid_statutory_days": paid_statutory_days,
        "company_welfare_days": company_welfare_days,
        "expected_attendance_days": expected_attendance_days,
    }


def load_schedule_context(schedule_file: str) -> dict:
    if not os.path.exists(schedule_file):
        raise FileNotFoundError(f"未找到作息表文件：{schedule_file}")

    wb = load_workbook(schedule_file)
    blocks = []
    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            title = ws.cell(row_idx, 1).value
            if isinstance(title, str) and "作息时间表" in title:
                block = _parse_schedule_block(ws, row_idx)
                if block:
                    blocks.append(block)

    if not blocks:
        raise ValueError("作息表中未解析到有效的作息时间表。")

    month_keys = {(item["year"], item["month"]) for item in blocks}
    if len(month_keys) > 1:
        raise ValueError(f"作息表中包含多个年月：{sorted(month_keys)}，请保留单个月份后再运行。")

    year, month = next(iter(month_keys))
    month_start = date(year, month, 1)
    next_month_start = _next_month_start(year, month)
    month_end = next_month_start - timedelta(days=1)

    main_days = None
    chengdu_days = None
    main_expected_days = None
    chengdu_expected_days = None
    for block in blocks:
        if block["is_chengdu"]:
            chengdu_days = block["working_days"]
            chengdu_expected_days = block["expected_attendance_days"]
        elif main_days is None:
            main_days = block["working_days"]
            main_expected_days = block["expected_attendance_days"]

    if main_days is None:
        main_days = blocks[0]["working_days"]
    if main_expected_days is None:
        main_expected_days = blocks[0]["expected_attendance_days"]
    if chengdu_days is None:
        chengdu_days = set(main_days)
    if chengdu_expected_days is None:
        chengdu_expected_days = set(main_expected_days)

    return {
        "year": year,
        "month": month,
        "month_start": month_start,
        "month_end": month_end,
        "next_month_start": next_month_start,
        "next_month_label": f"{MONTH_LABELS.get(next_month_start.month, next_month_start.month)}月考勤",
        "main_working_days": main_days,
        "chengdu_working_days": chengdu_days,
        "main_expected_attendance_days": main_expected_days,
        "chengdu_expected_attendance_days": chengdu_expected_days,
    }


def working_hours_on_day(s_time, e_time):
    """
    计算某一工作日内 [s_time, e_time] 范围的实际工作小时数。
    自动排除午休（12:30-14:00），并限制在工作时间（9:00-18:30）内。
    s_time / e_time 为 (hour, minute) 元组。
    """
    ws = (max(t2min(s_time), t2min(WORK_START)),)
    we = (min(t2min(e_time),  t2min(WORK_END)),)

    def _clamp(mn): return mn[0]
    ws_min = _clamp(ws)
    we_min = _clamp(we)

    if we_min <= ws_min:
        return 0.0

    # 上午段：ws_min ～ min(we_min, LUNCH_START_MIN)
    ls = t2min(LUNCH_START)
    le = t2min(LUNCH_END)
    morning_end = min(we_min, ls)
    morning = max(0.0, (morning_end - ws_min) / 60)

    # 下午段：max(ws_min, LUNCH_END_MIN) ～ we_min
    af_start = max(ws_min, le)
    afternoon = max(0.0, (we_min - af_start) / 60)

    return morning + afternoon


def calc_target_month_working_hours(dt_start, dt_end, working_days, month_start, month_end):
    """
    计算 [dt_start, dt_end] 区间中，在指定工作日历（working_days）
    且在2月范围内的实际工作小时数。
    """
    d_start = dt_start.date()
    d_end   = dt_end.date()

    # 裁剪到2月范围
    clip_start = max(d_start, month_start)
    clip_end   = min(d_end, month_end)

    if clip_start > clip_end:
        return 0.0

    total = 0.0
    cur = clip_start
    while cur <= clip_end:
        if cur in working_days:
            # 确定当天的请假起止时间
            s = (dt_start.hour, dt_start.minute) if cur == d_start else WORK_START
            e = (dt_end.hour,   dt_end.minute)   if cur == d_end   else WORK_END
            total += working_hours_on_day(s, e)
        cur += timedelta(days=1)

    return total


def _round2(v):
    """四舍五入保留2位小数，与 Excel =ROUND(x,2) 行为一致（ROUND_HALF_UP）。"""
    return float(
        decimal.Decimal(str(v)).quantize(
            decimal.Decimal('0.01'), rounding=decimal.ROUND_HALF_UP
        )
    )


def _round_to_half_hour(hours):
    """向上取整到0.5小时的倍数。不够半小时按半小时算。"""
    if hours <= 0:
        return 0
    import math
    return math.ceil(hours * 2) / 2


def _dept_text(value):
    return str(value).strip() if value is not None else ""


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  核心计算函数                                                            ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def should_use_system_duration(row):
    """指定部门直接使用系统时长，不按作息表重新计算。"""
    return (
        _dept_text(row[ROW_DEPT_2]) in SYSTEM_DURATION_DEPTS
        or _dept_text(row[ROW_DEPT_3]) in SYSTEM_DURATION_DEPTS
    )


# 按自然日计算的假期类型
NATURAL_DAY_LEAVE_TYPES = {"陪产假", "丧假"}

# 按作息表应出勤日计算的假期类型
EXPECTED_ATTENDANCE_DAY_LEAVE_TYPES = {"产假"}


def is_natural_day_leave(leave_type):
    """检查是否为按自然日计算的假期类型（陪产假、丧假）。"""
    return str(leave_type).strip() in NATURAL_DAY_LEAVE_TYPES


def is_expected_attendance_day_leave(leave_type):
    """检查是否为按作息表应出勤日计算的假期类型（产假）。"""
    return str(leave_type).strip() in EXPECTED_ATTENDANCE_DAY_LEAVE_TYPES


def calc_from_system_duration(sys_val):
    if sys_val is None:
        return 0, 0, None

    days = parse_day_text(sys_val)
    if days is not None:
        final_h = days * 8
    else:
        final_h = _round2(float(sys_val))

    final_h = _round_to_half_hour(final_h)
    return (final_h, _round2(final_h / 8), None) if final_h else (0, 0, None)


def calc_final_fields(row, schedule_ctx, is_chengdu=False):
    """
    row: (工号, 姓名, 一级部门, 二级部门, 三级部门,
          请假类型, 开始时间, 结束时间, 系统时长)

    返回: (最终请假时长, 最终请假天数, 备注)
    """
    _, _, _, _, _, leave_type, raw_start, raw_end, sys_val = row
    working_days = schedule_ctx["chengdu_working_days"] if is_chengdu else schedule_ctx["main_working_days"]
    expected_attendance_key = (
        "chengdu_expected_attendance_days" if is_chengdu else "main_expected_attendance_days"
    )
    expected_attendance_days = schedule_ctx.get(expected_attendance_key) or working_days
    month_start = schedule_ctx["month_start"]
    month_end = schedule_ctx["month_end"]
    next_month_start = schedule_ctx["next_month_start"]

    dt_start = to_datetime(raw_start)
    dt_end   = to_datetime(raw_end)

    if dt_start is None or dt_end is None:
        return None, None, '日期解析失败'

    d_start = dt_start.date()
    d_end   = dt_end.date()

    # 陪产假、丧假按自然日计算
    if is_natural_day_leave(leave_type):
        # 计算目标月份内的自然日天数
        clip_start = max(d_start, month_start)
        clip_end = min(d_end, month_end)
        if clip_start <= clip_end:
            # 计算目标月份内的天数
            final_days = (clip_end - clip_start).days + 1
            final_h = final_days * 8
            final_h = _round_to_half_hour(final_h)  # 保持一致性
            return (final_h, final_days, None) if final_h else (0, 0, None)
        else:
            return 0, 0, None

    # 日期值（无时间）==> 开始默认09:00，结束默认18:30
    if not has_time_component(raw_start):
        dt_start = dt_start.replace(hour=9, minute=0)
    if not has_time_component(raw_end):
        dt_end = dt_end.replace(hour=18, minute=30)

    # ── 情形0：请假完全在上月或更早 ─────────────────────────────────────────────
    if d_end < month_start:
        label = f"{d_start.month}月请假"
        return label, label, label

    # ── 情形1：请假纯在3月及以后 ─────────────────────────────────────────────
    if d_start >= next_month_start:
        return schedule_ctx["next_month_label"], schedule_ctx["next_month_label"], None

    expected_attendance_day_leave = is_expected_attendance_day_leave(leave_type)

    # ── 情形2+3：请假完全在目标月内 ──────────────────────────────────────
    entirely_in_month = (d_start >= month_start and d_end <= month_end)
    if entirely_in_month:
        # 产假按作息表应出勤日计算，不能使用系统自然日时长
        if expected_attendance_day_leave:
            final_h = calc_target_month_working_hours(
                dt_start, dt_end, expected_attendance_days, month_start, month_end)
            final_h = _round_to_half_hour(final_h)
            return (final_h, _round2(final_h / 8), None) if final_h else (0, 0, None)
        # 指定运营/客服部门直接使用系统时长，不重新计算
        if should_use_system_duration(row):
            return calc_from_system_duration(sys_val)
        # 其他部门按作息表计算
        final_h = calc_target_month_working_hours(dt_start, dt_end, working_days, month_start, month_end)
        final_h = _round_to_half_hour(final_h)
        return (final_h, _round2(final_h / 8), None) if final_h else (0, 0, None)

    # ── 情形4：请假跨越（上月→本月 或 本月→下月 或 更长跨度） ───────────────
    # sys 为文字天数（陪产/婚假等）→ 按结束日是否落在作息工作日选择口径：
    #   落在作息工作日  → 按作息表工作日计算（与普通跨月假一致）
    #   落在非作息日（节假日/周末）→ 标准工作日计算，顺延一天（含当天后下一工作日）
    if expected_attendance_day_leave:
        final_h = calc_target_month_working_hours(
            dt_start, dt_end, expected_attendance_days, month_start, month_end)
        final_h = _round_to_half_hour(final_h)
        return (final_h, _round2(final_h / 8), None) if final_h else (0, 0, None)

    n = parse_day_text(sys_val)
    if n is not None:
        if d_end in working_days:
            # 结束日是作息表工作日（如节前陪产假）→ 直接按作息表
            final_h    = calc_target_month_working_hours(
                dt_start, dt_end, working_days, month_start, month_end)
            final_h    = _round_to_half_hour(final_h)
            final_days = _round2(final_h / 8)
        else:
            # 结束日落在节假日/周末（如春节期间的产假）→ 标准工作日+顺延
            # count_standard_workdays_in_month 内部已将 d_end+1 作为闭区间上限
            std_days   = count_standard_workdays_in_month(d_start, d_end, month_start, month_end)
            final_h    = std_days * 8
            final_days = std_days
        return (final_h, final_days, None) if final_h else (0, 0, None)

    # sys 为数值或其他文本 → 按作息表工作日计算本月部分
    final_h = calc_target_month_working_hours(dt_start, dt_end, working_days, month_start, month_end)
    final_h = _round_to_half_hour(final_h)

    if final_h == 0:
        return 0, 0, None

    final_days = _round2(final_h / 8)
    return final_h, final_days, None


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  步骤一：清洗导出表                                                      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def parse_export_duration(val):
    """
    将请假系统导出表的"时长"字段规范化为请假副本所需的"系统时长"格式：
      'N小时'  → float N          （如 '1小时' → 1.0，'3.5小时' → 3.5）
      'N天'    → 原字符串          （如 '4天' → '4天'，供 parse_day_text 消费）
      其他     → 原值（透传）
    """
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r'^([\d.]+)\s*小时$', s)
    if m:
        return float(m.group(1))
    if re.match(r'^\d+\s*天$', s):
        return s
    return val


def split_dept(dept_str):
    """
    将 '一级-二级-三级' 形式的部门字符串拆分为三元组。
    不足三级的用 None 补齐；超过三级时多余部分合并到第三级。
    """
    if not dept_str:
        return None, None, None
    parts = str(dept_str).split('-')
    d1 = parts[0] or None
    d2 = parts[1] if len(parts) > 1 else None
    d3 = '-'.join(parts[2:]) if len(parts) > 2 else None
    return d1, d2, (d3 or None)


def normalize_employee_name(name):
    if name is None:
        return None
    normalized = str(name).strip()
    normalized = re.sub(r'\s*[\(（]\s*已离职\s*[\)）]\s*$', '', normalized)
    return normalized


def normalize_employee_id(emp_id):
    if emp_id is None:
        return ""
    if isinstance(emp_id, int):
        return str(emp_id)
    if isinstance(emp_id, float):
        if emp_id.is_integer():
            return str(int(emp_id))
        return format(emp_id, 'f').rstrip('0').rstrip('.')

    text = str(emp_id).strip()
    if re.fullmatch(r'\d+\.0+', text):
        return text.split('.', 1)[0]
    return text


def parse_special_employee_names(text: str | None) -> tuple[str, ...]:
    if not text:
        return ()

    employee_names: list[str] = []
    seen: set[str] = set()
    for raw_item in SPECIAL_EMPLOYEE_SPLIT_RE.split(str(text)):
        emp_name = normalize_employee_name(raw_item) or ""
        if not emp_name or emp_name in seen:
            continue
        seen.add(emp_name)
        employee_names.append(emp_name)
    return tuple(employee_names)


def find_intern_column(col_map: dict) -> str | None:
    for col_name in INTERN_HINT_COLUMNS:
        if col_name in col_map:
            return col_name
    normalized_hints = tuple(
        normalized for hint in INTERN_HINT_COLUMNS if (normalized := normalize_header_name(hint))
    )
    for actual_name in col_map:
        normalized_name = normalize_header_name(actual_name)
        if any(hint in actual_name or hint in normalized_name for hint in INTERN_HINT_COLUMNS):
            return actual_name
        if any(hint in normalized_name for hint in normalized_hints):
            return actual_name
    return None


def is_intern_from_position(value) -> bool:
    if value is None:
        return False
    return "实习" in str(value)


def is_intern_from_employee_type(value) -> bool:
    if value is None:
        return False
    text = str(value)
    return any(keyword in text for keyword in EMPLOYEE_TYPE_TARGET_KEYWORDS)


def find_column_by_candidates(col_map: dict, candidates: tuple[str, ...]) -> str | None:
    for col_name in candidates:
        if col_name in col_map:
            return col_name

    normalized_candidates = tuple(
        normalized for candidate in candidates if (normalized := normalize_header_name(candidate))
    )
    normalized_to_actual = {
        normalize_header_name(actual_name): actual_name
        for actual_name in col_map
        if normalize_header_name(actual_name)
    }
    for candidate in normalized_candidates:
        if candidate in normalized_to_actual:
            return normalized_to_actual[candidate]

    for actual_name in col_map:
        normalized_name = normalize_header_name(actual_name)
        if any(candidate != "时长" and candidate in actual_name for candidate in candidates):
            return actual_name
        if any(candidate != "时长" and candidate in normalized_name for candidate in normalized_candidates):
            return actual_name
    return None


def find_offsite_duration_column(col_map: dict) -> str | None:
    matched = find_column_by_candidates(col_map, OFFSITE_DURATION_COLUMNS)
    if matched:
        return matched

    duration_context_keywords = ("请假", "申请系统", "系统")
    normalized_context_keywords = tuple(normalize_header_name(keyword) for keyword in duration_context_keywords)
    for actual_name in col_map:
        normalized_name = normalize_header_name(actual_name)
        if "时长" in normalized_name and any(
            keyword in normalized_name for keyword in normalized_context_keywords
        ):
            return actual_name
    return None


def normalize_header_name(value) -> str:
    if value is None:
        return ""
    return _HEADER_NOISE_RE.sub("", str(value).strip())


def compact_header_row(row: tuple) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx, name in enumerate(row):
        normalized = normalize_header_name(name)
        if normalized and normalized not in col_map:
            col_map[normalized] = idx
    return col_map


def _column_match_strength(col_map: dict, candidates: tuple[str, ...]) -> int:
    if any(col_name in col_map for col_name in candidates):
        return 3

    normalized_candidates = tuple(
        normalized for candidate in candidates if (normalized := normalize_header_name(candidate))
    )
    normalized_names = tuple(
        normalize_header_name(actual_name)
        for actual_name in col_map
        if normalize_header_name(actual_name)
    )
    if any(candidate in normalized_names for candidate in normalized_candidates):
        return 3

    for actual_name in col_map:
        normalized_name = normalize_header_name(actual_name)
        if any(candidate in actual_name for candidate in candidates):
            return 2
        if any(candidate in normalized_name for candidate in normalized_candidates):
            return 2
    return 0


def score_employee_type_header(col_map: dict) -> int:
    emp_id_score = _column_match_strength(col_map, EMPLOYEE_TYPE_ID_COLUMNS)
    emp_type_score = _column_match_strength(col_map, EMPLOYEE_TYPE_COLUMNS)
    if not emp_id_score or not emp_type_score:
        return 0
    return emp_id_score * 10 + emp_type_score * 10 + min(len(col_map), 10)


def score_offsite_duration_header(col_map: dict) -> int:
    emp_id_score = _column_match_strength(col_map, OFFSITE_DURATION_ID_COLUMNS)
    emp_name_score = _column_match_strength(col_map, OFFSITE_DURATION_NAME_COLUMNS)
    if not emp_id_score and not emp_name_score:
        return 0

    score = max(emp_id_score, emp_name_score) * 10 + min(len(col_map), 10)
    if emp_id_score and emp_name_score:
        score += 8
    if find_offsite_duration_column(col_map):
        score += 4
    for candidates in (
        OFFSITE_DURATION_LEAVE_TYPE_COLUMNS,
        OFFSITE_DURATION_START_COLUMNS,
        OFFSITE_DURATION_END_COLUMNS,
        OFFSITE_DURATION_APPROVAL_COLUMNS,
        OFFSITE_ATTENDANCE_GROUP_COLUMNS,
    ):
        if find_column_by_candidates(col_map, candidates):
            score += 2
    return score


def _matcher_score(col_map: dict, matcher) -> int:
    result = matcher(col_map)
    if isinstance(result, bool):
        return 1 if result else 0
    if isinstance(result, (int, float)):
        return int(result)
    return 0


def _following_data_score(rows: list[tuple], header_row_idx: int, col_map: dict) -> int:
    data_rows = 0
    filled_cells = 0
    col_indexes = tuple(col_map.values())
    for row in rows[header_row_idx + 1:header_row_idx + 11]:
        row_filled_cells = sum(
            1
            for col_idx in col_indexes
            if col_idx < len(row) and not _is_blank(row[col_idx])
        )
        if row_filled_cells:
            data_rows += 1
            filled_cells += row_filled_cells
    return min(data_rows, 5) * 3 + min(filled_cells, 10)


def _find_table_header(rows: list[tuple], matcher, max_scan_rows: int | None = HEADER_SCAN_LIMIT):
    best_row = None
    best_col = None
    best_score = 0
    scan_rows = rows if max_scan_rows is None else rows[:max_scan_rows]
    for row_idx, row in enumerate(scan_rows):
        current_col = compact_header_row(row)
        score = _matcher_score(current_col, matcher) if current_col else 0
        if score:
            score += _following_data_score(rows, row_idx, current_col)
        if score > best_score:
            best_row = row_idx
            best_col = current_col
            best_score = score
    return best_row, best_col, best_score


def _find_table_in_workbook(wb_src, matcher, max_scan_rows: int | None = HEADER_SCAN_LIMIT):
    first_rows = []
    best_sheet = None
    best_rows = []
    best_header_row = None
    best_col = None
    best_score = 0
    for ws_src in wb_src.worksheets:
        rows = list(ws_src.iter_rows(values_only=True))
        if rows and not first_rows:
            first_rows = rows
        header_row, col, score = _find_table_header(rows, matcher, max_scan_rows)
        if score > best_score:
            best_sheet = ws_src.title
            best_rows = rows
            best_header_row = header_row
            best_col = col
            best_score = score
    if best_col is not None:
        return best_sheet, best_rows, best_header_row, best_col
    return None, first_rows, None, None


def is_offsite_attendance_group(value) -> bool:
    if _is_blank(value):
        return False
    text = str(value).strip()
    return "异地外勤" in text and ("免打卡" in text or "不打卡" in text)


def _is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def parse_duration_hours(value) -> float | None:
    """解析用户上传的请假时长，数字默认按小时，文本支持 N小时 / N天。"""
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    number = r"\d+(?:\.\d+)?"

    match = re.fullmatch(rf"({number})\s*(?:小时|钟头|h|H)", text)
    if match:
        return float(match.group(1))

    match = re.fullmatch(rf"({number})\s*(?:天|日|d|D)", text)
    if match:
        return float(match.group(1)) * 8

    if re.fullmatch(number, text):
        return float(text)

    raise ValueError(f"无法解析请假时长：{value}")


def calc_from_duration_hours(hours: float | int | None):
    if hours is None:
        return 0, 0, None
    final_h = _round_to_half_hour(_round2(float(hours)))
    return (final_h, _round2(final_h / 8), None) if final_h else (0, 0, None)


def _normalize_match_datetime(value) -> str:
    if _is_blank(value):
        return ""
    dt = to_datetime(value)
    if dt is None:
        return str(value).strip()
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _offsite_person_keys(emp_id, emp_name) -> list[tuple[str, str]]:
    keys = []
    norm_id = normalize_employee_id(emp_id)
    norm_name = normalize_employee_name(emp_name)
    if norm_id:
        keys.append(("id", norm_id))
    if norm_name:
        keys.append(("name", norm_name))
    return keys


def _offsite_row_keys(emp_id, emp_name, leave_type, start_time, end_time) -> list[tuple[str, str, str, str, str]]:
    leave_type_text = str(leave_type or "").strip()
    start_key = _normalize_match_datetime(start_time)
    end_key = _normalize_match_datetime(end_time)
    if not leave_type_text or not start_key or not end_key:
        return []
    return [
        (*person_key, leave_type_text, start_key, end_key)
        for person_key in _offsite_person_keys(emp_id, emp_name)
    ]


def load_offsite_duration_overrides(offsite_duration_file: str) -> dict:
    """
    读取异地不打卡人员请假时长表。

    支持列：
      - 工号或姓名：用于匹配人员，至少需要一个
      - 请假时长/申请系统时长/系统时长/时长：可选；缺失时匹配行直接使用原请假系统时长
      - 审批编号，或 请假类型+开始时间+结束时间：可选；存在时优先精确匹配单条申请
      - 考勤组：可选；若存在“异地外勤（免打卡）”，只载入该考勤组人员
    """
    if not os.path.exists(offsite_duration_file):
        raise FileNotFoundError(f"未找到异地不打卡人员请假时长表：{offsite_duration_file}")

    wb_src = load_workbook(offsite_duration_file, data_only=True)
    try:
        sheet_name, rows, header_row, col = _find_table_in_workbook(
            wb_src,
            score_offsite_duration_header,
        )
    finally:
        wb_src.close()

    if not rows:
        raise ValueError("异地不打卡人员请假时长表为空。")

    if header_row is None or col is None:
        raise ValueError("异地不打卡人员请假时长表缺少“工号”或“姓名”字段。")

    emp_id_col = find_column_by_candidates(col, OFFSITE_DURATION_ID_COLUMNS)
    emp_name_col = find_column_by_candidates(col, OFFSITE_DURATION_NAME_COLUMNS)
    duration_col = find_offsite_duration_column(col)
    leave_type_col = find_column_by_candidates(col, OFFSITE_DURATION_LEAVE_TYPE_COLUMNS)
    start_col = find_column_by_candidates(col, OFFSITE_DURATION_START_COLUMNS)
    end_col = find_column_by_candidates(col, OFFSITE_DURATION_END_COLUMNS)
    approval_col = find_column_by_candidates(col, OFFSITE_DURATION_APPROVAL_COLUMNS)
    attendance_group_col = find_column_by_candidates(col, OFFSITE_ATTENDANCE_GROUP_COLUMNS)

    use_attendance_group_filter = False
    if attendance_group_col:
        attendance_group_idx = col[attendance_group_col]
        use_attendance_group_filter = any(
            len(row) > attendance_group_idx
            and is_offsite_attendance_group(row[attendance_group_idx])
            for row in rows[header_row + 1:]
            if any(row)
        )

    overrides = {
        "by_approval": {},
        "by_row": {},
        "by_employee": {},
    }

    loaded_rows = 0
    skipped_identity = 0
    specific_rules = 0
    employee_rules = 0
    upload_duration_rows = 0
    duplicate_rules = 0
    skipped_non_offsite_group = 0

    for source_row_num, row in enumerate(rows[header_row + 1:], start=header_row + 2):
        if not any(row):
            continue

        if use_attendance_group_filter:
            attendance_group = row[col[attendance_group_col]]
            if not is_offsite_attendance_group(attendance_group):
                skipped_non_offsite_group += 1
                continue

        emp_id = row[col[emp_id_col]] if emp_id_col else None
        emp_name = row[col[emp_name_col]] if emp_name_col else None
        person_keys = _offsite_person_keys(emp_id, emp_name)
        if not person_keys:
            skipped_identity += 1
            continue

        duration_raw = row[col[duration_col]] if duration_col else None
        duration_hours = None
        if not _is_blank(duration_raw):
            try:
                duration_hours = parse_duration_hours(duration_raw)
                upload_duration_rows += 1
            except ValueError as exc:
                raise ValueError(f"第 {source_row_num} 行{exc}") from exc

        leave_type = row[col[leave_type_col]] if leave_type_col else None
        start_time = row[col[start_col]] if start_col else None
        end_time = row[col[end_col]] if end_col else None
        approval_id = str(row[col[approval_col]]).strip() if approval_col and not _is_blank(row[col[approval_col]]) else ""
        row_keys = _offsite_row_keys(emp_id, emp_name, leave_type, start_time, end_time)

        rule = {
            "hours": duration_hours,
            "raw_duration": duration_raw,
            "source_row": source_row_num,
        }

        has_specific_rule = False
        if approval_id:
            if approval_id in overrides["by_approval"]:
                duplicate_rules += 1
            overrides["by_approval"][approval_id] = rule
            specific_rules += 1
            has_specific_rule = True

        for row_key in row_keys:
            if row_key in overrides["by_row"]:
                duplicate_rules += 1
            overrides["by_row"][row_key] = rule
            specific_rules += 1
            has_specific_rule = True

        if not has_specific_rule:
            for person_key in person_keys:
                if person_key in overrides["by_employee"]:
                    duplicate_rules += 1
                overrides["by_employee"][person_key] = rule
                employee_rules += 1

        loaded_rows += 1

    if loaded_rows == 0:
        raise ValueError("异地不打卡人员请假时长表未读取到有效人员。")

    print(f'[异地不打卡] 识别工作表：{sheet_name}，表头第 {header_row + 1} 行')
    print(f'[异地不打卡] 识别工号字段：{emp_id_col or "未提供"}，姓名字段：{emp_name_col or "未提供"}，时长字段：{duration_col or "未提供"}，考勤组字段：{attendance_group_col or "未提供"}')
    print(f'[异地不打卡] 共载入 {loaded_rows} 行，上传时长 {upload_duration_rows} 行，精确规则 {specific_rules} 条，人员名单规则 {employee_rules} 条')
    if use_attendance_group_filter:
        print(f'[异地不打卡] 已按“{attendance_group_col}”筛选异地外勤（免打卡），跳过非免打卡 {skipped_non_offsite_group} 行')
    if duration_col is None:
        print('[异地不打卡] 未找到时长字段，匹配人员将直接使用请假系统导出表中的系统时长')
    if skipped_identity:
        print(f'[异地不打卡] 跳过缺少工号/姓名的行 {skipped_identity} 行')
    if duplicate_rules:
        print(f'[异地不打卡] 存在 {duplicate_rules} 条重复匹配规则，已按最后一条记录覆盖')

    return overrides


def find_offsite_duration_override(row: tuple, overrides: dict | None) -> dict | None:
    if not overrides:
        return None

    approval_id = str(row[ROW_APPROVAL_ID] or "").strip()
    if approval_id and approval_id in overrides.get("by_approval", {}):
        rule = dict(overrides["by_approval"][approval_id])
        rule["match_type"] = "审批编号"
        return rule

    for row_key in _offsite_row_keys(
        row[ROW_EMP_ID],
        row[ROW_EMP_NAME],
        row[ROW_LEAVE_TYPE],
        row[ROW_START],
        row[ROW_END],
    ):
        if row_key in overrides.get("by_row", {}):
            rule = dict(overrides["by_row"][row_key])
            rule["match_type"] = "请假时间"
            return rule

    for person_key in _offsite_person_keys(row[ROW_EMP_ID], row[ROW_EMP_NAME]):
        if person_key in overrides.get("by_employee", {}):
            rule = dict(overrides["by_employee"][person_key])
            rule["match_type"] = "人员名单"
            return rule

    return None


def calc_offsite_duration_fields(row: tuple, rule: dict):
    if rule.get("hours") is not None:
        final_h, final_days, _ = calc_from_duration_hours(rule["hours"])
        return final_h, final_days, "异地不打卡-按上传时长"

    final_h, final_days, _ = calc_from_system_duration(row[ROW_SYS_DURATION])
    return final_h, final_days, "异地不打卡-按系统时长"


def load_employee_type_map(employee_type_file: str) -> dict[str, str]:
    if not os.path.exists(employee_type_file):
        raise FileNotFoundError(f"未找到员工类型表文件：{employee_type_file}")

    wb_src = load_workbook(employee_type_file, data_only=True)
    try:
        sheet_name, rows, header_row, col = _find_table_in_workbook(
            wb_src,
            score_employee_type_header,
        )
    finally:
        wb_src.close()

    if not rows:
        raise ValueError("员工类型表为空。")

    if header_row is None or col is None:
        raise ValueError('员工类型表缺少“工号”或“员工类型”字段。')

    emp_id_col = find_column_by_candidates(col, EMPLOYEE_TYPE_ID_COLUMNS)
    emp_type_col = find_column_by_candidates(col, EMPLOYEE_TYPE_COLUMNS)
    assert emp_id_col is not None and emp_type_col is not None

    employee_type_map: dict[str, str] = {}
    duplicate_count = 0
    for row in rows[header_row + 1:]:
        if not any(row):
            continue
        if col[emp_id_col] >= len(row) or col[emp_type_col] >= len(row):
            continue
        emp_id = normalize_employee_id(row[col[emp_id_col]])
        emp_type = str(row[col[emp_type_col]]).strip() if row[col[emp_type_col]] is not None else ""
        if not emp_id or not emp_type:
            continue
        if emp_id in employee_type_map and employee_type_map[emp_id] != emp_type:
            duplicate_count += 1
        employee_type_map[emp_id] = emp_type

    intern_count = sum(1 for emp_type in employee_type_map.values() if is_intern_from_employee_type(emp_type))
    print(f'[员工类型] 识别工作表：{sheet_name}，表头第 {header_row + 1} 行')
    print(f'[员工类型] 识别工号字段：{emp_id_col}，员工类型字段：{emp_type_col}')
    print(f'[员工类型] 共载入 {len(employee_type_map)} 个工号，其中实习员工 {intern_count} 个')
    if duplicate_count:
        print(f'[员工类型] 存在 {duplicate_count} 个重复工号，已按最后一条记录覆盖')
    return employee_type_map



def _approval_result_rank(result) -> int:
    text = str(result or "").strip()
    if text == "同意":
        return 2
    if text:
        return 1
    return 0


def _approval_status_rank(status) -> int:
    text = str(status or "").strip()
    if text == "完成":
        return 2
    if text == "已修改":
        return 3
    if text == "审批中":
        return 0
    return -1


def has_modified_status(*rows: tuple) -> bool:
    return any(str(row[ROW_APPROVAL_STATUS] or "").strip() == "已修改" for row in rows)


def get_modified_row_priority(row: tuple) -> tuple:
    finish_time = to_datetime(row[ROW_FINISH_TIME]) or datetime.min
    launch_time = to_datetime(row[ROW_LAUNCH_TIME]) or datetime.min
    return finish_time, launch_time, row[ROW_SOURCE_ROW]


def get_row_priority(row: tuple) -> tuple:
    finish_time = to_datetime(row[ROW_FINISH_TIME]) or datetime.min
    launch_time = to_datetime(row[ROW_LAUNCH_TIME]) or datetime.min
    return (
        _approval_result_rank(row[ROW_APPROVAL_RESULT]),
        _approval_status_rank(row[ROW_APPROVAL_STATUS]),
        finish_time,
        launch_time,
        row[ROW_SOURCE_ROW],
    )


def is_better_leave_row(current_row: tuple, candidate_row: tuple) -> bool:
    if has_modified_status(current_row, candidate_row):
        return get_modified_row_priority(candidate_row) > get_modified_row_priority(current_row)
    return get_row_priority(candidate_row) > get_row_priority(current_row)


def clean_export(export_file: str) -> list[tuple]:
    """
    读取请假系统导出文件，按状态过滤并整理为标准内存格式，返回行列表。
    第0行为表头元组，后续为数据行元组。

    保留条件：
      审批状态 in {'完成', '审批中', '已修改'}
      且 NOT (审批状态 == '完成' AND 审批结果 == '拒绝')

    返回列顺序：
      发起人工号 / 发起人姓名 / 一级部门 / 二级部门 / 三级部门
      / 请假类型 / 开始时间 / 结束时间 / 系统时长
      / 发起时间 / 完成时间 / 审批编号 / 审批状态 / 审批结果
      / 是否实习生 / 源文件行号
    """
    if not os.path.exists(export_file):
        raise FileNotFoundError(f"未找到请假系统导出文件：{export_file}")

    wb_src = load_workbook(export_file, data_only=True)
    ws_src = wb_src.active
    rows = list(ws_src.iter_rows(values_only=True))
    wb_src.close()

    if not rows:
        raise ValueError("请假系统导出文件为空。")

    # 按表头动态定位列索引，兼容列顺序变动
    header = rows[0]
    col = compact_header_row(header)
    required_cols = {"审批状态", "审批结果", "发起人工号", "发起人姓名",
                     "发起人部门", "请假类型", "开始时间", "结束时间", "时长", "审批编号"}
    missing = {name for name in required_cols if normalize_header_name(name) not in col}
    if missing:
        raise ValueError(f"请假系统导出文件缺少必要字段：{missing}")

    intern_col_name = find_intern_column(col)
    out_rows: list[tuple] = [
        ('发起人工号', '发起人姓名', '一级部门', '二级部门', '三级部门',
         '请假类型', '开始时间', '结束时间', '系统时长',
         '发起时间', '完成时间', '审批编号', '审批状态', '审批结果',
         '是否实习生', '源文件行号')
    ]

    kept = skipped = 0
    revoked = 0
    approval_deduped = 0
    staged_by_approval: dict[str, tuple] = {}
    for source_row_num, row in enumerate(rows[1:], start=2):
        if not any(row):       # 全空行跳过
            continue
        status = row[col[normalize_header_name("审批状态")]]
        result = row[col[normalize_header_name("审批结果")]]

        # 过滤：剔除已撤销
        if str(status or "").strip() == "已撤销":
            skipped += 1
            revoked += 1
            continue
        # 过滤：仅保留指定审批状态
        if status not in KEEP_STATUSES:
            skipped += 1
            continue
        # 过滤：完成 + 拒绝 也剔除
        if status == "完成" and result == "拒绝":
            skipped += 1
            continue

        emp_id     = normalize_employee_id(row[col[normalize_header_name("发起人工号")]])
        emp_name   = normalize_employee_name(row[col[normalize_header_name("发起人姓名")]])
        d1, d2, d3 = split_dept(row[col[normalize_header_name("发起人部门")]])
        leave_type = row[col[normalize_header_name("请假类型")]]
        start_time = row[col[normalize_header_name("开始时间")]]
        end_time   = row[col[normalize_header_name("结束时间")]]
        duration   = parse_export_duration(row[col[normalize_header_name("时长")]])
        launch_key = normalize_header_name("发起时间")
        finish_key = normalize_header_name("完成时间")
        launch_time = row[col[launch_key]] if launch_key in col else None
        finish_time = row[col[finish_key]] if finish_key in col else None
        approval_id = row[col[normalize_header_name("审批编号")]]
        is_intern = False
        if intern_col_name:
            is_intern = is_intern_from_position(row[col[intern_col_name]])

        cleaned_row = (
            emp_id, emp_name, d1, d2, d3,
            leave_type, start_time, end_time, duration,
            launch_time, finish_time, approval_id, status, result,
            is_intern, source_row_num,
        )
        approval_key = str(approval_id).strip()
        existing_row = staged_by_approval.get(approval_key)
        if existing_row is None:
            staged_by_approval[approval_key] = cleaned_row
        elif is_better_leave_row(existing_row, cleaned_row):
            staged_by_approval[approval_key] = cleaned_row
            approval_deduped += 1
        else:
            approval_deduped += 1
        kept += 1

    deduped_rows = sorted(staged_by_approval.values(), key=lambda item: item[ROW_SOURCE_ROW])
    out_rows.extend(deduped_rows)
    print(f'[清洗] 保留 {kept} 行，剔除 {skipped} 行（含不符合审批状态及"完成+拒绝"数据）')
    if revoked:
        print(f'[清洗] 其中已撤销记录 {revoked} 行')
    if approval_deduped:
        print(f'[清洗] 已按审批编号保留更有效记录，去重 {approval_deduped} 行')
    if intern_col_name:
        print(f'[清洗] 实习生识别字段：{intern_col_name}')
    else:
        print('[清洗] 未找到职位相关字段，本次不拆分实习生子表数据')
    return out_rows


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  成都员工判断                                                            ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def is_chengdu_row(row, special_chengdu_names: set[str] | tuple[str, ...] | None = None):
    """检查一/二/三级部门是否含"成都"，或姓名是否在成都作息名单内。"""
    emp_name = normalize_employee_name(row[ROW_EMP_NAME]) or ""
    if special_chengdu_names and emp_name in special_chengdu_names:
        return True
    for col in (2, 3, 4):
        val = row[col]
        if val and '成都' in str(val):
            return True
    return False


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  主处理：读取源文件，计算三字段，写出明细表                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def decide_intern_target(
    row: tuple,
    employee_type_map: dict[str, str] | None = None,
    special_employee_names: set[str] | tuple[str, ...] | None = None,
) -> tuple[bool, str | None]:
    emp_id = normalize_employee_id(row[ROW_EMP_ID])
    emp_name = normalize_employee_name(row[ROW_EMP_NAME]) or ""
    if special_employee_names and emp_name in special_employee_names:
        return True, "special_list"
    if employee_type_map and emp_id in employee_type_map:
        return is_intern_from_employee_type(employee_type_map[emp_id]), "employee_type"
    return bool(row[ROW_IS_INTERN]), None


def process(src_rows: list[tuple], out_file: str, schedule_ctx: dict,
            employee_type_map: dict[str, str] | None = None,
            offsite_duration_overrides: dict | None = None,
            special_employee_names: tuple[str, ...] | set[str] | None = None,
            special_chengdu_names: tuple[str, ...] | set[str] | None = None):
    """清洗后的内存行列表 → 计算三字段 → 写入请假明细表。"""
    wb_out = openpyxl.Workbook()
    ws_fulltime = wb_out.active
    ws_fulltime.title = '请假明细'
    ws_intern = wb_out.create_sheet('实习生请假明细')

    print(f'        当前作息月份：{schedule_ctx["year"]}年{schedule_ctx["month"]}月')

    special_employee_name_set = set(special_employee_names or ())
    special_chengdu_name_set = set(special_chengdu_names or ())
    special_chengdu_matched = 0
    processed = 0
    fulltime_count = 0
    intern_count = 0
    special_list_matched = 0
    employee_type_matched = 0
    employee_type_intern = 0
    fallback_intern = 0
    offsite_matched = 0
    offsite_upload_duration = 0
    offsite_system_duration = 0

    for i, row in enumerate(src_rows):
        if i == 0:
            # 写表头，追加三列
            new_header = list(row[:9]) + ['最终请假时长', '最终请假天数', '备注']
            ws_fulltime.append(new_header)
            ws_intern.append(new_header)
            continue

        if not row[ROW_EMP_ID]:  # 空行跳过
            continue

        offsite_rule = find_offsite_duration_override(row, offsite_duration_overrides)
        if offsite_rule:
            final_h, final_days, remark = calc_offsite_duration_fields(row, offsite_rule)
            offsite_matched += 1
            if offsite_rule.get("hours") is not None:
                offsite_upload_duration += 1
            else:
                offsite_system_duration += 1
        else:
            is_cd = is_chengdu_row(row, special_chengdu_name_set)
            if is_cd and (normalize_employee_name(row[ROW_EMP_NAME]) or "") in special_chengdu_name_set:
                special_chengdu_matched += 1
            final_h, final_days, remark = calc_final_fields(row[:9], schedule_ctx, is_chengdu=is_cd)
        is_intern, intern_match_source = decide_intern_target(
            row, employee_type_map, special_employee_name_set
        )
        target_ws = ws_intern if is_intern else ws_fulltime
        target_ws.append(list(row[:9]) + [final_h, final_days, remark])
        processed += 1
        if intern_match_source == "special_list":
            special_list_matched += 1
        elif intern_match_source == "employee_type":
            employee_type_matched += 1
            if is_intern:
                employee_type_intern += 1
        elif is_intern:
            fallback_intern += 1
        if is_intern:
            intern_count += 1
        else:
            fulltime_count += 1

    wb_out.save(out_file)
    print(f'\n[完成] 已写出 → {out_file}')
    print(f'       共处理 {processed} 行')
    print(f'       全职表 {fulltime_count} 行，实习生子表 {intern_count} 行')
    if offsite_duration_overrides is not None:
        print(f'       异地不打卡匹配 {offsite_matched} 行，其中按上传时长 {offsite_upload_duration} 行，按系统时长 {offsite_system_duration} 行')
    if special_employee_names is not None:
        print(f'       特殊名单设置 {len(special_employee_name_set)} 个姓名，匹配 {special_list_matched} 行，已写入实习生子表')
    if special_chengdu_names is not None:
        print(f'       成都作息名单设置 {len(special_chengdu_name_set)} 个姓名，匹配 {special_chengdu_matched} 行，按成都作息计算')
    if employee_type_map is not None:
        print(f'       员工类型表匹配 {employee_type_matched} 行，其中实习生 {employee_type_intern} 行')
        print(f'       未匹配员工类型表、沿用原始识别的实习生 {fallback_intern} 行')


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  主入口                                                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def parse_args():
    parser = argparse.ArgumentParser(
        description='请假字段自动计算工具（按作息表动态解析工作日）'
    )
    parser.add_argument('--export',   default=EXPORT_FILE,   help=f'请假系统导出文件（默认: {EXPORT_FILE}）')
    parser.add_argument('--output',   default=OUT_FILE,      help=f'请假明细表输出（默认: {OUT_FILE}）')
    parser.add_argument('--schedule', default=SCHEDULE_FILE, help=f'作息表文件（默认: {SCHEDULE_FILE}）')
    parser.add_argument('--special-employees', help='特殊名单姓名，支持换行、逗号、顿号、空格分隔（可选）')
    parser.add_argument('--special-employees-file', help='特殊名单文本文件，支持 UTF-8/GBK（可选）')
    parser.add_argument('--special-chengdu-names', help='成都作息名单姓名，支持换行、逗号、顿号、空格分隔（可选）')
    parser.add_argument('--employee-type', help='员工类型表文件（可选）')
    parser.add_argument('--offsite-duration', help='异地不打卡人员请假时长表文件（可选）')
    return parser.parse_args()


def load_special_employee_names(special_text: str | None = None, special_file: str | None = None) -> tuple[str, ...]:
    parts: list[str] = []
    if special_text:
        parts.append(special_text)

    if special_file:
        if not os.path.exists(special_file):
            raise FileNotFoundError(f"未找到特殊名单文件：{special_file}")
        last_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "gbk"):
            try:
                with open(special_file, "r", encoding=encoding) as fh:
                    parts.append(fh.read())
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise ValueError(f"特殊名单文件编码无法识别：{last_error}") from last_error

    if parts:
        return parse_special_employee_names("\n".join(parts))
    return DEFAULT_SPECIAL_EMPLOYEE_NAMES


def load_special_chengdu_names(special_text: str | None = None) -> tuple[str, ...]:
    if special_text:
        return parse_special_employee_names(special_text)
    return DEFAULT_CHENGDU_WORK_LOCATION_NAMES


if __name__ == '__main__':
    args = parse_args()

    print('── 请假字段自动计算工具 ─────────────────────────────────────────\n')

    print('── 步骤一：清洗请假系统导出表 ────────────────────────────────────')
    src_rows = clean_export(args.export)
    print()

    print('── 步骤二：计算请假字段 ──────────────────────────────────────────')
    schedule_ctx = load_schedule_context(args.schedule)
    special_employee_names = load_special_employee_names(
        args.special_employees,
        args.special_employees_file,
    )
    special_chengdu_names = load_special_chengdu_names(args.special_chengdu_names)
    print(f'── 步骤二点四：特殊名单 {len(special_employee_names)} 个姓名 ─────────────────────────────')
    print(f'── 步骤二点四点五：成都作息名单 {len(special_chengdu_names)} 个姓名 ─────────────────────────')
    employee_type_map = None
    if args.employee_type:
        print('── 步骤二点五：加载员工类型表 ───────────────────────────────────')
        employee_type_map = load_employee_type_map(args.employee_type)
        print()
    offsite_duration_overrides = None
    if args.offsite_duration:
        print('── 步骤二点六：加载异地不打卡人员请假时长表 ───────────────────────')
        offsite_duration_overrides = load_offsite_duration_overrides(args.offsite_duration)
        print()
    process(
        src_rows,
        args.output,
        schedule_ctx,
        employee_type_map,
        offsite_duration_overrides,
        special_employee_names,
        special_chengdu_names,
    )
