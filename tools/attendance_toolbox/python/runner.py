from __future__ import annotations

import argparse
import contextlib
import io
import json
import os
import re
import sys
import traceback
from datetime import date, datetime
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
# Add BASE_DIR for dingtalk_sync.py which lives at the root level
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))
for subdir in ("leave", "overtime", "subsidy", "finally", "parttime"):
    path = str(BASE_DIR / subdir)
    if path not in sys.path:
        sys.path.insert(0, path)

import calc_leave
import fill_overtime_fields as ot
from rules_engine import get_default_config, save_config
import calc_subsidy_deduction as sub
import calc_finally as fin
import calc_parttime_summary as part
import dingtalk_sync
from excel_compat import audit_upload
from rules_adapter import config_to_preview_dict, resolve_overtime_config
import templates


NAME_SEPARATOR_RE = re.compile(r"[,，、;；\r\n]+")


def path_or_empty(config: dict, key: str) -> str:
    value = str(config.get(key) or "").strip()
    return value if value and os.path.exists(value) else ""


def paths(config: dict, key: str) -> list[str]:
    result = []
    for value in config.get(key) or []:
        value = str(value or "").strip()
        if value and os.path.exists(value):
            result.append(value)
    return result


def names(config: dict, key: str) -> list[str]:
    raw = config.get(key)
    if isinstance(raw, str):
        return [item.strip() for item in NAME_SEPARATOR_RE.split(raw) if item.strip()]
    if isinstance(raw, list):
        result = []
        for value in raw:
            result.extend(item.strip() for item in NAME_SEPARATOR_RE.split(str(value)) if item.strip())
        return result
    return []


def names_or_default(config: dict, key: str, default: list[str] | tuple[str, ...] | set[str]) -> list[str]:
    if key in config:
        return names(config, key)
    return list(default)


def default_text_values() -> dict[str, list[str]]:
    return {
        "leave_special_names": list(calc_leave.DEFAULT_SPECIAL_EMPLOYEE_NAMES),
        "chengdu_schedule_names": list(calc_leave.DEFAULT_CHENGDU_WORK_LOCATION_NAMES),
        "sub_dept_keywords": list(sub.RD_DEPT_KEYWORDS),
        "sub_late22_names": list(sub.LATE22_INCLUDED_NAMES),
        "part_special_names": list(part.DEFAULT_SPECIAL_DEFAULT_NAMES),
    }


def parse_target_month(value, label: str = "处理月份") -> tuple[int, int] | None:
    """Parse an optional YYYY-MM / YYYY-MM-DD month lock."""
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text.replace("/", "-")
    for fmt in ("%Y-%m", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            return parsed.year, parsed.month
        except ValueError:
            continue
    raise ValueError(f"{label}格式不正确，请使用 YYYY-MM。")


def validate_target_month_match(
    value,
    detected_month: tuple[int, int] | None,
    *,
    label: str,
    detected_label: str = "作息表",
) -> tuple[int, int] | None:
    locked_month = parse_target_month(value, label)
    if locked_month and detected_month and locked_month != detected_month:
        raise ValueError(
            f"{label}为 {locked_month[0]}年{locked_month[1]}月，"
            f"但{detected_label}识别为 {detected_month[0]}年{detected_month[1]}月；请保持一致。"
        )
    return locked_month


def select_subsidy_legal_holidays(
    schedule_holidays: set[date],
    override_holidays: set[date],
    year: int,
    month: int,
) -> tuple[set[date], dict]:
    """Choose one holiday source and describe conflicts for audit output."""
    schedule_dates = sub.filter_dates_to_month(schedule_holidays, year, month)
    override_dates = sub.filter_dates_to_month(override_holidays, year, month)
    if override_dates:
        selected = override_dates
        source = "custom_rules"
    else:
        selected = schedule_dates
        source = "schedule"
    only_in_rules = sorted(override_dates - schedule_dates)
    only_in_schedule = sorted(schedule_dates - override_dates) if override_dates else []
    return selected, {
        "source": source,
        "schedule_count": len(schedule_dates),
        "override_count": len(override_dates),
        "selected_count": len(selected),
        "only_in_rules": [item.isoformat() for item in only_in_rules],
        "only_in_schedule": [item.isoformat() for item in only_in_schedule],
        "conflict_count": len(only_in_rules) + len(only_in_schedule),
    }


def build_subsidy_audit_meta(
    *,
    year: int,
    month: int,
    holiday_audit: dict,
    missing_attendance: list[str],
    intern_names: list[str],
) -> dict:
    missing_names = sorted({str(name).strip() for name in missing_attendance if str(name).strip()})
    excluded_interns = sorted({str(name).strip() for name in intern_names if str(name).strip()})
    return {
        "subsidy_audit": {
            "target_month": f"{year:04d}-{month:02d}",
            "holiday_source": holiday_audit.get("source", "schedule"),
            "holiday_count": int(holiday_audit.get("selected_count", 0) or 0),
            "holiday_conflict_count": int(holiday_audit.get("conflict_count", 0) or 0),
            "holiday_only_in_rules": list(holiday_audit.get("only_in_rules") or []),
            "holiday_only_in_schedule": list(holiday_audit.get("only_in_schedule") or []),
            "missing_attendance_count": len(missing_names),
            "missing_attendance_names": missing_names,
            "intern_excluded_count": len(excluded_interns),
            "intern_excluded_names": excluded_interns,
        },
        "counts": {
            "missing_attendance": len(missing_names),
            "intern_excluded": len(excluded_interns),
            "holiday_conflicts": int(holiday_audit.get("conflict_count", 0) or 0),
        },
    }


def print_json(payload: dict) -> None:
    print(json.dumps(payload))


def capture(fn, *args, **kwargs):
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        result = fn(*args, **kwargs)
    return result, buf.getvalue()


def merge_overtime_schedule_map(target: dict, source: dict) -> None:
    for name, group_map in (source or {}).items():
        target_group = target.setdefault(name, {})
        for group_name, dates in (group_map or {}).items():
            target_group.setdefault(group_name, set()).update(dates or set())


def run_leave(config: dict, output_dir: Path) -> list[dict]:
    export_path = path_or_empty(config, "leave_src")
    schedule_path = path_or_empty(config, "leave_schedule")
    if not export_path or not schedule_path:
        raise ValueError("请上传请假系统导出表和作息表。")

    out_path = output_dir / "请假明细表.xlsx"
    src_rows, _ = capture(calc_leave.clean_export, export_path)
    schedule_ctx, _ = capture(calc_leave.load_schedule_context, schedule_path)
    src_rows, _ = capture(
        calc_leave.merge_maternity_leave_overrides,
        src_rows,
        config.get("maternity_leave_overrides"),
        schedule_ctx,
    )
    offsite_overrides = None
    offsite_path = path_or_empty(config, "leave_offsite_duration")
    if offsite_path:
        offsite_overrides, _ = capture(calc_leave.load_offsite_duration_overrides, offsite_path)
    _, _ = capture(
        calc_leave.process,
        src_rows,
        str(out_path),
        schedule_ctx,
        None,
        offsite_overrides,
        tuple(names_or_default(config, "leave_special_names", calc_leave.DEFAULT_SPECIAL_EMPLOYEE_NAMES)),
        tuple(names_or_default(config, "chengdu_schedule_names", calc_leave.DEFAULT_CHENGDU_WORK_LOCATION_NAMES)),
    )
    return [{"path": str(out_path), "file_name": out_path.name}]


def format_roster_department_error(diagnostic: dict | None, roster_path: str) -> str:
    """根据花名册解析诊断信息生成明确错误，避免只返回笼统提示。"""
    prefix = f"花名册/员工信息表未识别到可用部门映射（{roster_path}）。"
    if not diagnostic:
        return prefix + " 请检查表头是否包含员工编号/姓名及一级/二级/三级部门或部门路径/部门名称等字段。"
    matched = diagnostic.get("matched_sheets") or []
    headers_preview = diagnostic.get("headers_preview") or []
    missing = diagnostic.get("missing", "员工标识字段或部门字段缺失")
    sheets_msg = f"实际识别到的工作表：{matched}" if matched else "未识别到任何包含员工标识和部门字段的表头行。"
    header_lines = []
    for item in headers_preview[:3]:
        sheet = item.get("sheet")
        headers = item.get("headers")
        header_lines.append(f"  - [{sheet}] 表头：{headers}")
    headers_msg = "实际表头示例：\n" + "\n".join(header_lines) if header_lines else ""
    parts = [prefix, missing, sheets_msg]
    if headers_msg:
        parts.append(headers_msg)
    return " ".join(parts[:3]) + ("\n" + headers_msg if headers_msg else "")


def run_overtime(config: dict, output_dir: Path) -> list[dict]:
    export_path = path_or_empty(config, "overtime_src")
    if not export_path:
        raise ValueError("请上传加班系统导出表。")

    out_path = output_dir / "加班明细_回填.xlsx"
    # Adapter: resolve custom rules_json / rules_file, else default.
    rules_config, rules_meta = resolve_overtime_config(config)
    print(f"[规则] 本次计算使用：{rules_meta.get('source', 'default')}")
    attendance_path = path_or_empty(config, "overtime_attendance")
    calendar_path = path_or_empty(config, "overtime_calendar")
    roster_path = path_or_empty(config, "overtime_roster")
    schedule_paths = paths(config, "overtime_schedules")

    name_group_map = None
    if attendance_path:
        name_group_map, _ = capture(ot.load_attendance_name_group_map_if_available, attendance_path)

    employee_department_map = {}
    roster_diagnostic: dict | None = None
    if roster_path:
        (employee_department_map, roster_diagnostic), _ = capture(ot.parse_employee_department_map, roster_path)
        if not employee_department_map:
            raise ValueError(format_roster_department_error(roster_diagnostic, roster_path))

    src_rows, _ = capture(ot.clean_export_overtime, export_path)
    _, _ = capture(ot.load_work_calendar, calendar_path, rules_config)
    target_ym = ot.get_work_calendar_month_key() or ot._infer_target_month_key(src_rows)

    # Optional explicit target month from config (YYYY-MM or YYYY-MM-DD).
    target_month_anchor = None
    target_month_raw = str(config.get("overtime_target_month") or config.get("target_month") or "").strip()
    if target_month_raw:
        calendar_ym = ot.get_work_calendar_month_key()
        locked_ym = validate_target_month_match(
            target_month_raw,
            calendar_ym,
            label="加班处理月份",
        )
        target_month_anchor = date(locked_ym[0], locked_ym[1], 1)
        target_ym = locked_ym
        print(f"[月份] 使用请求指定目标月份：{target_ym[0]}年{target_ym[1]}月")

    schedule_map = {}
    if schedule_paths:
        for schedule_path in schedule_paths:
            loaded_result, _ = capture(
                ot.load_schedule_if_available,
                schedule_path,
                None,
                target_ym,
                set(),
                rules_config,
            )
            single_map, loaded = loaded_result
            if loaded:
                merge_overtime_schedule_map(schedule_map, single_map)
    else:
        loaded_result, _ = capture(ot.load_schedule_if_available, "", None, target_ym, set(), rules_config)
        schedule_map, _ = loaded_result

    attendance_missing_checkout_map, _ = capture(ot.load_attendance_if_available, attendance_path)
    attendance_missing_checkout_map, _ = attendance_missing_checkout_map
    _, _ = capture(
        ot.process_overtime,
        src_rows,
        str(out_path),
        schedule_map,
        attendance_missing_checkout_map,
        name_group_map,
        rules_config,
        target_month_anchor,
        tuple(names_or_default(config, "chengdu_schedule_names", calc_leave.DEFAULT_CHENGDU_WORK_LOCATION_NAMES)),
        employee_department_map,
    )
    # Persist rules meta for Go/frontend without changing business output files.
    meta_path = output_dir / "overtime_rules_meta.json"
    meta_path.write_text(json.dumps(rules_meta, ensure_ascii=False, indent=2), encoding="utf-8")
    return [
        {"path": str(out_path), "file_name": out_path.name},
        {"path": str(meta_path), "file_name": meta_path.name, "kind": "meta"},
    ]


def run_subsidy(config: dict, output_dir: Path) -> list[dict]:
    src_path = path_or_empty(config, "subsidy_src")
    checkin_path = path_or_empty(config, "subsidy_checkin")
    schedule_path = path_or_empty(config, "subsidy_schedule")
    if not src_path or not checkin_path or not schedule_path:
        raise ValueError("请上传补贴扣款表、签到表和作息表。")

    att_path = path_or_empty(config, "subsidy_attendance") or src_path
    att_result_path = path_or_empty(config, "subsidy_attendance_result")
    if att_path == src_path and not sub.looks_like_attendance_summary(src_path):
        raise ValueError("未上传考勤表，且补贴扣款表中未识别到日期级考勤结果列。")

    rd_keywords = names_or_default(config, "sub_dept_keywords", sub.RD_DEPT_KEYWORDS)
    late22_names = names_or_default(config, "sub_late22_names", sub.LATE22_INCLUDED_NAMES)
    out_path = output_dir / "补贴扣款表_核对.xlsx"
    diff_path = output_dir / "补贴扣款表_差异清单.xlsx"

    period_year, period_month, period_source = sub.resolve_schedule_period(schedule_path)
    locked_ym = validate_target_month_match(
        config.get("subsidy_target_month") or config.get("target_month"),
        (period_year, period_month),
        label="补贴考勤月份",
    )
    if locked_ym:
        print(f"[月份] 使用请求指定补贴月份：{locked_ym[0]}年{locked_ym[1]}月")
    source_records = sub.parse_source_table(src_path, rd_dept_keywords=rd_keywords, year=period_year, month=period_month)
    activity_days = sub.parse_activity_checkin(checkin_path)
    employees = sub.parse_attendance(
        att_path,
        rd_dept_keywords=rd_keywords,
        year=period_year,
        month=period_month,
        period_source=period_source,
    )
    schedule_holidays = sub.load_statutory_holidays_from_schedule(
        schedule_path,
        target_year=period_year,
        target_month=period_month,
    )
    override_holidays: set[date] = set()
    rules_source = "default"
    if config.get("rules_json") or config.get("rules_file"):
        rules_config, rules_meta = resolve_overtime_config(config)
        override_holidays = set(rules_config.legal_holidays_override or set())
        rules_source = str(rules_meta.get("source") or "custom")
    legal_holidays, holiday_audit = select_subsidy_legal_holidays(
        set(schedule_holidays),
        override_holidays,
        period_year,
        period_month,
    )
    holiday_audit["rules_source"] = rules_source
    if holiday_audit["source"] == "custom_rules":
        print(f"[法定节假日] 使用加班规则配置，共 {len(legal_holidays)} 天")
        if holiday_audit["conflict_count"]:
            print(
                f"[法定节假日] 与作息表存在 {holiday_audit['conflict_count']} 个日期差异，"
                "已记录到异常审计。"
            )
    else:
        print(f"[法定节假日] 使用作息表，共 {len(legal_holidays)} 天")
    approved_dates_by_name = None
    if att_result_path:
        approved_dates_by_name = sub.parse_attendance_result(
            att_result_path,
            year=period_year,
            month=period_month,
        )

    intern_names = [
        str(item.get("name") or "").strip()
        for item in sub.split_intern_records(source_records)[1]
        if str(item.get("name") or "").strip()
    ]
    mismatches, missing_attendance = sub.write_output(
        src_path,
        source_records,
        employees,
        activity_days,
        str(out_path),
        approved_dates_by_name=approved_dates_by_name,
        legal_holidays=legal_holidays,
        late22_included_names=late22_names,
        audit_context={
            "target_month": f"{period_year:04d}-{period_month:02d}",
            "holiday_source": holiday_audit["source"],
            "holiday_count": holiday_audit["selected_count"],
            "holiday_conflict_count": holiday_audit["conflict_count"],
            "holiday_only_in_rules": holiday_audit["only_in_rules"],
            "holiday_only_in_schedule": holiday_audit["only_in_schedule"],
        },
    )
    outputs = [{"path": str(out_path), "file_name": out_path.name}]
    if mismatches:
        sub.write_mismatch_report(mismatches, str(diff_path))
        outputs.append({"path": str(diff_path), "file_name": diff_path.name})
    audit_meta = build_subsidy_audit_meta(
        year=period_year,
        month=period_month,
        holiday_audit=holiday_audit,
        missing_attendance=missing_attendance,
        intern_names=intern_names,
    )
    audit_meta_path = output_dir / "subsidy_audit_meta.json"
    audit_meta_path.write_text(
        json.dumps(audit_meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    outputs.append({"path": str(audit_meta_path), "file_name": audit_meta_path.name, "kind": "meta"})
    return outputs


def run_final(config: dict, output_dir: Path) -> list[dict]:
    active_path = path_or_empty(config, "final_active")
    schedule_path = path_or_empty(config, "final_schedule")
    leave_path = path_or_empty(config, "final_leave")
    overtime_path = path_or_empty(config, "final_overtime")
    subsidy_path = path_or_empty(config, "final_subsidy")
    if not all([active_path, schedule_path, leave_path, overtime_path, subsidy_path]):
        raise ValueError("请上传在职花名册、作息表、请假明细表、加班明细表和补贴扣款表。")

    out_path = output_dir / "最终表.xlsx"
    resign_path = path_or_empty(config, "final_resign") or None
    transfer_path = path_or_empty(config, "final_transfer")

    employees = fin.parse_roster(active_path, resign_path)
    attendance_records = fin.parse_attendance_identity(subsidy_path)
    employees = fin.apply_attendance_identity(employees, attendance_records)
    transfer_map = fin.parse_transfer(transfer_path) if transfer_path else {}
    schedule_ctx = fin.parse_schedule(schedule_path)
    chengdu_names = tuple(names_or_default(config, "chengdu_schedule_names", calc_leave.DEFAULT_CHENGDU_WORK_LOCATION_NAMES))
    leave_map = fin.parse_leave_summary(leave_path, schedule_ctx)
    leave_day_details = fin.parse_leave_day_details(leave_path, schedule_ctx, chengdu_names)
    overtime_map = fin.parse_overtime_summary(
        overtime_path,
        schedule_ctx["year"],
        schedule_ctx["month"],
        employees,
        schedule_ctx,
    )
    absent_by_no, absent_by_name = fin.parse_subsidy_absent(subsidy_path)
    absent_day_details = fin.parse_subsidy_absent_day_details(
        subsidy_path,
        schedule_ctx["year"],
        schedule_ctx["month"],
    )
    fin.generate(
        employees,
        transfer_map,
        schedule_ctx,
        leave_map,
        overtime_map,
        absent_by_no,
        absent_by_name,
        str(out_path),
        leave_day_details,
        chengdu_names,
        absent_day_details,
    )
    return [{"path": str(out_path), "file_name": out_path.name}]


def run_parttime(config: dict, output_dir: Path) -> list[dict]:
    out_path = output_dir / "兼职汇总.xlsx"
    default_schedule_path = path_or_empty(config, "parttime_default_schedule")
    if not default_schedule_path:
        raise ValueError("请上传默认作息表。")
    part.generate_parttime_summary(
        output_path=str(out_path),
        attendance_detail_path=path_or_empty(config, "parttime_attendance_detail") or None,
        monthly_summary_paths=paths(config, "parttime_monthly"),
        schedule_paths=paths(config, "parttime_schedules"),
        default_schedule_path=default_schedule_path,
        special_default_names=names_or_default(config, "part_special_names", part.DEFAULT_SPECIAL_DEFAULT_NAMES),
    )
    return [{"path": str(out_path), "file_name": out_path.name}]


def run_dingtalk_sync(config: dict, output_dir: Path) -> list[dict]:
    """Run DingTalk sync and generate intermediate tables + structured meta."""
    from datetime import datetime

    # Parse config parameters
    start_date_str = config.get("dingtalk_sync_start_date")
    end_date_str = config.get("dingtalk_sync_end_date")
    if not start_date_str or not end_date_str:
        raise ValueError("请提供同步开始日期和结束日期。")

    start_date = datetime.strptime(str(start_date_str), "%Y-%m-%d").date()
    end_date = datetime.strptime(str(end_date_str), "%Y-%m-%d").date()

    # Flow keys: leave, overtime, attendance_correction, position_transfer
    flow_keys = config.get("dingtalk_sync_flow_keys") or ["leave", "overtime", "attendance_correction", "position_transfer"]
    if isinstance(flow_keys, str):
        flow_keys = [k.strip() for k in flow_keys.split(",") if k.strip()]

    # Max instances per flow (None means unlimited)
    max_instances = config.get("dingtalk_sync_max_instances")
    if max_instances is not None:
        max_instances = int(max_instances)

    # Query window padding days
    padding_days = int(config.get("dingtalk_sync_padding_days", 31))

    # DingTalk config from environment or config (never log secrets)
    dingtalk_config = dingtalk_sync.config_from_mapping(config.get("dingtalk", {}))

    # Run sync
    result = dingtalk_sync.sync_date_range(
        dingtalk_config,
        start_date,
        end_date,
        max_instances_per_flow=max_instances,
        flow_keys=flow_keys,
        query_window_padding_days=padding_days,
    )

    # Write exports to output_dir
    outputs: list[dict] = []
    file_index: list[dict] = []
    for key, export_data in result.exports.items():
        if export_data and export_data.data:
            out_path = output_dir / export_data.file_name
            out_path.write_bytes(export_data.data)
            item = {
                "path": str(out_path),
                "file_name": export_data.file_name,
                "flow_key": key,
                "row_count": int(getattr(export_data, "row_count", 0) or 0),
                "source_count": int(getattr(export_data, "source_count", 0) or 0),
                "skipped_count": int(getattr(export_data, "skipped_count", 0) or 0),
                "kind": "export",
            }
            outputs.append(item)
            file_index.append({k: v for k, v in item.items() if k != "path"})

    # Write audit report if available
    if result.audit_export and result.audit_export.data:
        audit_path = output_dir / result.audit_export.file_name
        audit_path.write_bytes(result.audit_export.data)
        item = {
            "path": str(audit_path),
            "file_name": result.audit_export.file_name,
            "kind": "audit",
            "row_count": int(getattr(result.audit_export, "row_count", 0) or 0),
        }
        outputs.append(item)
        file_index.append({k: v for k, v in item.items() if k != "path"})

    for msg in result.messages or []:
        print(msg)

    meta = {
        "start_date": str(start_date),
        "end_date": str(end_date),
        "flow_keys": list(flow_keys),
        "counts": dict(result.counts or {}),
        "messages": list(result.messages or []),
        "files": file_index,
        "padding_days": padding_days,
        "max_instances": max_instances,
    }
    meta_path = output_dir / "dingtalk_sync_meta.json"
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs.append({"path": str(meta_path), "file_name": meta_path.name, "kind": "meta"})
    return outputs


def run_quick_workflow(config: dict, output_dir: Path) -> list[dict]:
    """
    One-shot workflow: DingTalk sync → leave/overtime using intermediate files.
    Equivalent to D:\\app quick linkage without re-uploading intermediate tables.
    """
    run_leave_flag = str(config.get("run_leave", "true")).lower() not in {"0", "false", "no"}
    run_overtime_flag = str(config.get("run_overtime", "true")).lower() not in {"0", "false", "no"}
    if not run_leave_flag and not run_overtime_flag:
        raise ValueError("一键联动至少选择请假或加班其中一项。")

    schedule_path = path_or_empty(config, "leave_schedule") or path_or_empty(config, "overtime_calendar")
    if not schedule_path:
        raise ValueError("一键联动必须上传作息表。")

    # Ensure both leave/overtime can read the shared schedule.
    config = dict(config)
    config.setdefault("leave_schedule", schedule_path)
    config.setdefault("overtime_calendar", schedule_path)

    # Sync first into this output_dir.
    sync_outputs = run_dingtalk_sync(config, output_dir)
    leave_src = ""
    overtime_src = ""
    for item in sync_outputs:
        name = str(item.get("file_name") or "")
        path = str(item.get("path") or "")
        if name == "请假系统导出.xlsx":
            leave_src = path
        elif name == "加班系统导出.xlsx":
            overtime_src = path

    outputs = list(sync_outputs)
    workflow_meta: dict = {
        "run_leave": run_leave_flag,
        "run_overtime": run_overtime_flag,
        "leave_src": bool(leave_src),
        "overtime_src": bool(overtime_src),
    }

    if run_leave_flag:
        if not leave_src:
            raise ValueError("钉钉同步未生成请假系统导出，无法计算请假明细。")
        leave_cfg = dict(config)
        leave_cfg["leave_src"] = leave_src
        leave_cfg["leave_schedule"] = schedule_path
        leave_outputs = run_leave(leave_cfg, output_dir)
        outputs.extend(leave_outputs)
        workflow_meta["leave_files"] = [o.get("file_name") for o in leave_outputs]

    if run_overtime_flag:
        if not overtime_src:
            raise ValueError("钉钉同步未生成加班系统导出，无法计算加班明细。")
        ot_cfg = dict(config)
        ot_cfg["overtime_src"] = overtime_src
        ot_cfg["overtime_calendar"] = schedule_path
        ot_outputs = run_overtime(ot_cfg, output_dir)
        outputs.extend(ot_outputs)
        workflow_meta["overtime_files"] = [o.get("file_name") for o in ot_outputs]

    meta_path = output_dir / "quick_workflow_meta.json"
    meta_path.write_text(json.dumps(workflow_meta, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs.append({"path": str(meta_path), "file_name": meta_path.name, "kind": "meta"})
    return outputs


RUNNERS = {
    "leave": run_leave,
    "overtime": run_overtime,
    "subsidy": run_subsidy,
    "final": run_final,
    "parttime": run_parttime,
    "dingtalk_sync": run_dingtalk_sync,
    "quick": run_quick_workflow,
}


# ── Action: export-rules ─────────────────────────────────────────────────────
# 将加班规则配置导出为 Excel 文件

def action_export_rules(config: dict, output_dir: Path) -> list[dict]:
    out_path = output_dir / "加班规则配置.xlsx"
    rules_config, _ = resolve_overtime_config(config) if config.get("rules_json") or config.get("rules_file") else (get_default_config(), {})
    save_config(rules_config, str(out_path))
    return [{"path": str(out_path), "file_name": out_path.name}]


# ── Action: import-rules-preview ─────────────────────────────────────────────
# 导入加班规则 Excel/JSON，返回 JSON 预览（不写入业务结果，只解析）

def action_import_rules_preview(config: dict, output_dir: Path) -> list[dict]:
    rules_path = path_or_empty(config, "rules_file")
    if not rules_path and not config.get("rules_json"):
        raise ValueError("请上传加班规则配置文件，或提供 rules_json。")

    loaded_config, meta = resolve_overtime_config(config if rules_path or config.get("rules_json") else {"rules_file": rules_path})
    preview = config_to_preview_dict(loaded_config)
    preview["meta"] = meta
    # 写入 preview JSON 文件供 Go 层读取
    preview_path = output_dir / "rules_preview.json"
    preview_path.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")
    return [{"path": str(preview_path), "file_name": preview_path.name}]


# ── Action: validate ─────────────────────────────────────────────────────────
# 校验上传的 Excel 文件表头是否匹配预期

EXPECTED_HEADERS = {
    "leave": {
        "export": ["姓名", "工号", "开始时间", "结束时间", "请假类型", "审批状态"],
        "schedule": ["日期"],
    },
    "overtime": {
        "export": ["姓名", "工号", "开始时间", "结束时间"],
    },
    "subsidy": {
        "source": ["姓名", "工号"],
    },
    "final": {
        "roster": ["姓名", "工号"],
        "schedule": ["日期"],
    },
    "parttime": {
        "detail": ["姓名", "工号"],
    },
}


def _check_headers(file_path: str, expected: list[str], sheet_index: int = 0) -> dict:
    """检查 Excel 文件的表头是否包含预期列。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        missing = [h for h in expected if h not in headers]
        return {"ok": len(missing) == 0, "headers": headers, "missing": missing}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def action_validate(config: dict, output_dir: Path) -> list[dict]:
    module = config.get("validate_module", "")
    results = {}

    if module == "leave":
        export_path = path_or_empty(config, "leave_src")
        schedule_path = path_or_empty(config, "leave_schedule")
        if export_path:
            results["leave_export"] = _check_headers(export_path, EXPECTED_HEADERS["leave"]["export"])
        if schedule_path:
            results["leave_schedule"] = _check_headers(schedule_path, EXPECTED_HEADERS["leave"]["schedule"])
    elif module == "overtime":
        export_path = path_or_empty(config, "overtime_src")
        if export_path:
            results["overtime_export"] = _check_headers(export_path, EXPECTED_HEADERS["overtime"]["export"])
    elif module == "subsidy":
        src_path = path_or_empty(config, "subsidy_src")
        if src_path:
            results["subsidy_source"] = _check_headers(src_path, EXPECTED_HEADERS["subsidy"]["source"])
    elif module == "final":
        roster_path = path_or_empty(config, "final_active")
        schedule_path = path_or_empty(config, "final_schedule")
        if roster_path:
            results["final_roster"] = _check_headers(roster_path, EXPECTED_HEADERS["final"]["roster"])
        if schedule_path:
            results["final_schedule"] = _check_headers(schedule_path, EXPECTED_HEADERS["final"]["schedule"])
    elif module == "parttime":
        detail_path = path_or_empty(config, "parttime_attendance_detail")
        if detail_path:
            results["parttime_detail"] = _check_headers(detail_path, EXPECTED_HEADERS["parttime"]["detail"])

    all_ok = all(r.get("ok", False) for r in results.values()) if results else True
    preview_path = output_dir / "validation_result.json"
    preview_path.write_text(json.dumps({"ok": all_ok, "results": results}, ensure_ascii=False, indent=2), encoding="utf-8")
    return [{"path": str(preview_path), "file_name": preview_path.name}]


# ── Action: preview ──────────────────────────────────────────────────────────
# 运行计算，但只返回前 200 行 JSON 预览

def _xlsx_to_preview_json(xlsx_path: str, max_rows: int = 200) -> list[dict]:
    """将 Excel 文件转换为前 N 行的 JSON 列表。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or f"col_{i}") for i, h in enumerate(next(rows_iter, []))]
    data = []
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            break
        data.append({headers[j]: str(cell) if cell is not None else "" for j, cell in enumerate(row)})
    wb.close()
    return data


def action_preview(config: dict, output_dir: Path) -> list[dict]:
    module = config.get("preview_module") or config.get("module", "")
    if not module or module not in RUNNERS:
        raise ValueError(f"未知模块：{module}")

    # 运行实际计算，生成输出文件
    outputs = RUNNERS[module](config, output_dir)

    # 读取第一个输出文件的前 200 行
    preview_data = []
    if outputs:
        first_output = outputs[0]["path"]
        if first_output.endswith(".xlsx"):
            preview_data = _xlsx_to_preview_json(first_output, max_rows=200)

    preview_path = output_dir / "preview.json"
    preview_path.write_text(
        json.dumps({"ok": True, "preview": preview_data, "total_outputs": len(outputs)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return [{"path": str(preview_path), "file_name": preview_path.name}] + outputs


def action_preview_existing(config: dict, output_dir: Path) -> list[dict]:
    """Preview first N rows of an existing xlsx without re-running calculation."""
    src = path_or_empty(config, "preview_file")
    if not src:
        raise ValueError("preview_file is required")
    max_rows = int(config.get("max_rows") or 200)
    if max_rows <= 0:
        max_rows = 200
    if max_rows > 500:
        max_rows = 500
    rows = _xlsx_to_preview_json(src, max_rows=max_rows)
    preview_path = output_dir / "preview.json"
    preview_path.write_text(
        json.dumps({"ok": True, "rows": rows, "preview": rows, "row_count": len(rows)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return [{"path": str(preview_path), "file_name": preview_path.name}]


# ── Action: export-templates ─────────────────────────────────────────────────
# 输出工具箱用到的 Excel 空白模板（单个或全部 zip）


def action_export_templates(config: dict, output_dir: Path) -> list[dict]:
    """Export one template or the full template set.

    config:
        template_id: optional TEMPLATE_REGISTRY id; if missing, returns all 16.
        templates_meta: optional bool (default true) — include template list.
    """
    requested = config.get("template_id")
    requested_str = str(requested).strip() if requested is not None else ""

    if requested_str:
        if requested_str not in templates.TEMPLATE_BY_ID:
            raise ValueError(f"未知模板：{requested_str}")
        filename, builder = templates.TEMPLATE_BY_ID[requested_str]
        data = builder()
        out_path = output_dir / filename
        out_path.write_bytes(data)
        outputs = [{"path": str(out_path), "file_name": filename}]
    else:
        outputs = []
        for template_id, filename, builder in templates.TEMPLATE_REGISTRY:
            data = builder()
            out_path = output_dir / filename
            out_path.write_bytes(data)
            outputs.append({"path": str(out_path), "file_name": filename})

    if str(config.get("templates_meta", "true")).lower() in {"true", "1", "yes"}:
        meta_path = output_dir / "templates_meta.json"
        meta_path.write_text(
            json.dumps(
                {
                    "templates": [
                        {"id": tid, "file_name": fname}
                        for tid, fname, _ in templates.TEMPLATE_REGISTRY
                    ],
                    "count": len(templates.TEMPLATE_REGISTRY),
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        outputs.append({"path": str(meta_path), "file_name": meta_path.name})

    return outputs


# ── Action: audit ─────────────────────────────────────────────────────────────
# 接收文件名列表和对应的字节数，返回观测警告（支持扫描内容）

MAX_AUDIT_BYTES = 1 * 1024 * 1024  # 内容扫描时最多读取的前 N 字节


def action_audit(config: dict, output_dir: Path) -> list[dict]:
    """Audit uploaded files before running computation.

    config shape (all fields optional):
        files: list of { "name": str, "size": int, "path": str (absolute, on disk) }
        max_warn_mb: int (default 50)
        max_warn_rows: int (default 50000)
        max_warn_cols: int (default 500)
        scan_content: bool (default true)
    """
    files_meta = config.get("files") or []
    if isinstance(files_meta, dict):
        files_meta = files_meta.get("items") or []
    if not isinstance(files_meta, list):
        raise ValueError("files 字段需要是 list")

    max_warn_mb = int(config.get("max_warn_mb", 50))
    max_warn_rows = int(config.get("max_warn_rows", 50000))
    max_warn_cols = int(config.get("max_warn_cols", 500))
    scan_content = str(config.get("scan_content", "true")).lower() not in {"0", "false", "no"}

    all_warnings: list[str] = []
    per_file: list[dict] = []
    audit_log: list[dict] = []

    for raw in files_meta:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or raw.get("file_name") or "uploaded file")
        size = int(raw.get("size") or raw.get("size_bytes") or 0)
        raw_path = str(raw.get("path") or "").strip()

        data = b""
        if raw_path and os.path.exists(raw_path):
            try:
                with open(raw_path, "rb") as fh:
                    data = fh.read(MAX_AUDIT_BYTES)
            except Exception:
                data = b""
        elif "bytes" in raw:
            supplied = raw["bytes"]
            if isinstance(supplied, (bytes, bytearray)):
                data = bytes(supplied[:MAX_AUDIT_BYTES])
            elif isinstance(supplied, str):
                try:
                    data = __import__("base64").b64decode(supplied)[:MAX_AUDIT_BYTES]
                except Exception:
                    data = b""

        result = audit_upload(
            name,
            size,
            data,
            max_warn_mb=max_warn_mb,
            max_warn_rows=max_warn_rows,
            max_warn_cols=max_warn_cols,
        )
        all_warnings.extend(result.get("warnings", []))
        audit_log.extend(result.get("audit", []))
        per_file.append({"file": result.get("file_name", name), "warnings": result.get("warnings", [])})

    out = {
        "ok": True,
        "warnings": all_warnings,
        "audit": audit_log,
        "files": per_file,
        "thresholds": {
            "max_warn_mb": max_warn_mb,
            "max_warn_rows": max_warn_rows,
            "max_warn_cols": max_warn_cols,
            "scan_content": scan_content,
        },
    }
    audit_path = output_dir / "audit_result.json"
    audit_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    return [{"path": str(audit_path), "file_name": audit_path.name}]


def action_parttime_monthly_punch(config: dict, output_dir: Path) -> list[dict]:
    """Render the part-time monthly punch grid + audit sheet (req: 兼职月度打卡记录)."""
    from parttime_monthly_punch import render as render_parttime_monthly_punch
    # render() writes to workdir/outputs/<file>; runner already passes that path
    # as output_dir, so pass it through directly.
    result = render_parttime_monthly_punch(output_dir, config)
    return [result]


# ── Action: generate-roster ─────────────────────────────────────────────────
# 由 Go 后端传入当前组织的花名册数据（JSON），生成标准在职花名册 xlsx。

_ROSTER_HEADERS = [
    "工号", "姓名", "合同主体", "一级部门", "二级部门", "三级部门",
    "岗位", "员工类型", "人员分类", "入职日期", "离职日期", "转正日期",
]


def _roster_str(val) -> str:
    if val is None:
        return ""
    value = str(val).strip()
    if value.endswith(".0"):
        value = value[:-2]
    return value


def _roster_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    value = str(val).strip()
    if not value:
        return ""
    for fmt in (
        "%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y.%m.%d",
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return value


def action_generate_roster(config: dict, output_dir: Path) -> list[dict]:
    """Generate an active-employee roster consumable by overtime and final modules.

    config:
        org_name: str（用于文件名/日志，不写入合同主体列）
        employees: list of {
            emp_no, name, contract_entity?, dept1?, dept2?, dept3?,
            position?, emp_type?, category?, hire_date?, resign_date?, confirm_date?
        }
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment

    org_name = str(config.get("org_name") or "").strip()
    raw_employees = config.get("employees", [])
    if raw_employees is None:
        raw_employees = []
    if not isinstance(raw_employees, list):
        raise ValueError("generate-roster: employees 字段需要是 list")

    normalized_employees: list[dict] = []
    missing_emp_no = 0
    missing_name = 0
    missing_dept_path = 0
    for idx, raw in enumerate(raw_employees, 1):
        if not isinstance(raw, dict):
            raise ValueError(f"generate-roster: employees[{idx}] 需要是 object")
        employee = dict(raw)
        employee["emp_no"] = _roster_str(raw.get("emp_no"))
        employee["name"] = _roster_str(raw.get("name"))
        employee["dept1"] = _roster_str(raw.get("dept1"))
        employee["dept2"] = _roster_str(raw.get("dept2"))
        employee["dept3"] = _roster_str(raw.get("dept3"))
        if not employee["emp_no"]:
            missing_emp_no += 1
        if not employee["name"]:
            missing_name += 1
        if not any((employee["dept1"], employee["dept2"], employee["dept3"])):
            missing_dept_path += 1
        normalized_employees.append(employee)
    if missing_emp_no:
        raise ValueError(
            f"generate-roster: {missing_emp_no} 名在职员工缺少业务工号（EmployeeID），已拒绝生成不完整花名册"
        )
    if missing_name:
        raise ValueError(
            f"generate-roster: {missing_name} 名在职员工缺少姓名，已拒绝生成不完整花名册"
        )
    if missing_dept_path:
        raise ValueError(
            f"generate-roster: {missing_dept_path} 名在职员工缺少有效部门路径，已拒绝生成不完整花名册"
        )
    if not normalized_employees:
        raise ValueError("generate-roster: 当前组织没有在职员工")

    safe_org = re.sub(r'[\\/:*?"<>|]', "_", org_name) or "组织"
    filename = f"花名册_{safe_org}.xlsx"
    out_path = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "在职花名册"

    for ci, header in enumerate(_ROSTER_HEADERS, 1):
        cell = ws.cell(1, ci, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    missing_contract = 0
    for raw in normalized_employees:
        contract_entity = _roster_str(raw.get("contract_entity"))
        if not contract_entity:
            missing_contract += 1
        ws.append([
            raw["emp_no"],
            raw["name"],
            contract_entity,
            raw["dept1"],
            raw["dept2"],
            raw["dept3"],
            _roster_str(raw.get("position")),
            _roster_str(raw.get("emp_type")),
            _roster_str(raw.get("category")),
            _roster_date(raw.get("hire_date")),
            _roster_date(raw.get("resign_date")),
            _roster_date(raw.get("confirm_date")),
        ])

    wb.save(str(out_path))
    wb.close()
    row_count = len(normalized_employees)
    print(f"[花名册] 已生成 {filename}: {row_count} 人，缺工号 0，缺部门路径 0，缺合同主体 {missing_contract}")
    return [{"path": str(out_path), "file_name": filename, "kind": "export", "row_count": row_count}]


ACTIONS = {
    "export-rules": action_export_rules,
    "import-rules-preview": action_import_rules_preview,
    "validate": action_validate,
    "preview": action_preview,
    "preview-existing": action_preview_existing,
    "audit": action_audit,
    "export-templates": action_export_templates,
    "parttime-monthly-punch": action_parttime_monthly_punch,
    "generate-roster": action_generate_roster,
}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--module", choices=sorted(RUNNERS))
    parser.add_argument("--action", choices=sorted(ACTIONS))
    parser.add_argument("--workdir")
    parser.add_argument("--defaults", action="store_true")
    parser.add_argument("--config-json", default="{}")
    args = parser.parse_args()

    if args.defaults:
        print_json({
            "ok": True,
            "defaults": default_text_values(),
        })
        return 0

    # Actions that don't require --module
    if args.action and args.action in ACTIONS:
        if not args.workdir:
            parser.error("--workdir is required for --action")
        workdir = Path(args.workdir).resolve()
        output_dir = workdir / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        try:
            config = json.loads(args.config_json)
            log_buffer = io.StringIO()
            with contextlib.redirect_stdout(log_buffer):
                outputs = ACTIONS[args.action](config, output_dir)
            print_json({
                "ok": True,
                "outputs": outputs,
                "log": log_buffer.getvalue(),
            })
            return 0
        except Exception as exc:
            print_json({
                "ok": False,
                "error": str(exc),
                "traceback": traceback.format_exc(),
            })
            return 1

    if not args.module or not args.workdir:
        parser.error("--module and --workdir are required unless --defaults or --action is used")

    workdir = Path(args.workdir).resolve()
    output_dir = workdir / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        config = json.loads(args.config_json)
        log_buffer = io.StringIO()
        with contextlib.redirect_stdout(log_buffer):
            outputs = RUNNERS[args.module](config, output_dir)
        print_json({
            "ok": True,
            "outputs": outputs,
            "log": log_buffer.getvalue(),
        })
        return 0
    except Exception as exc:
        print_json({
            "ok": False,
            "error": str(exc),
            "traceback": traceback.format_exc(),
        })
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
