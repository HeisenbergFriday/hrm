# -*- coding: utf-8 -*-
"""
Attendance toolbox equivalence / holiday / dingtalk tests.

Business coverage:
  [covered] leave/subsidy/final/parttime: isolated subprocess runs against
            D:\\app and toolbox code with the same synthetic Excel inputs.
  [covered] overtime: isolated subprocess output fingerprint parity.
  [covered] custom rules change overtime Config via rules_adapter.
  [covered] holiday year validation for 2003/2004/2099/2100.
  [covered] dingtalk_sync.sync_date_range with a mocked client.

Not claimed by this offline suite:
  - live DingTalk network behavior
  - equivalence for every production-specific workbook variation
"""
from __future__ import annotations

import hashlib
import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

APP_ROOT = Path(os.environ.get("ATTENDANCE_APP_SOURCE_ROOT", r"D:/app"))
TBX_ROOT = Path(__file__).resolve().parent

# Checked-in SHA256 (LF-normalized) for core modules — update only when intentional.
# Used by CI without D:\\app. Values filled at first successful local hash generation
# and must be regenerated via: python -c "..." when modules change.
CORE_HASH_FILE = TBX_ROOT / "CI_SOURCE_HASHES.json"


def _app_available() -> bool:
    return APP_ROOT.is_dir() and (APP_ROOT / "overtime" / "rules_engine.py").is_file()


def _lf_sha256(path: Path) -> str:
    raw = path.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
    return hashlib.sha256(raw).hexdigest()


def _workbook_fingerprint(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            for cell in row:
                fill_rgb = None
                if cell.fill and cell.fill.fgColor is not None:
                    try:
                        fill_rgb = str(cell.fill.fgColor.rgb)
                    except Exception:  # noqa: BLE001
                        fill_rgb = None
                cells.append(
                    {
                        "coord": cell.coordinate,
                        "value": cell.value,
                        "number_format": cell.number_format,
                        "font": {
                            "name": cell.font.name,
                            "size": cell.font.sz,
                            "bold": bool(cell.font.bold),
                            "italic": bool(cell.font.italic),
                            "underline": cell.font.underline,
                            "color": str(cell.font.color) if cell.font.color else None,
                        },
                        "fill": fill_rgb,
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text,
                            "text_rotation": cell.alignment.text_rotation,
                            "shrink_to_fit": cell.alignment.shrink_to_fit,
                        },
                        "border": {
                            "left": cell.border.left.style,
                            "right": cell.border.right.style,
                            "top": cell.border.top.style,
                            "bottom": cell.border.bottom.style,
                        },
                        "protection": {
                            "locked": cell.protection.locked,
                            "hidden": cell.protection.hidden,
                        },
                    }
                )
        sheets.append(
            {
                "name": name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged": [str(r) for r in ws.merged_cells.ranges],
                "sheet_state": ws.sheet_state,
                "freeze_panes": getattr(ws.freeze_panes, "coordinate", ws.freeze_panes),
                "auto_filter": ws.auto_filter.ref,
                "hidden_rows": sorted(idx for idx, dim in ws.row_dimensions.items() if dim.hidden),
                "hidden_columns": sorted(key for key, dim in ws.column_dimensions.items() if dim.hidden),
                "row_heights": {idx: dim.height for idx, dim in ws.row_dimensions.items() if dim.height is not None},
                "column_widths": {key: dim.width for key, dim in ws.column_dimensions.items() if dim.width is not None},
                "tab_color": str(ws.sheet_properties.tabColor) if ws.sheet_properties.tabColor else None,
                "cells": cells,
            }
        )
    wb.close()
    return {"sheet_names": [s["name"] for s in sheets], "sheets": sheets}


def _make_minimal_overtime_export(path: Path) -> None:
    """Minimal desensitized overtime export that process_overtime can load."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "加班"
    headers = [
        "发起人工号", "发起人姓名", "一级部门", "二级部门", "三级部门",
        "开始时间", "结束时间", "明细", "加班时间", "加班时长", "时长",
        "是否包含法定节假日期", "审批状态", "审批结果",
    ]
    ws.append(headers)
    ws.append([
        "E01", "测试甲", "研发中心", "平台组", "",
        "2026-03-02 18:30:00", "2026-03-02 21:00:00", "", "2026-03-02", "2.5", "2.5",
        "否", "完成", "同意",
    ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _make_minimal_calendar(path: Path) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "作息时间表"
    # Minimal schedule sheet — yellow workday cells (engine may soft-fail if unused)
    ws["A1"] = "周数"
    ws["B1"] = "日期"
    yellow = PatternFill("solid", fgColor="FFFF00")
    for i, day in enumerate(range(1, 8), start=2):
        ws.cell(row=i, column=1, value=1)
        cell = ws.cell(row=i, column=2, value=f"2026-03-0{day}" if day < 10 else f"2026-03-{day}")
        cell.fill = yellow
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()

def _make_minimal_schedule(path: Path) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "schedule"
    ws["A1"] = "2026\u5e743\u6708\u4f5c\u606f\u65f6\u95f4\u8868"
    ws["A2"] = "\u5468\u6570"
    yellow = PatternFill("solid", fgColor="FFFF00")
    for column, day in enumerate(range(2, 9), start=2):
        ws.cell(row=2, column=column, value=f"W{column - 1}")
        cell = ws.cell(row=3, column=column, value=day)
        cell.fill = yellow
    ws["A3"] = 1
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _make_minimal_leave_export(path: Path) -> None:
    import openpyxl
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "leave"
    ws.append([
        "\u53d1\u8d77\u4eba\u5de5\u53f7", "\u53d1\u8d77\u4eba\u59d3\u540d", "\u53d1\u8d77\u4eba\u90e8\u95e8",
        "\u8bf7\u5047\u7c7b\u578b", "\u5f00\u59cb\u65f6\u95f4", "\u7ed3\u675f\u65f6\u95f4", "\u65f6\u957f",
        "\u53d1\u8d77\u65f6\u95f4", "\u5b8c\u6210\u65f6\u95f4", "\u5ba1\u6279\u7f16\u53f7", "\u5ba1\u6279\u72b6\u6001", "\u5ba1\u6279\u7ed3\u679c",
    ])
    ws.append([
        "MT001", "\u6d4b\u8bd5\u7532", "\u7814\u53d1\u4e2d\u5fc3-\u5e73\u53f0\u7ec4", "\u4e8b\u5047",
        datetime(2026, 3, 2, 9, 0), datetime(2026, 3, 2, 18, 30), 8,
        datetime(2026, 3, 1, 10, 0), datetime(2026, 3, 1, 11, 0),
        "LEAVE-001", "\u5b8c\u6210", "\u540c\u610f",
    ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _make_minimal_subsidy_inputs(source_path: Path, checkin_path: Path) -> None:
    import openpyxl
    from datetime import date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "\u8003\u52e4\u6708\u5ea6\u6c47\u603b"
    headers = [
        "\u59d3\u540d", "\u5de5\u53f7", "\u4e00\u7ea7\u90e8\u95e8", "\u4e8c\u7ea7\u90e8\u95e8", "\u4e09\u7ea7\u90e8\u95e8", "\u804c\u4f4d",
        "15-30\u5206\u949f\u8fdf\u5230\u6263\u6b3e", "15-30\u5206\u949f\u65e9\u9000\u6263\u6b3e", "\u65f7\u5de5\u5929\u6570", "\u7f3a\u5361\u6b21\u6570",
        "\u665a\u4e8e22\u70b9\u6253\u5361\u5929\u6570", "\u665a\u8d70\u8865\u8d34", "\u4ea7\u7814\u4f11\u606f\u65e5\u52a0\u73ed>4\u5c0f\u65f6\u5929\u6570", "\u4ea7\u7814\u4f11\u606f\u65e5\u52a0\u73ed\u8865\u8d34",
        1, 2, 3, 4, 5,
    ]
    ws.append(headers)
    ws.append([
        "\u6d4b\u8bd5\u4e59", "MT002", "\u7814\u53d1\u4e2d\u5fc3", "\u5e73\u53f0\u7ec4", "", "\u5de5\u7a0b\u5e08",
        0, 0, 0, 0, 0, 0, 0, 0,
        "09:00 18:30", "09:00 18:30", "09:00 18:30", "09:00 18:30", "09:00 18:30",
    ])
    source_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(source_path)
    wb.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "\u7fbd\u6bdb\u7403"
    ws.append(["\u59d3\u540d", date(2026, 3, 2)])
    ws.append(["\u6d4b\u8bd5\u4e59", "\u672a\u53c2\u52a0"])
    wb.save(checkin_path)
    wb.close()


def _make_minimal_parttime_schedule(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "parttime"
    ws["A1"] = "2026\u5e743\u6708\u517c\u804c\u6392\u73ed"
    ws.append(["\u59d3\u540d", "\u516c\u53f8", 1, 2, 3, 4, 5])
    ws.append(["\u6d4b\u8bd5\u4e01", "\u6d4b\u8bd5\u516c\u53f8", "\u73ed", "\u73ed", "\u4f11", "\u73ed", "\u73ed"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _run_app_toolbox_pair(
    testcase: unittest.TestCase,
    root: Path,
    build_script,
    app_output: Path,
    toolbox_output: Path,
    toolbox_extra_sheets: tuple[str, ...] = (),
) -> None:
    app_script = root / "run_app.py"
    toolbox_script = root / "run_toolbox.py"
    app_script.write_text(build_script(APP_ROOT, app_output), encoding="utf-8")
    toolbox_script.write_text(build_script(TBX_ROOT, toolbox_output), encoding="utf-8")
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    app_run = subprocess.run([sys.executable, str(app_script)], capture_output=True, env=env)
    toolbox_run = subprocess.run([sys.executable, str(toolbox_script)], capture_output=True, env=env)

    def tail(raw: bytes | None) -> str:
        return (raw or b"").decode("utf-8", "replace")[-1200:]

    testcase.assertEqual(app_run.returncode, 0, f"D:/app run failed: {tail(app_run.stderr)}")
    testcase.assertEqual(toolbox_run.returncode, 0, f"toolbox run failed: {tail(toolbox_run.stderr)}")
    testcase.assertTrue(app_output.is_file(), f"missing D:/app output: {app_output}")
    testcase.assertTrue(toolbox_output.is_file(), f"missing toolbox output: {toolbox_output}")
    app_fingerprint = _workbook_fingerprint(app_output)
    toolbox_fingerprint = _workbook_fingerprint(toolbox_output)
    if toolbox_extra_sheets:
        for sheet_name in toolbox_extra_sheets:
            testcase.assertIn(sheet_name, toolbox_fingerprint["sheet_names"])
        toolbox_fingerprint["sheet_names"] = [
            name for name in toolbox_fingerprint["sheet_names"]
            if name not in toolbox_extra_sheets
        ]
        toolbox_fingerprint["sheets"] = [
            sheet for sheet in toolbox_fingerprint["sheets"]
            if sheet["name"] not in toolbox_extra_sheets
        ]
    testcase.assertEqual(app_fingerprint, toolbox_fingerprint)


class CIHashTests(unittest.TestCase):
    """CI-safe: hashes must match checked-in expected digests (not mere length)."""

    CORES = [
        "overtime/rules_engine.py",
        "leave/calc_leave.py",
        "subsidy/calc_subsidy_deduction.py",
        "parttime/calc_parttime_summary.py",
        "dingtalk_sync.py",
    ]

    def test_core_modules_match_checked_in_hashes_or_bootstrap(self) -> None:
        actual = {}
        for rel in self.CORES:
            path = TBX_ROOT / rel
            self.assertTrue(path.is_file(), rel)
            actual[rel] = _lf_sha256(path)

        self.assertTrue(
            CORE_HASH_FILE.is_file(),
            f"missing checked-in hash pins: {CORE_HASH_FILE}",
        )

        data = json.loads(CORE_HASH_FILE.read_text(encoding="utf-8"))
        expected = data.get("files") or {}
        for rel in self.CORES:
            self.assertIn(rel, expected, f"missing pin for {rel}")
            self.assertEqual(expected[rel], actual[rel], f"hash drift for {rel}")


class CustomRulesTests(unittest.TestCase):
    def test_custom_rules_change_overtime_config(self) -> None:
        if str(TBX_ROOT) not in sys.path:
            sys.path.insert(0, str(TBX_ROOT))
        if str(TBX_ROOT / "overtime") not in sys.path:
            sys.path.insert(0, str(TBX_ROOT / "overtime"))
        from rules_adapter import resolve_overtime_config  # type: ignore

        default_cfg, default_meta = resolve_overtime_config({})
        self.assertEqual(default_meta.get("source"), "default")
        custom = {
            "premium_rules": [
                {
                    "priority": 1,
                    "date_type": "LEGAL_HOLIDAY",
                    "department_group": "DEFAULT",
                    "action": "加班工资",
                    "multiplier": 9.9,
                }
            ],
            "department_rules": [],
            "params": {
                "standard_hours_per_day": 7.5,
                "no_punch_mark": "无",
                "schedule_augment_holidays": False,
                "schedule_augment_rest_dept_group": "",
                "chengdu_use_separate_calendar": False,
                "rest_premium_excluded_names": [],
                "rest_premium_excluded_codes": [],
            },
        }
        cfg, meta = resolve_overtime_config({"rules_json": json.dumps(custom, ensure_ascii=False)})
        self.assertNotEqual(meta.get("source"), "default")
        self.assertTrue(any(abs(getattr(r, "multiplier", 0) - 9.9) < 1e-9 for r in cfg.premium_rules))
        self.assertNotEqual(
            [getattr(r, "multiplier", 0) for r in default_cfg.premium_rules],
            [getattr(r, "multiplier", 0) for r in cfg.premium_rules],
        )


class HolidayBoundaryTests(unittest.TestCase):
    def setUp(self) -> None:
        if str(TBX_ROOT / "overtime") not in sys.path:
            sys.path.insert(0, str(TBX_ROOT / "overtime"))

    def test_validate_holiday_years_2003_2004_fail_or_raise(self) -> None:
        import rules_engine as re  # type: ignore

        with self.assertRaises(Exception):
            re.validate_holiday_years_available([2003])
        with self.assertRaises(Exception):
            re.validate_holiday_years_available([2004])

    def test_validate_holiday_years_current_range(self) -> None:
        import rules_engine as re  # type: ignore

        # 2026 should be within chinesecalendar coverage in this project.
        re.validate_holiday_years_available([2026])

    def test_validate_holiday_years_far_future_raises(self) -> None:
        import rules_engine as re  # type: ignore

        with self.assertRaises(Exception):
            re.validate_holiday_years_available([2099])
        with self.assertRaises(Exception):
            re.validate_holiday_years_available([2100])


class DingtalkMockTests(unittest.TestCase):
    def test_sync_date_range_mocked_client(self) -> None:
        if str(TBX_ROOT) not in sys.path:
            sys.path.insert(0, str(TBX_ROOT))
        import dingtalk_sync as ds  # type: ignore
        from datetime import date

        self.assertTrue(hasattr(ds, "sync_date_range"))
        self.assertTrue(hasattr(ds, "DingTalkClient"))

        # Build a minimal config object expected by sync_date_range.
        config = mock.MagicMock()
        config.process_codes = {"leave": "PROC_LEAVE"}
        config.client_id = "x"
        config.client_secret = "y"

        fetch_result = mock.MagicMock()
        fetch_result.instances = []
        fetch_result.truncated = False
        fetch_result.page_count = 1
        fetch_result.duplicate_instance_id_count = 0
        fetch_result.query_start = "2026-03-01"
        fetch_result.query_end = "2026-03-07"

        export = mock.MagicMock()
        export.row_count = 0
        export.skipped_count = 0
        export.file_name = "请假系统导出.xlsx"

        mock_client = mock.MagicMock()
        mock_client.fetch_process_instances_by_time.return_value = fetch_result
        mock_exporter = mock.MagicMock()
        mock_exporter.build_leave_export.return_value = export
        mock_exporter.build_overtime_export.return_value = export
        mock_exporter.build_attendance_correction_export.return_value = export
        mock_exporter.build_position_transfer_export.return_value = export

        with mock.patch.object(ds, "DingTalkClient", return_value=mock_client):
            with mock.patch.object(ds, "DingTalkExporter", return_value=mock_exporter):
                result = ds.sync_date_range(
                    config,
                    date(2026, 3, 1),
                    date(2026, 3, 7),
                    max_instances_per_flow=10,
                    flow_keys=["leave"],
                    query_window_padding_days=0,
                )

        self.assertIsNotNone(result)
        self.assertTrue(hasattr(result, "counts") or isinstance(result, dict))
        counts = result.counts if hasattr(result, "counts") else result.get("counts")
        self.assertIn("leave", counts)
        messages = result.messages if hasattr(result, "messages") else result.get("messages")
        self.assertTrue(any("leave" in str(m) for m in messages))
        mock_client.fetch_process_instances_by_time.assert_called()



class RemainingModuleAppParityTests(unittest.TestCase):
    """Synthetic end-to-end output parity for the remaining business modules."""

    def setUp(self) -> None:
        if not _app_available():
            self.skipTest("D:/app not available; local source parity only")

    def test_leave_output_fingerprint_vs_app(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            export = root / "leave_export.xlsx"
            schedule = root / "schedule.xlsx"
            _make_minimal_leave_export(export)
            _make_minimal_schedule(schedule)

            def build(source_root: Path, output: Path) -> str:
                return f'''
import sys
sys.path.insert(0, r"{source_root / "leave"}")
sys.path.insert(0, r"{source_root}")
import calc_leave
src_rows = calc_leave.clean_export(r"{export}")
schedule_ctx = calc_leave.load_schedule_context(r"{schedule}")
calc_leave.process(src_rows, r"{output}", schedule_ctx, None, None, (), ())
'''

            _run_app_toolbox_pair(
                self,
                root,
                build,
                root / "app_leave.xlsx",
                root / "toolbox_leave.xlsx",
            )

    def test_subsidy_output_fingerprint_vs_app(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "subsidy_source.xlsx"
            checkin = root / "checkin.xlsx"
            schedule = root / "schedule.xlsx"
            _make_minimal_subsidy_inputs(source, checkin)
            _make_minimal_schedule(schedule)

            def build(source_root: Path, output: Path) -> str:
                return f'''
import sys
sys.path.insert(0, r"{source_root / "subsidy"}")
sys.path.insert(0, r"{source_root}")
import calc_subsidy_deduction as sub
year, month, period_source = sub.resolve_schedule_period(r"{schedule}")
source_records = sub.parse_source_table(r"{source}")
activity_days = sub.parse_activity_checkin(r"{checkin}")
employees = sub.parse_attendance(r"{source}", year=year, month=month, period_source=period_source)
legal_holidays = sub.load_statutory_holidays_from_schedule(
    r"{schedule}", target_year=year, target_month=month
)
sub.write_output(
    r"{source}", source_records, employees, activity_days, r"{output}",
    legal_holidays=legal_holidays, late22_included_names=(),
)
'''

            _run_app_toolbox_pair(
                self,
                root,
                build,
                root / "app_subsidy.xlsx",
                root / "toolbox_subsidy.xlsx",
                toolbox_extra_sheets=("异常审计",),
            )

    def test_final_output_fingerprint_vs_app(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)

            def build(source_root: Path, output: Path) -> str:
                return f'''
import sys
from datetime import date
sys.path.insert(0, r"{source_root / "finally"}")
sys.path.insert(0, r"{source_root / "leave"}")
sys.path.insert(0, r"{source_root}")
import calc_finally as fin
working_days = {{date(2026, 3, day) for day in (2, 3, 4, 5, 6)}}
schedule_ctx = {{
    "year": 2026,
    "month": 3,
    "month_start": date(2026, 3, 1),
    "month_end": date(2026, 3, 31),
    "main_working_days": working_days,
    "chengdu_working_days": set(working_days),
    "main_attendance_days": len(working_days),
    "chengdu_attendance_days": len(working_days),
    "main_payable_days": set(working_days),
    "chengdu_payable_days": set(working_days),
    "main_statutory_holidays": set(),
    "chengdu_statutory_holidays": set(),
    "main_company_welfare_days": set(),
    "chengdu_company_welfare_days": set(),
}}
employees = [{{
    "emp_no": "MT003",
    "name": "\u6d4b\u8bd5\u4e19",
    "contract_entity": "\u6d4b\u8bd5\u4e3b\u4f53",
    "dept1": "\u7814\u53d1\u4e2d\u5fc3",
    "dept2": "\u5e73\u53f0\u7ec4",
    "dept3": "",
    "position": "\u5de5\u7a0b\u5e08",
    "emp_type": "\u6b63\u5f0f",
    "category": "\u5728\u804c",
    "hire_date": date(2025, 1, 1),
    "resign_date": None,
    "confirm_date": date(2025, 4, 1),
}}]
fin.generate(employees, {{}}, schedule_ctx, {{}}, {{}}, {{}}, {{}}, r"{output}", {{}}, (), {{}})
'''

            _run_app_toolbox_pair(
                self,
                root,
                build,
                root / "app_final.xlsx",
                root / "toolbox_final.xlsx",
            )

    def test_parttime_output_fingerprint_vs_app(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            schedule = root / "schedule.xlsx"
            matrix = root / "parttime_2026-03.xlsx"
            _make_minimal_schedule(schedule)
            _make_minimal_parttime_schedule(matrix)

            def build(source_root: Path, output: Path) -> str:
                return f'''
import sys
sys.path.insert(0, r"{source_root / "parttime"}")
sys.path.insert(0, r"{source_root / "leave"}")
sys.path.insert(0, r"{source_root}")
import calc_parttime_summary as part
part.generate_parttime_summary(
    output_path=r"{output}",
    schedule_paths=[r"{matrix}"],
    default_schedule_path=r"{schedule}",
    special_default_names=(),
)
'''

            _run_app_toolbox_pair(
                self,
                root,
                build,
                root / "app_parttime.xlsx",
                root / "toolbox_parttime.xlsx",
            )
class OvertimeAppParityTests(unittest.TestCase):
    """Real end-to-end overtime compare when D:\\app is present (isolated subprocess)."""

    def test_overtime_default_rules_fingerprint_vs_app(self) -> None:
        if not _app_available():
            self.skipTest("D:/app not available; local source parity only")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            export = root / "export.xlsx"
            calendar = root / "calendar.xlsx"
            app_out = root / "app_out"
            tbx_out = root / "tbx_out"
            app_out.mkdir()
            tbx_out.mkdir()
            _make_minimal_overtime_export(export)
            _make_minimal_calendar(calendar)

            # Isolated PYTHONPATH so app/toolbox modules do not pollute each other.
            app_script = root / "run_app.py"
            tbx_script = root / "run_tbx.py"
            # process_overtime expects cleaned in-memory rows, not file paths.
            # Parity entrypoint for file-based runs is clean_export_overtime + process_overtime.
            runner_body = r'''
import sys
sys.path.insert(0, r"{ot_dir}")
sys.path.insert(0, r"{leave_dir}")
sys.path.insert(0, r"{root_dir}")
import fill_overtime_fields as ot
from rules_engine import get_default_config
export = r"{export}"
out = r"{out}"
calendar = r"{calendar}"
src_rows = ot.clean_export_overtime(export)
# optional calendar load for month inference
schedule_map = {{}}
if hasattr(ot, "load_work_calendar"):
    try:
        ot.load_work_calendar(calendar, get_default_config())
    except Exception as exc:
        print("calendar_load_soft_fail", exc)
ot.process_overtime(src_rows, out, schedule_map, {{}}, rules_config=get_default_config())
print("ok")
'''
            app_script.write_text(
                runner_body.format(
                    ot_dir=str(APP_ROOT / "overtime"),
                    leave_dir=str(APP_ROOT / "leave"),
                    root_dir=str(APP_ROOT),
                    export=str(export),
                    out=str(app_out / "加班明细_回填.xlsx"),
                    calendar=str(calendar),
                ),
                encoding="utf-8",
            )
            tbx_script.write_text(
                runner_body.format(
                    ot_dir=str(TBX_ROOT / "overtime"),
                    leave_dir=str(TBX_ROOT / "leave"),
                    root_dir=str(TBX_ROOT),
                    export=str(export),
                    out=str(tbx_out / "加班明细_回填.xlsx"),
                    calendar=str(calendar),
                ),
                encoding="utf-8",
            )

            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            r1 = subprocess.run([sys.executable, str(app_script)], capture_output=True, env=env)
            r2 = subprocess.run([sys.executable, str(tbx_script)], capture_output=True, env=env)
            def _tail(raw: bytes | None) -> str:
                if not raw:
                    return ""
                return raw.decode("utf-8", "replace")[-400:]

            self.assertEqual(
                r1.returncode,
                0,
                f"D:/app overtime run failed: {_tail(r1.stderr)}",
            )
            self.assertEqual(
                r2.returncode,
                0,
                f"toolbox overtime run failed: {_tail(r2.stderr)}",
            )

            app_files = list(app_out.glob("*.xlsx"))
            tbx_files = list(tbx_out.glob("*.xlsx"))
            self.assertTrue(app_files and tbx_files, "missing outputs")
            fp_app = _workbook_fingerprint(app_files[0])
            fp_tbx = _workbook_fingerprint(tbx_files[0])
            self.assertEqual(fp_app["sheet_names"], fp_tbx["sheet_names"])
            self.assertEqual(fp_app, fp_tbx)


if __name__ == "__main__":
    unittest.main()
