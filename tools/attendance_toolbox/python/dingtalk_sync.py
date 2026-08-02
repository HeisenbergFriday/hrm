from __future__ import annotations

import io
import json
import re
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_CN_TZ = timezone(timedelta(hours=8))
_HEADER_FILL = PatternFill("solid", fgColor="FFE2E8F0")
_OAPI_RATE_LIMIT_PER_SECOND = 20
_OAPI_MAX_RETRIES = 8
_OAPI_PROCESS_INSTANCE_QUERY_MAX_DAYS = 120
_OAPI_QUERY_CLOCK_SKEW_SECONDS = 60
_DINGTALK_API_HOSTS = {"api.dingtalk.com", "oapi.dingtalk.com"}
_RATE_LIMIT_END_RE = re.compile(r"限制将在\s*(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\s*结束")
_FLOW_LABELS = {
    "leave": "请假",
    "overtime": "加班",
    "attendance_correction": "补卡",
    "position_transfer": "岗位异动",
}
_INSTANCE_AUDIT_HEADERS = [
    "process_instance_id",
    "审批编号",
    "发起人userid",
    "发起人部门",
    "审批状态(原始)",
    "审批状态(映射)",
    "审批结果(原始)",
    "审批结果(映射)",
    "发起时间",
    "完成时间",
    "写出行数",
    "写入状态",
    "备注",
]


@dataclass
class DingTalkConfig:
    client_id: str
    client_secret: str
    process_codes: dict[str, str]


@dataclass
class DingTalkExport:
    file_name: str
    data: bytes
    row_count: int
    source_count: int = 0
    skipped_count: int = 0
    audit_rows: list[list[Any]] | None = None
    mapping_rows: list[list[Any]] | None = None


@dataclass
class DingTalkFetchResult:
    process_code: str
    instances: list[dict[str, Any]]
    instance_ids: list[str]
    raw_instance_id_count: int
    duplicate_instance_id_count: int
    page_count: int
    query_start: str
    query_end: str
    max_instances: int | None
    truncated: bool
    repeated_cursor_stopped: bool
    detail_count: int


@dataclass
class DingTalkSyncResult:
    exports: dict[str, DingTalkExport]
    counts: dict[str, int]
    messages: list[str]
    audit_export: DingTalkExport | None = None


class DingTalkApiError(RuntimeError):
    pass


class _SlidingWindowRateLimiter:
    def __init__(self, max_calls: int, period_seconds: float) -> None:
        self.max_calls = max_calls
        self.period_seconds = period_seconds
        self._calls: deque[float] = deque()
        self._lock = threading.Lock()

    def wait(self) -> None:
        while True:
            with self._lock:
                now = time.monotonic()
                while self._calls and now - self._calls[0] >= self.period_seconds:
                    self._calls.popleft()
                if len(self._calls) < self.max_calls:
                    self._calls.append(now)
                    return
                sleep_seconds = self.period_seconds - (now - self._calls[0]) + 0.05
            time.sleep(max(sleep_seconds, 0.05))


def config_from_mapping(raw: Any) -> DingTalkConfig:
    section = _mapping_to_plain_dict(raw)
    process_codes = _mapping_to_plain_dict(section.get("process_codes", {}))
    client_id = str(section.get("client_id") or section.get("app_key") or "").strip()
    client_secret = str(section.get("client_secret") or section.get("app_secret") or "").strip()
    if not client_id or not client_secret:
        raise ValueError("钉钉配置缺少 client_id/client_secret")
    return DingTalkConfig(
        client_id=client_id,
        client_secret=client_secret,
        process_codes={str(k): str(v) for k, v in process_codes.items() if v},
    )


def sync_month(
    config: DingTalkConfig,
    year: int,
    month: int,
    max_instances_per_flow: int | None = None,
    flow_keys: list[str] | tuple[str, ...] | None = None,
    query_window_padding_months: int = 1,
) -> DingTalkSyncResult:
    month_start = date(year, month, 1)
    next_year, next_month = _add_months(year, month, 1)
    month_end = date(next_year, next_month, 1) - timedelta(days=1)
    return sync_date_range(
        config,
        month_start,
        month_end,
        max_instances_per_flow=max_instances_per_flow,
        flow_keys=flow_keys,
        query_window_padding_days=max(query_window_padding_months, 0) * 31,
    )


def sync_date_range(
    config: DingTalkConfig,
    start_date: date,
    end_date: date,
    max_instances_per_flow: int | None = None,
    flow_keys: list[str] | tuple[str, ...] | None = None,
    query_window_padding_days: int = 31,
) -> DingTalkSyncResult:
    if start_date > end_date:
        raise ValueError("开始日期不能晚于结束日期")

    client = DingTalkClient(config)
    exporter = DingTalkExporter(client)
    exports: dict[str, DingTalkExport] = {}
    counts: dict[str, int] = {}
    messages: list[str] = []
    fetch_results: dict[str, DingTalkFetchResult] = {}

    process_jobs = [
        ("leave", "请假系统导出.xlsx", exporter.build_leave_export),
        ("overtime", "加班系统导出.xlsx", exporter.build_overtime_export),
        ("attendance_correction", "补卡审批摘要.xlsx", exporter.build_attendance_correction_export),
        ("position_transfer", "岗位异动流程表.xlsx", exporter.build_position_transfer_export),
    ]
    selected_flow_keys = set(flow_keys or [job[0] for job in process_jobs])

    for code_key, file_name, build_func in process_jobs:
        if code_key not in selected_flow_keys:
            continue
        process_code = config.process_codes.get(code_key)
        if not process_code:
            messages.append(f"[{code_key}] 未配置流程码，已跳过")
            continue
        query_start, query_end_exclusive = _date_datetime_range(
            start_date,
            end_date,
            query_window_padding_days,
        )
        fetch_result = client.fetch_process_instances_by_time(
            process_code,
            query_start,
            query_end_exclusive,
            max_instances=max_instances_per_flow,
        )
        export = build_func(
            fetch_result.instances,
            file_name,
            filter_start=start_date,
            filter_end=end_date,
        )
        fetch_results[code_key] = fetch_result
        exports[code_key] = export
        counts[code_key] = export.row_count
        limit_text = "已按上限截断" if fetch_result.truncated else "未截断"
        messages.append(
            f"[{code_key}] 业务范围 {start_date} 至 {end_date}，"
            f"查询 {fetch_result.query_start} 至 {fetch_result.query_end}，"
            f"分页 {fetch_result.page_count} 页，拉取 {len(fetch_result.instances)} 条审批，"
            f"去重 {fetch_result.duplicate_instance_id_count} 条，"
            f"写出 {export.row_count} 行，跳过 {export.skipped_count} 条，{limit_text}"
        )

    range_label = f"{start_date}_{end_date}"
    audit_export = _build_sync_audit_report(range_label, fetch_results, exports) if fetch_results else None
    return DingTalkSyncResult(
        exports=exports,
        counts=counts,
        messages=messages,
        audit_export=audit_export,
    )


class DingTalkClient:
    def __init__(self, config: DingTalkConfig) -> None:
        self.config = config
        self._access_token = ""
        self._access_token_expire_at = 0.0
        self._user_cache: dict[str, dict[str, Any]] = {}
        self._user_cache_lock = threading.Lock()
        self._oapi_rate_limiter = _SlidingWindowRateLimiter(_OAPI_RATE_LIMIT_PER_SECOND, 1.0)

    def fetch_process_instances(
        self,
        process_code: str,
        year: int,
        month: int,
        max_instances: int | None = None,
        query_window_padding_months: int = 0,
    ) -> DingTalkFetchResult:
        start_dt, end_dt_exclusive = _month_datetime_range(year, month, query_window_padding_months)
        return self.fetch_process_instances_by_time(
            process_code,
            start_dt,
            end_dt_exclusive,
            max_instances=max_instances,
        )

    def fetch_process_instances_by_time(
        self,
        process_code: str,
        query_start: datetime,
        query_end_exclusive: datetime,
        max_instances: int | None = None,
    ) -> DingTalkFetchResult:
        query_windows = _process_instance_query_windows(query_start, query_end_exclusive)
        ids: list[str] = []
        seen_instance_ids: set[str] = set()
        raw_instance_id_count = 0
        duplicate_instance_id_count = 0
        page_count = 0
        truncated = False
        repeated_cursor_stopped = False
        stop_fetching = False

        for window_index, (window_start, window_end_exclusive) in enumerate(query_windows):
            cursor: int | str | None = 0
            seen_cursors: set[str] = set()
            start_ms = int(window_start.timestamp() * 1000)
            end_ms = int(window_end_exclusive.timestamp() * 1000) - 1

            while True:
                payload = {
                    "process_code": process_code,
                    "start_time": start_ms,
                    "end_time": end_ms,
                    "size": 20,
                    "cursor": cursor or 0,
                }
                data = self.oapi_post("topapi/processinstance/listids", payload)
                page_count += 1
                result = data.get("result") or {}
                page_ids = result.get("list") or result.get("process_instance_ids") or []
                for item in page_ids:
                    if not item:
                        continue
                    raw_instance_id_count += 1
                    instance_id = str(item)
                    if instance_id in seen_instance_ids:
                        duplicate_instance_id_count += 1
                        continue
                    seen_instance_ids.add(instance_id)
                    ids.append(instance_id)
                next_cursor = result.get("next_cursor")
                has_next_page = bool(next_cursor and len(page_ids) >= 20)
                if max_instances is not None and len(ids) >= max_instances:
                    has_more_windows = window_index < len(query_windows) - 1
                    truncated = len(ids) > max_instances or has_next_page or has_more_windows
                    ids = ids[:max_instances]
                    stop_fetching = True
                    break
                if not next_cursor or len(page_ids) < 20:
                    break
                cursor_key = str(next_cursor)
                if cursor_key in seen_cursors:
                    repeated_cursor_stopped = True
                    break
                seen_cursors.add(cursor_key)
                cursor = next_cursor
            if stop_fetching:
                break

        def load_detail(instance_id: str) -> dict[str, Any]:
            detail = self.oapi_post(
                "topapi/processinstance/get",
                {"process_instance_id": instance_id},
            )
            instance = detail.get("process_instance") or detail.get("result") or {}
            if instance:
                instance["_sync_process_instance_id"] = instance_id
                instance["_sync_process_code"] = process_code
            return instance

        instances: list[dict[str, Any]] = []
        if ids:
            max_workers = min(6, max(1, len(ids)))
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_map = {
                    executor.submit(load_detail, instance_id): instance_id
                    for instance_id in ids
                }
                by_id: dict[str, dict[str, Any]] = {}
                for future in as_completed(future_map):
                    instance_id = future_map[future]
                    instance = future.result()
                    if instance:
                        by_id[instance_id] = instance
                instances = [by_id[instance_id] for instance_id in ids if instance_id in by_id]
        effective_query_start = query_windows[0][0] if query_windows else query_start
        effective_query_end_exclusive = query_windows[-1][1] if query_windows else query_start
        return DingTalkFetchResult(
            process_code=process_code,
            instances=instances,
            instance_ids=ids,
            raw_instance_id_count=raw_instance_id_count,
            duplicate_instance_id_count=duplicate_instance_id_count,
            page_count=page_count,
            query_start=effective_query_start.strftime("%Y-%m-%d %H:%M:%S"),
            query_end=(effective_query_end_exclusive - timedelta(milliseconds=1)).strftime("%Y-%m-%d %H:%M:%S"),
            max_instances=max_instances,
            truncated=truncated,
            repeated_cursor_stopped=repeated_cursor_stopped,
            detail_count=len(instances),
        )

    def get_user(self, userid: str | None) -> dict[str, Any]:
        userid = str(userid or "").strip()
        if not userid:
            return {}
        with self._user_cache_lock:
            if userid in self._user_cache:
                return self._user_cache[userid]
        token = self.access_token()
        data = self._post_oapi_json(
            "topapi/v2/user/get",
            f"https://oapi.dingtalk.com/topapi/v2/user/get?access_token={token}",
            {"userid": userid},
        )
        errcode = data.get("errcode")
        if errcode == 60121:
            with self._user_cache_lock:
                self._user_cache[userid] = {}
            return {}
        if errcode not in (None, 0):
            raise DingTalkApiError(f"topapi/v2/user/get 调用失败：{errcode} {data.get('errmsg')}")
        result = data.get("result") or {}
        with self._user_cache_lock:
            self._user_cache[userid] = result
        return result

    def prefetch_users(self, userids: list[str]) -> None:
        unique_userids = []
        seen = set()
        for userid in userids:
            userid = str(userid or "").strip()
            if not userid or userid in seen:
                continue
            seen.add(userid)
            with self._user_cache_lock:
                cached = userid in self._user_cache
            if not cached:
                unique_userids.append(userid)
        if not unique_userids:
            return
        max_workers = min(6, len(unique_userids))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(self.get_user, userid) for userid in unique_userids]
            for future in as_completed(futures):
                future.result()

    def oapi_post(self, path: str, payload: dict[str, Any]) -> dict[str, Any]:
        token = self.access_token()
        url = f"https://oapi.dingtalk.com/{path}?access_token={token}"
        data = self._post_oapi_json(path, url, payload)
        errcode = data.get("errcode")
        if errcode not in (None, 0):
            raise DingTalkApiError(f"{path} 调用失败：{errcode} {data.get('errmsg')}")
        return data

    def _post_oapi_json(self, path: str, url: str, payload: dict[str, Any]) -> dict[str, Any]:
        last_error: str | None = None
        for attempt in range(_OAPI_MAX_RETRIES):
            self._oapi_rate_limiter.wait()
            data = self._post_json(url, payload)
            errcode = data.get("errcode")
            errmsg = str(data.get("errmsg") or "")
            if errcode not in (None, 0) and _is_dingtalk_rate_limited(data):
                last_error = f"{errcode} {errmsg}"
                time.sleep(_rate_limit_sleep_seconds(errmsg, attempt))
                continue
            return data
        raise DingTalkApiError(f"{path} 调用失败：钉钉接口持续限流，已重试 {_OAPI_MAX_RETRIES} 次。最后错误：{last_error}")

    def access_token(self) -> str:
        if self._access_token and time.time() < self._access_token_expire_at:
            return self._access_token
        data = self._post_json(
            "https://api.dingtalk.com/v1.0/oauth2/accessToken",
            {
                "appKey": self.config.client_id,
                "appSecret": self.config.client_secret,
            },
        )
        token = data.get("accessToken") or data.get("access_token")
        if not token:
            raise DingTalkApiError("钉钉 access_token 返回为空")
        expire_in = int(data.get("expireIn") or data.get("expires_in") or 7200)
        self._access_token = str(token)
        self._access_token_expire_at = time.time() + max(expire_in - 120, 60)
        return self._access_token

    @staticmethod
    def _post_json(url: str, payload: dict[str, Any]) -> dict[str, Any]:
        parsed = urllib.parse.urlsplit(url)
        if parsed.scheme != "https" or parsed.hostname not in _DINGTALK_API_HOSTS:
            raise DingTalkApiError("DingTalk API URL must use a trusted HTTPS host")
        request = urllib.request.Request(
            url,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=30) as response:  # nosec B310
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise DingTalkApiError(f"HTTP {exc.code}: {body[:500]}") from exc


class DingTalkExporter:
    def __init__(self, client: DingTalkClient) -> None:
        self.client = client

    def build_leave_export(
        self,
        instances: list[dict[str, Any]],
        file_name: str,
        filter_start: date | None = None,
        filter_end: date | None = None,
    ) -> DingTalkExport:
        headers = [
            "审批状态", "审批结果", "发起人工号", "发起人姓名", "发起人部门",
            "请假类型", "开始时间", "结束时间", "时长", "审批编号",
            "发起时间", "完成时间", "岗位名称",
        ]
        rows: list[list[Any]] = []
        write_counts: dict[str, int] = {}
        skip_reasons: dict[str, str] = {}
        self.client.prefetch_users([str(instance.get("originator_userid") or "") for instance in instances])
        for instance in instances:
            key = _instance_key(instance)
            write_counts.setdefault(key, 0)
            holiday = _find_form_value(instance, component_type="DDHolidayField")
            parsed = _loads_maybe_json(holiday.get("value")) if holiday else None
            if not isinstance(parsed, list) or len(parsed) < 5:
                skip_reasons[key] = "未识别到请假控件 DDHolidayField 或字段数量不足"
                continue
            if not _date_interval_overlaps(parsed[0], parsed[1], filter_start, filter_end):
                skip_reasons[key] = "请假时间不在筛选日期区间"
                continue
            user = self.client.get_user(instance.get("originator_userid"))
            rows.append([
                _approval_status(instance),
                _approval_result(instance),
                _job_number(user),
                user.get("name") or _name_from_title(instance.get("title")),
                instance.get("originator_dept_name") or "",
                parsed[4],
                parsed[0],
                parsed[1],
                _duration_text(parsed[2], parsed[3] if len(parsed) > 3 else ""),
                instance.get("business_id"),
                instance.get("create_time"),
                instance.get("finish_time"),
                user.get("title") or "",
            ])
            write_counts[key] += 1
        mapping_rows = [
            ["审批状态", "process_instance.status", "转换为当前模板中文状态"],
            ["审批结果", "process_instance.result", "agree/refuse 转换为同意/拒绝"],
            ["发起人工号", "topapi/v2/user/get.job_number", "按 originator_userid 查询员工详情"],
            ["发起人姓名", "员工详情 name / 审批标题兜底", "员工详情缺失时从标题截取"],
            ["发起人部门", "originator_dept_name", "钉钉审批实例发起部门"],
            ["请假类型", "DDHolidayField[4]", "请假控件解析"],
            ["开始时间", "DDHolidayField[0]", "请假控件解析"],
            ["结束时间", "DDHolidayField[1]", "请假控件解析"],
            ["时长", "DDHolidayField[2]+[3]", "数值和单位合并"],
            ["审批编号", "business_id", "钉钉审批编号"],
            ["发起时间", "create_time", "钉钉审批实例字段"],
            ["完成时间", "finish_time", "钉钉审批实例字段"],
            ["岗位名称", "员工详情 title", "按 originator_userid 查询员工详情"],
        ]
        return _build_export(
            file_name,
            "请假系统导出",
            headers,
            rows,
            instances,
            write_counts,
            skip_reasons,
            mapping_rows,
        )

    def build_overtime_export(
        self,
        instances: list[dict[str, Any]],
        file_name: str,
        filter_start: date | None = None,
        filter_end: date | None = None,
    ) -> DingTalkExport:
        headers = [
            "审批状态", "审批结果", "发起人工号", "发起人姓名", "发起人部门",
            "开始时间", "结束时间", "时长", "审批编号",
            "明细", "加班时间", "加班时长",
            "是否包含法定节假日期", "2026法定节假日如下：",
        ]
        rows: list[list[Any]] = []
        write_counts: dict[str, int] = {}
        skip_reasons: dict[str, str] = {}
        self.client.prefetch_users([str(instance.get("originator_userid") or "") for instance in instances])
        for instance in instances:
            key = _instance_key(instance)
            write_counts.setdefault(key, 0)
            user = self.client.get_user(instance.get("originator_userid"))
            form = _form_values_by_name(instance)
            suite = _find_biz_suite(instance, aliases={"startTime", "finishTime", "duration"})
            by_alias = _suite_components_by_alias(suite)
            start_time = _component_value(by_alias.get("startTime"))
            end_time = _component_value(by_alias.get("finishTime"))
            total_duration = _component_value(by_alias.get("duration"))
            detail_rows = _table_rows(by_alias.get("everyDayDuration"))
            if not detail_rows:
                detail_rows = [{}]
            wrote_instance = False
            for detail in detail_rows:
                overtime_date = detail.get("overtimeDate") or detail.get("加班时间") or ""
                if overtime_date:
                    if not _date_in_range(overtime_date, filter_start, filter_end):
                        continue
                elif not _date_interval_overlaps(start_time, end_time, filter_start, filter_end):
                    continue
                overtime_duration = detail.get("overtimeDuration") or detail.get("加班时长") or ""
                rows.append([
                    _approval_status(instance),
                    _approval_result(instance),
                    _job_number(user),
                    user.get("name") or _name_from_title(instance.get("title")),
                    instance.get("originator_dept_name") or "",
                    start_time,
                    end_time,
                    total_duration,
                    instance.get("business_id"),
                    overtime_date,
                    overtime_date,
                    overtime_duration,
                    form.get("是否包含法定节假日期") or "",
                    _join_multi_value(form.get("2026法定节假日如下：")),
                ])
                write_counts[key] += 1
                wrote_instance = True
            if not wrote_instance:
                skip_reasons[key] = "加班时间不在筛选日期区间"
        mapping_rows = [
            ["审批状态", "process_instance.status", "转换为当前模板中文状态"],
            ["审批结果", "process_instance.result", "agree/refuse 转换为同意/拒绝"],
            ["发起人工号", "topapi/v2/user/get.job_number", "按 originator_userid 查询员工详情"],
            ["发起人姓名", "员工详情 name / 审批标题兜底", "员工详情缺失时从标题截取"],
            ["发起人部门", "originator_dept_name", "钉钉审批实例发起部门"],
            ["开始时间", "DDBizSuite bizAlias=startTime", "加班套件字段"],
            ["结束时间", "DDBizSuite bizAlias=finishTime", "加班套件字段"],
            ["时长", "DDBizSuite bizAlias=duration", "加班套件字段"],
            ["审批编号", "business_id", "钉钉审批编号"],
            ["明细", "DDBizSuite everyDayDuration.overtimeDate", "当前模板要求，与加班时间保持一致"],
            ["加班时间", "DDBizSuite everyDayDuration.overtimeDate", "每日加班明细行"],
            ["加班时长", "DDBizSuite everyDayDuration.overtimeDuration", "每日加班明细行"],
            ["是否包含法定节假日期", "表单字段：是否包含法定节假日期", "按字段名读取"],
            ["2026法定节假日如下：", "表单字段：2026法定节假日如下：", "多选值用分号拼接"],
        ]
        return _build_export(
            file_name,
            "加班系统导出",
            headers,
            rows,
            instances,
            write_counts,
            skip_reasons,
            mapping_rows,
        )

    def build_attendance_correction_export(
        self,
        instances: list[dict[str, Any]],
        file_name: str,
        filter_start: date | None = None,
        filter_end: date | None = None,
    ) -> DingTalkExport:
        headers = [
            "审批状态", "审批结果", "发起人工号", "发起人姓名", "发起人部门",
            "补卡时间", "补卡理由", "缺卡证明", "证明人",
            "审批编号", "发起时间", "完成时间",
        ]
        rows: list[list[Any]] = []
        write_counts: dict[str, int] = {}
        skip_reasons: dict[str, str] = {}
        self.client.prefetch_users([str(instance.get("originator_userid") or "") for instance in instances])
        for instance in instances:
            key = _instance_key(instance)
            write_counts.setdefault(key, 0)
            user = self.client.get_user(instance.get("originator_userid"))
            form = _form_values_by_name(instance)
            suite = _find_biz_suite(instance, aliases={"userCheckTime"})
            by_alias = _suite_components_by_alias(suite)
            repair_time = _component_value(by_alias.get("userCheckTime")) or _timestamp_to_text(form.get("repairCheckTime"))
            if not _date_in_range(repair_time, filter_start, filter_end):
                skip_reasons[key] = "补卡时间不在筛选日期区间"
                continue
            rows.append([
                _approval_status(instance),
                _approval_result(instance),
                _job_number(user),
                user.get("name") or _name_from_title(instance.get("title")),
                instance.get("originator_dept_name") or "",
                repair_time,
                form.get("补卡理由") or "",
                form.get("缺卡证明") or "",
                form.get("证明人") or "",
                instance.get("business_id"),
                instance.get("create_time"),
                instance.get("finish_time"),
            ])
            write_counts[key] += 1
        mapping_rows = [
            ["审批状态", "process_instance.status", "转换为当前模板中文状态"],
            ["审批结果", "process_instance.result", "agree/refuse 转换为同意/拒绝"],
            ["发起人工号", "topapi/v2/user/get.job_number", "按 originator_userid 查询员工详情"],
            ["发起人姓名", "员工详情 name / 审批标题兜底", "员工详情缺失时从标题截取"],
            ["发起人部门", "originator_dept_name", "钉钉审批实例发起部门"],
            ["补卡时间", "DDBizSuite bizAlias=userCheckTime / repairCheckTime", "优先读取补卡套件字段，失败时按时间戳兜底"],
            ["补卡理由", "表单字段：补卡理由", "按字段名读取"],
            ["缺卡证明", "表单字段：缺卡证明", "按字段名读取"],
            ["证明人", "表单字段：证明人", "按字段名读取"],
            ["审批编号", "business_id", "钉钉审批编号"],
            ["发起时间", "create_time", "钉钉审批实例字段"],
            ["完成时间", "finish_time", "钉钉审批实例字段"],
        ]
        return _build_export(
            file_name,
            "补卡审批摘要",
            headers,
            rows,
            instances,
            write_counts,
            skip_reasons,
            mapping_rows,
        )

    def build_position_transfer_export(
        self,
        instances: list[dict[str, Any]],
        file_name: str,
        filter_start: date | None = None,
        filter_end: date | None = None,
    ) -> DingTalkExport:
        headers = [
            "实际申请人工号", "实际申请人", "发起人工号", "发起人姓名",
            "异动日期", "异动类型", "生效日期",
        ]
        rows: list[list[Any]] = []
        write_counts: dict[str, int] = {}
        skip_reasons: dict[str, str] = {}
        userids = [str(instance.get("originator_userid") or "") for instance in instances]
        for instance in instances:
            suite = _find_biz_suite(instance, aliases={"actualApplicant", "effectiveDay"})
            by_alias = _suite_components_by_alias(suite)
            _actual_name, actual_userid = _contact_value(by_alias.get("actualApplicant"))
            if actual_userid:
                userids.append(actual_userid)
        self.client.prefetch_users(userids)
        for instance in instances:
            launcher = self.client.get_user(instance.get("originator_userid"))
            suite = _find_biz_suite(instance, aliases={"actualApplicant", "effectiveDay"})
            by_alias = _suite_components_by_alias(suite)
            actual_name, actual_userid = _contact_value(by_alias.get("actualApplicant"))
            actual_user = self.client.get_user(actual_userid)
            effective_day = _component_value(by_alias.get("effectiveDay"))
            if not _date_in_range(effective_day, filter_start, filter_end):
                skip_reasons[_instance_key(instance)] = "异动生效日期不在筛选日期区间"
                continue
            rows.append([
                _job_number(actual_user),
                actual_name,
                _job_number(launcher),
                launcher.get("name") or _name_from_title(instance.get("title")),
                effective_day,
                "岗位异动",
                effective_day,
            ])
            write_counts[_instance_key(instance)] = write_counts.get(_instance_key(instance), 0) + 1
        mapping_rows = [
            ["实际申请人工号", "DDBizSuite actualApplicant -> 员工详情 job_number", "按实际申请人 userid 查询"],
            ["实际申请人", "DDBizSuite bizAlias=actualApplicant", "联系人控件名称"],
            ["发起人工号", "topapi/v2/user/get.job_number", "按 originator_userid 查询员工详情"],
            ["发起人姓名", "员工详情 name / 审批标题兜底", "员工详情缺失时从标题截取"],
            ["异动日期", "DDBizSuite bizAlias=effectiveDay", "岗位异动生效日期"],
            ["异动类型", "固定值：岗位异动", "当前流程只有岗位异动"],
            ["生效日期", "DDBizSuite bizAlias=effectiveDay", "岗位异动生效日期"],
        ]
        return _build_export(
            file_name,
            "岗位异动流程表",
            headers,
            rows,
            instances,
            write_counts,
            skip_reasons,
            mapping_rows,
        )


def _workbook_bytes(
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    extra_sheets: list[tuple[str, list[str], list[list[Any]]]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _populate_worksheet(ws, headers, rows)

    for extra_name, extra_headers, extra_rows in extra_sheets or []:
        extra_ws = wb.create_sheet(extra_name[:31])
        _populate_worksheet(extra_ws, extra_headers, extra_rows)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _populate_worksheet(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append([_excel_cell_value(value) for value in headers])
    for row in rows:
        ws.append([_excel_cell_value(value) for value in row])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        column = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 32))
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"


def _excel_cell_value(value: Any) -> Any:
    value = _loads_maybe_json(value)
    if value in (None, ""):
        return ""
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    if isinstance(value, list):
        return ";".join(
            text for text in (_excel_cell_text(item) for item in value) if text
        )
    if isinstance(value, dict):
        return _excel_cell_text(value)
    return str(value)


def _excel_cell_text(value: Any) -> str:
    value = _loads_maybe_json(value)
    if value in (None, ""):
        return ""
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return str(value)
    if isinstance(value, list):
        return ";".join(
            text for text in (_excel_cell_text(item) for item in value) if text
        )
    if isinstance(value, dict):
        for key in ("name", "label", "value", "fileName", "filename", "title", "text", "emplId", "userid"):
            if value.get(key) not in (None, ""):
                return str(value.get(key))
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _build_export(
    file_name: str,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    instances: list[dict[str, Any]],
    write_counts: dict[str, int],
    skip_reasons: dict[str, str],
    mapping_rows: list[list[Any]],
) -> DingTalkExport:
    audit_rows = _build_instance_audit_rows(instances, write_counts, skip_reasons)
    skipped_count = sum(1 for instance in instances if write_counts.get(_instance_key(instance), 0) == 0)
    data = _workbook_bytes(
        sheet_name,
        headers,
        rows,
        extra_sheets=[
            ("同步审计", _INSTANCE_AUDIT_HEADERS, audit_rows),
            ("字段映射", ["输出字段", "钉钉来源", "说明"], mapping_rows),
        ],
    )
    return DingTalkExport(
        file_name=file_name,
        data=data,
        row_count=len(rows),
        source_count=len(instances),
        skipped_count=skipped_count,
        audit_rows=audit_rows,
        mapping_rows=mapping_rows,
    )


def _build_instance_audit_rows(
    instances: list[dict[str, Any]],
    write_counts: dict[str, int],
    skip_reasons: dict[str, str],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for instance in instances:
        key = _instance_key(instance)
        write_count = write_counts.get(key, 0)
        if write_count > 0:
            write_status = "已写入"
        elif skip_reasons.get(key):
            write_status = "已跳过"
        else:
            write_status = "未写入"
        rows.append([
            instance.get("_sync_process_instance_id") or instance.get("process_instance_id") or "",
            instance.get("business_id") or "",
            instance.get("originator_userid") or "",
            instance.get("originator_dept_name") or "",
            instance.get("status") or "",
            _approval_status(instance),
            instance.get("result") or "",
            _approval_result(instance),
            _format_time_cell(instance.get("create_time")),
            _format_time_cell(instance.get("finish_time")),
            write_count,
            write_status,
            skip_reasons.get(key, ""),
        ])
    return rows


def _build_sync_audit_report(
    range_label: str,
    fetch_results: dict[str, DingTalkFetchResult],
    exports: dict[str, DingTalkExport],
) -> DingTalkExport:
    summary_headers = [
        "流程",
        "流程键",
        "流程码",
        "同步范围",
        "查询开始",
        "查询结束",
        "分页页数",
        "listids返回实例数",
        "去重后实例数",
        "重复实例数",
        "详情实例数",
        "写出行数",
        "跳过审批数",
        "是否截断",
        "条数上限",
        "备注",
    ]
    detail_headers = ["流程", "流程键"] + _INSTANCE_AUDIT_HEADERS
    mapping_headers = ["流程", "流程键", "输出字段", "钉钉来源", "说明"]
    summary_rows: list[list[Any]] = []
    detail_rows: list[list[Any]] = []
    mapping_rows: list[list[Any]] = []

    for flow_key, fetch_result in fetch_results.items():
        export = exports.get(flow_key)
        flow_label = _FLOW_LABELS.get(flow_key, flow_key)
        output_count = export.row_count if export else 0
        skipped_count = export.skipped_count if export else 0
        notes = []
        if fetch_result.repeated_cursor_stopped:
            notes.append("钉钉返回重复游标，已停止继续分页")
        if fetch_result.truncated:
            notes.append("按页面上限截断，仅用于小样本测试")
        if fetch_result.duplicate_instance_id_count:
            notes.append("已按 process_instance_id 去重")
        summary_rows.append([
            flow_label,
            flow_key,
            fetch_result.process_code,
            range_label,
            fetch_result.query_start,
            fetch_result.query_end,
            fetch_result.page_count,
            fetch_result.raw_instance_id_count,
            len(fetch_result.instance_ids),
            fetch_result.duplicate_instance_id_count,
            fetch_result.detail_count,
            output_count,
            skipped_count,
            "是" if fetch_result.truncated else "否",
            fetch_result.max_instances if fetch_result.max_instances is not None else "不限",
            "；".join(notes),
        ])
        if export and export.audit_rows:
            for row in export.audit_rows:
                detail_rows.append([flow_label, flow_key] + row)
        if export and export.mapping_rows:
            for row in export.mapping_rows:
                mapping_rows.append([flow_label, flow_key] + row)

    data = _workbook_bytes(
        "同步汇总",
        summary_headers,
        summary_rows,
        extra_sheets=[
            ("审批实例", detail_headers, detail_rows),
            ("字段映射", mapping_headers, mapping_rows),
        ],
    )
    return DingTalkExport(
        file_name=f"钉钉同步审计报告_{range_label}.xlsx",
        data=data,
        row_count=sum(export.row_count for export in exports.values()),
        source_count=sum(result.detail_count for result in fetch_results.values()),
        skipped_count=sum(export.skipped_count for export in exports.values()),
    )


def _instance_key(instance: dict[str, Any]) -> str:
    return str(
        instance.get("_sync_process_instance_id")
        or instance.get("process_instance_id")
        or instance.get("business_id")
        or id(instance)
    )


def _format_time_cell(value: Any) -> Any:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return _timestamp_to_text(value)
    text = str(value)
    if text.isdigit():
        return _timestamp_to_text(text)
    return text


def _date_interval_overlaps(
    start_value: Any,
    end_value: Any,
    filter_start: date | None,
    filter_end: date | None,
) -> bool:
    if filter_start is None or filter_end is None:
        return True
    start = _coerce_date(start_value)
    end = _coerce_date(end_value)
    if start is None or end is None:
        return False
    if start > end:
        start, end = end, start
    return start <= filter_end and end >= filter_start


def _date_in_range(
    value: Any,
    filter_start: date | None,
    filter_end: date | None,
) -> bool:
    if filter_start is None or filter_end is None:
        return True
    parsed = _coerce_date(value)
    if parsed is None:
        return False
    return filter_start <= parsed <= filter_end


def _coerce_date(value: Any) -> date | None:
    parsed = _coerce_datetime(value)
    return parsed.date() if parsed else None


def _coerce_datetime(value: Any) -> datetime | None:
    value = _loads_maybe_json(value)
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=_CN_TZ)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=_CN_TZ)
    if isinstance(value, (int, float)):
        timestamp = float(value)
        if timestamp > 10_000_000_000:
            timestamp = timestamp / 1000
        return datetime.fromtimestamp(timestamp, tz=_CN_TZ)
    if isinstance(value, dict):
        for key in ("value", "timeStamp", "timestamp", "date", "label"):
            parsed = _coerce_datetime(value.get(key))
            if parsed:
                return parsed
        return None
    if isinstance(value, list):
        for item in value:
            parsed = _coerce_datetime(item)
            if parsed:
                return parsed
        return None

    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return _coerce_datetime(int(text))

    iso_text = text.replace("Z", "+00:00")
    try:
        parsed_iso = datetime.fromisoformat(iso_text)
        return parsed_iso if parsed_iso.tzinfo else parsed_iso.replace(tzinfo=_CN_TZ)
    except ValueError:
        pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y年%m月%d日 %H:%M:%S",
        "%Y年%m月%d日 %H:%M",
        "%Y年%m月%d日",
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.replace(tzinfo=_CN_TZ)
        except ValueError:
            continue

    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})(?:\D+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?", text)
    if not match:
        return None
    year, month, day = (int(match.group(i)) for i in range(1, 4))
    hour = int(match.group(4) or 0)
    minute = int(match.group(5) or 0)
    second = int(match.group(6) or 0)
    try:
        return datetime(year, month, day, hour, minute, second, tzinfo=_CN_TZ)
    except ValueError:
        return None


def _is_dingtalk_rate_limited(data: dict[str, Any]) -> bool:
    errcode = data.get("errcode")
    errmsg = str(data.get("errmsg") or "")
    return errcode == 88 and (
        "subcode=90018" in errmsg
        or "接口次数过多" in errmsg
        or "限流" in errmsg
    )


def _rate_limit_sleep_seconds(errmsg: str, attempt: int) -> float:
    match = _RATE_LIMIT_END_RE.search(errmsg)
    if match:
        try:
            end_at = datetime.strptime(match.group(1), "%Y-%m-%d %H:%M:%S").replace(tzinfo=_CN_TZ)
            wait_seconds = (end_at - datetime.now(_CN_TZ)).total_seconds()
            return max(wait_seconds + 0.3, 0.8)
        except ValueError:
            pass
    return min(0.8 * (2 ** attempt), 8.0)


def _mapping_to_plain_dict(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, dict):
        return {
            str(k): _mapping_to_plain_dict(v) if _looks_like_mapping(v) else v
            for k, v in value.items()
        }
    if hasattr(value, "to_dict"):
        return _mapping_to_plain_dict(value.to_dict())
    if hasattr(value, "items"):
        return {str(k): v for k, v in value.items()}
    return {}


def _looks_like_mapping(value: Any) -> bool:
    return isinstance(value, dict) or hasattr(value, "items") or hasattr(value, "to_dict")


def _month_datetime_range(year: int, month: int, padding_months: int = 0) -> tuple[datetime, datetime]:
    start_year, start_month = _add_months(year, month, -max(padding_months, 0))
    end_year, end_month = _add_months(year, month, max(padding_months, 0) + 1)
    start = datetime(start_year, start_month, 1, tzinfo=_CN_TZ)
    end = datetime(end_year, end_month, 1, tzinfo=_CN_TZ)
    return start, end


def _date_datetime_range(start_date: date, end_date: date, padding_days: int = 0) -> tuple[datetime, datetime]:
    padding = max(int(padding_days or 0), 0)
    query_start_date = start_date - timedelta(days=padding)
    query_end_date = end_date + timedelta(days=padding + 1)
    start = datetime(query_start_date.year, query_start_date.month, query_start_date.day, tzinfo=_CN_TZ)
    end = datetime(query_end_date.year, query_end_date.month, query_end_date.day, tzinfo=_CN_TZ)
    return start, end


def _now_cn() -> datetime:
    return datetime.now(tz=_CN_TZ)


def _process_instance_query_windows(
    query_start: datetime,
    query_end_exclusive: datetime,
) -> list[tuple[datetime, datetime]]:
    """Return DingTalk-compatible [start, end) windows.

    listids rejects oversized ranges and end timestamps ahead of DingTalk's
    clock. Keep a small skew buffer, split long ranges, and let the caller
    de-duplicate instances across adjacent windows.
    """
    start = query_start if query_start.tzinfo else query_start.replace(tzinfo=_CN_TZ)
    end = query_end_exclusive if query_end_exclusive.tzinfo else query_end_exclusive.replace(tzinfo=_CN_TZ)
    latest_end = _now_cn() - timedelta(seconds=_OAPI_QUERY_CLOCK_SKEW_SECONDS)
    effective_end = min(end, latest_end)
    if start >= effective_end:
        return []

    windows: list[tuple[datetime, datetime]] = []
    window_start = start
    max_window = timedelta(days=_OAPI_PROCESS_INSTANCE_QUERY_MAX_DAYS)
    while window_start < effective_end:
        window_end = min(window_start + max_window, effective_end)
        windows.append((window_start, window_end))
        window_start = window_end
    return windows


def _month_ms_range(year: int, month: int, padding_months: int = 0) -> tuple[int, int]:
    start, end = _month_datetime_range(year, month, padding_months)
    return int(start.timestamp() * 1000), int(end.timestamp() * 1000) - 1


def _add_months(year: int, month: int, offset: int) -> tuple[int, int]:
    month_index = year * 12 + (month - 1) + offset
    return month_index // 12, month_index % 12 + 1


def _loads_maybe_json(value: Any) -> Any:
    if value in (None, "", "null"):
        return None
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return value
    return value


def _form_values_by_name(instance: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for item in instance.get("form_component_values") or []:
        name = item.get("name")
        if not name:
            continue
        result[str(name)] = _loads_maybe_json(item.get("value"))
    return result


def _find_form_value(
    instance: dict[str, Any],
    *,
    component_type: str | None = None,
    name: str | None = None,
) -> dict[str, Any] | None:
    for item in instance.get("form_component_values") or []:
        if component_type and item.get("component_type") != component_type:
            continue
        if name and item.get("name") != name:
            continue
        return item
    return None


def _find_biz_suite(instance: dict[str, Any], aliases: set[str]) -> list[dict[str, Any]]:
    for item in instance.get("form_component_values") or []:
        if item.get("component_type") != "DDBizSuite":
            continue
        suite = _loads_maybe_json(item.get("value"))
        if not isinstance(suite, list):
            continue
        by_alias = _suite_components_by_alias(suite)
        if aliases & set(by_alias):
            return suite
    return []


def _suite_components_by_alias(suite: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for component in suite or []:
        props = component.get("props") or {}
        alias = props.get("bizAlias")
        if alias:
            result[str(alias)] = component
    return result


def _component_value(component: dict[str, Any] | None) -> Any:
    if not component:
        return None
    value = _loads_maybe_json(component.get("value"))
    if value not in (None, ""):
        return value
    ext = _loads_maybe_json(component.get("extValue"))
    if isinstance(ext, dict):
        for key in ("label", "value", "timeStamp", "userCheckTime"):
            if ext.get(key) not in (None, ""):
                return ext.get(key)
    return None


def _table_rows(component: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not component:
        return []
    raw_rows = _loads_maybe_json(component.get("value"))
    if not isinstance(raw_rows, list):
        return []
    rows: list[dict[str, Any]] = []
    for raw_row in raw_rows:
        row: dict[str, Any] = {}
        for cell in raw_row.get("rowValue") or []:
            alias = cell.get("bizAlias") or cell.get("label") or cell.get("key")
            if alias:
                row[str(alias)] = _loads_maybe_json(cell.get("value"))
        if row:
            rows.append(row)
    return rows


def _contact_value(component: dict[str, Any] | None) -> tuple[str, str]:
    if not component:
        return "", ""
    name = str(_component_value(component) or "")
    ext = _loads_maybe_json(component.get("extValue"))
    if isinstance(ext, list) and ext:
        first = ext[0]
        if isinstance(first, dict):
            return str(first.get("name") or name or ""), str(first.get("emplId") or first.get("userid") or "")
    return name, ""


def _approval_status(instance: dict[str, Any]) -> str:
    status = str(instance.get("status") or "").upper()
    if status == "COMPLETED":
        return "完成"
    if status in {"RUNNING", "NEW"}:
        return "审批中"
    if status in {"TERMINATED", "CANCELED", "CANCELLED"}:
        return "已撤销"
    return status or ""


def _approval_result(instance: dict[str, Any]) -> str:
    result = str(instance.get("result") or "").lower()
    if result == "agree":
        return "同意"
    if result in {"refuse", "reject"}:
        return "拒绝"
    return result


def _job_number(user: dict[str, Any]) -> str:
    return str(user.get("job_number") or user.get("jobNumber") or "").strip()


def _duration_text(value: Any, unit: Any) -> Any:
    if value in (None, ""):
        return None
    unit_text = str(unit or "").lower()
    if "day" in unit_text:
        return f"{value}天"
    if "hour" in unit_text:
        return f"{value}小时"
    return value


def _join_multi_value(value: Any) -> str:
    value = _loads_maybe_json(value)
    if value in (None, ""):
        return ""
    if isinstance(value, list):
        return ";".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def _timestamp_to_text(value: Any) -> str:
    try:
        timestamp = int(str(value))
    except Exception:
        return ""
    # DingTalk timestamps are milliseconds.
    if timestamp > 10_000_000_000:
        timestamp = timestamp // 1000
    return datetime.fromtimestamp(timestamp, tz=_CN_TZ).strftime("%Y-%m-%d %H:%M")


def _name_from_title(title: Any) -> str:
    text = str(title or "")
    if "提交的" in text:
        return text.split("提交的", 1)[0].strip()
    return ""
