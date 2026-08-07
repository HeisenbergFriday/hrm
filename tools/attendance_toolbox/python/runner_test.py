from datetime import date, datetime
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from overtime import fill_overtime_fields as overtime
from subsidy import calc_subsidy_deduction as subsidy
import runner
import templates as toolbox_templates


RUNNER_PATH = Path(__file__).resolve().with_name("runner.py")


class RunnerNameParsingTest(unittest.TestCase):
    def test_names_accepts_display_separators(self):
        self.assertEqual(
            runner.names({"names": "梁伯林、陈秋宇，蔡依诺,王天钦\n周灵京；秦洋"}, "names"),
            ["梁伯林", "陈秋宇", "蔡依诺", "王天钦", "周灵京", "秦洋"],
        )

    def test_names_accepts_form_value_list(self):
        self.assertEqual(
            runner.names({"names": ["费婷玉、张莹", "陈星雨，伏鸣"]}, "names"),
            ["费婷玉", "张莹", "陈星雨", "伏鸣"],
        )


class OvertimeValidationTest(unittest.TestCase):
    def test_holiday_validation_ignores_rows_outside_target_month(self):
        old_row = [None] * len(overtime.CLEAN_OUTPUT_COLS)
        old_row[5] = datetime(2025, 10, 1, 9, 0)
        old_row[6] = datetime(2025, 10, 1, 18, 0)
        old_row[7] = date(2025, 10, 1)

        target_row = [None] * len(overtime.CLEAN_OUTPUT_COLS)
        target_row[5] = datetime(2026, 4, 1, 9, 0)
        target_row[6] = datetime(2026, 4, 1, 18, 0)
        target_row[7] = date(2026, 4, 1)

        filtered_rows, skipped = overtime.filter_rows_by_target_month(
            [overtime.CLEAN_OUTPUT_COLS, tuple(old_row), tuple(target_row)],
            date(2026, 4, 1),
        )

        self.assertEqual(skipped, 1)
        self.assertEqual(overtime._collect_holiday_years(filtered_rows), {2026})


class ToolboxMonthAndHolidayLinkageTest(unittest.TestCase):
    def test_parse_target_month_accepts_month_and_date(self):
        self.assertEqual(runner.parse_target_month("2026-07"), (2026, 7))
        self.assertEqual(runner.parse_target_month("2026/07/18"), (2026, 7))
        self.assertIsNone(runner.parse_target_month(""))

    def test_parse_target_month_rejects_invalid_value(self):
        with self.assertRaisesRegex(ValueError, "YYYY-MM"):
            runner.parse_target_month("2026-13", "补贴考勤月份")

    def test_locked_month_rejects_schedule_conflict(self):
        with self.assertRaisesRegex(ValueError, "但作息表识别为 2026年6月"):
            runner.validate_target_month_match(
                "2026-07",
                (2026, 6),
                label="补贴考勤月份",
            )

    def test_custom_holidays_override_schedule_and_record_conflicts(self):
        selected, audit = runner.select_subsidy_legal_holidays(
            {date(2026, 7, 1), date(2026, 7, 2)},
            {date(2026, 7, 1), date(2026, 7, 3), date(2025, 10, 1)},
            2026,
            7,
        )
        self.assertEqual(selected, {date(2026, 7, 1), date(2026, 7, 3)})
        self.assertEqual(audit["source"], "custom_rules")
        self.assertEqual(audit["only_in_rules"], ["2026-07-03"])
        self.assertEqual(audit["only_in_schedule"], ["2026-07-02"])
        self.assertEqual(audit["conflict_count"], 2)

    def test_schedule_holidays_are_used_without_override(self):
        selected, audit = runner.select_subsidy_legal_holidays(
            {date(2026, 7, 1)},
            set(),
            2026,
            7,
        )
        self.assertEqual(selected, {date(2026, 7, 1)})
        self.assertEqual(audit["source"], "schedule")
        self.assertEqual(audit["conflict_count"], 0)

    def test_exception_audit_sheet_lists_missing_attendance(self):
        wb = Workbook()
        subsidy.append_exception_audit_sheet(
            wb,
            missing_attendance=["李睿", "李睿", "任澳辉"],
            intern_names=["实习员工"],
            audit_context={
                "target_month": "2026-07",
                "holiday_source": "custom_rules",
                "holiday_count": 2,
                "holiday_conflict_count": 1,
                "holiday_only_in_rules": ["2026-07-03"],
            },
        )
        ws = wb["异常审计"]
        values = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
        flattened = "\n".join(str(item) for row in values for item in row if item is not None)
        self.assertIn("李睿", flattened)
        self.assertIn("任澳辉", flattened)
        self.assertIn("2026-07-03", flattened)
        self.assertEqual(sum(1 for row in values if row[1] == "李睿"), 1)

    def test_audit_meta_keeps_missing_names_and_counts(self):
        meta = runner.build_subsidy_audit_meta(
            year=2026,
            month=7,
            holiday_audit={"source": "schedule", "selected_count": 1, "conflict_count": 0},
            missing_attendance=["丁俊", "丁俊", "任澳辉"],
            intern_names=[],
        )
        audit = meta["subsidy_audit"]
        self.assertEqual(audit["missing_attendance_count"], 2)
        self.assertEqual(audit["missing_attendance_names"], ["丁俊", "任澳辉"])


class GenerateRosterCLITest(unittest.TestCase):
    def run_cli(self, workdir: str, config: dict, action: str = "generate-roster"):
        return subprocess.run(
            [
                sys.executable,
                str(RUNNER_PATH),
                "--action",
                action,
                "--workdir",
                workdir,
                "--config-json",
                json.dumps(config, ensure_ascii=False),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=False,
        )

    def test_generate_roster_cli_creates_valid_workbook(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {
                "org_name": '测试/组织:华南?*',
                "employees": [
                    {
                        "emp_no": "MT9999", "name": "  测试运维  ",
                        "dept1": "运营管理中心", "dept2": "运营支撑部", "dept3": "智慧寄存运维组",
                        "position": "运维工程师", "emp_type": "正式",
                        "hire_date": "2026/01/02", "confirm_date": "2026-04-02 00:00:00",
                    },
                    {"emp_no": "MT1000", "name": "李四", "dept1": "研发中心"},
                ],
            })

            self.assertEqual(completed.returncode, 0, completed.stderr or completed.stdout)
            payload = json.loads(completed.stdout)
            self.assertTrue(payload["ok"])
            self.assertEqual(len(payload["outputs"]), 1)
            output = payload["outputs"][0]
            self.assertEqual(output["kind"], "export")
            self.assertEqual(output["row_count"], 2)
            self.assertNotRegex(output["file_name"], r'[\\/:*?"<>|]')

            output_path = Path(output["path"])
            self.assertTrue(output_path.is_file())
            workbook = load_workbook(output_path, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["在职花名册"])
                sheet = workbook["在职花名册"]
                headers = [cell.value for cell in sheet[1]]
                self.assertEqual(headers, runner._ROSTER_HEADERS)
                self.assertEqual(sheet.max_column, 12)
                self.assertEqual(
                    [sheet.cell(2, column).value for column in range(1, 13)],
                    [
                        "MT9999", "测试运维", None, "运营管理中心", "运营支撑部",
                        "智慧寄存运维组", "运维工程师", "正式", None, "2026-01-02", None,
                        "2026-04-02",
                    ],
                )
                self.assertEqual(sheet.cell(3, 1).value, "MT1000")
            finally:
                workbook.close()

    def test_generate_roster_cli_rejects_non_list_employees(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {"org_name": "测试组织", "employees": "not-a-list"})
        self.assertNotEqual(completed.returncode, 0)
        payload = json.loads(completed.stdout)
        self.assertFalse(payload["ok"])
        self.assertIn("employees 字段需要是 list", payload["error"])

    def test_generate_roster_cli_rejects_missing_employee_ids(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {
                "org_name": "测试组织",
                "employees": [
                    {"emp_no": "", "name": "甲", "dept1": "总部"},
                    {"name": "乙", "dept1": "总部"},
                    {"emp_no": "MT0003", "name": "丙", "dept1": "总部"},
                ],
            })
        self.assertNotEqual(completed.returncode, 0)
        payload = json.loads(completed.stdout)
        self.assertFalse(payload["ok"])
        self.assertIn("2 名在职员工缺少业务工号", payload["error"])

    def test_generate_roster_cli_rejects_mixed_empty_and_whitespace_names(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {
                "org_name": "测试组织",
                "employees": [
                    {"emp_no": "MT0001", "name": "有效员工", "dept1": "总部"},
                    {"emp_no": "MT0002", "name": "", "dept1": "总部"},
                    {"emp_no": "MT0003", "name": "   ", "dept1": "总部"},
                ],
            })
            generated = list(Path(workdir).glob("*.xlsx"))
        self.assertNotEqual(completed.returncode, 0)
        payload = json.loads(completed.stdout)
        self.assertFalse(payload["ok"])
        self.assertIn("2 名在职员工缺少姓名", payload["error"])
        self.assertEqual(generated, [])

    def test_generate_roster_cli_rejects_missing_department_path(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {
                "org_name": "测试组织",
                "employees": [{"emp_no": "MT0001", "name": "甲"}],
            })
        self.assertNotEqual(completed.returncode, 0)
        payload = json.loads(completed.stdout)
        self.assertIn("1 名在职员工缺少有效部门路径", payload["error"])

    def test_generate_roster_cli_keeps_duplicate_names_separate_by_employee_id(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {
                "org_name": "测试组织",
                "employees": [
                    {"emp_no": "MT0001", "name": "同名员工", "dept1": "总部"},
                    {"emp_no": "MT0002", "name": "同名员工", "dept1": "总部"},
                ],
            })
        self.assertEqual(completed.returncode, 0, completed.stderr or completed.stdout)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["outputs"][0]["row_count"], 2)

    def test_runner_cli_rejects_unregistered_action(self):
        with tempfile.TemporaryDirectory() as workdir:
            completed = self.run_cli(workdir, {}, action="generate-roster-unregistered")
        self.assertNotEqual(completed.returncode, 0)
        self.assertIn("invalid choice", completed.stderr)


class FinalRosterContractEndToEndTest(unittest.TestCase):
    def test_name_roster_and_dingtalk_monthly_summary_generate_final_workbook(self):
        with tempfile.TemporaryDirectory() as workdir:
            root = Path(workdir)
            output_dir = root / "outputs"
            output_dir.mkdir()

            roster_path = root / "roster.xlsx"
            roster = Workbook()
            roster.active.title = "在职花名册"
            roster.active.append(["姓名"])
            roster.active.append(["张三"])
            roster.save(roster_path)
            roster.close()

            inputs = {
                "final_schedule": toolbox_templates.build_schedule_template(),
                "final_leave": toolbox_templates.build_final_leave_detail_template(),
                "final_overtime": toolbox_templates.build_final_overtime_detail_template(),
                "final_subsidy": toolbox_templates.build_subsidy_source_template(),
            }
            config = {"final_active": str(roster_path)}
            for field, content in inputs.items():
                path = root / f"{field}.xlsx"
                path.write_bytes(content)
                config[field] = str(path)

            outputs = runner.run_final(config, output_dir)
            self.assertEqual(len(outputs), 1)
            workbook = load_workbook(outputs[0]["path"], data_only=False)
            try:
                sheet = workbook.active
                headers = [cell.value for cell in sheet[2]]
                values = {
                    header: sheet.cell(3, index + 1).value
                    for index, header in enumerate(headers)
                }
            finally:
                workbook.close()

            self.assertEqual(values["姓名"], "张三")
            self.assertEqual(values["工号"], "MT0001")
            self.assertEqual(values["考勤组"], "默认考勤组")
            self.assertEqual(
                (values["一级部门"], values["二级部门"], values["三级部门"]),
                ("总部", "产品技术部", "后端组"),
            )
            self.assertEqual(values["岗位"], "工程师")
            for field in ("合同主体", "员工类型", "人员分类", "入职日期", "离职日期", "转正日期"):
                self.assertIsNone(values[field], field)


if __name__ == "__main__":
    unittest.main()
