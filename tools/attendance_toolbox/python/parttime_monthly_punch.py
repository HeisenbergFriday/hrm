# -*- coding: utf-8 -*-
"""Render the part-time "月度打卡记录" grid + audit sheet.

The generated workbook mirrors the shape that
``parttime/calc_parttime_summary.parse_attendance_detail`` consumes:

  - Row 1 (header): 姓名 | 考勤组 | 部门 | 工号 | 职位 | (reserved) | 1 | 2 | … | N
  - Row 2 (day header): dates for each day of the month
  - Row 3..: one row per matched employee; day columns hold status labels such
    as "正常 (08:30,18:00)" / "迟到 (08:35,18:00)" / "旷工" that the parser reads.

A second "审计" sheet lists unmatched employees and anomalies so the operator
can see who was kept without a punch record and why.
"""
from __future__ import annotations

import json
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER_FILL = PatternFill("solid", fgColor="FFF2F2F2")
TITLE_FILL = PatternFill("solid", fgColor="FFD9EAF7")
UNMATCHED_FILL = PatternFill("solid", fgColor="FFFFE599")
FONT_BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)

IDENTITY_COLUMNS = ["姓名", "考勤组", "部门", "工号", "职位", "匹配方式"]
RESERVED_COLUMN = ""  # column 6 reserved (parser starts day columns at col 7)

# 兼职 rows must carry a role keyword in 职位, otherwise the part-time
# summary's scope filter (_is_attendance_row_in_scope) drops them. We default
# to "兼职" when the org roster supplies no position.
DEFAULT_POSITION = "兼职"


def render(workdir: Path, config: dict) -> dict:
    year = int(config.get("year"))
    month = int(config.get("month"))
    days_in_month = int(config.get("days_in_month"))
    matched = config.get("matched", [])
    unmatched = config.get("unmatched", [])
    anomalies = config.get("anomalies", [])

    validate_input(year, month, days_in_month, matched)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月度打卡记录"
    write_header(ws, year, month, days_in_month)
    write_day_header(ws, year, month, days_in_month)
    write_matched_rows(ws, matched, days_in_month)

    audit = wb.create_sheet("审计")
    write_audit(audit, matched, unmatched, anomalies, year, month)

    out_path = workdir / "outputs" / f"兼职月度打卡记录_{year}{month:02d}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    wb.close()
    print(f"[兼职月度打卡记录] 已输出：{out_path}")
    return {"path": str(out_path), "file_name": out_path.name}


def validate_input(year: int, month: int, days_in_month: int, matched: Any) -> None:
    if not (1 <= month <= 12):
        raise ValueError(f"月份无效：{month}")
    if year < 2000 or year > 2100:
        raise ValueError(f"年份无效：{year}")
    if days_in_month < 28 or days_in_month > 31:
        raise ValueError(f"月份天数无效：{days_in_month}")
    if not isinstance(matched, list):
        raise ValueError("matched 必须是数组")


def day_date(year: int, month: int, day: int) -> date:
    return date(year, month, min(day, days_in_month_safe(year, month)))


def days_in_month_safe(year: int, month: int) -> int:
    if month == 12:
        return (date(year + 1, 1, 1) - timedelta(days=1)).day
    return (date(year, month + 1, 1) - timedelta(days=1)).day


def write_header(ws, year: int, month: int, days_in_month: int) -> None:
    # NOTE: the part-time parser locates the header row by searching for cells
    # whose value is exactly "姓名" or "工号" (see calc_parttime_summary._find_col).
    # This row MUST therefore contain both verbatim, and MUST NOT use merged
    # cells (merged cells read back as None and break header detection).
    headers = IDENTITY_COLUMNS + [RESERVED_COLUMN] + [str(d) for d in range(1, days_in_month + 1)]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = FONT_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    # Month label lives on the audit sheet; keep the grid header machine-readable.
    _ = year


def write_day_header(ws, year: int, month: int, days_in_month: int) -> None:
    ws.cell(row=2, column=1, value="日期").font = FONT_BOLD
    ws.cell(row=2, column=1).alignment = CENTER
    for day in range(1, days_in_month + 1):
        label = f"{month:02d}-{day:02d}"
        cell = ws.cell(row=2, column=6 + day, value=label)
        cell.alignment = CENTER
        cell.font = Font(bold=True, size=10)


def write_matched_rows(ws, matched: list[dict], days_in_month: int) -> None:
    for row_idx, emp in enumerate(matched, start=3):
        name = str(emp.get("name") or "").strip()
        code = str(emp.get("employee_no") or "").strip()
        days = emp.get("days") or {}
        position = str(emp.get("position") or "").strip()
        department = str(emp.get("department") or "").strip()
        matched_by = emp.get("matched_by") == "employee_no" and "工号" or "姓名"
        if not position:
            position = DEFAULT_POSITION

        ws.cell(row=row_idx, column=1, value=name).alignment = LEFT
        ws.cell(row=row_idx, column=2, value="").alignment = CENTER  # 考勤组
        ws.cell(row=row_idx, column=3, value=department).alignment = CENTER  # 部门
        ws.cell(row=row_idx, column=4, value=code).alignment = CENTER
        ws.cell(row=row_idx, column=5, value=position).alignment = CENTER  # 职位
        ws.cell(row=row_idx, column=6, value=matched_by).alignment = CENTER

        for day in range(1, days_in_month + 1):
            value = str(days.get(day, days.get(str(day), "")))
            cell = ws.cell(row=row_idx, column=6 + day, value=value)
            cell.alignment = LEFT
            if not value:
                cell.fill = UNMATCHED_FILL


def write_audit(ws, matched, unmatched, anomalies, year: int, month: int) -> None:
    ws.cell(row=1, column=1, value="审计项").font = FONT_BOLD
    ws.cell(row=1, column=2, value="内容").font = FONT_BOLD
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    row = 2
    summary = (
        f"{year}-{month:02d}: 花名册 {len(matched) + len(unmatched)} 人，"
        f"已匹配 {len(matched)} 人，未匹配 {len(unmatched)} 人，异常 {len(anomalies)} 项。"
    )
    ws.cell(row=row, column=1, value="汇总")
    ws.cell(row=row, column=2, value=summary)
    row += 1

    if unmatched:
        ws.cell(row=row, column=1, value="未匹配到打卡记录（已保留）")
        names = "、".join(
            f"{str(e.get('name') or '').strip()}({str(e.get('employee_no') or '').strip() or '无工号'})"
            for e in unmatched
        )
        ws.cell(row=row, column=2, value=names)
        row += 1

    for anomaly in anomalies:
        ws.cell(row=row, column=1, value="异常提示")
        ws.cell(row=row, column=2, value=str(anomaly))
        row += 1

    if row == 2:
        ws.cell(row=row, column=1, value="无异常")
        ws.cell(row=row, column=2, value="全部花名册人员均已匹配到打卡记录。")


def main() -> None:
    """Allow running standalone for tests: renders a demo workbook."""
    import sys

    cfg = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {
        "year": 2026, "month": 7, "days_in_month": 31,
        "matched": [
            {"name": "示例-张三", "employee_no": "MT001", "matched_by": "employee_no",
             "days": {1: "正常 (08:30,18:00)", 2: "迟到 (08:35,18:00)"}},
        ],
        "unmatched": [{"name": "示例-李四", "employee_no": ""}],
        "anomalies": [],
    }
    out = render(Path("."), cfg)
    print(json.dumps(out, ensure_ascii=False))


if __name__ == "__main__":
    main()
