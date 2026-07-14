"""Blank Excel template builders for the attendance toolbox.

Each builder mirrors a ``_build_*_template`` helper from the original
``D:/app/app.py``.  HR staff download these from the toolbox UI, fill in
their own data, then re-upload them to the calculation runner.
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Shared styling for all built templates.
_HEADER_FILL = PatternFill("solid", fgColor="FFE2E8F0")
_SAMPLE_FILL = PatternFill("solid", fgColor="FFF8FAFC")
_WORKDAY_FILL = PatternFill("solid", fgColor="FFCCFFCC")
_HOLIDAY_FILL = PatternFill("solid", fgColor="FFFFC7CE")
_WELFARE_FILL = PatternFill("solid", fgColor="FFD6EAF8")


# ── Column tuples reused by several final-table templates ─────────────────────

_LEAVE_EXPORT_REQUIRED_COLUMNS = (
    "审批状态", "审批结果", "发起人工号", "发起人姓名", "发起人部门",
    "请假类型", "开始时间", "结束时间", "时长", "审批编号",
)
_LEAVE_EXPORT_OPTIONAL_COLUMNS = ("发起时间", "完成时间", "岗位名称")
_OT_EXPORT_REQUIRED_COLUMNS = (
    "审批状态", "审批结果", "发起人工号", "发起人姓名", "发起人部门",
    "开始时间", "结束时间", "时长",
)
_OT_EXPORT_OPTIONAL_COLUMNS = (
    "审批编号", "明细", "加班时间", "加班时长", "是否包含法定节假日期", "2026法定节假日如下：",
)
_SUB_SOURCE_COLUMNS = (
    "姓名", "工号", "考勤组", "一级部门", "二级部门", "三级部门", "职位",
    "15-30分钟迟到扣款", "15-30分钟早退扣款", "旷工天数", "缺卡次数",
    "晚于22点打卡天数", "晚走补贴", "产研休息日加班>4小时天数", "产研休息日加班补贴",
)
_FIN_ROSTER_COLUMNS = (
    "工号", "姓名", "合同主体", "一级部门", "二级部门", "三级部门", "岗位",
    "员工类型", "人员分类", "入职日期", "离职日期", "转正日期",
)
_FIN_TRANSFER_COLUMNS = (
    "实际申请人工号", "实际申请人", "发起人工号", "发起人姓名",
    "异动日期", "异动类型", "生效日期",
)
_FIN_LEAVE_DETAIL_COLUMNS = (
    "发起人工号", "发起人姓名", "一级部门", "二级部门", "三级部门", "请假类型",
    "开始时间", "结束时间", "系统时长", "最终请假时长", "最终请假天数", "备注",
)
_FIN_OVERTIME_DETAIL_COLUMNS = (
    "发起人工号", "发起人姓名", "一级部门", "二级部门", "三级部门",
    "开始时间", "结束时间", "明细", "加班时间",
    "2倍加班小时", "3倍加班小时", "2倍加班天数", "3倍加班天数",
)


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        column = column_cells[0].column_letter
        width = 10
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, 28))
        ws.column_dimensions[column].width = width


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = __import__("io").BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _notes_sheet(wb: Workbook, notes: list[str]) -> None:
    note_ws = wb.create_sheet("填写说明")
    for note in notes:
        note_ws.append([note])
    note_ws.column_dimensions["A"].width = 90
    for row in note_ws.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")


def _build_table_template(
    sheet_name: str,
    headers: tuple[str, ...],
    sample_rows: list[tuple],
    notes: list[str] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for sample_row in sample_rows:
        ws.append(sample_row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = _SAMPLE_FILL
    ws.freeze_panes = "A2"
    _autosize_worksheet(ws)
    if notes:
        _notes_sheet(wb, notes)
    return _workbook_to_bytes(wb)


# ── Individual templates (each one mirrors a builder in the legacy app.py) ─────

def build_leave_export_template() -> bytes:
    headers = _LEAVE_EXPORT_REQUIRED_COLUMNS + _LEAVE_EXPORT_OPTIONAL_COLUMNS
    notes = [
        "请假系统导出表的表头必须放在第 1 行，字段名请保持和模板一致。",
        "审批状态会保留：完成、审批中、已修改；已撤销、完成且拒绝的记录会被排除。",
        "岗位名称为可选字段，用于辅助识别实习生；也可以使用职位、岗位、员工类型等字段名。",
    ]
    return _build_table_template(
        "请假系统导出",
        headers,
        [(
            "完成", "同意", "10001", "张三", "总部-产品技术部-后端组", "年假",
            "2026-06-03 09:00:00", "2026-06-03 18:30:00", "8小时",
            "DD202606030001", "2026-06-01 10:00:00", "2026-06-02 09:00:00", "正式员工",
        )],
        notes,
    )


def build_offsite_duration_template() -> bytes:
    return _build_table_template(
        "异地不打卡人员",
        ("工号", "姓名", "请假时长", "审批编号", "请假类型", "开始时间", "结束时间", "考勤组"),
        [(
            "10003", "李四", "8小时", "DD202606030002", "事假",
            "2026-06-04 09:00:00", "2026-06-04 18:30:00", "异地外勤（免打卡）",
        )],
        [
            "工号或姓名至少填写一个；有审批编号时会优先按审批编号精确匹配。",
            "请假时长可填数字小时，也可填 8小时、1天；不提供时长时会按请假系统导出的时长计算。",
            "如果存在考勤组字段且包含异地外勤（免打卡）人员，系统只载入该考勤组人员。",
        ],
    )


def build_schedule_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "作息表"
    ws["A1"] = "2026年6月作息时间表"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["周数", "周一", "周二", "周三", "周四", "周五", "周六", "周日"])
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    weeks = [
        (1, 1, 2, 3, 4, 5, 6, 7),
        (2, 8, 9, 10, 11, 12, 13, 14),
        (3, 15, 16, 17, 18, 19, 20, 21),
        (4, 22, 23, 24, 25, 26, 27, 28),
        (5, 29, 30, "", "", "", "", ""),
    ]
    for row in weeks:
        ws.append(row)
    for row in ws.iter_rows(min_row=3, max_row=7, min_col=2, max_col=8):
        for cell in row:
            if cell.value == "":
                continue
            if cell.value in {19}:
                cell.fill = _HOLIDAY_FILL
            elif cell.value in {20}:
                cell.fill = _WELFARE_FILL
            elif cell.column <= 6:
                cell.fill = _WORKDAY_FILL
            cell.alignment = Alignment(horizontal="center")
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "标题单元格需要包含\"作息时间表\"和年月，例如：2026年6月作息时间表。",
            "标题下方 8 行内需要出现\"周数\"这一行；日期填写在第 2 到第 8 列。",
            "黄色表示工作日，红色表示法定节假日，蓝色表示公司福利假；系统按单元格底色识别。",
            "同一份作息表请只保留一个月份；如有成都独立作息，可在另一张表放\"成都作息时间表\"。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_overtime_export_template() -> bytes:
    headers = _OT_EXPORT_REQUIRED_COLUMNS + _OT_EXPORT_OPTIONAL_COLUMNS
    return _build_table_template(
        "加班系统导出",
        headers,
        [(
            "完成", "同意", "MT0001", "张三", "总部-运营管理中心-运营支撑部",
            "2026-06-19 09:00:00", "2026-06-19 18:30:00", "8小时",
            "OT202606190001", "2026-06-19", "2026-06-19", "8小时", "是", "2026-06-19",
        )],
        [
            "加班系统导出表表头必须放在第 1 行；可包含多个 sheet，系统会自动合并字段完整的 sheet。",
            "审批状态会保留：完成、审批中、已修改；完成且审批结果为拒绝的记录会被排除。",
            "发起人部门会按\"-\"拆分为一级部门、二级部门、三级部门，用于判断加班口径。",
        ],
    )


def build_overtime_schedule_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "排班表"
    ws.append([
        None,
        date(2026, 6, 1), date(2026, 6, 2), date(2026, 6, 3), date(2026, 6, 4),
        date(2026, 6, 5), date(2026, 6, 6), date(2026, 6, 7),
    ])
    ws.append(["姓名", "周一", "周二", "周三", "周四", "周五", "周六", "周日"])
    ws.append(["张三", "班", "班", "班", "班", "班", "OFF", "班"])
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")
        if cell.column > 1:
            cell.number_format = "yyyy-mm-dd"
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.fill = _SAMPLE_FILL
            cell.alignment = Alignment(horizontal="center")
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "日期行第一格请留空，从 B 列开始填写日期；下一行为星期；员工行第一列填写姓名。",
            "员工在某日期有排班时填写任意非空且非 OFF 的内容；OFF 或空白不会生成排班补录。",
            "排班表仅用于补录审批导出缺失的加班记录，不参与 2 倍 / 3 倍主规则判定。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_overtime_attendance_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤打卡明细"
    ws.append(["姓名", "工号", "考勤组", "考勤结果", None, None, None, None, None])
    ws.append([None, None, None, None, 1, 2, 3, 4, 5])
    ws.append(["张三", "MT0001", "运营支撑部", "", "09:00-18:30", "(-)", "09:00-18:35", "09:00-22:10", "09:00-18:30"])
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[3]:
        cell.fill = _SAMPLE_FILL
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "第 1 行需要包含\"姓名\"和\"工号\"；第 2 行填写日期序号，员工数据从第 3 行开始。",
            "考勤组为可选字段，用于辅助识别特殊部门；日期单元格中出现 (-) 或单边卡时，调休记录会标记为未加。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_subsidy_source_template() -> bytes:
    return _build_table_template(
        "补贴扣款",
        _SUB_SOURCE_COLUMNS,
        [("张三", "MT0001", "默认考勤组", "总部", "产品技术部", "后端组", "工程师",
          0, 0, 0, 0, 1, 50, 1, 50)],
        [
            "补贴扣款表可以是钉钉导出表；系统会在前几行自动识别这些字段。",
            "迟到/早退扣款沿用源表；旷工天数、缺卡次数、晚走补贴、产研休息日加班补贴会按考勤明细核对。",
            "工号字段可使用：工号、员工工号、员工编号。",
        ],
    )


def build_activity_checkin_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "羽毛球签到"
    ws.append(["姓名", date(2026, 6, 5), date(2026, 6, 12), date(2026, 6, 19)])
    ws.append(["张三", "已参加", "", "已参加"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        if cell.column > 1:
            cell.number_format = "yyyy-mm-dd"
    for cell in ws[2]:
        cell.fill = _SAMPLE_FILL
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "sheet 名需要包含\"羽毛球\"或\"篮球\"；日期行中同时包含\"姓名\"和活动日期。",
            "人员在对应活动日期单元格填写\"已参加\"时，该日 22 点后打卡不计晚走补贴。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_subsidy_attendance_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤"
    ws.append(["姓名", "工号", "一级部门", "二级部门", "三级部门", "职位", 1, 2, 3, 4, 5, 6, 7])
    ws.append([
        "张三", "MT0001", "总部", "产品技术部", "后端组", "工程师",
        "09:00-18:30", "09:00-22:10", "休息 5小时", "09:00-18:30", "缺卡", "旷工", "09:00-18:30",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.fill = _SAMPLE_FILL
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "前 5 行内需要包含\"工号\"字段；日期行需要有连续日期序号，员工数据从日期行下一行开始。",
            "工号需使用有效前缀：MT、TXB、WB、JZ；日期单元格用于识别打卡时间、缺卡、旷工和休息日加班时长。",
            "当补贴扣款.xlsx 本身已包含日期级考勤结果列时，可不单独上传此表。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_attendance_result_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤结果表"
    ws.append(["考勤结果表"])
    ws.append(["批注含\"已补流程\"的旷工日期会被排除"])
    ws.append(["姓名", "工号", "考勤组", 1, 2, 3, 4, 5])
    ws.append(["张三", "MT0001", "默认考勤组", "旷工", "", "", "", ""])
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[4]:
        cell.fill = _SAMPLE_FILL
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "系统固定读取第 3 行作为日期行、第 4 行起作为员工数据。",
            "需要排除已补流程旷工时，请在对应日期单元格批注中写入\"已补流程\"。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_roster_template() -> bytes:
    return _build_table_template(
        "花名册",
        _FIN_ROSTER_COLUMNS,
        [
            ("MT0001", "张三", "示例科技有限公司", "总部", "产品技术部", "后端组", "工程师",
             "正式员工", "全职", "2025-08-01", "", "2025-11-01"),
            ("WB0001", "李四", "示例服务有限公司", "总部", "运营管理中心", "客服组", "客服专员",
             "外包", "外包", "2025-09-01", "2026-06-20", ""),
        ],
        [
            "在职花名册和离职花名册可使用同一模板；系统会在前 10 行内查找包含\"工号\"和\"姓名\"的表头行。",
            "工号和姓名至少保留一项；合同主体、部门、岗位、员工类型、人员分类、入职/离职/转正日期用于最终表展示和口径判断。",
            "合同主体为空的行会被跳过，不会出现在最终表。",
            "员工类型、人员分类或岗位包含\"兼职\"\"实习\"的人员会被剔除，不会进入最终表。",
            "有离职日期的人员仅保留作息表月份的上月、本月、下月；更早或更晚离职的人员不会进入最终表。",
            "部门字段支持一级/二级/三级部门，也兼容 1级/2级/3级部门；如果只有\"部门\"路径，系统会按\"-\"或\"/\"拆分。",
        ],
    )


def build_transfer_template() -> bytes:
    return _build_table_template(
        "异动流程表",
        _FIN_TRANSFER_COLUMNS,
        [
            ("MT0001", "张三", "HR0001", "代发起人", "2026-06-15", "部门调整", "2026-06-15"),
            ("WB0001", "李四", "HR0001", "代发起人", "2026-06-10", "岗位调整", "2026-06-10"),
        ],
        [
            "异动流程表为可选；系统会在前 10 行内查找实际申请人/实际异动人员和异动日期字段。",
            "人员字段优先使用：实际申请人工号、实际申请人、实际异动人员；仅缺失时才回退发起人工号/发起人姓名。",
            "同一实际异动人员多条异动记录时，最终表会采用最近一次异动日期。",
        ],
    )


def build_final_leave_detail_template() -> bytes:
    return _build_table_template(
        "请假明细表",
        _FIN_LEAVE_DETAIL_COLUMNS,
        [("MT0001", "张三", "总部", "产品技术部", "后端组", "年假",
          "2026-06-03 09:00:00", "2026-06-03 18:30:00", "8小时", "8小时", 1, "")],
        [
            "最终表模块通常上传\"请假明细计算\"生成的请假明细表.xlsx。",
            "必需字段为请假类型和最终请假天数；建议保留工号/姓名、开始时间、结束时间，便于离职计薪口径拆分。",
            "支持多 sheet，系统会自动读取包含发起人工号/工号、姓名、请假类型等字段的明细表。",
        ],
    )


def build_final_overtime_detail_template() -> bytes:
    return _build_table_template(
        "加班明细表",
        _FIN_OVERTIME_DETAIL_COLUMNS,
        [
            ("MT0001", "张三", "总部", "运营管理中心", "运营支撑部",
             "2026-06-19 09:00:00", "2026-06-19 18:30:00",
             "2026-06-19", "2026-06-19", "", 8, 0, 1),
            ("MT0002", "王五", "总部", "产品技术部", "前端组",
             "2026-06-10 19:00:00", "2026-06-10 22:00:00",
             "2026-06-10", "2026-06-10", "", "", "", ""),
        ],
        [
            "最终表模块通常上传\"加班明细回填\"生成的加班明细_回填.xlsx。",
            "最终表优先汇总 2倍/3倍加班小时，并按 8 小时/天派生展示天数；旧模板只有天数字段时仍兼容。",
            "开始/结束时间、明细或加班时间用于判断是否属于作息表月份；无法识别日期的行不会计入本月汇总。",
        ],
    )


def build_parttime_monthly_summary_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "月度汇总"
    ws["A1"] = "2026年6月腾小宝月度汇总"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["姓名", "工号", "考勤组", "部门", "职位", "出勤天数", "考勤结果",
              None, None, None, None, None, None])
    ws.append([None, None, None, None, None, None, 1, 2, 3, 4, 5, 6, 7])
    ws.append([
        "张三", "JZ0001", "兼职考勤组", "上海景区兼职组", "兼职", 5,
        "正常(09:00-18:00)", "休息", "迟到10分钟(09:10-18:00)",
        "请假4小时(09:00-13:00)", "缺卡(09:00)", "外勤(09:00-18:00)", "休息日加班5小时(09:00-14:00)",
    ])
    for cell in ws[2]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
    for cell in ws[3]:
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[4]:
        cell.fill = _SAMPLE_FILL
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "A1 或文件名需要包含年月，例如：2026年6月腾小宝月度汇总。",
            "表头行需要包含\"姓名\"\"出勤天数\"\"考勤结果\"；下一行填写日期序号，员工数据从日期行下一行开始。",
            "考勤组、部门、工号、职位用于筛选兼职/外包/实习等人员；支持多份月度汇总同时上传。",
        ],
    )
    return _workbook_to_bytes(wb)


def build_parttime_schedule_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "排班表"
    ws["A1"] = "2026年6月兼职排班表"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["姓名", "所属公司", 1, 2, 3, 4, 5, 6, 7])
    ws.append(["张三", "示例服务有限公司", "班", "休", "班", "班", "班", "", "班"])
    ws.append(["李四", "示例文旅有限公司", "1", "", "班", "休", "班", "班", ""])
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.fill = _SAMPLE_FILL
            cell.alignment = Alignment(horizontal="center")
    _autosize_worksheet(ws)
    _notes_sheet(
        wb,
        [
            "A1 或文件名需要包含年月，例如：2026年6月兼职排班表。",
            "前 10 行内需要有一行同时包含\"姓名\"和至少 3 个日期列；所属公司可选，用于同名人员或多主体拆分。",
            "排班日期下填写\"班\"、数字、√、是等会计 1 天；空白、0、/、-、休不会计出勤。",
        ],
    )
    return _workbook_to_bytes(wb)


# ── Public summary: the full list of templates the toolbox UI can download ─────

TEMPLATE_REGISTRY: list[tuple[str, str, callable]] = [
    ("leave_export", "请假系统导出表模板.xlsx", build_leave_export_template),
    ("offsite_duration", "异地不打卡人员请假时长表模板.xlsx", build_offsite_duration_template),
    ("schedule", "作息表模板.xlsx", build_schedule_template),
    ("overtime_export", "加班系统导出表模板.xlsx", build_overtime_export_template),
    ("overtime_schedule", "加班排班表模板.xlsx", build_overtime_schedule_template),
    ("overtime_attendance", "加班考勤打卡明细表模板.xlsx", build_overtime_attendance_template),
    ("subsidy_source", "补贴扣款表模板.xlsx", build_subsidy_source_template),
    ("activity_checkin", "活动签到表模板.xlsx", build_activity_checkin_template),
    ("subsidy_attendance", "考勤表模板.xlsx", build_subsidy_attendance_template),
    ("attendance_result", "考勤结果表模板.xlsx", build_attendance_result_template),
    ("roster", "花名册模板.xlsx", build_roster_template),
    ("transfer", "异动流程表模板.xlsx", build_transfer_template),
    ("final_leave_detail", "请假明细表模板.xlsx", build_final_leave_detail_template),
    ("final_overtime_detail", "加班明细表模板.xlsx", build_final_overtime_detail_template),
    ("parttime_monthly_summary", "腾小宝月度汇总模板.xlsx", build_parttime_monthly_summary_template),
    ("parttime_schedule", "兼职排班表模板.xlsx", build_parttime_schedule_template),
]

TEMPLATE_BY_ID: dict[str, tuple[str, callable]] = {
    template_id: (filename, builder)
    for template_id, filename, builder in TEMPLATE_REGISTRY
}
