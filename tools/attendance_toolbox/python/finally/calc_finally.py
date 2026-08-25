"""
calc_finally.py
──────────────────────────────────────────────────────────────────────────────
功能：
  汇总花名册、异动流程、作息表、请假明细、加班明细、补贴扣款表的数据，
  生成最终考勤汇总表。

输入：
  1. 在职花名册.xlsx + 离职花名册.xlsx（可选）→ 最终表员工名单
  2. 异动流程表.xlsx（可选）→ 异动日期
  3. 作息表.xlsx → 应出勤天数 / 法定节假日 / 公司福利假
  4. 请假明细表.xlsx → 各类请假天数
  5. 加班明细_回填.xlsx → 2倍/3倍加班小时，派生展示天数（兼容旧天数字段）
  6. 补贴扣款表_核对.xlsx → 旷工天数

输出：
  最终表.xlsx — 包含所有员工的考勤汇总
──────────────────────────────────────────────────────────────────────────────
"""
from __future__ import annotations

import calendar
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 确保可以导入兄弟模块 ──────────────────────────────────────────────────
_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)
_leave_path = os.path.join(_BASE, "leave")
if _leave_path not in sys.path:
    sys.path.insert(0, _leave_path)

import calc_leave  # noqa: E402  # 复用作息表解析
from excel_compat import load_workbook_compat  # noqa: E402

# ── 输出列定义 ────────────────────────────────────────────────────────────
OUTPUT_HEADERS = [
    "序号", "月份", "工号", "姓名", "考勤组", "合同主体",
    "一级部门", "二级部门", "三级部门", "岗位",
    "员工类型", "人员分类",
    "入职日期", "离职日期", "转正日期", "异动日期",
    "基本工资", "加班工资", "绩效工资", "季度津贴",
    "调薪工资", "管理工资", "调薪绩效额", "工资总额", "缺勤扣款",
    "应出勤\n天数", "计薪\n出勤天数", "实际\n出勤天数",
    "法定\n节假日", "公司\n福利假",
    "年假", "调休", "婚假", "陪产假", "丧假", "产检假", "工伤假", "产假",
    "病假", "事假",
    "旷工天数",
    "入/离职\n缺勤天数", "当月转正天数",
    "2倍加班（小时）", "3倍加班（小时）",
    "2倍加班（天）", "3倍加班（天）",
    "开户行", "发放银行", "银行卡号", "身份证号码", "手机号", "备注",
]

RETAINED_OUTPUT_HEADERS = {
    "基本工资", "加班工资", "绩效工资", "季度津贴",
    "调薪工资", "管理工资", "调薪绩效额", "工资总额", "缺勤扣款",
    "开户行", "发放银行", "银行卡号", "身份证号码", "手机号", "备注",
}

MANUAL_SOURCE_OUTPUT_HEADERS = {
    "员工类型", "人员分类", "离职日期", "异动日期",
}

LEAVE_TYPES = [
    "年假", "调休", "婚假", "陪产假", "丧假",
    "产检假", "工伤假", "产假", "病假", "事假",
]

STATUTORY_HOLIDAY_COLOR = "FFFF0000"
COMPANY_WELFARE_COLOR = "FF0070C0"
UNPAID_LEAVE_TYPES = {"事假"}
DEFAULT_CHENGDU_WORK_LOCATION_NAMES = calc_leave.DEFAULT_CHENGDU_WORK_LOCATION_NAMES
FIXED_MONTHLY_REST_DAYS = 4
FIXED_MONTHLY_REST_NAMES = {
    calc_leave.normalize_employee_name(name) or name
    for name in ("杨艳群", "杨永菊")
}
REST_PREMIUM_EXCLUDED_NAMES = {"陈星雨"}
REST_PREMIUM_EXCLUDED_CODES = {"MT0019"}
STANDARD_HOURS_PER_DAY = 8.0
OVERTIME_HOUR_OUTPUT_NUMBER_FORMAT = "0.00"
OVERTIME_DAY_OUTPUT_NUMBER_FORMAT = "0.00"


# ══════════════════════════════════════════════════════════════════════════
#  工具函数
# ══════════════════════════════════════════════════════════════════════════

def _normalize_emp_no(val) -> str:
    """清洗工号：去 .0 / 空白 / 统一大写"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper()


def _field_key(name) -> str:
    """字段匹配时忽略模板里的换行、空白和常见不可见字符。"""
    text = str(name or "")
    return re.sub(r"[\s\u00a0\u200b-\u200f\ufeff]+", "", text).lower()


def _clean_name(val) -> str:
    """清洗姓名：去除离职后缀"""
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"[（(]已?离职[）)]", "", s).strip()
    return s


def _normalize_name_key(val) -> str:
    """姓名匹配键：清理离职后缀并忽略首尾及内部空白。"""
    return re.sub(r"\s+", "", _clean_name(val))


def _is_rest_premium_excluded(emp_no: str, name: str) -> bool:
    """不汇总调休休息日 2 倍加班工资的人员例外。"""
    return (
        _normalize_emp_no(emp_no) in REST_PREMIUM_EXCLUDED_CODES
        or _clean_name(name) in REST_PREMIUM_EXCLUDED_NAMES
    )


def _build_employee_lookup(employees: list[dict] | tuple[dict, ...] | None) -> tuple[dict[str, dict], dict[str, dict]]:
    by_emp_no: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    for emp in employees or ():
        emp_no = _normalize_emp_no(emp.get("emp_no"))
        name = _clean_name(emp.get("name"))
        if emp_no and emp_no not in by_emp_no:
            by_emp_no[emp_no] = emp
        if name and name not in by_name:
            by_name[name] = emp
    return by_emp_no, by_name


def _lookup_employee(
    emp_no: str,
    name: str,
    employee_lookup: tuple[dict[str, dict], dict[str, dict]] | None,
) -> dict | None:
    if not employee_lookup:
        return None
    by_emp_no, by_name = employee_lookup
    normalized_emp_no = _normalize_emp_no(emp_no)
    cleaned_name = _clean_name(name)
    if normalized_emp_no and normalized_emp_no in by_emp_no:
        return by_emp_no[normalized_emp_no]
    if cleaned_name and cleaned_name in by_name:
        return by_name[cleaned_name]
    return None


def _is_operations_support_employee(emp: dict | None) -> bool:
    if not emp:
        return False
    values = [
        emp.get("dept1"),
        emp.get("dept2"),
        emp.get("dept3"),
        emp.get("department"),
        emp.get("department_path"),
        emp.get("attendance_group"),
    ]
    extra_depts = emp.get("extra_depts")
    if isinstance(extra_depts, (list, tuple, set)):
        values.extend(extra_depts)
    elif extra_depts:
        values.append(extra_depts)
    return any("运营支撑部" in str(value or "") for value in values)


def _holiday_adjust_rest_days_for_employee(schedule_ctx: dict | None, emp: dict | None) -> set[date]:
    if not schedule_ctx or not emp:
        return set()
    is_chengdu = _is_chengdu(
        emp.get("dept1"),
        emp.get("dept2"),
        emp.get("dept3"),
        emp.get("name"),
    )
    selected_key = "chengdu_holiday_adjust_rest_days" if is_chengdu else "main_holiday_adjust_rest_days"
    selected_days = set(schedule_ctx.get(selected_key) or set())
    if selected_days:
        return selected_days
    return set(schedule_ctx.get("main_holiday_adjust_rest_days") or set()) | set(
        schedule_ctx.get("chengdu_holiday_adjust_rest_days") or set()
    )


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y.%m.%d",
                "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_datetime(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    s = str(val).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y年%m月%d日",
        "%Y.%m.%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    try:
        return float(s)
    except ValueError:
        return None


def _is_nonempty(val) -> bool:
    return val not in (None, "")


def _first_text(*values) -> str | None:
    for val in values:
        if val is None:
            continue
        text = str(val).strip()
        if text:
            return text
    return None


def _clean_final_status_label(val) -> str | None:
    """最终表状态展示去掉括号及括号内说明。"""
    if val is None:
        return None
    text = str(val).strip()
    if not text:
        return None

    previous = None
    while previous != text:
        previous = text
        text = re.sub(r"[（(][^（）()]*[）)]", "", text).strip()
    return re.sub(r"\s+", " ", text) or None


def _split_department_text(text, contract_entity=None) -> tuple[str | None, str | None, str | None]:
    """Split a single department path into up to three department levels."""
    if text is None:
        return None, None, None
    raw = str(text).strip()
    if not raw:
        return None, None, None

    parts = [
        part.strip()
        for part in re.split(r"\s*(?:-|/|／|\\|>|＞|》|→)\s*", raw)
        if part and part.strip()
    ]
    if not parts:
        return None, None, None

    contract_text = str(contract_entity or "").strip()
    if contract_text and len(parts) > 1 and parts[0] == contract_text:
        parts = parts[1:]

    return (
        parts[0] if len(parts) > 0 else None,
        parts[1] if len(parts) > 1 else None,
        parts[2] if len(parts) > 2 else None,
    )


def _month_bounds(year: int, month: int) -> tuple[date, date]:
    _, days_in_month = calendar.monthrange(year, month)
    return date(year, month, 1), date(year, month, days_in_month)


def _shift_year_month(year: int, month: int, offset: int) -> tuple[int, int]:
    month_index = (year * 12 + month - 1) + offset
    return month_index // 12, month_index % 12 + 1


def _in_resign_keep_window(
    resign_date: date | None,
    year: int,
    month: int,
    hire_date: date | None = None,
) -> bool:
    """Keep employees who overlapped the target month (plus a one-month edge).

    Historical stats must not depend on whether the person is still active today.
    Rule:
      - no resign date → always keep
      - still employed during the target month (resign_date >= month_start) → keep
      - resigned in previous/current/next month of the target → keep (edge buffer
        for late payroll adjustments)
      - otherwise drop (left long before the target month)
    """
    if not resign_date:
        return True
    month_start, month_end = _month_bounds(year, month)
    # Active for any day of the target month.
    if resign_date >= month_start:
        if hire_date and hire_date > month_end:
            return False
        return True
    # One-month lookback buffer after resignation (legacy ±1 month behavior).
    for offset in (-1, 0, 1):
        window_year, window_month = _shift_year_month(year, month, offset)
        w_start, w_end = _month_bounds(window_year, window_month)
        if w_start <= resign_date <= w_end:
            return True
    return False


def _in_month(day: date | None, month_start: date, month_end: date) -> bool:
    return bool(day and month_start <= day <= month_end)


def calc_statutory_holiday_days(
    hire_date: date | None,
    resign_date: date | None,
    statutory_holidays: set[date],
    month_start: date,
    month_end: date,
) -> int:
    """Count paid statutory holidays inside the employee active window.

    The hire date is inclusive; the resignation date is exclusive.
    """
    if not statutory_holidays:
        return 0

    active_start = max(month_start, hire_date) if hire_date else month_start
    active_end = month_end
    if resign_date:
        active_end = min(month_end, resign_date - timedelta(days=1))
    if active_start > active_end:
        return 0

    return sum(1 for day in statutory_holidays if active_start <= day <= active_end)


def calc_maternity_statutory_overlap_days(
    leave_day_detail: dict[date, dict[str, float]] | None,
    statutory_holidays: set[date],
    hire_date: date | None,
    resign_date: date | None,
    month_start: date,
    month_end: date,
) -> int:
    """统计本月在职区间内同时落在产假的法定节假日，避免重复计薪。"""
    if not leave_day_detail or not statutory_holidays:
        return 0
    active_start = max(month_start, hire_date) if hire_date else month_start
    active_end = (
        min(month_end, resign_date - timedelta(days=1))
        if resign_date else month_end
    )
    if active_start > active_end:
        return 0
    return sum(
        1
        for day in statutory_holidays
        if active_start <= day <= active_end
        and float((leave_day_detail.get(day) or {}).get("产假") or 0) > 0
    )


def calc_active_day_count(
    days: set[date],
    hire_date: date | None,
    resign_date: date | None,
    month_start: date,
    month_end: date,
) -> int:
    """统计日期集合中落在员工本月在职区间内的天数。"""
    if not days:
        return 0

    active_start = max(month_start, hire_date) if hire_date else month_start
    active_end = min(month_end, resign_date) if resign_date else month_end
    if active_start > active_end:
        return 0

    return sum(1 for day in days if active_start <= day <= active_end)


def _derive_emp_type(emp: dict, month_end: date) -> str | None:
    source_text = " ".join(
        str(emp.get(key) or "")
        for key in ("emp_type", "type_hint", "contract_entity", "position", "dept1", "dept2", "dept3")
    )
    is_labor_dispatch = any(word in source_text for word in ("劳务", "派遣", "外包"))
    confirm_date = emp.get("confirm_date")
    is_probation = not confirm_date or confirm_date > month_end

    if is_labor_dispatch and is_probation:
        return "劳务派遣（试用期）"
    if is_labor_dispatch:
        return "劳务派遣"
    if is_probation:
        return "试用期"
    return "全职"


def _derive_category(emp: dict, month_start: date, month_end: date) -> str | None:
    if _is_nonempty(emp.get("category")):
        return emp.get("category")

    hire_date = emp.get("hire_date")
    resign_date = emp.get("resign_date")
    confirm_date = emp.get("confirm_date")

    hired_this_month = _in_month(hire_date, month_start, month_end)
    resigned_this_month = _in_month(resign_date, month_start, month_end)
    if hired_this_month and resigned_this_month:
        return "当月入职后离职"
    if resigned_this_month:
        return "当月离职"
    if hired_this_month:
        return "当月入职"
    if _in_month(confirm_date, month_start, month_end):
        return "当月转正"
    if (hire_date is None or hire_date <= month_end) and (
        resign_date is None or resign_date >= month_start
    ):
        return "当月在职"
    return "其他状态（入离职时间非本月）"


def _is_final_table_excluded_employee(emp: dict) -> bool:
    """兼职、实习和劳务外包人员不进入最终考勤汇总表。"""
    source_text = " ".join(
        str(emp.get(key) or "")
        for key in ("emp_type", "category", "position", "type_hint")
    )
    return any(word in source_text for word in ("兼职", "实习", "劳务外包"))


def _same_year_month(day: date | None, year: int, month: int) -> bool:
    """
    判断日期是否属于指定年月。
    注意：对于跨月加班（如 4月30日 23:00 - 5月1日 01:00），
    明细日期可能是 4月30日，但实际应该算在 5月份的加班记录中。
    因此这里只检查 year 和 month，不严格要求完全匹配。
    """
    if not day:
        return False
    # 允许前一个月的最后几天（跨月加班场景）
    if day.year == year and day.month == month:
        return True
    # 检查是否是上个月的最后几天（跨月加班）
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    if day.year == prev_year and day.month == prev_month:
        # 只允许上个月的最后 3 天（覆盖跨月加班场景）
        import calendar
        last_day = calendar.monthrange(prev_year, prev_month)[1]
        if day.day >= last_day - 2:  # 最后 3 天
            return True

    return False


def _resolve_overtime_row_date(
    ws,
    row_idx: int,
    col_detail: int | None,
    col_overtime_date: int | None,
    col_start: int | None,
    col_end: int | None,
) -> date | None:
    detail_date = _to_date(ws.cell(row_idx, col_detail + 1).value) if col_detail is not None else None
    overtime_date = _to_date(ws.cell(row_idx, col_overtime_date + 1).value) if col_overtime_date is not None else None
    start_dt = _to_datetime(ws.cell(row_idx, col_start + 1).value) if col_start is not None else None
    end_dt = _to_datetime(ws.cell(row_idx, col_end + 1).value) if col_end is not None else None

    if detail_date:
        if start_dt and end_dt:
            start_day = start_dt.date()
            end_day = end_dt.date()
            if end_day < start_day:
                start_day, end_day = end_day, start_day
            if start_day <= detail_date <= end_day:
                return detail_date
        else:
            return detail_date

    if overtime_date:
        return overtime_date
    if start_dt:
        return start_dt.date()
    if end_dt:
        return end_dt.date()
    return None


def _round2(val) -> float | None:
    """四舍五入到 2 位小数"""
    if val is None:
        return None
    return float(Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _overtime_days_from_hours(hours: float | int | None) -> float:
    if not hours:
        return 0.0
    return round(float(hours) / STANDARD_HOURS_PER_DAY, 6)


def _overtime_hours_from_days(days: float | int | None) -> float:
    if not days:
        return 0.0
    return round(float(days) * STANDARD_HOURS_PER_DAY, 6)


def _contains_cjk(val) -> bool:
    if val is None:
        return False
    return bool(re.search(r"[\u4e00-\u9fff]", str(val)))


def _normalize_leave_dedupe_part(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()


def _find_header_row(ws, *keywords, max_rows: int = 10) -> tuple[int | None, list]:
    """
    在前 max_rows 行中查找包含任一关键字的行作为表头。
    返回 (行号 1-based, [单元格值列表])。
    使用子串匹配，保持请假、加班等模块的全局兼容性。
    """
    for r in range(1, min(max_rows, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        text_vals = [_field_key(v) for v in vals]
        keyword_vals = [_field_key(kw) for kw in keywords]
        for kw in keyword_vals:
            if kw and any(kw in t for t in text_vals):
                return r, vals
    return None, []


def _find_col(header_vals: list, *keywords) -> int | None:
    """在表头行值列表中查找包含任一关键字的列索引（0-based）"""
    keyword_vals = [_field_key(kw) for kw in keywords]
    for idx, val in enumerate(header_vals):
        if val is None:
            continue
        text = _field_key(val)
        for kw in keyword_vals:
            if kw and kw in text:
                return idx
    return None


def _find_col_exact(header_vals: list, *keywords) -> int | None:
    """在表头行值列表中查找与任一关键字精确匹配的列索引（0-based）。
    用于花名册身份列（工号/姓名），避免"发起人工号"误匹配"工号"。"""
    keyword_vals = [_field_key(kw) for kw in keywords]
    for idx, val in enumerate(header_vals):
        if val is None:
            continue
        text = _field_key(val)
        for kw in keyword_vals:
            if kw and kw == text:
                return idx
    return None


def _find_roster_header_row(ws, max_rows: int = 10) -> tuple[int | None, list]:
    """查找具有精确姓名列的花名册表头。

    该规则只服务于 ``parse_roster``，不得改变请假、加班等解析器依赖的
    ``_find_header_row`` 子串匹配行为。
    """
    name_headers = ("姓名", "员工姓名")
    for row_idx in range(1, min(max_rows, ws.max_row) + 1):
        header_vals = [
            ws.cell(row_idx, col_idx).value
            for col_idx in range(1, ws.max_column + 1)
        ]
        if _find_col_exact(header_vals, *name_headers) is not None:
            return row_idx, header_vals
    return None, []


def _find_col_excluding(header_vals: list, keywords: tuple[str, ...], excludes: tuple[str, ...]) -> int | None:
    """Find a header column by keyword while skipping launcher/applicant columns."""
    keyword_vals = [_field_key(kw) for kw in keywords]
    exclude_vals = [_field_key(kw) for kw in excludes]
    for idx, val in enumerate(header_vals):
        if val is None:
            continue
        text = _field_key(val)
        if any(ex and ex in text for ex in exclude_vals):
            continue
        if any(kw and kw in text for kw in keyword_vals):
            return idx
    return None


def _find_col_ws(ws, *keywords, max_rows: int = 4) -> int | None:
    """在 worksheet 前 max_rows 行中找包含任一关键字的列号（1-based）"""
    keyword_vals = [_field_key(kw) for kw in keywords]
    for row in range(1, max_rows + 1):
        for col in range(1, ws.max_column + 1):
            val = _field_key(ws.cell(row, col).value)
            if any(k and k in val for k in keyword_vals):
                return col
    return None


def _is_chengdu(
    dept1,
    dept2,
    dept3,
    name: str | None = None,
    special_chengdu_names: set[str] | tuple[str, ...] | None = None,
) -> bool:
    if name:
        normalized_name = calc_leave.normalize_employee_name(name) or ""
        if special_chengdu_names and normalized_name in special_chengdu_names:
            return True
    for d in (dept1, dept2, dept3):
        if d and "成都" in str(d):
            return True
    return False


def _is_fixed_monthly_rest_employee(name: str | None) -> bool:
    normalized_name = calc_leave.normalize_employee_name(name) or ""
    return normalized_name in FIXED_MONTHLY_REST_NAMES


def _fixed_monthly_rest_attendance_days(month_end: date) -> int:
    return max(0, month_end.day - FIXED_MONTHLY_REST_DAYS)


def _parse_schedule_color_days(path: str, target_color: str) -> tuple[set[date], set[date]]:
    """按颜色解析作息表中的日期集合，返回 (主作息, 成都作息)。"""
    wb = openpyxl.load_workbook(path)
    main_days = None
    chengdu_days = None

    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            title = str(ws.cell(row_idx, 1).value or "").strip()
            if "作息时间表" not in title:
                continue

            year, month = calc_leave._parse_title(title)
            if not year or not month:
                continue

            header_row_idx = None
            week_key = _field_key("周数")
            for idx in range(row_idx + 1, min(ws.max_row, row_idx + 8) + 1):
                if _field_key(ws.cell(idx, 1).value) == week_key:
                    header_row_idx = idx
                    break
            if header_row_idx is None:
                continue

            days = set()
            for data_row_idx in range(header_row_idx + 1, ws.max_row + 1):
                first_val = ws.cell(data_row_idx, 1).value
                if isinstance(first_val, (int, float)):
                    for col_idx in range(2, min(8, ws.max_column) + 1):
                        cell = ws.cell(data_row_idx, col_idx)
                        if calc_leave._cell_fg_rgb(cell) != target_color:
                            continue
                        current = calc_leave._fix_date(cell.value, year, month)
                        if current:
                            days.add(current)
                    continue
                if first_val is not None:
                    break

            is_chengdu = "成都" in title or "成都" in ws.title
            if is_chengdu:
                chengdu_days = days
            elif main_days is None:
                main_days = days

    wb.close()

    if main_days is None:
        main_days = set()
    if chengdu_days is None:
        chengdu_days = set(main_days)
    return main_days, chengdu_days


def _parse_schedule_summary_days(path: str, summary_label: str) -> tuple[int | None, int | None]:
    """读取作息表汇总区的天数字段，返回 (主作息, 成都作息)。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    main_value = None
    chengdu_value = None

    for ws in wb.worksheets:
        current_is_chengdu = False
        for row_idx in range(1, ws.max_row + 1):
            title = str(ws.cell(row_idx, 1).value or "").strip()
            if "作息时间表" in title:
                current_is_chengdu = "成都" in title or "成都" in ws.title

            label = str(ws.cell(row_idx, 2).value or "").strip()
            if summary_label not in label:
                continue

            value = _to_float(ws.cell(row_idx, 3).value)
            if value is None:
                continue

            value = int(value) if float(value).is_integer() else value
            if current_is_chengdu:
                chengdu_value = value
            elif main_value is None:
                main_value = value

    wb.close()
    return main_value, chengdu_value


def _collect_deduped_leave_rows(path: str) -> list[dict]:
    """读取并按相同请假时间去重后的请假明细行。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    deduped_rows: dict[tuple, dict] = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        header_row_idx, header_vals = _find_header_row(ws, "发起人工号", "工号", "请假类型")
        if header_row_idx is None:
            continue

        col_emp_no = _find_col(header_vals, "发起人工号", "工号")
        col_name = _find_col(header_vals, "发起人姓名", "姓名")
        col_dept1 = _find_col(header_vals, "一级部门")
        col_dept2 = _find_col(header_vals, "二级部门")
        col_dept3 = _find_col(header_vals, "三级部门")
        col_leave_type = _find_col(header_vals, "请假类型")
        col_start = _find_col(header_vals, "开始时间")
        col_end = _find_col(header_vals, "结束时间")
        col_sys_hours = _find_col(header_vals, "系统时长", "时长", "请假时长")
        col_final_hours = _find_col(
            header_vals,
            "最终请假时长",
            "最终请假时长（小时）",
            "最终请假时长(小时)",
            "最终时长（小时）",
            "最终时长(小时)",
            "最终小时",
        )
        col_days = _find_col(
            header_vals,
            "最终请假天数",
            "最终请假时长（天）",
            "最终请假时长(天)",
            "最终时长（天）",
            "最终时长(天)",
            "最终天数",
        )

        if col_leave_type is None or col_days is None:
            missing = []
            if col_leave_type is None:
                missing.append("请假类型")
            if col_days is None:
                missing.append("最终请假天数/最终时长（天）")
            print(f"[请假明细] {sn} 未找到 {'、'.join(missing)} 列，已跳过")
            continue

        for r in range(header_row_idx + 1, ws.max_row + 1):
            emp_no = _normalize_emp_no(
                ws.cell(r, col_emp_no + 1).value
            ) if col_emp_no is not None else ""
            name = _clean_name(
                ws.cell(r, col_name + 1).value
            ) if col_name is not None else ""

            leave_type = str(ws.cell(r, col_leave_type + 1).value or "").strip()
            raw_days = ws.cell(r, col_days + 1).value
            raw_start = ws.cell(r, col_start + 1).value if col_start is not None else None
            raw_end = ws.cell(r, col_end + 1).value if col_end is not None else None
            key = emp_no or name

            if key and leave_type and raw_start is not None and raw_end is not None:
                dedupe_key = (
                    key,
                    leave_type,
                    _normalize_leave_dedupe_part(raw_start),
                    _normalize_leave_dedupe_part(raw_end),
                )
            else:
                dedupe_key = ("__row__", sn, r)

            current_row = {
                "emp_no": emp_no,
                "name": name,
                "dept1": ws.cell(r, col_dept1 + 1).value if col_dept1 is not None else None,
                "dept2": ws.cell(r, col_dept2 + 1).value if col_dept2 is not None else None,
                "dept3": ws.cell(r, col_dept3 + 1).value if col_dept3 is not None else None,
                "leave_type": leave_type,
                "raw_start": raw_start,
                "raw_end": raw_end,
                "sys_hours": ws.cell(r, col_sys_hours + 1).value if col_sys_hours is not None else None,
                "final_hours": (
                    ws.cell(r, col_final_hours + 1).value
                    if col_final_hours is not None else None
                ),
                "raw_days": raw_days,
            }

            existing_row = deduped_rows.get(dedupe_key)
            if existing_row is None:
                deduped_rows[dedupe_key] = current_row
                continue

            existing_has_cjk = _contains_cjk(existing_row["raw_days"])
            current_has_cjk = _contains_cjk(raw_days)
            if existing_has_cjk and not current_has_cjk:
                deduped_rows[dedupe_key] = current_row

    wb.close()
    return list(deduped_rows.values())


def _normalize_paid_statutory_holidays(days: set[date]) -> set[date]:
    """
    将红底假期日期收敛为法定计薪日。
    具体节假日规则复用请假模块，避免最终表和作息表口径分叉。
    """
    return calc_leave._normalize_paid_statutory_holidays(days)


# ══════════════════════════════════════════════════════════════════════════
#  1. 花名册解析
# ══════════════════════════════════════════════════════════════════════════

def parse_roster(
    active_path: str,
    resigned_path: str | None = None,
) -> list[dict]:
    """
    解析在职花名册 + 离职花名册，返回员工列表。
    每个员工 dict 键：emp_no, name, contract_entity, dept1, dept2, dept3,
                     position, emp_type, category, hire_date, resign_date, confirm_date

    字段契约（输入端与解析端必须一致）：
      - 姓名列识别：精确的姓名 / 员工姓名（必需）
      - 工号及其他旧版扩展字段均为可选，仅用于兼容和匹配。
      - 在职源无可用表头或有效员工为 0 时，必须中止并返回明确中文错误。
    """
    employees: list[dict] = []
    paths: list[tuple[str, bool]] = [(active_path, False)]
    if resigned_path:
        paths.append((resigned_path, True))

    seen_emp_nos: set[str] = set()
    seen_name_indexes: dict[str, list[int]] = defaultdict(list)

    for path, is_resigned_roster in paths:
        wb = load_workbook_compat(path, data_only=True)
        matched_sheets = 0
        file_valid_employees = 0
        file_skipped_blank_row = 0
        file_skipped_duplicate = 0
        file_missing_contract = 0
        file_missing_emp_no = 0
        file_source_rows = 0

        for ws in wb.worksheets:
            # 花名册局部使用精确身份列表头，禁止把审批流程的发起人/申请人字段当成员工身份列。
            header_row_idx, header_vals = _find_roster_header_row(ws)
            if header_row_idx is None:
                continue

            # 列映射。兼容钉钉员工导出的 1级部门/2级部门/3级部门 命名。
            # 工号和姓名使用精确匹配，避免"发起人工号"等流程字段误匹配。
            col_emp_no = _find_col_exact(header_vals, "员工编号", "员工工号", "工号")
            col_name = _find_col_exact(header_vals, "员工姓名", "姓名")
            col_contract = _find_col(header_vals, "合同主体", "所属公司", "公司主体", "主体")
            col_dept1 = _find_col(header_vals, "一级部门", "1级部门", "一级组织", "1级组织")
            col_dept2 = _find_col(header_vals, "二级部门", "2级部门", "二级组织", "2级组织")
            col_dept3 = _find_col(header_vals, "三级部门", "3级部门", "三级组织", "3级组织")
            col_dept4 = _find_col(header_vals, "四级部门", "4级部门", "四级组织", "4级组织")
            col_dept5 = _find_col(header_vals, "五级部门", "5级部门", "五级组织", "5级组织")
            col_dept6 = _find_col(header_vals, "六级部门", "6级部门", "六级组织", "6级组织")
            col_dept_single = _find_col(
                header_vals, "部门路径", "完整部门", "主部门", "所属部门", "部门"
            )
            if col_dept_single in {col_dept1, col_dept2, col_dept3}:
                col_dept_single = None
            col_position = _find_col(header_vals, "岗位名称", "岗位", "职位")
            col_emp_type = _find_col(
                header_vals, "员工类型", "用工类型", "员工性质", "人员类型", "员工类别", "用工性质"
            )
            col_category = _find_col(header_vals, "人员分类", "人员类别", "员工状态", "人员状态")
            col_hire = _find_col(header_vals, "入职日期", "入职时间")
            col_resign = _find_col(header_vals, "实际离职日期", "离职日期", "离职时间", "最后工作日")
            col_confirm = _find_col(header_vals, "实际转正日期", "转正日期", "转正时间")

            if col_name is None:
                print(f"[花名册] 未找到姓名列: {os.path.basename(path)} / {ws.title}")
                continue

            matched_sheets += 1
            source_label = f"{os.path.basename(path)} / {ws.title}"
            print(
                "[花名册] 部门列识别 "
                f"{source_label}: "
                f"一级={col_dept1 + 1 if col_dept1 is not None else '-'}, "
                f"二级={col_dept2 + 1 if col_dept2 is not None else '-'}, "
                f"三级={col_dept3 + 1 if col_dept3 is not None else '-'}, "
                f"部门路径={col_dept_single + 1 if col_dept_single is not None else '-'}"
            )
            print(
                "[花名册] 其他列识别 "
                f"{source_label}: "
                f"工号={col_emp_no + 1 if col_emp_no is not None else '-'}, "
                f"岗位={col_position + 1 if col_position is not None else '-'}, "
                f"员工类型={col_emp_type + 1 if col_emp_type is not None else '-'}, "
                f"人员分类={col_category + 1 if col_category is not None else '-'}, "
                f"入职日期={col_hire + 1 if col_hire is not None else '-'}, "
                f"离职日期={col_resign + 1 if col_resign is not None else '-'}, "
                f"转正日期={col_confirm + 1 if col_confirm is not None else '-'}"
            )
            if col_emp_type is None or col_category is None:
                missing = []
                if col_emp_type is None:
                    missing.append("员工类型")
                if col_category is None:
                    missing.append("人员分类")
                print(f"[花名册] {source_label} 缺少 {'、'.join(missing)}，最终表对应字段保持为空")

            def _cell(row_idx: int, col_idx: int | None):
                if col_idx is None:
                    return None
                return ws.cell(row_idx, col_idx + 1).value

            dept2_nonempty = 0
            dept3_nonempty = 0
            optional_nonempty = {
                "岗位": 0,
                "员工类型": 0,
                "人员分类": 0,
                "入职日期": 0,
                "离职日期": 0,
                "转正日期": 0,
            }

            for r in range(header_row_idx + 1, ws.max_row + 1):
                emp_no = _normalize_emp_no(_cell(r, col_emp_no))
                name = _clean_name(_cell(r, col_name))

                # 花名册只以姓名限定人员；无有效姓名的行一律跳过。
                if not name:
                    file_skipped_blank_row += 1
                    continue

                file_source_rows += 1

                # 同工号已存在：离职花名册只更新离职日期，不重复添加
                if emp_no and emp_no in seen_emp_nos:
                    file_skipped_duplicate += 1
                    if is_resigned_roster:
                        resign_val = _to_date(_cell(r, col_resign))
                        if resign_val:
                            for e in employees:
                                if e["emp_no"] == emp_no:
                                    e["resign_date"] = resign_val
                                    break
                    continue

                name_key = _normalize_name_key(name)
                existing_indexes = seen_name_indexes.get(name_key, [])
                existing_without_emp_no = any(not employees[index]["emp_no"] for index in existing_indexes)
                if existing_indexes and (not emp_no or existing_without_emp_no):
                    raise ValueError(
                        f"花名册存在重复姓名“{name}”（至少 {len(existing_indexes) + 1} 条），"
                        "且无法全部通过工号唯一识别，请处理重名后重试。"
                    )

                contract_entity = _first_text(_cell(r, col_contract))
                if not contract_entity:
                    file_missing_contract += 1
                if not emp_no:
                    file_missing_emp_no += 1

                # 部门解析
                dept1 = _first_text(_cell(r, col_dept1))
                dept2 = _first_text(_cell(r, col_dept2))
                dept3 = _first_text(_cell(r, col_dept3))
                if dept1 is None and col_dept_single is not None:
                    dept1, dept2, dept3 = _split_department_text(
                        _cell(r, col_dept_single), contract_entity
                    )
                elif dept1 and not dept2 and not dept3:
                    split_dept1, split_dept2, split_dept3 = _split_department_text(
                        dept1, contract_entity
                    )
                    dept1 = split_dept1 or dept1
                    dept2 = split_dept2
                    dept3 = split_dept3
                if dept2:
                    dept2_nonempty += 1
                if dept3:
                    dept3_nonempty += 1

                emp = {
                    "emp_no":          emp_no,
                    "name":            name,
                    "contract_entity": contract_entity,
                    "dept1":           dept1,
                    "dept2":           dept2,
                    "dept3":           dept3,
                    "position":        _first_text(_cell(r, col_position)),
                    "emp_type":        _first_text(_cell(r, col_emp_type)),
                    "type_hint":       " ".join(
                        text for text in (
                            _first_text(_cell(r, col_dept4)),
                            _first_text(_cell(r, col_dept5)),
                            _first_text(_cell(r, col_dept6)),
                        )
                        if text
                    ),
                    "category":        _first_text(_cell(r, col_category)),
                    "hire_date":       _to_date(_cell(r, col_hire)),
                    "resign_date":     _to_date(_cell(r, col_resign)),
                    "confirm_date":    _to_date(_cell(r, col_confirm)),
                }
                if emp["position"]:
                    optional_nonempty["岗位"] += 1
                if emp["emp_type"]:
                    optional_nonempty["员工类型"] += 1
                if emp["category"]:
                    optional_nonempty["人员分类"] += 1
                if emp["hire_date"]:
                    optional_nonempty["入职日期"] += 1
                if emp["resign_date"]:
                    optional_nonempty["离职日期"] += 1
                if emp["confirm_date"]:
                    optional_nonempty["转正日期"] += 1
                employees.append(emp)
                seen_name_indexes[name_key].append(len(employees) - 1)
                file_valid_employees += 1
                if emp_no:
                    seen_emp_nos.add(emp_no)

            print(
                f"[花名册] {source_label} 二级部门非空 {dept2_nonempty} 条，"
                f"三级部门非空 {dept3_nonempty} 条"
            )
            print(
                f"[花名册] {source_label} 其他字段非空："
                + "，".join(f"{name} {count} 条" for name, count in optional_nonempty.items())
            )

        wb.close()

        label = "离职" if is_resigned_roster else "在职"
        print(
            f"[花名册] {label}源 {os.path.basename(path)}: "
            f"匹配工作表 {matched_sheets}，"
            f"源数据 {file_source_rows} 行，"
            f"有效员工 {file_valid_employees} 人，"
            f"空行跳过 {file_skipped_blank_row}，"
            f"重复跳过 {file_skipped_duplicate}，"
            f"缺工号 {file_missing_emp_no}，"
            f"缺合同主体 {file_missing_contract}"
        )

        # 零在职保护（仅对在职源生效）：
        # 只要有效在职员工为 0，或未识别到可用花名册表头，就必须中止。
        if not is_resigned_roster:
            if matched_sheets == 0:
                raise ValueError(
                    f"在职花名册未找到可用表头：文件 {os.path.basename(path)} 的所有工作表中"
                    f"均未找到精确表头\"姓名\"或\"员工姓名\"。"
                    f"请确认上传的是花名册而非其他类型文件。"
                )
            if file_valid_employees == 0:
                raise ValueError(
                    f"在职花名册解析为 0 人：文件 {os.path.basename(path)} 有 {file_source_rows} 行数据，"
                    f"但有效姓名为 0（"
                    f"重复跳过 {file_skipped_duplicate}，"
                    f"缺合同主体 {file_missing_contract}）。"
                    f"请检查花名册是否包含精确表头 姓名/员工姓名，以及至少一条非空姓名。"
                )

    active_count = sum(1 for e in employees if not e.get("resign_date"))
    resigned_count = len(employees) - active_count
    print(f"[花名册] 汇总: 合并去重后 {len(employees)} 人（在职 {active_count}，离职 {resigned_count}）")
    return employees


def parse_attendance_identity(path: str) -> list[dict]:
    """从钉钉月度考勤/补贴核对业务表提取员工身份与组织字段。"""
    wb = load_workbook_compat(path, data_only=True)
    records: list[dict] = []
    matched_sheets = 0
    try:
        for ws in wb.worksheets:
            header_row_idx, header_vals = _find_roster_header_row(ws)
            if header_row_idx is None:
                continue
            col_name = _find_col_exact(header_vals, "姓名", "员工姓名")
            col_emp_no = _find_col_exact(header_vals, "工号", "员工工号", "员工编号")
            col_group = _find_col_exact(header_vals, "考勤组", "考勤组名称")
            col_dept1 = _find_col_exact(header_vals, "一级部门", "1级部门", "一级组织", "1级组织")
            col_dept2 = _find_col_exact(header_vals, "二级部门", "2级部门", "二级组织", "2级组织")
            col_dept3 = _find_col_exact(header_vals, "三级部门", "3级部门", "三级组织", "3级组织")
            col_department = _find_col_exact(header_vals, "部门", "所属部门", "部门名称", "部门路径")
            col_position = _find_col_exact(header_vals, "岗位", "岗位名称", "职位")
            if col_name is None:
                continue
            matched_sheets += 1

            def _cell(row_idx: int, col_idx: int | None):
                return ws.cell(row_idx, col_idx + 1).value if col_idx is not None else None

            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                name = _clean_name(_cell(row_idx, col_name))
                if not name or name in {"姓名", "合计"}:
                    continue
                dept1 = _first_text(_cell(row_idx, col_dept1))
                dept2 = _first_text(_cell(row_idx, col_dept2))
                dept3 = _first_text(_cell(row_idx, col_dept3))
                if not any((dept1, dept2, dept3)) and col_department is not None:
                    dept1, dept2, dept3 = _split_department_text(_cell(row_idx, col_department))
                records.append({
                    "emp_no": _normalize_emp_no(_cell(row_idx, col_emp_no)),
                    "name": name,
                    "attendance_group": _first_text(_cell(row_idx, col_group)),
                    "dept1": dept1,
                    "dept2": dept2,
                    "dept3": dept3,
                    "position": _first_text(_cell(row_idx, col_position)),
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                })
    finally:
        wb.close()
    if matched_sheets == 0:
        raise ValueError(
            f"钉钉月度考勤记录未找到精确表头\"姓名\"或\"员工姓名\"：{os.path.basename(path)}"
        )
    if not records:
        raise ValueError(f"钉钉月度考勤记录有效员工为 0：{os.path.basename(path)}")
    return records


def apply_attendance_identity(employees: list[dict], attendance_records: list[dict]) -> list[dict]:
    """用钉钉考勤身份覆盖花名册扩展字段；花名册只保留名单作用。"""
    identity_fields = (
        "attendance_group", "dept1", "dept2", "dept3", "position",
    )

    # 打卡明细可能同一员工有多行；先按业务工号收敛为一个身份记录。
    # 没有工号的行无法证明属于同一员工，保留为独立身份供后续歧义检查。
    identities: list[dict] = []
    by_emp_no: dict[str, dict] = {}
    for record in attendance_records:
        emp_no = _normalize_emp_no(record.get("emp_no"))
        if emp_no:
            identity = by_emp_no.get(emp_no)
            if identity is None:
                identity = dict(record)
                identity["emp_no"] = emp_no
                by_emp_no[emp_no] = identity
                identities.append(identity)
            else:
                if not _clean_name(identity.get("name")):
                    identity["name"] = _clean_name(record.get("name"))
                for field in identity_fields:
                    if not _first_text(identity.get(field)):
                        identity[field] = _first_text(record.get(field))
            continue

        identity = dict(record)
        identity["emp_no"] = ""
        identities.append(identity)

    by_name: dict[str, list[dict]] = defaultdict(list)
    for identity in identities:
        name_key = _normalize_name_key(identity.get("name"))
        if name_key:
            by_name[name_key].append(identity)

    enriched_employees: list[dict] = []
    for employee in employees:
        roster_emp_no = _normalize_emp_no(employee.get("emp_no"))
        roster_name = _clean_name(employee.get("name"))
        matched = by_emp_no.get(roster_emp_no) if roster_emp_no else None
        if matched is not None:
            # 精确工号是最高优先级；即使存在其他同名身份也不得改走姓名匹配。
            candidates = [matched]
        else:
            # 花名册无工号或旧工号未命中时，都必须按规范化姓名回退。
            candidates = by_name.get(_normalize_name_key(roster_name), [])
        if len(candidates) > 1:
            raise ValueError(
                f"钉钉考勤记录中姓名“{roster_name}”匹配到 {len(candidates)} 个员工身份，"
                "无法唯一对应花名册员工。"
            )

        enriched = dict(employee)
        enriched.update({
            "emp_no": "",
            "name": roster_name,
            "attendance_group": None,
            "contract_entity": None,
            "dept1": None,
            "dept2": None,
            "dept3": None,
            "position": None,
            "emp_type": None,
            "type_hint": "",
            "category": None,
            "hire_date": None,
            "resign_date": None,
            "confirm_date": None,
        })
        if candidates:
            record = candidates[0]
            enriched.update({
                "emp_no": _normalize_emp_no(record.get("emp_no")),
                "name": _clean_name(record.get("name")) or roster_name,
                "attendance_group": _first_text(record.get("attendance_group")),
                "dept1": _first_text(record.get("dept1")),
                "dept2": _first_text(record.get("dept2")),
                "dept3": _first_text(record.get("dept3")),
                "position": _first_text(record.get("position")),
            })
        enriched_employees.append(enriched)
    return enriched_employees


# ══════════════════════════════════════════════════════════════════════════
#  2. 异动流程表解析
# ══════════════════════════════════════════════════════════════════════════

def parse_transfer(path: str) -> dict[str, date]:
    """解析异动流程表，返回 {实际异动人员工号/姓名: 最近异动日期}"""
    result: dict[str, date] = {}
    wb = openpyxl.load_workbook(path, data_only=True)
    scanned_sheets = 0
    raw_records = 0
    skipped_records = 0
    fallback_launcher_records = 0

    def _cell(ws, row_idx: int, col_idx: int | None):
        if col_idx is None:
            return None
        return ws.cell(row_idx, col_idx + 1).value

    def _resolve_identity_columns(header_vals: list) -> tuple[int | None, int | None, int | None, int | None]:
        actual_emp_no = _find_col(
            header_vals,
            "实际申请人工号", "实际异动人工号", "实际异动人员工号",
            "被异动人工号", "异动人工号",
        )
        actual_name = _find_col_excluding(
            header_vals,
            ("实际申请人", "实际异动人", "实际异动人员", "被异动人", "异动人员"),
            ("工号", "编号", "id"),
        )
        generic_emp_no = _find_col_excluding(
            header_vals,
            ("员工工号", "异动人工号", "工号"),
            ("发起人", "申请人"),
        )
        generic_name = _find_col_excluding(
            header_vals,
            ("员工姓名", "异动人", "姓名"),
            ("发起人", "申请人", "工号"),
        )
        fallback_emp_no = _find_col(header_vals, "发起人工号")
        fallback_name = _find_col(header_vals, "发起人姓名")
        return (
            actual_emp_no if actual_emp_no is not None else generic_emp_no,
            actual_name if actual_name is not None else generic_name,
            fallback_emp_no,
            fallback_name,
        )

    def _find_transfer_header_row(ws) -> tuple[int | None, list]:
        for r in range(1, min(10, ws.max_row) + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            col_date = _find_col(vals, "异动日期", "异动时间", "生效日期")
            col_emp_no, col_name, fallback_emp_no, fallback_name = _resolve_identity_columns(vals)
            if col_date is not None and any(
                col is not None for col in (col_emp_no, col_name, fallback_emp_no, fallback_name)
            ):
                return r, vals
        return None, []

    def _store_latest(keys: list[str], dt: date) -> None:
        for key in keys:
            if key not in result or dt > result[key]:
                result[key] = dt

    for ws in wb.worksheets:
        header_row_idx, header_vals = _find_transfer_header_row(ws)
        if header_row_idx is None:
            continue

        col_emp_no, col_name, fallback_emp_no, fallback_name = _resolve_identity_columns(header_vals)
        col_date = _find_col(header_vals, "异动日期", "异动时间", "生效日期")
        col_status = _find_col(header_vals, "审批状态", "状态")
        col_result = _find_col(header_vals, "审批结果", "结果")

        if col_date is None or all(
            col is None for col in (col_emp_no, col_name, fallback_emp_no, fallback_name)
        ):
            print(f"[异动表] {ws.title} 未找到实际异动人员列或异动日期列")
            continue

        scanned_sheets += 1
        for r in range(header_row_idx + 1, ws.max_row + 1):
            status_text = str(_cell(ws, r, col_status) or "")
            result_text = str(_cell(ws, r, col_result) or "")
            if (
                "撤销" in status_text
                or "拒绝" in status_text
                or "拒绝" in result_text
                or "不同意" in result_text
            ):
                skipped_records += 1
                continue

            emp_no = _normalize_emp_no(_cell(ws, r, col_emp_no))
            name = _clean_name(_cell(ws, r, col_name))
            used_fallback = False
            if not emp_no and not name:
                emp_no = _normalize_emp_no(_cell(ws, r, fallback_emp_no))
                name = _clean_name(_cell(ws, r, fallback_name))
                used_fallback = bool(emp_no or name)

            dt = _to_date(_cell(ws, r, col_date))
            keys = []
            if emp_no:
                keys.append(emp_no)
            if name:
                keys.append(name)
            if keys and dt:
                raw_records += 1
                if used_fallback:
                    fallback_launcher_records += 1
                _store_latest(keys, dt)

    if scanned_sheets == 0:
        print("[异动表] 未找到表头行")
    print(
        f"[异动表] 扫描 {scanned_sheets} 个 sheet，解析 {raw_records} 条有效记录，"
        f"跳过 {skipped_records} 条撤销/拒绝记录，最终匹配 {len(result)} 个工号/姓名键"
    )
    if fallback_launcher_records:
        print(f"[异动表] {fallback_launcher_records} 条记录未找到实际异动人员，已回退使用发起人")
    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════
#  3. 作息表解析（复用 calc_leave 并扩展节假日统计）
# ══════════════════════════════════════════════════════════════════════════

def parse_schedule(path: str) -> dict:
    """
    解析作息表，返回 dict 包含：
    - year, month, month_start, month_end
    - main_working_days, chengdu_working_days (set[date])
    - main_attendance_days, chengdu_attendance_days (int)
    - holidays (int)：法定节假日+公司福利假合计
    """
    ctx = calc_leave.load_schedule_context(path)

    year  = ctx["year"]
    month = ctx["month"]

    main_wd    = ctx["main_working_days"]
    chengdu_wd = ctx["chengdu_working_days"]
    main_attendance_summary, chengdu_attendance_summary = _parse_schedule_summary_days(
        path, "应出勤天数",
    )
    main_welfare_days, chengdu_welfare_days = _parse_schedule_color_days(
        path, COMPANY_WELFARE_COLOR,
    )
    main_statutory, chengdu_statutory = _parse_schedule_color_days(
        path, STATUTORY_HOLIDAY_COLOR,
    )
    main_statutory_raw = set(main_statutory)
    chengdu_statutory_raw = set(chengdu_statutory)
    main_statutory = _normalize_paid_statutory_holidays(main_statutory)
    chengdu_statutory = _normalize_paid_statutory_holidays(chengdu_statutory)
    main_holiday_adjust_rest_days = (
        main_statutory_raw - set(main_statutory) - set(main_welfare_days)
    )
    chengdu_holiday_adjust_rest_days = (
        chengdu_statutory_raw - set(chengdu_statutory) - set(chengdu_welfare_days)
    )

    # 法定节假日 + 公司福利假，不再用工作日差集推导，避免把调休休息日算成法定日。
    main_holidays    = len(main_statutory) + len(main_welfare_days)
    chengdu_holidays = len(chengdu_statutory) + len(chengdu_welfare_days)

    result = {
        **ctx,
        "main_attendance_days":    (
            main_attendance_summary
            if main_attendance_summary is not None else len(main_wd)
        ),
        "chengdu_attendance_days": (
            chengdu_attendance_summary
            if chengdu_attendance_summary is not None else len(chengdu_wd)
        ),
        "main_holidays":           main_holidays,
        "chengdu_holidays":        chengdu_holidays,
        "main_company_welfare_days": main_welfare_days,
        "chengdu_company_welfare_days": chengdu_welfare_days,
        "main_statutory_holidays": main_statutory,
        "chengdu_statutory_holidays": chengdu_statutory,
        "main_holiday_adjust_rest_days": main_holiday_adjust_rest_days,
        "chengdu_holiday_adjust_rest_days": chengdu_holiday_adjust_rest_days,
        "main_payable_days": set(main_wd) | set(main_welfare_days) | set(main_statutory),
        "chengdu_payable_days": set(chengdu_wd) | set(chengdu_welfare_days) | set(chengdu_statutory),
    }

    print(f"[作息表] {year}年{month}月")
    print(f"  深圳应出勤 {result['main_attendance_days']} 天, 节假日 {main_holidays} 天")
    print(f"  成都应出勤 {result['chengdu_attendance_days']} 天, 节假日 {chengdu_holidays} 天")
    return result


# ══════════════════════════════════════════════════════════════════════════
#  4. 请假明细表解析
# ══════════════════════════════════════════════════════════════════════════

def parse_leave_summary(
    path: str,
    schedule_ctx: dict | None = None,
) -> dict[str, dict[str, float]]:
    """
    解析请假明细表（支持多 Sheet），
    返回 {工号或姓名: {请假类型: 天数}}。
    汇总时直接使用请假明细中的“最终请假天数”。
    """
    result: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    hours_fallback_rows = 0
    for row in _collect_deduped_leave_rows(path):
        emp_no = row["emp_no"]
        name = row["name"]
        leave_type = row["leave_type"]
        days = _to_float(row["raw_days"])
        # Defensive: prefer final hours → days when day column is missing/blank.
        # Only fall back when days is None (not when explicitly 0), so true zero
        # outside-month rows stay zero. Prefer 最终请假时长 over 系统时长.
        if days is None:
            final_hours = _to_float(row.get("final_hours"))
            sys_hours = _to_float(row.get("sys_hours"))
            hours = final_hours if final_hours and final_hours > 0 else sys_hours
            if hours and hours > 0:
                days = _overtime_days_from_hours(hours)
                hours_fallback_rows += 1

        key = emp_no or name
        if not key or not leave_type or days is None:
            continue
        if days == 0:
            continue

        # 标准化请假类型
        matched = None
        for lt in LEAVE_TYPES:
            if lt in leave_type:
                matched = lt
                break
        if matched:
            result[key][matched] += days

    print(f"[请假明细] 共解析 {len(result)} 名员工的请假数据")
    if hours_fallback_rows:
        print(f"[请假明细] 其中 {hours_fallback_rows} 行按系统时长小时折算天数（天数字段为空/0）")
    return dict(result)


def parse_leave_day_details(
    path: str,
    schedule_ctx: dict,
    special_chengdu_names: set[str] | tuple[str, ...] | None = None,
) -> dict[str, dict[date, dict[str, float]]]:
    """将请假明细按员工+日期拆分，供离职计薪口径使用。"""
    result: dict[str, dict[date, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(float))
    )
    month_start = schedule_ctx["month_start"]
    month_end = schedule_ctx["month_end"]
    special_chengdu_name_set = set(special_chengdu_names or ())

    for row in _collect_deduped_leave_rows(path):
        key = row["emp_no"] or row["name"]
        leave_type = row["leave_type"]
        if not key or not leave_type:
            continue

        matched = None
        for lt in LEAVE_TYPES:
            if lt in leave_type:
                matched = lt
                break
        if matched is None:
            continue

        dt_start = calc_leave.to_datetime(row["raw_start"])
        dt_end = calc_leave.to_datetime(row["raw_end"])
        if dt_start is None or dt_end is None:
            continue

        if not calc_leave.has_time_component(row["raw_start"]):
            dt_start = dt_start.replace(hour=9, minute=0)
        if not calc_leave.has_time_component(row["raw_end"]):
            dt_end = dt_end.replace(hour=18, minute=30)

        if (
            not calc_leave.is_natural_day_leave(matched)
            and not calc_leave.is_expected_attendance_day_leave(matched)
        ):
            cross_month_daily_hours, _ = (
                calc_leave.allocate_adjacent_cross_month_system_daily_hours(
                    dt_start,
                    dt_end,
                    row.get("sys_hours"),
                )
            )
            if cross_month_daily_hours is not None:
                for leave_date, hours in cross_month_daily_hours.items():
                    if month_start <= leave_date <= month_end and hours > 0:
                        result[key][leave_date][matched] += _round2(hours / 8)
                continue

        is_cd = _is_chengdu(
            row["dept1"],
            row["dept2"],
            row["dept3"],
            row["name"],
            special_chengdu_name_set,
        )
        working_days = (
            schedule_ctx["chengdu_working_days"]
            if is_cd else schedule_ctx["main_working_days"]
        )
        if matched == "产假":
            expected_key = (
                "chengdu_expected_attendance_days"
                if is_cd else "main_expected_attendance_days"
            )
            day_calendar = schedule_ctx.get(expected_key) or working_days
        else:
            day_calendar = working_days

        cur = max(dt_start.date(), month_start)
        end = min(dt_end.date(), month_end)
        while cur <= end:
            hours = calc_leave.calc_target_month_working_hours(
                dt_start, dt_end, day_calendar, cur, cur,
            )
            day_value = _round2(hours / 8) if hours else None
            if day_value:
                result[key][cur][matched] += day_value
            cur += timedelta(days=1)

    return {
        key: {
            day: dict(type_map) for day, type_map in day_map.items()
        }
        for key, day_map in result.items()
    }


# ══════════════════════════════════════════════════════════════════════════
#  5. 加班明细表解析
# ══════════════════════════════════════════════════════════════════════════

def parse_overtime_summary(
    path: str,
    target_year: int | None = None,
    target_month: int | None = None,
    employees: list[dict] | None = None,
    schedule_ctx: dict | None = None,
) -> dict[str, dict[str, float]]:
    """
    解析加班明细表，
    返回 {工号或姓名: {"2x_hours": 小时, "3x_hours": 小时, "2x": 天, "3x": 天}}。
    - 新表优先累加 2倍/3倍加班小时，再派生天数
    - 旧表没有小时列时，兼容累加 2倍/3倍加班天数
    - 回填表漏写运营支撑部调休休息日 2倍时，可用最终表花名册和作息表兜底补算
    """
    result: dict[str, dict[str, float]] = defaultdict(
        lambda: {"2x_hours": 0.0, "3x_hours": 0.0, "2x": 0.0, "3x": 0.0}
    )
    wb = openpyxl.load_workbook(path, data_only=True)
    filtered_out_rows = 0
    unresolved_rows = 0
    excluded_2x_rows = 0
    fallback_2x_rows = 0
    employee_lookup = _build_employee_lookup(employees) if employees else None

    for sn in wb.sheetnames:
        ws = wb[sn]
        header_row_idx, header_vals = _find_header_row(
            ws, "发起人工号", "工号", "2倍加班", "3倍加班"
        )
        if header_row_idx is None:
            continue

        col_emp_no = _find_col(header_vals, "发起人工号", "工号")
        col_name   = _find_col(header_vals, "发起人姓名", "姓名")
        col_start  = _find_col(header_vals, "开始时间")
        col_end    = _find_col(header_vals, "结束时间")
        col_detail = _find_col(header_vals, "明细")
        col_ot_day = _find_col(header_vals, "加班时间")
        col_type = _find_col(header_vals, "加班类型")
        col_final_hours = _find_col(header_vals, "最终加班时长（小时）", "最终加班时长", "最终加班小时")
        col_2x_hours = _find_col(header_vals, "2倍加班小时", "2倍加班（小时）")
        col_3x_hours = _find_col(header_vals, "3倍加班小时", "3倍加班（小时）")
        col_2x_days  = _find_col(header_vals, "2倍加班天数", "2倍加班（天）")
        col_3x_days  = _find_col(header_vals, "3倍加班天数", "3倍加班（天）")
        if col_2x_hours is None and col_2x_days is None:
            col_2x_days = _find_col(header_vals, "2倍加班")
        if col_3x_hours is None and col_3x_days is None:
            col_3x_days = _find_col(header_vals, "3倍加班")

        has_premium_cols = not all(
            col is None
            for col in (col_2x_hours, col_3x_hours, col_2x_days, col_3x_days)
        )
        can_fallback_2x = bool(employee_lookup and schedule_ctx and col_final_hours is not None)
        if not has_premium_cols and not can_fallback_2x:
            continue

        for r in range(header_row_idx + 1, ws.max_row + 1):
            overtime_date = None
            if target_year is not None and target_month is not None:
                overtime_date = _resolve_overtime_row_date(
                    ws, r, col_detail, col_ot_day, col_start, col_end,
                )
                if overtime_date is None:
                    unresolved_rows += 1
                    continue
                if not _same_year_month(overtime_date, target_year, target_month):
                    filtered_out_rows += 1
                    continue

            emp_no = _normalize_emp_no(
                ws.cell(r, col_emp_no + 1).value
            ) if col_emp_no is not None else ""
            name = _clean_name(
                ws.cell(r, col_name + 1).value
            ) if col_name is not None else ""

            key = emp_no or name
            if not key:
                continue

            val_2x_hours = _to_float(
                ws.cell(r, col_2x_hours + 1).value
            ) if col_2x_hours is not None else None
            val_3x_hours = _to_float(
                ws.cell(r, col_3x_hours + 1).value
            ) if col_3x_hours is not None else None
            val_2x_days = _to_float(
                ws.cell(r, col_2x_days + 1).value
            ) if col_2x_days is not None else None
            val_3x_days = _to_float(
                ws.cell(r, col_3x_days + 1).value
            ) if col_3x_days is not None else None

            row_has_premium = any(
                value for value in (val_2x_hours, val_2x_days, val_3x_hours, val_3x_days)
            )
            if can_fallback_2x and not row_has_premium:
                if overtime_date is None:
                    overtime_date = _resolve_overtime_row_date(
                        ws, r, col_detail, col_ot_day, col_start, col_end,
                    )
                overtime_type = _first_text(
                    ws.cell(r, col_type + 1).value
                ) if col_type is not None else None
                emp = _lookup_employee(emp_no, name, employee_lookup)
                if (
                    overtime_date
                    and (not overtime_type or "调休" in overtime_type)
                    and _is_operations_support_employee(emp)
                    and overtime_date in _holiday_adjust_rest_days_for_employee(schedule_ctx, emp)
                ):
                    fallback_hours = _to_float(ws.cell(r, col_final_hours + 1).value)
                    if fallback_hours and fallback_hours > 0:
                        if _is_rest_premium_excluded(emp_no, name):
                            excluded_2x_rows += 1
                        else:
                            result[key]["2x_hours"] += fallback_hours
                            result[key]["2x"] += _overtime_days_from_hours(fallback_hours)
                            fallback_2x_rows += 1

            # 新表优先按小时汇总；旧表只有天数时再按天数兼容。
            if val_2x_hours or val_2x_days:
                if _is_rest_premium_excluded(emp_no, name):
                    excluded_2x_rows += 1
                elif val_2x_hours:
                    result[key]["2x_hours"] += val_2x_hours
                    result[key]["2x"] += _overtime_days_from_hours(val_2x_hours)
                else:
                    result[key]["2x"] += val_2x_days
                    result[key]["2x_hours"] += _overtime_hours_from_days(val_2x_days)
            if val_3x_hours:
                result[key]["3x_hours"] += val_3x_hours
                result[key]["3x"] += _overtime_days_from_hours(val_3x_hours)
            elif val_3x_days:
                result[key]["3x"] += val_3x_days
                result[key]["3x_hours"] += _overtime_hours_from_days(val_3x_days)

    wb.close()
    if target_year is not None and target_month is not None:
        print(
            f"[加班明细] 仅汇总 {target_year}年{target_month}月；"
            f"已忽略 {filtered_out_rows} 行其他月份记录"
        )
        if unresolved_rows:
            print(f"[加班明细] 另有 {unresolved_rows} 行因无法识别日期，未计入本月汇总")
    if excluded_2x_rows:
        print(f"[加班明细] 已忽略 {excluded_2x_rows} 行 2倍加班例外人员记录")
    if fallback_2x_rows:
        print(f"[加班明细] 已按花名册/作息表兜底补算 {fallback_2x_rows} 行运营支撑部调休休息日 2倍")
    print(f"[加班明细] 共解析 {len(result)} 名员工的加班数据")
    return dict(result)


# ══════════════════════════════════════════════════════════════════════════
#  6. 补贴扣款表解析（旷工天数）
# ══════════════════════════════════════════════════════════════════════════

def parse_subsidy_absent(
    path: str,
) -> tuple[dict[str, float], dict[str, float]]:
    """
    解析补贴扣款表，提取旷工天数。
    返回 (by_emp_no: {工号: 天数}, by_name: {姓名: 天数})
    """
    by_emp_no: dict[str, float] = {}
    by_name: dict[str, float] = {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 用列搜索方式（兼容合并表头）
    col_name   = _find_col_ws(ws, "姓名")
    col_emp_no = _find_col_ws(ws, "工号", "员工工号")
    col_absent = _find_col_ws(ws, "旷工天数", "旷工")

    if col_absent is None:
        print("[补贴扣款表] 未找到旷工天数列")
        wb.close()
        return by_emp_no, by_name

    if col_name is None:
        print("[补贴扣款表] 未找到姓名列")
        wb.close()
        return by_emp_no, by_name

    # 跳过表头找到数据起始行
    data_start = None
    for r in range(1, min(8, ws.max_row + 1)):
        val = ws.cell(r, col_name).value
        if val and isinstance(val, str):
            name = val.strip()
            if name not in ("姓名", "合计", ""):
                data_start = r
                break
    if data_start is None:
        wb.close()
        return by_emp_no, by_name

    for r in range(data_start, ws.max_row + 1):
        raw_name = ws.cell(r, col_name).value
        if not raw_name or not isinstance(raw_name, str):
            continue
        name = _clean_name(raw_name.strip())
        if name in ("姓名", "合计", ""):
            continue

        emp_no = _normalize_emp_no(
            ws.cell(r, col_emp_no).value
        ) if col_emp_no else ""

        absent = _to_float(ws.cell(r, col_absent).value)

        if absent is not None:
            by_name[name] = absent
            if emp_no:
                by_emp_no[emp_no] = absent

    wb.close()
    print(f"[补贴扣款表] 共解析 {len(by_name)} 名员工的旷工数据")
    return by_emp_no, by_name


def parse_subsidy_absent_day_details(
    path: str,
    year: int,
    month: int,
) -> dict[str, set[date]]:
    """解析补贴扣款表的日期级旷工记录，供核对口径补充个人排班旷工日。"""
    result: dict[str, set[date]] = defaultdict(set)
    _, days_in_month = calendar.monthrange(year, month)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    col_name = _find_col_ws(ws, "姓名")
    col_emp_no = _find_col_ws(ws, "工号", "员工工号")
    col_attendance_result = _find_col_ws(ws, "考勤结果")

    if col_name is None or col_attendance_result is None:
        wb.close()
        print("[补贴扣款表] 未找到日期级考勤结果列，核对口径不补充个人排班旷工日")
        return {}

    data_start = None
    for r in range(1, min(8, ws.max_row + 1)):
        val = ws.cell(r, col_name).value
        if val and isinstance(val, str):
            name = val.strip()
            if name not in ("姓名", "合计", ""):
                data_start = r
                break
    if data_start is None:
        wb.close()
        return {}

    for r in range(data_start, ws.max_row + 1):
        raw_name = ws.cell(r, col_name).value
        if not raw_name or not isinstance(raw_name, str):
            continue
        name = _clean_name(raw_name.strip())
        if name in ("姓名", "合计", ""):
            continue

        emp_no = _normalize_emp_no(
            ws.cell(r, col_emp_no).value
        ) if col_emp_no else ""
        keys = [key for key in (emp_no, name) if key]
        if not keys:
            continue

        absent_days = {
            date(year, month, day)
            for day in range(1, days_in_month + 1)
            if "旷工" in str(ws.cell(r, col_attendance_result + day - 1).value or "")
        }
        if not absent_days:
            continue

        for key in keys:
            result[key].update(absent_days)

    wb.close()
    print(f"[补贴扣款表] 共解析 {len(result)} 个工号/姓名键的日期级旷工记录")
    return dict(result)


# ══════════════════════════════════════════════════════════════════════════
#  7. 缺勤 / 转正 计算
# ══════════════════════════════════════════════════════════════════════════

def calc_entry_leave_absence(
    hire_date: date | None,
    resign_date: date | None,
    payable_days: set[date],
    month_start: date,
    month_end: date,
) -> float | None:
    """
    入/离职缺勤天数：
    - 月中入职 → 入职日前的应计出勤日
    - 月中离职 → 离职日后的应计出勤日
    """
    absence = 0

    # 入职缺勤：入职日期在月内且不是月初
    if hire_date and month_start < hire_date <= month_end:
        absence += sum(
            1 for d in payable_days if month_start <= d < hire_date
        )

    # 离职缺勤：离职日期在月内且不是月末
    if resign_date and month_start <= resign_date < month_end:
        absence += sum(
            1 for d in payable_days if resign_date < d <= month_end
        )

    return absence if absence > 0 else None


def calc_entry_leave_absence_audit(
    hire_date: date | None,
    resign_date: date | None,
    payable_days: set[date],
    statutory_holidays: set[date],
    attendance_days: float | int | None,
    month_start: date,
    month_end: date,
    absent_days: set[date] | None = None,
) -> float | int | None:
    """
    入/离职缺勤天数（核对口径）：
    - 上月已离职但保留到本月表 → 按本月应出勤整月缺勤
    - 月中入职 → 入职日前的核对出勤日
    - 月中离职 → 离职日后的核对出勤日
    - 员工未在职期间的应计薪法定日也纳入核对出勤日，避免实际出勤虚增
    """
    if resign_date and resign_date < month_start:
        if attendance_days and attendance_days > 0:
            return attendance_days
        return None

    audit_days = set(payable_days)
    absence_days: set[date] = set()

    if hire_date and month_start < hire_date <= month_end:
        absence_days.update(d for d in audit_days if month_start <= d < hire_date)

    if resign_date and month_start <= resign_date < month_end:
        absence_days.update(d for d in audit_days if resign_date < d <= month_end)
        active_start = max(month_start, hire_date) if hire_date else month_start
        absence_days.update(
            d
            for d in (absent_days or set())
            if active_start <= d <= resign_date and d not in audit_days
        )

    return len(absence_days) if absence_days else None




def calc_resign_absence_by_payroll(
    hire_date: date | None,
    resign_date: date | None,
    attendance_days: float | int | None,
    payable_days: set[date],
    leave_day_detail: dict[date, dict[str, float]] | None,
    month_start: date,
    month_end: date,
) -> float | None:
    """
    当月离职缺勤天数：
    直接按“应出勤天数 - 计薪出勤天数”口径计算。

    计薪出勤天数口径：
    - 离职日前（不含离职日）的应出勤日，默认计薪
    - 离职日当天仅累计带薪请假
    - 离职日前的事假从计薪天数中扣除
    """
    if not resign_date or attendance_days is None:
        return None
    if not (month_start <= resign_date <= month_end):
        return None

    active_start = max(month_start, hire_date) if hire_date else month_start
    payroll_days = sum(
        1 for d in payable_days
        if active_start <= d < resign_date
    )

    for day, type_map in (leave_day_detail or {}).items():
        if day < resign_date:
            payroll_days -= sum(
                val for lt, val in type_map.items()
                if lt in UNPAID_LEAVE_TYPES
            )
        elif day == resign_date:
            payroll_days += sum(
                val for lt, val in type_map.items()
                if lt not in UNPAID_LEAVE_TYPES
            )

    absence = _round2(attendance_days - payroll_days)
    return absence if absence and absence > 0 else None


def calc_probation_days(
    confirm_date: date | None,
    working_days: set[date],
    statutory_holidays: set[date],
    month_start: date,
    month_end: date,
) -> float | None:
    """Count post-confirmation days for an employee confirmed this month.

    Uses a closed interval [confirm_date, month_end]: the confirmation day
    itself is included if it is a working day or a statutory holiday.
    """
    if not confirm_date:
        return None
    if not (month_start <= confirm_date <= month_end):
        return None

    countable_days = {
        day for day in working_days
        if confirm_date < day <= month_end
    } | {
        day for day in statutory_holidays
        if confirm_date <= day <= month_end
    }
    count = len(countable_days)
    return count if count > 0 else None


# ══════════════════════════════════════════════════════════════════════════
#  8. 生成最终表
# ══════════════════════════════════════════════════════════════════════════

def generate(
    employees: list[dict],
    transfer_map: dict[str, date],
    schedule_ctx: dict,
    leave_map: dict[str, dict[str, float]],
    overtime_map: dict[str, dict[str, float]],
    absent_by_no: dict[str, float],
    absent_by_name: dict[str, float],
    out_path: str,
    leave_day_details: dict[str, dict[date, dict[str, float]]] | None = None,
    special_chengdu_names: set[str] | tuple[str, ...] | None = None,
    absent_day_details: dict[str, set[date]] | None = None,
) -> None:
    """生成最终考勤汇总表 xlsx"""
    year        = schedule_ctx["year"]
    month       = schedule_ctx["month"]
    month_start = schedule_ctx["month_start"]
    month_end   = schedule_ctx["month_end"]
    main_wd     = schedule_ctx["main_working_days"]
    chengdu_wd  = schedule_ctx["chengdu_working_days"]
    main_attendance_days = schedule_ctx["main_attendance_days"]
    chengdu_attendance_days = schedule_ctx["chengdu_attendance_days"]
    main_payable_days = schedule_ctx["main_payable_days"]
    chengdu_payable_days = schedule_ctx["chengdu_payable_days"]
    special_chengdu_name_set = set(special_chengdu_names or ())
    special_chengdu_matched = 0

    # ── 查找辅助 ──────────────────────────────────────────────────────
    def _lookup(mapping: dict, emp: dict):
        """优先用工号匹配，回退姓名"""
        emp_no = emp["emp_no"]
        name   = emp["name"]
        if emp_no and emp_no in mapping:
            return mapping[emp_no]
        if name and name in mapping:
            return mapping[name]
        return None

    def _lookup_leave(emp):
        return _lookup(leave_map, emp) or {}

    def _lookup_overtime(emp):
        return _lookup(overtime_map, emp) or {}

    def _lookup_absent(emp):
        emp_no = emp["emp_no"]
        name   = emp["name"]
        if emp_no and emp_no in absent_by_no:
            return absent_by_no[emp_no]
        if name and name in absent_by_name:
            return absent_by_name[name]
        return None

    def _lookup_leave_detail(emp):
        return _lookup(leave_day_details or {}, emp) or {}

    def _lookup_absent_days(emp):
        return _lookup(absent_day_details or {}, emp) or set()

    # ── Excel 构建 ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month:02d}考勤"

    month_label = f"{year}年{month}月"

    # 标题行
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=len(OUTPUT_HEADERS),
    )
    title_cell = ws.cell(1, 1, f"{month_label}员工考勤汇总表")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # 样式
    header_font  = Font(bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align   = Alignment(horizontal="center", vertical="center")
    calculated_fill = PatternFill("solid", fgColor="FF00B050")
    manual_fill = PatternFill("solid", fgColor="FFF4B183")
    retained_fill = PatternFill("solid", fgColor="FF5B9BD5")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    retained_header_keys = {_field_key(h) for h in RETAINED_OUTPUT_HEADERS}
    manual_header_keys = {_field_key(h) for h in MANUAL_SOURCE_OUTPUT_HEADERS}
    output_col_letters = {
        _field_key(header): get_column_letter(ci)
        for ci, header in enumerate(OUTPUT_HEADERS, 1)
    }

    def _col(field: str) -> str:
        return output_col_letters[_field_key(field)]

    tracked_output_fields = (
        "二级部门",
        "三级部门",
        "员工类型",
        "人员分类",
        "入职日期",
        "离职日期",
        "转正日期",
        "异动日期",
    )
    written_nonempty_counts = {field: 0 for field in tracked_output_fields}
    normalized_emp_type_count = 0
    derived_field_counts = {"人员分类": 0}
    fixed_monthly_rest_matched_names: set[str] = set()

    # 表头行
    for ci, header in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(2, ci, header)
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border
        header_key = _field_key(header)
        if header_key in retained_header_keys:
            cell.fill = retained_fill
        elif header_key in manual_header_keys:
            cell.fill = manual_fill
        else:
            cell.fill = calculated_fill

    window_employees = [
        emp for emp in employees
        if _in_resign_keep_window(
            emp.get("resign_date"), year, month, emp.get("hire_date"),
        )
    ]
    skipped_resigned_outside_window = len(employees) - len(window_employees)
    excluded_final_table_employees = [
        emp for emp in window_employees
        if _is_final_table_excluded_employee(emp)
    ]
    output_employees = [
        emp for emp in window_employees
        if not _is_final_table_excluded_employee(emp)
    ]

    # ── 逐员工写数据 ─────────────────────────────────────────────────
    for idx, emp in enumerate(output_employees, 1):
        row_num = idx + 2

        is_cd = _is_chengdu(
            emp["dept1"],
            emp["dept2"],
            emp["dept3"],
            emp["name"],
            special_chengdu_name_set,
        )
        if is_cd and emp["name"] in special_chengdu_name_set:
            special_chengdu_matched += 1
        working_days    = chengdu_wd if is_cd else main_wd
        statutory_holidays = (
            schedule_ctx["chengdu_statutory_holidays"]
            if is_cd else schedule_ctx["main_statutory_holidays"]
        )
        attendance_days = chengdu_attendance_days if is_cd else main_attendance_days
        payable_days = chengdu_payable_days if is_cd else main_payable_days
        is_fixed_monthly_rest = _is_fixed_monthly_rest_employee(emp["name"])
        if is_fixed_monthly_rest:
            attendance_days = _fixed_monthly_rest_attendance_days(month_end)
            fixed_monthly_rest_matched_names.add(emp["name"])

        # 节假日拆分：法定节假日 + 公司福利假
        # 法定节假日沿用作息表“计薪法定假日”口径，避免把调休形成的休息日计入。
        company_welfare_days = (
            schedule_ctx["chengdu_company_welfare_days"]
            if is_cd else schedule_ctx["main_company_welfare_days"]
        )
        company_welfare = calc_active_day_count(
            company_welfare_days, emp["hire_date"], emp["resign_date"],
            month_start, month_end,
        )
        holidays = calc_statutory_holiday_days(
            emp["hire_date"], emp["resign_date"],
            statutory_holidays, month_start, month_end,
        )
        if is_fixed_monthly_rest:
            company_welfare = 0
            payable_days = set(payable_days) - set(company_welfare_days)

        leave_data = _lookup_leave(emp)
        leave_day_detail = _lookup_leave_detail(emp)
        holidays = max(
            0,
            holidays - calc_maternity_statutory_overlap_days(
                leave_day_detail, statutory_holidays,
                emp["hire_date"], emp["resign_date"],
                month_start, month_end,
            ),
        )
        ot_data    = _lookup_overtime(emp)
        absent     = _lookup_absent(emp)
        absent_day_detail = _lookup_absent_days(emp)
        # 员工类型、人员分类和日期不是月度考勤数据字段；缺失时必须保持为空。
        display_emp_type = _clean_final_status_label(emp.get("emp_type"))
        display_category = _clean_final_status_label(emp.get("category"))
        display_dept1 = emp["dept1"]
        display_dept2 = emp["dept2"]
        display_dept3 = emp["dept3"]

        entry_leave_absence = calc_entry_leave_absence_audit(
            emp["hire_date"],
            emp["resign_date"],
            payable_days,
            statutory_holidays,
            attendance_days,
            month_start,
            month_end,
            absent_day_detail,
        )
        probation = calc_probation_days(
            emp["confirm_date"],
            working_days, statutory_holidays,
            month_start, month_end,
        )
        actual_deductions = [
            "法定节假日", "公司福利假",
            "年假", "调休", "婚假", "陪产假", "丧假", "产检假", "工伤假", "产假",
            "病假", "事假",
            "旷工天数", "入/离职缺勤天数",
        ]
        actual_formula = (
            f"={_col('应出勤天数')}{row_num}"
            + "".join(f"-{_col(field)}{row_num}" for field in actual_deductions)
        )
        payroll_additions = [
            "法定节假日", "公司福利假",
            "年假", "调休", "婚假", "陪产假", "丧假", "产检假", "工伤假", "产假",
            "病假",
        ]
        payroll_formula = (
            f"={_col('实际出勤天数')}{row_num}"
            + "".join(f"+{_col(field)}{row_num}" for field in payroll_additions)
        )

        # 组装行
        row_values = {
            "序号": idx,
            "月份": f"{month}月",
            "工号": emp["emp_no"],
            "姓名": emp["name"],
            "考勤组": emp.get("attendance_group"),
            "合同主体": emp["contract_entity"],
            "一级部门": display_dept1,
            "二级部门": display_dept2,
            "三级部门": display_dept3,
            "岗位": emp["position"],
            "员工类型": display_emp_type,
            "人员分类": display_category,
            "入职日期": emp["hire_date"],
            "离职日期": emp["resign_date"],
            "转正日期": emp["confirm_date"],
            "异动日期": _lookup(transfer_map, emp),
            "应出勤天数": attendance_days,
            "计薪出勤天数": payroll_formula,
            "实际出勤天数": actual_formula,
            "法定节假日": holidays if holidays else None,
            "公司福利假": company_welfare if company_welfare else None,
        }

        # 请假各类型
        for lt in LEAVE_TYPES:
            val = leave_data.get(lt)
            row_values[lt] = _round2(val) if val else None

        row_values.update({
            "旷工天数": _round2(absent) if absent else None,
            "入/离职缺勤天数": entry_leave_absence,
            "当月转正天数": probation,
            "2倍加班（小时）": ot_data.get("2x_hours") if ot_data.get("2x_hours") else None,
            "3倍加班（小时）": ot_data.get("3x_hours") if ot_data.get("3x_hours") else None,
            "2倍加班（天）": ot_data.get("2x") if ot_data.get("2x") else None,
            "3倍加班（天）": ot_data.get("3x") if ot_data.get("3x") else None,
        })
        row_values_by_key = {
            _field_key(field): value
            for field, value in row_values.items()
        }

        for ci, header in enumerate(OUTPUT_HEADERS, 1):
            header_key = _field_key(header)
            val = None if header_key in retained_header_keys else row_values_by_key.get(header_key)
            cell = ws.cell(row_num, ci, val)
            cell.alignment = data_align
            cell.border    = thin_border
            if isinstance(val, date):
                cell.number_format = "YYYY-MM-DD"
            elif header_key in {
                _field_key("2倍加班（小时）"),
                _field_key("3倍加班（小时）"),
            }:
                cell.number_format = OVERTIME_HOUR_OUTPUT_NUMBER_FORMAT
            elif header_key in {
                _field_key("2倍加班（天）"),
                _field_key("3倍加班（天）"),
            }:
                cell.number_format = OVERTIME_DAY_OUTPUT_NUMBER_FORMAT

        for field in tracked_output_fields:
            value = row_values_by_key.get(_field_key(field))
            if value not in (None, ""):
                written_nonempty_counts[field] += 1

    # ── 列宽 ─────────────────────────────────────────────────────────
    for ci in range(1, len(OUTPUT_HEADERS) + 1):
        header_text = OUTPUT_HEADERS[ci - 1]
        # 中文字符约 2 字节宽
        width = max(len(header_text) * 2, 8)
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 20)

    wb.save(out_path)
    print(
        "[最终表] 关键字段写出非空："
        + "，".join(
            f"{field} {count}/{len(output_employees)}"
            for field, count in written_nonempty_counts.items()
        )
    )
    if any(derived_field_counts.values()):
        print(
            "[最终表] 缺失字段兜底推导："
            + "，".join(f"{field} {count} 条" for field, count in derived_field_counts.items())
        )
    if normalized_emp_type_count:
        print(f"[最终表] 员工类型标准化 {normalized_emp_type_count} 条")
    if skipped_resigned_outside_window:
        print(
            "[最终表] 已跳过离职日期不在上月/本月/下月范围内 "
            f"{skipped_resigned_outside_window} 人"
        )
    if excluded_final_table_employees:
        excluded_names = [
            emp.get("name") or emp.get("emp_no") or "(未命名)"
            for emp in excluded_final_table_employees
        ]
        shown_names = "、".join(excluded_names[:30])
        if len(excluded_names) > 30:
            shown_names += f" 等 {len(excluded_names)} 人"
        print(
            f"[最终表] 已剔除兼职/实习/劳务外包人员 {len(excluded_final_table_employees)} 人："
            f"{shown_names}"
        )
    if special_chengdu_names is not None:
        print(
            f"[成都作息名单] 设置 {len(special_chengdu_name_set)} 个姓名，"
            f"最终表匹配 {special_chengdu_matched} 人"
        )
    if fixed_monthly_rest_matched_names:
        print(
            f"[月休4天名单] 最终表匹配 {len(fixed_monthly_rest_matched_names)} 人："
            + "、".join(sorted(fixed_monthly_rest_matched_names))
        )
    print(f"[最终表] 已生成: {out_path}, 共 {len(output_employees)} 行数据")
