"""Regression tests for final-table data issues (leave/OT/subsidy path).

Covers the high-confidence code fixes for:
  - 3.5h triple-OT hour parsing and day conversion
  - revoked leave status filtering
  - 1-hour 调休 decimal precision
  - late22 exclusion for sales staff
  - RD subsidy keyword match across dept levels
  - resigned employee keep window for target-month history
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest
from unittest.mock import patch
from datetime import date
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
for sub in ("leave", "overtime", "subsidy", "finally"):
    p = str(BASE / sub)
    if p not in sys.path:
        sys.path.insert(0, p)
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))

import calc_leave  # noqa: E402
import calc_finally as fin  # noqa: E402
import fill_overtime_fields as ot  # noqa: E402
import calc_subsidy_deduction as sub  # noqa: E402


class ParseFloatDurationTests(unittest.TestCase):
    def test_plain_number(self):
        self.assertEqual(ot.parse_float(3.5), 3.5)
        self.assertEqual(ot.parse_float("3.5"), 3.5)

    def test_hour_text(self):
        self.assertEqual(ot.parse_float("3.5小时"), 3.5)
        self.assertEqual(ot.parse_float("1 小时"), 1.0)
        self.assertEqual(ot.parse_float("0.5h"), 0.5)

    def test_day_text_converted_to_hours(self):
        self.assertEqual(ot.parse_float("0.5天"), 4.0)
        self.assertEqual(ot.parse_float("1天"), 8.0)

    def test_empty_and_invalid(self):
        self.assertIsNone(ot.parse_float(None))
        self.assertIsNone(ot.parse_float(""))
        self.assertIsNone(ot.parse_float("半天"))


class PremiumHourAndDayTests(unittest.TestCase):
    def test_3_5_hours_is_not_one_day(self):
        hours = ot._calc_premium_hour_value(3.5, None, 8.0)
        days = ot._calc_premium_days_from_hours(hours, 8.0)
        self.assertAlmostEqual(hours, 3.5, places=6)
        self.assertAlmostEqual(days, 3.5 / 8.0, places=6)
        self.assertNotEqual(days, 1.0)

    def test_one_hour_fraction(self):
        hours = ot._calc_premium_hour_value(1.0, None, 8.0)
        days = ot._calc_premium_days_from_hours(hours, 8.0)
        self.assertAlmostEqual(days, 0.125, places=6)

    def test_eight_hours_is_one_day(self):
        hours = ot._calc_premium_hour_value(8.0, None, 8.0)
        days = ot._calc_premium_days_from_hours(hours, 8.0)
        self.assertAlmostEqual(days, 1.0, places=6)

    def test_missing_raw_hours_does_not_default_to_full_day(self):
        # Previously raw_hours=None silently became standard_hours (8) → 1 day.
        hours = ot._calc_premium_hour_value(None, None, 8.0)
        self.assertEqual(hours, 0.0)
        days = ot._calc_premium_days_from_hours(hours, 8.0)
        self.assertEqual(days, 0.0)

    def test_hour_text_through_parse_float_pipeline(self):
        raw = ot.parse_float("3.5小时")
        hours = ot._calc_premium_hour_value(raw, None, 8.0)
        days = ot._calc_premium_days_from_hours(hours, 8.0)
        self.assertAlmostEqual(days, 0.4375, places=6)


class LeaveRevokeAndDurationTests(unittest.TestCase):
    def test_revoked_status_exact(self):
        self.assertTrue(calc_leave.is_revoked_approval("已撤销", None))
        self.assertTrue(calc_leave.is_revoked_approval("已撤销", "同意"))

    def test_revoked_status_keywords(self):
        self.assertTrue(calc_leave.is_revoked_approval("流程已终止", None))
        self.assertTrue(calc_leave.is_revoked_approval("作废", None))
        self.assertTrue(calc_leave.is_revoked_approval("完成", "申请人撤销"))

    def test_keep_valid_status(self):
        self.assertFalse(calc_leave.is_revoked_approval("完成", "同意"))
        self.assertFalse(calc_leave.is_revoked_approval("审批中", None))
        self.assertFalse(calc_leave.is_revoked_approval("已修改", "同意"))

    def test_one_hour_comp_leave_days_precision(self):
        # 1 hour 调休 must not collapse to 0 after hour→day conversion.
        final_h, final_days, _ = calc_leave.calc_from_system_duration(1.0)
        self.assertAlmostEqual(final_h, 1.0, places=2)
        self.assertGreater(final_days, 0)
        self.assertAlmostEqual(final_days, 0.13, places=2)  # ROUND_HALF_UP(0.125, 2)

    def test_parse_export_duration_hour_text(self):
        self.assertEqual(calc_leave.parse_export_duration("1小时"), 1.0)
        self.assertEqual(calc_leave.parse_export_duration("3.5小时"), 3.5)


class ResignKeepWindowTests(unittest.TestCase):
    def test_active_employee_kept(self):
        self.assertTrue(fin._in_resign_keep_window(None, 2026, 5))

    def test_resigned_during_target_month_kept(self):
        self.assertTrue(
            fin._in_resign_keep_window(date(2026, 5, 10), 2026, 5)
        )

    def test_resigned_after_target_month_kept(self):
        # Still employed throughout May → keep for May stats.
        self.assertTrue(
            fin._in_resign_keep_window(date(2026, 7, 1), 2026, 5)
        )

    def test_resigned_long_before_target_dropped(self):
        self.assertFalse(
            fin._in_resign_keep_window(date(2025, 12, 1), 2026, 5)
        )

    def test_resigned_previous_month_buffer_kept(self):
        self.assertTrue(
            fin._in_resign_keep_window(date(2026, 4, 20), 2026, 5)
        )


class OvertimeDaysFromHoursTests(unittest.TestCase):
    def test_finally_overtime_days_from_hours(self):
        self.assertAlmostEqual(fin._overtime_days_from_hours(3.5), 3.5 / 8.0, places=6)
        self.assertAlmostEqual(fin._overtime_days_from_hours(1.0), 0.125, places=6)
        self.assertEqual(fin._overtime_days_from_hours(0), 0.0)
        self.assertEqual(fin._overtime_days_from_hours(None), 0.0)


class Late22SalesExclusionTests(unittest.TestCase):
    def test_sales_dept_keyword_excluded(self):
        record = {
            "name": "李睿",
            "dept1": "AI智慧文创事业部",
            "dept2": "销售组",
            "dept3": None,
            "pos": "销售经理",
        }
        self.assertTrue(sub._should_exclude_late22_count(record, included_names=[]))

    def test_sales_position_only_excluded(self):
        record = {
            "name": "李睿",
            "dept1": "业务中心",
            "dept2": "华北区",
            "pos": "销售专员",
        }
        self.assertTrue(sub._should_exclude_late22_count(record, included_names=[]))

    def test_rd_not_excluded_by_sales_rule(self):
        record = {
            "name": "丁俊",
            "dept1": "智慧行政事业部",
            "dept2": "研发小组",
            "pos": "后端开发",
        }
        self.assertFalse(sub._should_exclude_late22_count(record, included_names=[]))

    def test_include_list_overrides(self):
        record = {
            "name": "崔利华",
            "dept1": "运营管理中心-运营支撑部",
            "pos": "",
        }
        self.assertFalse(
            sub._should_exclude_late22_count(record, included_names=["崔利华"])
        )


class RdDeptMatchTests(unittest.TestCase):
    def test_full_path_in_dept1(self):
        self.assertTrue(
            sub._is_rd_dept("智慧行政事业部-研发小组", keywords=sub.RD_DEPT_KEYWORDS)
        )

    def test_split_across_dept_levels(self):
        # Real bug: leaf group lives in dept2 while keyword is full path.
        self.assertTrue(
            sub._is_rd_dept(
                "智慧行政事业部",
                "研发小组",
                None,
                keywords=sub.RD_DEPT_KEYWORDS,
            )
        )

    def test_keyword_substring_match(self):
        self.assertTrue(
            sub._is_rd_dept("产品中心", keywords=sub.RD_DEPT_KEYWORDS)
        )
        self.assertTrue(
            sub._is_rd_dept("研发中心-平台组", keywords=sub.RD_DEPT_KEYWORDS)
        )

    def test_non_rd_excluded(self):
        self.assertFalse(
            sub._is_rd_dept(
                "AI智慧文创事业部",
                "销售组",
                None,
                keywords=sub.RD_DEPT_KEYWORDS,
            )
        )
        self.assertFalse(
            sub._is_rd_dept(
                "行政部",
                "前台",
                None,
                keywords=sub.RD_DEPT_KEYWORDS,
            )
        )

    def test_similar_but_not_rd(self):
        # 反例：名称相似但不在产研关键字列表
        self.assertFalse(
            sub._is_rd_dept(
                "智慧行政事业部",
                "综合支持组",
                None,
                keywords=sub.RD_DEPT_KEYWORDS,
            )
        )


class MaternityLeaveTypeTests(unittest.TestCase):
    def test_maternity_is_expected_attendance_day_leave(self):
        self.assertTrue(calc_leave.is_expected_attendance_day_leave("产假"))
        self.assertFalse(calc_leave.is_expected_attendance_day_leave("事假"))


class MaternityLeaveOverrideTests(unittest.TestCase):
    def setUp(self):
        expected_days = {
            date(2026, 6, 1), date(2026, 6, 2), date(2026, 6, 8),
            date(2026, 6, 10), date(2026, 6, 15), date(2026, 6, 16),
            date(2026, 6, 29),
        }
        working_days = expected_days - {date(2026, 6, 10)}
        self.schedule_ctx = {
            "year": 2026,
            "month": 6,
            "month_start": date(2026, 6, 1),
            "month_end": date(2026, 6, 30),
            "next_month_start": date(2026, 7, 1),
            "next_month_label": "七月考勤",
            "main_working_days": working_days,
            "chengdu_working_days": set(working_days),
            "main_expected_attendance_days": expected_days,
            "chengdu_expected_attendance_days": set(expected_days),
        }
        self.header = (
            "发起人工号", "发起人姓名", "一级部门", "二级部门", "三级部门",
            "请假类型", "开始时间", "结束时间", "系统时长", "发起时间",
            "完成时间", "审批编号", "审批状态", "审批结果", "是否实习生", "源文件行号",
        )

    def test_parse_json_and_optional_employee_number(self):
        parsed = calc_leave.parse_maternity_leave_overrides(
            '[{"name":"长期产假员工","start_date":"2026-05-20","end_date":"2026-07-10"}]'
        )
        self.assertEqual(parsed[0]["employee_no"], "")
        self.assertEqual(parsed[0]["start_date"], date(2026, 5, 20))
        self.assertEqual(parsed[0]["end_date"], date(2026, 7, 10))

    def test_add_when_dingtalk_has_no_maternity_leave(self):
        merged = calc_leave.merge_maternity_leave_overrides(
            [self.header],
            [{"name": "长期产假员工", "start_date": "2026-05-20", "end_date": "2026-07-10"}],
            self.schedule_ctx,
        )
        self.assertEqual(len(merged), 2)
        self.assertEqual(merged[1][calc_leave.ROW_LEAVE_TYPE], "产假")
        self.assertTrue(
            str(merged[1][calc_leave.ROW_APPROVAL_ID]).startswith(
                calc_leave.MATERNITY_OVERRIDE_APPROVAL_PREFIX
            )
        )

    def test_skip_when_dingtalk_maternity_leave_already_exists(self):
        existing = (
            "001", "长期产假员工", "", "", "", "产假",
            "2026-05-20 09:00", "2026-07-10 18:30", None,
            None, None, "DING-001", "完成", "同意", False, 2,
        )
        merged = calc_leave.merge_maternity_leave_overrides(
            [self.header, existing],
            [{"employee_no": "001", "name": "长期产假员工", "start_date": "2026-05-20", "end_date": "2026-07-10"}],
            self.schedule_ctx,
        )
        self.assertEqual(merged, [self.header, existing])

    def test_full_first_and_last_month_use_expected_attendance_dates(self):
        full = ("", "员工", "", "", "", "产假", "2026-05-01", "2026-07-31", None)
        first = ("", "员工", "", "", "", "产假", "2026-06-15", "2026-07-31", None)
        last = ("", "员工", "", "", "", "产假", "2026-05-01", "2026-06-10", None)
        self.assertEqual(calc_leave.calc_final_fields(full, self.schedule_ctx)[1], 7)
        self.assertEqual(calc_leave.calc_final_fields(first, self.schedule_ctx)[1], 3)
        self.assertEqual(calc_leave.calc_final_fields(last, self.schedule_ctx)[1], 4)

    def test_name_only_manual_leave_is_written(self):
        merged = calc_leave.merge_maternity_leave_overrides(
            [self.header],
            [{"name": "长期产假员工", "start_date": "2026-05-20", "end_date": "2026-07-10"}],
            self.schedule_ctx,
        )
        workbook = calc_leave.openpyxl.Workbook()
        with patch.object(calc_leave.openpyxl, "Workbook", return_value=workbook), patch.object(workbook, "save"):
            calc_leave.process(merged, "unused.xlsx", self.schedule_ctx)
        worksheet = workbook["请假明细"]
        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet.cell(2, 2).value, "长期产假员工")
        self.assertEqual(worksheet.cell(2, 12).value, "长期产假人工兜底")

    def test_maternity_day_detail_uses_expected_attendance_calendar(self):
        leave_rows = [{
            "emp_no": "",
            "name": "长期产假员工",
            "dept1": "",
            "dept2": "",
            "dept3": "",
            "leave_type": "产假",
            "raw_start": "2026-05-20 09:00",
            "raw_end": "2026-07-10 18:30",
        }]
        with patch.object(fin, "_collect_deduped_leave_rows", return_value=leave_rows):
            detail = fin.parse_leave_day_details("unused.xlsx", self.schedule_ctx)
        self.assertEqual(detail["长期产假员工"][date(2026, 6, 10)]["产假"], 1.0)


class EmploymentStatutoryBoundaryTests(unittest.TestCase):
    def setUp(self):
        self.month_start = date(2026, 6, 1)
        self.month_end = date(2026, 6, 30)

    def test_confirmation_day_same_as_statutory_holiday_is_counted(self):
        count = fin.calc_probation_days(
            date(2026, 6, 10),
            {date(2026, 6, 10), date(2026, 6, 11)},
            {date(2026, 6, 10)},
            self.month_start,
            self.month_end,
        )
        self.assertEqual(count, 2)

    def test_confirmation_day_same_as_weekend_statutory_holiday_is_counted(self):
        count = fin.calc_probation_days(
            date(2026, 6, 14),
            set(),
            {date(2026, 6, 14)},
            self.month_start,
            self.month_end,
        )
        self.assertEqual(count, 1)

    def test_statutory_holiday_before_confirmation_is_not_counted(self):
        count = fin.calc_probation_days(
            date(2026, 6, 11),
            set(),
            {date(2026, 6, 10)},
            self.month_start,
            self.month_end,
        )
        self.assertIsNone(count)

    def test_resignation_day_same_as_statutory_holiday_is_not_counted(self):
        count = fin.calc_statutory_holiday_days(
            None,
            date(2026, 6, 10),
            {date(2026, 6, 10)},
            self.month_start,
            self.month_end,
        )
        self.assertEqual(count, 0)

    def test_statutory_holiday_before_resignation_day_is_counted(self):
        count = fin.calc_statutory_holiday_days(
            None,
            date(2026, 6, 11),
            {date(2026, 6, 10), date(2026, 6, 11)},
            self.month_start,
            self.month_end,
        )
        self.assertEqual(count, 1)


class MaternityStatutoryHolidayTests(unittest.TestCase):
    def test_statutory_holiday_inside_maternity_is_removed(self):
        overlap = fin.calc_maternity_statutory_overlap_days(
            {
                date(2026, 6, 10): {"产假": 1.0},
                date(2026, 6, 11): {"事假": 1.0},
            },
            {date(2026, 6, 10), date(2026, 6, 11)},
            None,
            None,
            date(2026, 6, 1),
            date(2026, 6, 30),
        )
        self.assertEqual(overlap, 1)

    def test_statutory_holiday_outside_maternity_is_kept(self):
        overlap = fin.calc_maternity_statutory_overlap_days(
            {date(2026, 6, 9): {"产假": 1.0}},
            {date(2026, 6, 10)},
            None,
            None,
            date(2026, 6, 1),
            date(2026, 6, 30),
        )
        self.assertEqual(overlap, 0)


# ── 在职花名册字段契约回归测试 ──────────────────────────────────────────────


def _write_roster(rows, sheet_name: str = "在职花名册") -> str:
    """把 rows（rows[0] 为表头）写入临时 xlsx，返回路径。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(list(row))
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


class RosterFieldContractTests(unittest.TestCase):
    """在职花名册字段契约：姓名必填、旧格式兼容、零在职保护和歧义拒绝。"""

    def test_name_only_roster_parses(self):
        path = _write_roster([["姓名"], [" 张三 "], ["李四"]])
        try:
            employees = fin.parse_roster(path)
        finally:
            os.remove(path)
        self.assertEqual([(row["emp_no"], row["name"]) for row in employees], [("", "张三"), ("", "李四")])

    def test_employee_name_only_roster_parses(self):
        path = _write_roster([["员工姓名"], ["王五"]])
        try:
            employees = fin.parse_roster(path)
        finally:
            os.remove(path)
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0]["name"], "王五")

    def test_approval_identity_headers_are_not_roster_headers(self):
        for header in ("发起人姓名", "申请人姓名", "实际申请人"):
            with self.subTest(header=header):
                path = _write_roster([[header], ["张三"]], "审批流程")
                try:
                    with self.assertRaisesRegex(ValueError, "未找到可用表头"):
                        fin.parse_roster(path)
                finally:
                    os.remove(path)

    def test_name_only_duplicate_is_rejected(self):
        path = _write_roster([["姓名"], ["张 三"], ["张三"]])
        try:
            with self.assertRaises(ValueError) as ctx:
                fin.parse_roster(path)
            self.assertIn("重复姓名“张三”", str(ctx.exception))
            self.assertIn("2 条", str(ctx.exception))
        finally:
            os.remove(path)

    def test_legacy_multicolumn_roster_without_contract_entity_parses(self):
        """旧版多列表花名册缺少合同主体时仍须兼容解析。"""
        rows = [
            ["员工编号", "员工姓名", "一级部门", "二级部门", "三级部门"],
            ["MT0001", "张三", "总部", "产品技术部", "后端组"],
            ["MT0002", "李四", "总部", "运营管理中心", "客服组"],
            ["MT0003", "王五", "成都分公司", "研发中心", "算法组"],
        ]
        path = _write_roster(rows)
        try:
            employees = fin.parse_roster(path)
        finally:
            os.remove(path)
        self.assertEqual(len(employees), 3)
        emp_nos = [e["emp_no"] for e in employees]
        self.assertIn("MT0001", emp_nos)
        self.assertIn("MT0002", emp_nos)
        self.assertIn("MT0003", emp_nos)
        self.assertTrue(all(not e["contract_entity"] for e in employees))

    def test_position_transfer_headers_as_active_path_raises(self):
        """岗位异动流程表头作为在职花名册 → 必须抛错（零在职保护），禁止返回只有离职人员的结果。"""
        pt_rows = [
            ["实际申请人工号", "实际申请人", "发起人工号", "发起人姓名",
             "异动日期", "异动类型", "生效日期"],
            ["MT0001", "张三", "MT0001", "张三", "2026-06-15", "岗位异动", "2026-06-15"],
        ]
        resign_rows = [
            ["工号", "姓名", "一级部门", "离职日期"],
            ["MT9999", "远古离职", "总部", "2024-01-01"],
        ]
        pt_path = _write_roster(pt_rows, "岗位异动流程表")
        resign_path = _write_roster(resign_rows, "离职")
        try:
            with self.assertRaises(ValueError) as ctx:
                fin.parse_roster(pt_path, resign_path)
            err_msg = str(ctx.exception)
            # 必须是零在职保护错误（无可用表头或解析为 0 人），不能返回只有离职人员的结果
            self.assertTrue(
                "未找到可用表头" in err_msg or "解析为 0 人" in err_msg,
                f"expected zero-active protection error, got: {err_msg}",
            )
        finally:
            os.remove(pt_path)
            os.remove(resign_path)

    def test_non_roster_file_as_active_path_raises(self):
        """非花名册文件（无工号/姓名列）作为在职花名册 → 必须抛错（零在职保护），禁止返回只有离职人员的结果。"""
        # 使用真正无匹配表头的文件（列名不包含"工号"或"姓名"子串）
        non_roster_rows = [
            ["流程编号", "申请人名称", "审批状态", "审批结果"],
            ["FLOW-001", "张三", "已完成", "同意"],
        ]
        # 有效离职花名册
        resign_rows = [
            ["工号", "姓名", "一级部门", "离职日期"],
            ["MT9999", "远古离职", "总部", "2024-01-01"],
        ]
        non_roster_path = _write_roster(non_roster_rows, "审批流程表")
        resign_path = _write_roster(resign_rows, "离职")
        try:
            with self.assertRaises(ValueError) as ctx:
                fin.parse_roster(non_roster_path, resign_path)
            self.assertIn("未找到可用表头", str(ctx.exception))
        finally:
            os.remove(non_roster_path)
            os.remove(resign_path)

    def test_active_resigned_both_appear(self):
        """在职多人 + 离职少量 → 最终结果同时包含。"""
        active_rows = [
            ["工号", "姓名", "一级部门", "二级部门", "三级部门"],
            ["MT0001", "张三", "总部", "产品技术部", "后端组"],
            ["MT0002", "李四", "总部", "运营管理中心", "客服组"],
        ]
        resign_rows = [
            ["工号", "姓名", "一级部门", "二级部门", "三级部门", "离职日期"],
            ["MT0004", "赵六", "总部", "销售部", "华东组", "2026-06-15"],
        ]
        active_path = _write_roster(active_rows, "在职")
        resign_path = _write_roster(resign_rows, "离职")
        try:
            employees = fin.parse_roster(active_path, resign_path)
        finally:
            os.remove(active_path)
            os.remove(resign_path)
        emp_nos = [e["emp_no"] for e in employees]
        self.assertIn("MT0001", emp_nos)
        self.assertIn("MT0002", emp_nos)
        self.assertIn("MT0004", emp_nos)
        mt0004 = next(e for e in employees if e["emp_no"] == "MT0004")
        self.assertEqual(mt0004["resign_date"], date(2026, 6, 15))

    def test_duplicate_emp_no_merges_resign_date(self):
        """同一工号同时存在于在职和离职花名册 → 只输出一行，正确合并离职日期。"""
        active_rows = [
            ["工号", "姓名", "一级部门"],
            ["MT0001", "张三", "总部"],
        ]
        resign_rows = [
            ["工号", "姓名", "一级部门", "离职日期"],
            ["MT0001", "张三", "总部", "2026-06-20"],
        ]
        active_path = _write_roster(active_rows, "在职")
        resign_path = _write_roster(resign_rows, "离职")
        try:
            employees = fin.parse_roster(active_path, resign_path)
        finally:
            os.remove(active_path)
            os.remove(resign_path)
        mt0001 = [e for e in employees if e["emp_no"] == "MT0001"]
        self.assertEqual(len(mt0001), 1)
        self.assertEqual(mt0001[0]["resign_date"], date(2026, 6, 20))
        self.assertEqual(len(employees), 1)

    def test_contract_entity_missing_keeps_employee(self):
        """合同主体缺失 → 员工不被删除。"""
        rows = [
            ["工号", "姓名", "合同主体", "一级部门"],
            ["MT0001", "张三", "", "总部"],
        ]
        path = _write_roster(rows)
        try:
            employees = fin.parse_roster(path)
        finally:
            os.remove(path)
        self.assertEqual(len(employees), 1)
        self.assertFalse(employees[0]["contract_entity"])

    def test_zero_active_protection_no_header(self):
        """在职源无可用表头 → 抛明确错误。"""
        rows = [
            ["异动日期", "异动类型", "生效日期"],
            ["2026-06-15", "岗位异动", "2026-06-15"],
        ]
        path = _write_roster(rows, "岗位异动")
        try:
            with self.assertRaises(ValueError) as ctx:
                fin.parse_roster(path)
            self.assertIn("未找到可用表头", str(ctx.exception))
        finally:
            os.remove(path)

    def test_zero_active_protection_all_rows_filtered(self):
        """在职源数据行全部没有姓名时抛明确错误。"""
        # 表头有效但姓名全部为空 → source_rows=0, valid=0
        rows = [
            ["工号", "姓名", "一级部门"],
            [None, None, None],
            ["", "", ""],
        ]
        path = _write_roster(rows)
        try:
            with self.assertRaises(ValueError) as ctx:
                fin.parse_roster(path)
            self.assertIn("解析为 0 人", str(ctx.exception))
        finally:
            os.remove(path)

    def test_parttime_intern_exclusion_rule_preserved(self):
        """兼职、实习、劳务外包排除规则继续生效。"""
        intern = {"emp_no": "MT0001", "name": "实习生甲", "position": "实习开发", "emp_type": "实习",
                  "category": "", "dept1": "", "dept2": "", "dept3": "",
                  "contract_entity": "", "hire_date": None, "resign_date": None, "confirm_date": None}
        parttime = {"emp_no": "MT0002", "name": "兼职乙", "emp_type": "兼职", "position": "",
                    "category": "", "dept1": "", "dept2": "", "dept3": "",
                    "contract_entity": "", "hire_date": None, "resign_date": None, "confirm_date": None}
        normal = {"emp_no": "MT0003", "name": "正式丙", "emp_type": "正式员工", "position": "工程师",
                  "category": "", "dept1": "总部", "dept2": "", "dept3": "",
                  "contract_entity": "", "hire_date": None, "resign_date": None, "confirm_date": None}
        self.assertTrue(fin._is_final_table_excluded_employee(intern))
        self.assertTrue(fin._is_final_table_excluded_employee(parttime))
        self.assertFalse(fin._is_final_table_excluded_employee(normal))

    def test_emp_no_alias_员工编号(self):
        """工号列识别必须支持'员工编号'别名，且 emp_no 正确保留。"""
        rows = [
            ["员工编号", "员工姓名", "部门名称"],
            ["MT0001", "张三", "总部-产品技术部-后端组"],
        ]
        path = _write_roster(rows)
        try:
            employees = fin.parse_roster(path)
        finally:
            os.remove(path)
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0]["emp_no"], "MT0001")
        self.assertEqual(employees[0]["dept1"], "总部")
        self.assertEqual(employees[0]["dept2"], "产品技术部")
        self.assertEqual(employees[0]["dept3"], "后端组")


class AttendanceIdentityContractTests(unittest.TestCase):
    def _write_attendance(self, rows: list[list]) -> str:
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "月度汇总"
        for row in rows:
            ws.append(row)
        wb.save(path)
        wb.close()
        return path

    def test_attendance_fields_override_legacy_roster_fields(self):
        roster_path = _write_roster([
            ["工号", "姓名", "合同主体", "一级部门", "二级部门", "三级部门", "岗位", "员工类型", "入职日期"],
            ["OLD001", "张三", "本地公司", "旧一级", "旧二级", "旧三级", "旧岗位", "旧类型", "2020-01-01"],
        ])
        attendance_path = self._write_attendance([
            ["姓名", "工号", "考勤组", "一级部门", "二级部门", "三级部门", "职位"],
            ["张三", "OLD001", "总部考勤组", "总部", "研发中心", "平台部", "后端工程师"],
        ])
        try:
            roster = fin.parse_roster(roster_path)
            records = fin.parse_attendance_identity(attendance_path)
            employees = fin.apply_attendance_identity(roster, records)
        finally:
            os.remove(roster_path)
            os.remove(attendance_path)
        employee = employees[0]
        self.assertEqual(employee["emp_no"], "OLD001")
        self.assertEqual((employee["dept1"], employee["dept2"], employee["dept3"]), ("总部", "研发中心", "平台部"))
        self.assertEqual(employee["position"], "后端工程师")
        self.assertEqual(employee["attendance_group"], "总部考勤组")
        self.assertIsNone(employee["contract_entity"])
        self.assertIsNone(employee["emp_type"])
        self.assertIsNone(employee["hire_date"])

    def test_stale_roster_emp_no_uses_new_dingtalk_identity_by_name(self):
        employees = [{"emp_no": "OLD001", "name": "张三"}]
        records = [{
            "emp_no": "NEW001",
            "name": "张三",
            "attendance_group": "标准考勤组",
            "dept1": "总部",
            "position": "工程师",
        }]

        enriched = fin.apply_attendance_identity(employees, records)

        self.assertEqual(enriched[0]["emp_no"], "NEW001")
        self.assertEqual(enriched[0]["attendance_group"], "标准考勤组")
        self.assertEqual(enriched[0]["dept1"], "总部")
        self.assertEqual(enriched[0]["position"], "工程师")

    def test_exact_emp_no_wins_when_another_identity_has_same_name(self):
        employees = [{"emp_no": "E001", "name": "张三"}]
        records = [
            {"emp_no": "E001", "name": "张三", "dept1": "精确工号部门"},
            {"emp_no": "E002", "name": "张三", "dept1": "同名其他部门"},
        ]

        enriched = fin.apply_attendance_identity(employees, records)

        self.assertEqual(enriched[0]["emp_no"], "E001")
        self.assertEqual(enriched[0]["dept1"], "精确工号部门")

    def test_unmatched_emp_no_falls_back_to_unique_normalized_name(self):
        employees = [{"emp_no": "LEGACY001", "name": "张 三"}]
        records = [{"emp_no": "CURRENT001", "name": "张三", "dept1": "总部"}]

        enriched = fin.apply_attendance_identity(employees, records)

        self.assertEqual(enriched[0]["emp_no"], "CURRENT001")
        self.assertEqual(enriched[0]["name"], "张三")
        self.assertEqual(enriched[0]["dept1"], "总部")

    def test_unmatched_emp_no_rejects_ambiguous_name_identities(self):
        employees = [{"emp_no": "LEGACY001", "name": "张三"}]
        records = [
            {"emp_no": "CURRENT001", "name": "张三"},
            {"emp_no": "CURRENT002", "name": "张三"},
        ]

        with self.assertRaises(ValueError) as ctx:
            fin.apply_attendance_identity(employees, records)

        self.assertIn("张三", str(ctx.exception))
        self.assertIn("2 个员工身份", str(ctx.exception))

    def test_unmatched_employee_is_retained_with_empty_identity_fields(self):
        employees = [{
            "emp_no": "OLD001",
            "name": "张三",
            "attendance_group": "旧考勤组",
            "dept1": "旧部门",
            "dept2": "旧二级部门",
            "dept3": "旧三级部门",
            "position": "旧岗位",
        }]

        enriched = fin.apply_attendance_identity(
            employees,
            [{"emp_no": "NEW002", "name": "李四", "dept1": "其他部门"}],
        )

        self.assertEqual(len(enriched), 1)
        self.assertEqual(enriched[0]["name"], "张三")
        for field in ("emp_no", "attendance_group", "dept1", "dept2", "dept3", "position"):
            self.assertFalse(enriched[0][field], field)

    def test_legacy_roster_profile_fields_never_enter_enriched_result(self):
        employees = [{
            "emp_no": "OLD001",
            "name": "张三",
            "contract_entity": "旧合同主体",
            "emp_type": "旧员工类型",
            "category": "旧人员分类",
            "hire_date": date(2020, 1, 1),
            "resign_date": date(2026, 1, 1),
            "confirm_date": date(2020, 4, 1),
        }]

        enriched = fin.apply_attendance_identity(
            employees,
            [{"emp_no": "NEW001", "name": "张三"}],
        )

        for field in (
            "contract_entity", "emp_type", "category",
            "hire_date", "resign_date", "confirm_date",
        ):
            self.assertIsNone(enriched[0][field], field)

    def test_name_only_roster_rejects_ambiguous_attendance_match(self):
        employees = [{"emp_no": "", "name": "张三"}]
        records = [
            {"emp_no": "E001", "name": "张三"},
            {"emp_no": "E002", "name": "张三"},
        ]
        with self.assertRaises(ValueError) as ctx:
            fin.apply_attendance_identity(employees, records)
        self.assertIn("张三", str(ctx.exception))
        self.assertIn("2 个员工身份", str(ctx.exception))

    def test_repeated_punch_rows_for_same_employee_are_merged(self):
        employees = [{"emp_no": "", "name": "张三"}]
        records = [
            {"emp_no": "E001", "name": "张三", "dept1": "总部"},
            {"emp_no": "E001", "name": "张三", "position": "工程师"},
        ]
        enriched = fin.apply_attendance_identity(employees, records)
        self.assertEqual(len(enriched), 1)
        self.assertEqual(enriched[0]["emp_no"], "E001")
        self.assertEqual(enriched[0]["dept1"], "总部")
        self.assertEqual(enriched[0]["position"], "工程师")

    def test_name_roster_and_dingtalk_monthly_record_generate_final_table(self):
        roster_path = _write_roster([["姓名"], ["张三"]])
        attendance_path = self._write_attendance([
            ["姓名", "工号", "考勤组", "部门", "职位", "旷工天数"],
            ["张三", "MT1001", "标准考勤组", "总部-产品技术部-后端组", "高级工程师", 0],
        ])
        output_path = tempfile.mktemp(suffix=".xlsx")
        workdays = {date(2026, 7, day) for day in range(1, 32) if date(2026, 7, day).weekday() < 5}
        schedule_ctx = {
            "year": 2026,
            "month": 7,
            "month_start": date(2026, 7, 1),
            "month_end": date(2026, 7, 31),
            "main_working_days": workdays,
            "chengdu_working_days": workdays,
            "main_attendance_days": len(workdays),
            "chengdu_attendance_days": len(workdays),
            "main_payable_days": workdays,
            "chengdu_payable_days": workdays,
            "main_company_welfare_days": set(),
            "chengdu_company_welfare_days": set(),
            "main_statutory_holidays": set(),
            "chengdu_statutory_holidays": set(),
            "main_holiday_adjust_rest_days": set(),
            "chengdu_holiday_adjust_rest_days": set(),
        }
        try:
            employees = fin.apply_attendance_identity(
                fin.parse_roster(roster_path),
                fin.parse_attendance_identity(attendance_path),
            )
            fin.generate(employees, {}, schedule_ctx, {}, {}, {}, {}, output_path)
            wb = openpyxl.load_workbook(output_path, data_only=False)
            ws = wb.active
            headers = [cell.value for cell in ws[2]]
            values = {header: ws.cell(3, index + 1).value for index, header in enumerate(headers)}
            wb.close()
        finally:
            for path in (roster_path, attendance_path, output_path):
                if os.path.exists(path):
                    os.remove(path)
        self.assertEqual(values["姓名"], "张三")
        self.assertEqual(values["工号"], "MT1001")
        self.assertEqual(values["考勤组"], "标准考勤组")
        self.assertEqual((values["一级部门"], values["二级部门"], values["三级部门"]), ("总部", "产品技术部", "后端组"))
        self.assertEqual(values["岗位"], "高级工程师")
        for field in ("合同主体", "员工类型", "人员分类", "入职日期", "离职日期", "转正日期"):
            self.assertIsNone(values[field], field)

    def test_missing_department_levels_stay_empty(self):
        roster_path = _write_roster([["姓名"], ["张三"]])
        attendance_path = self._write_attendance([
            ["姓名", "工号", "一级部门"],
            ["张三", "MT1001", "总部"],
        ])
        output_path = tempfile.mktemp(suffix=".xlsx")
        workdays = {date(2026, 7, 1)}
        schedule_ctx = {
            "year": 2026,
            "month": 7,
            "month_start": date(2026, 7, 1),
            "month_end": date(2026, 7, 31),
            "main_working_days": workdays,
            "chengdu_working_days": workdays,
            "main_attendance_days": 1,
            "chengdu_attendance_days": 1,
            "main_payable_days": workdays,
            "chengdu_payable_days": workdays,
            "main_company_welfare_days": set(),
            "chengdu_company_welfare_days": set(),
            "main_statutory_holidays": set(),
            "chengdu_statutory_holidays": set(),
            "main_holiday_adjust_rest_days": set(),
            "chengdu_holiday_adjust_rest_days": set(),
        }
        try:
            employees = fin.apply_attendance_identity(
                fin.parse_roster(roster_path),
                fin.parse_attendance_identity(attendance_path),
            )
            fin.generate(employees, {}, schedule_ctx, {}, {}, {}, {}, output_path)
            wb = openpyxl.load_workbook(output_path, data_only=False)
            ws = wb.active
            headers = [cell.value for cell in ws[2]]
            values = {header: ws.cell(3, index + 1).value for index, header in enumerate(headers)}
            wb.close()
        finally:
            for path in (roster_path, attendance_path, output_path):
                if os.path.exists(path):
                    os.remove(path)
        self.assertEqual(values["一级部门"], "总部")
        self.assertIsNone(values["二级部门"])
        self.assertIsNone(values["三级部门"])


class LeaveOvertimeHeaderCompatTests(unittest.TestCase):
    """确保花名册精确匹配不影响请假/加班等模块的子串匹配兼容性。"""

    def test_leave_parser_accepts_historical_employee_headers(self):
        """请假表历史表头：员工工号、员工姓名、请假类型名称。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "请假明细"
        ws.append(["员工工号", "员工姓名", "一级部门", "请假类型名称", "开始时间", "结束时间",
                    "系统时长", "最终请假时长", "最终请假天数", "审批状态", "审批结果"])
        ws.append(["MT0001", "张三", "总部", "事假", "2026-06-15 09:00", "2026-06-15 18:00",
                    8.0, 8.0, 1.0, "完成", "同意"])
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        try:
            rows = fin._collect_deduped_leave_rows(path)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["emp_no"], "MT0001")
            self.assertEqual(rows[0]["name"], "张三")
            self.assertEqual(rows[0]["leave_type"], "事假")
        finally:
            os.remove(path)

    def test_leave_summary_parsing_unaffected_by_exact_match(self):
        """请假明细解析不受 _find_col_exact 影响（仍使用 _find_col 子串匹配）。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "请假明细"
        ws.append(["发起人工号", "发起人姓名", "一级部门", "二级部门", "三级部门",
                    "请假类型", "开始时间", "结束时间", "系统时长", "最终请假时长", "最终请假天数",
                    "备注", "审批编号", "审批状态", "审批结果", "是否实习生", "源文件行号"])
        ws.append(["MT0001", "张三", "总部", "产品技术部", "后端组",
                    "事假", "2026-06-15 09:00", "2026-06-15 18:00", 8.0, 8.0, 1.0,
                    "", "DING-001", "完成", "同意", False, 2])
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        try:
            # parse_leave_summary 应能正常解析（使用 _find_col 子串匹配）
            schedule_ctx = {
                "year": 2026, "month": 6,
                "month_start": date(2026, 6, 1), "month_end": date(2026, 6, 30),
                "main_working_days": set(), "chengdu_working_days": set(),
                "main_expected_attendance_days": set(), "chengdu_expected_attendance_days": set(),
            }
            leave_map = fin.parse_leave_summary(path, schedule_ctx)
            # 应解析出 MT0001 的事假
            self.assertIn("MT0001", leave_map)
            self.assertIn("事假", leave_map["MT0001"])
        finally:
            os.remove(path)

    def test_overtime_parser_accepts_historical_employee_headers(self):
        """加班表历史表头：员工工号、员工姓名、2倍加班（小时）。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "加班明细"
        ws.append(["员工工号", "员工姓名", "一级部门", "二级部门", "三级部门",
                    "开始时间", "结束时间", "2倍加班（小时）", "3倍加班（小时）"])
        ws.append(["MT0001", "张三", "总部", "产品技术部", "后端组",
                    "2026-06-15 09:00", "2026-06-15 21:00", 4.0, 0.0])
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        try:
            overtime = fin.parse_overtime_summary(path)
            self.assertIn("MT0001", overtime)
            self.assertEqual(overtime["MT0001"]["2x_hours"], 4.0)
        finally:
            os.remove(path)


class GenerateRosterActionTests(unittest.TestCase):
    """action_generate_roster 输出必须可被 parse_roster 解析（端到端契约）。"""

    def test_action_output_parsable_by_parse_roster(self):
        from runner import action_generate_roster
        out_dir = tempfile.mkdtemp(prefix="roster-action-")
        try:
            config = {
                "org_name": "测试组织",
                "employees": [
                    {"emp_no": "MT0001", "name": "张三", "dept1": "总部"},
                    {"emp_no": "MT0002", "name": "李四", "dept1": "总部"},
                    {"emp_no": "MT0003", "name": "王五", "dept1": "总部"},
                ],
            }
            outputs = action_generate_roster(config, Path(out_dir))
            self.assertEqual(len(outputs), 1)
            out_path = outputs[0]["path"]
            employees = fin.parse_roster(out_path)
            self.assertEqual(len(employees), 3)
            self.assertEqual([e["name"] for e in employees], ["张三", "李四", "王五"])
            self.assertEqual([e["emp_no"] for e in employees], ["MT0001", "MT0002", "MT0003"])
            self.assertTrue(all(e["dept1"] == "总部" for e in employees))
        finally:
            import shutil
            shutil.rmtree(out_dir, ignore_errors=True)

    def test_action_empty_employees_is_rejected(self):
        from runner import action_generate_roster
        out_dir = tempfile.mkdtemp(prefix="roster-empty-")
        try:
            config = {"org_name": "空组织", "employees": []}
            with self.assertRaisesRegex(ValueError, "没有姓名非空"):
                action_generate_roster(config, Path(out_dir))
        finally:
            import shutil
            shutil.rmtree(out_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
