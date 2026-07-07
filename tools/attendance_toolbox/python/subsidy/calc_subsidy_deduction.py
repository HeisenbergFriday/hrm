"""
calc_subsidy_deduction.py
──────────────────────────────────────────────────────────────────────────────
功能：读取现有补贴扣款表（钉钉导出），剔除羽毛球当日 22:00 后打卡的补贴，
      并补充产研加班补贴，输出核对后的表。

输入：
  --source      现有补贴扣款表 Excel（钉钉导出，含扣款明细4字段，默认: 补贴扣款表.xlsx）
  --checkin     签到表 Excel（默认: 签到表.xlsx）
  --attendance  考勤表 Excel（默认: 系统数据/考勤.xlsx，用于获取每日打卡时间以核算补贴）
  --schedule    作息表 Excel（可选，用于识别法定节假日）
  --output      输出文件路径（默认: 补贴扣款表_核对.xlsx）

处理逻辑：
  扣款明细：迟到/早退扣款沿用源表；旷工天数、缺卡次数从考勤明细重算。
  补贴明细（4列）：从考勤表重新计算，并剔除羽毛球签到当日 22:00 后打卡的人员。
  若提供作息表，会额外识别法定节假日，避免将三倍工资日误算为产研休息日加班补贴。
──────────────────────────────────────────────────────────────────────────────
"""

import argparse
import ast
import math
import os
import re
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.utils.cell import range_boundaries

# ── 默认路径 ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SOURCE     = os.path.join(_HERE, "补贴扣款.xlsx")         # 钉钉导出/现有表
DEFAULT_CHECKIN    = os.path.join(_HERE, "签到表.xlsx")
DEFAULT_ATTENDANCE = os.path.join(_HERE, "考勤.xlsx") # 用于计算补贴
DEFAULT_OUTPUT     = os.path.join(_HERE, "补贴扣款表_核对.xlsx")
DEFAULT_DIFF_OUTPUT = os.path.join(_HERE, "补贴扣款表_差异清单.xlsx")

# ── 产研部门关键字 ─────────────────────────────────────────────────────────────
RD_DEPT_KEYWORDS = [
    "产品中心",
    "研发中心",
    "成都分公司-成都研发",
    "智慧行政事业部-研发小组",
    "ai智慧文创事业部-研发小组",
    "AI数字员工部",
]

# ── 有效工号前缀 ─────────────────────────────────────────────────────────────
VALID_EMP_NO_PREFIXES = ("MT", "TXB", "WB", "JZ")


def _field_key(value) -> str:
    return re.sub(r"[\s\u00a0\u2000-\u200f\u2028-\u202f\u205f\u3000\ufeff]+", "", str(value or "")).strip()

# ── 活动签到表识别 ────────────────────────────────────────────────────────────
CHECKIN_SHEET_KEYWORDS = ("羽毛球", "篮球")

# ── 22:00 补贴门槛 ────────────────────────────────────────────────────────────
LATE_CHECKIN_HOUR = 22          # >= 22:00 → taxi subsidy
TAXI_SUBSIDY      = 50          # 元/次
RD_OT_SUBSIDY     = 50          # 元/次（产研休息日加班 > 4 小时）
RD_OT_THRESHOLD   = 4.0         # 小时
MONTHLY_EXEMPT_MINUTES = 120    # 分钟/人/月
DONATION_PER_MINUTE    = 2      # 元/分钟
WORKDAY_HOURS          = 8      # 小时/天
STATUTORY_HOLIDAY_COLOR = "FFFF0000"
_SCHEDULE_COLOR_NORMALIZE_DISTANCE = 32

# 22点补贴排除规则：
# 1. 客服/售后部门（部门包含"客服"或"售后"）
# 2. 不坐班业务人员：AI智慧文创事业部-销售组
# 3. 排班到22点部门：运营管理中心-运营支撑部
LATE22_EXCLUDED_DEPT_KEYWORDS = (
    "客服",
    "售后",
)
LATE22_EXCLUDED_DEPT_FULL = (
    "AI智慧文创事业部-销售组",        # 不坐班业务
    "运营管理中心-运营支撑部",        # 排班到22点
)
# 晚走补贴强制纳入人员（即使在排除部门中也允许按22点后打卡计算）
LATE22_INCLUDED_NAMES = (
    "崔利华",
    "郑楠杰",
    "徐小辉",
    "丘雨婷",
)

TARGET_FIELDS = (
    "deduct_late",
    "deduct_early",
    "absent_days",
    "missing_punch",
    "late22_days",
    "taxi_subsidy",
    "rd_ot_gt4_days",
    "rd_ot_subsidy",
)

SUBSIDY_FIELDS = frozenset({
    "late22_days",
    "taxi_subsidy",
    "rd_ot_gt4_days",
    "rd_ot_subsidy",
})

OVERWRITE_FIELDS = SUBSIDY_FIELDS | frozenset({
    "absent_days",
    "missing_punch",
})

FIELD_LABELS = {
    "deduct_late": "15-30分钟迟到扣款",
    "deduct_early": "15-30分钟早退扣款",
    "absent_days": "旷工天数",
    "missing_punch": "缺卡次数",
    "late22_days": "晚于22点打卡天数",
    "taxi_subsidy": "晚走补贴",
    "rd_ot_gt4_days": "产研休息日加班>4小时天数",
    "rd_ot_subsidy": "产研休息日加班补贴",
}


def _cell_fg_rgb(cell) -> str:
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb":
            return _normalize_schedule_color(fg.rgb)
    except Exception:
        pass
    return ""


def _rgb_parts(argb: str) -> tuple[int, int, int] | None:
    text = str(argb or "").strip().upper()
    if len(text) == 8:
        text = text[2:]
    if len(text) != 6 or not re.fullmatch(r"[0-9A-F]{6}", text):
        return None
    return int(text[0:2], 16), int(text[2:4], 16), int(text[4:6], 16)


def _normalize_schedule_color(argb: str) -> str:
    text = str(argb or "").strip().upper()
    if len(text) == 6:
        text = "FF" + text
    if text == STATUTORY_HOLIDAY_COLOR:
        return text

    source = _rgb_parts(text)
    target = _rgb_parts(STATUTORY_HOLIDAY_COLOR)
    if source is None or target is None:
        return text

    distance = sum((source[idx] - target[idx]) ** 2 for idx in range(3)) ** 0.5
    if distance <= _SCHEDULE_COLOR_NORMALIZE_DISTANCE:
        return STATUTORY_HOLIDAY_COLOR
    return text


def _parse_schedule_title(title_str: str) -> tuple[int | None, int | None]:
    match = re.search(r"(\d{4})\D*(\d{1,2})\D*月", title_str or "")
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _validate_year_month(year, month) -> tuple[int, int]:
    try:
        year_i = int(year)
        month_i = int(month)
    except (TypeError, ValueError):
        raise ValueError("考勤年月必须同时提供有效的年份和月份。")
    if not (1 <= month_i <= 12):
        raise ValueError(f"考勤月份无效：{month_i}")
    if not (1900 <= year_i <= 2100):
        raise ValueError(f"考勤年份无效：{year_i}")
    return year_i, month_i


def _parse_year_month_from_text(text: str) -> tuple[int, int] | None:
    text = str(text or "").strip()
    if not text:
        return None

    patterns = (
        r"(?<!\d)(20\d{2})\s*年\s*(0?[1-9]|1[0-2])\s*月",
        r"(?<!\d)(20\d{2})[-/.](0?[1-9]|1[0-2])[-/.](?:0?[1-9]|[12]\d|3[01])",
        r"(?<!\d)(20\d{2})[-/.](0?[1-9]|1[0-2])(?=\D|$)",
        r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(?:[0-3]\d)?(?!\d)",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return _validate_year_month(match.group(1), match.group(2))
    return None


def _period_from_cell_value(value) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month
    return _parse_year_month_from_text(str(value or ""))


def _dedupe_texts(items) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in items or []:
        if not item:
            continue
        text = str(item).strip()
        if not text or text in seen:
            continue
        result.append(text)
        seen.add(text)
    return result


def _same_year_month(current: date | None, year: int, month: int) -> bool:
    return current is not None and current.year == year and current.month == month


def filter_dates_to_month(dates, year: int, month: int) -> set[date]:
    """只保留目标年月内的日期。"""
    year, month = _validate_year_month(year, month)
    return {
        current
        for current in (dates or set())
        if _same_year_month(current, year, month)
    }


def _filter_date_map_to_month(date_map: dict[int, date], year: int, month: int) -> dict[int, date]:
    return {
        col: current
        for col, current in date_map.items()
        if _same_year_month(current, year, month)
    }


def resolve_attendance_period(
    filepath: str,
    *,
    source_name: str | None = None,
    context_names: list[str] | tuple[str, ...] | None = None,
    year: int | None = None,
    month: int | None = None,
) -> tuple[int, int, str]:
    """推断考勤统计年月；手动 year/month 传入时优先使用手动值。"""
    if year is not None or month is not None:
        year_i, month_i = _validate_year_month(year, month)
        return year_i, month_i, "手动指定"

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = wb.active

        # 1. 标题/表头：优先读取工作簿前几行的文字日期。
        for row_idx in range(1, min(6, ws.max_row + 1)):
            for col_idx in range(1, min(12, ws.max_column) + 1):
                period = _period_from_cell_value(ws.cell(row_idx, col_idx).value)
                if period:
                    return period[0], period[1], "标题/表头"

        # 2. 文件名：优先使用原始上传文件名，其次使用临时路径文件名。
        filename_candidates = _dedupe_texts([
            source_name,
            os.path.basename(filepath),
        ])
        for text in filename_candidates:
            period = _parse_year_month_from_text(text)
            if period:
                return period[0], period[1], f"文件名：{os.path.basename(text)}"

        # 3. 日期行：如果日期行本身含完整年月日，也可识别。
        for row_idx in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                period = _period_from_cell_value(ws.cell(row_idx, col_idx).value)
                if period:
                    return period[0], period[1], "日期行"
    finally:
        wb.close()

    # 4. 上传上下文：例如补贴扣款源表、签到表、作息表等文件名。
    for text in _dedupe_texts(context_names):
        period = _parse_year_month_from_text(text)
        if period:
            return period[0], period[1], f"上传上下文：{os.path.basename(text)}"

    raise ValueError(
        "未能自动识别考勤年月。请确认考勤表标题/文件名/日期行包含年月，"
        "或在页面“考勤月份”中手动选择年月后再计算。"
    )


def _find_schedule_header_row(ws, title_row_idx: int) -> int | None:
    week_key = _field_key("周数")
    for idx in range(title_row_idx + 1, min(ws.max_row, title_row_idx + 8) + 1):
        if _field_key(ws.cell(idx, 1).value) == week_key:
            return idx
    return None


def _iter_schedule_blocks(wb):
    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            title = str(ws.cell(row_idx, 1).value or "").strip()
            if "作息时间表" not in title:
                continue

            year, month = _parse_schedule_title(title)
            if not year or not month:
                continue

            header_row_idx = _find_schedule_header_row(ws, row_idx)
            if header_row_idx is None:
                continue

            yield {
                "ws": ws,
                "title": title,
                "title_row_idx": row_idx,
                "header_row_idx": header_row_idx,
                "year": year,
                "month": month,
            }


def resolve_schedule_period(filepath: str) -> tuple[int, int, str]:
    """从作息表标题解析本次处理年月，要求同一文件只包含一个年月。"""
    if not filepath or not os.path.exists(filepath):
        raise FileNotFoundError(f"未找到作息表文件：{filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        blocks = list(_iter_schedule_blocks(wb))
    finally:
        wb.close()

    if not blocks:
        raise ValueError("作息表中未解析到有效的作息时间表。")

    month_keys = {(item["year"], item["month"]) for item in blocks}
    if len(month_keys) > 1:
        raise ValueError(f"作息表中包含多个年月：{sorted(month_keys)}，请保留单个月份后再运行。")

    year, month = next(iter(month_keys))
    title = next(item["title"] for item in blocks if item["year"] == year and item["month"] == month)
    return year, month, f"作息表标题：{title}"


def _coerce_schedule_explicit_date(cell_val, target_year: int) -> date | None:
    if isinstance(cell_val, datetime):
        try:
            return date(target_year, cell_val.month, cell_val.day)
        except ValueError:
            return None
    if isinstance(cell_val, date):
        try:
            return date(target_year, cell_val.month, cell_val.day)
        except ValueError:
            return None
    if not isinstance(cell_val, str):
        return None

    text = cell_val.strip()
    match = re.search(r"(?:(20\d{2})\D*)?(0?[1-9]|1[0-2])\s*(?:月|[/\-.])\s*([0-3]?\d)", text)
    if not match:
        return None

    year = int(match.group(1)) if match.group(1) else target_year
    month = int(match.group(2))
    day = int(match.group(3))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _coerce_schedule_day_number(cell_val) -> int | None:
    if isinstance(cell_val, (datetime, date)):
        return cell_val.day
    if isinstance(cell_val, (int, float)) and float(cell_val).is_integer():
        day = int(cell_val)
    else:
        text = str(cell_val or "").strip()
        if not re.fullmatch(r"\d{1,2}", text):
            return None
        day = int(text)
    return day if 1 <= day <= 31 else None


def _schedule_block_date_map(block: dict) -> dict[tuple[int, int], date]:
    ws = block["ws"]
    target_year = block["year"]
    target_month = block["month"]
    header_row_idx = block["header_row_idx"]

    cells = []
    for data_row_idx in range(header_row_idx + 1, ws.max_row + 1):
        first_val = ws.cell(data_row_idx, 1).value
        if isinstance(first_val, (int, float)):
            for col_idx in range(2, min(8, ws.max_column) + 1):
                cell = ws.cell(data_row_idx, col_idx)
                if cell.value is not None and str(cell.value).strip():
                    cells.append((data_row_idx, col_idx, cell.value))
            continue
        if first_val is not None:
            break

    result: dict[tuple[int, int], date] = {}
    current_year = target_year
    current_month = target_month
    prev_day: int | None = None
    initialized = False

    for row_idx, col_idx, value in cells:
        explicit = _coerce_schedule_explicit_date(value, target_year)
        day = _coerce_schedule_day_number(value)
        if explicit is not None:
            current = explicit
            current_year = current.year
            current_month = current.month
            prev_day = current.day
            initialized = True
        elif day is not None:
            if not initialized:
                current_year = target_year
                current_month = target_month
                if day > 20:
                    current_month -= 1
                    if current_month < 1:
                        current_month = 12
                        current_year -= 1
                initialized = True
            elif prev_day is not None and day < prev_day and prev_day > 25:
                current_month += 1
                if current_month > 12:
                    current_month = 1
                    current_year += 1
            try:
                current = date(current_year, current_month, day)
            except ValueError:
                continue
            prev_day = day
        else:
            continue

        result[(row_idx, col_idx)] = current

    return result


def load_statutory_holidays_from_schedule(
    filepath: str,
    target_year: int | None = None,
    target_month: int | None = None,
) -> set[date]:
    """从作息表红底日期中提取法定节假日集合。"""
    if not filepath or not os.path.exists(filepath):
        return set()
    if target_year is not None or target_month is not None:
        target_year, target_month = _validate_year_month(target_year, target_month)

    wb = openpyxl.load_workbook(filepath)
    holidays: set[date] = set()
    try:
        for block in _iter_schedule_blocks(wb):
            year = block["year"]
            month = block["month"]
            if target_year is not None and (year != target_year or month != target_month):
                continue

            date_map = _schedule_block_date_map(block)
            ws = block["ws"]
            for (row_idx, col_idx), current in date_map.items():
                if not _same_year_month(current, year, month):
                    continue
                if _cell_fg_rgb(ws.cell(row_idx, col_idx)) == STATUTORY_HOLIDAY_COLOR:
                    holidays.add(current)
    finally:
        wb.close()

    if target_year is not None:
        holidays = filter_dates_to_month(holidays, target_year, target_month)

    if holidays:
        preview = ", ".join(d.isoformat() for d in sorted(holidays)[:10])
        print(f"[作息表] 已识别法定节假日 {len(holidays)} 天: {preview}")
    else:
        print("[作息表] 未识别到法定节假日红底日期")
    return holidays


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  1. 解析签到表 — 返回 {date: set(names)}                                 ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _find_col(ws, *keywords) -> int | None:
    """在前5行中找包含任意关键字的列号（1-based）。"""
    keyword_keys = [_field_key(k) for k in keywords]
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            val = _field_key(ws.cell(row, col).value)
            if any(k and k in val for k in keyword_keys):
                return col
    return None


def _find_source_detail_col(ws, *keywords) -> int | None:
    """
    源表有两层表头：第3行常有"4月需补回晚走补贴/5月共计..."等汇总列，
    第4行才是本次核对要读写的明细字段。优先读明细表头，避免拿到补回/共计列。
    """
    excluded = ("补回", "补扣", "共计")
    excluded_keys = tuple(_field_key(word) for word in excluded)
    keyword_keys = tuple(_field_key(keyword) for keyword in keywords)
    preferred_rows = range(4, min(8, ws.max_row) + 1)

    for exact_match in (True, False):
        for row in preferred_rows:
            for col in range(1, ws.max_column + 1):
                val = _field_key(ws.cell(row, col).value)
                if not val or any(word in val for word in excluded_keys):
                    continue
                if exact_match and any(val == keyword for keyword in keyword_keys):
                    return col
                if not exact_match and any(keyword and keyword in val for keyword in keyword_keys):
                    return col

    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            val = _field_key(ws.cell(row, col).value)
            if not val or any(word in val for word in excluded_keys):
                continue
            if any(keyword and keyword in val for keyword in keyword_keys):
                return col
    return _find_col(ws, *keywords)


def _find_data_start(ws, col_name: int) -> int | None:
    name_key = _field_key("姓名")
    for row in range(1, min(6, ws.max_row + 1)):
        val = ws.cell(row, col_name).value
        if val and isinstance(val, str) and _field_key(val) not in (name_key, ""):
            return row
    return None


_ATTENDANCE_WEEKDAY_MARKERS = {"一", "二", "三", "四", "五", "六", "日", "六)", "日)"}


def _coerce_day_number(value) -> int | None:
    if isinstance(value, datetime):
        return value.day
    if isinstance(value, date):
        return value.day
    if isinstance(value, (int, float)) and float(value).is_integer():
        day = int(value)
    else:
        text = str(value or "").strip()
        if not re.fullmatch(r"\d{1,2}", text):
            return None
        day = int(text)
    return day if 1 <= day <= 31 else None


def _is_attendance_date_marker(value) -> bool:
    if _coerce_day_number(value) is not None:
        return True
    return str(value or "").strip() in _ATTENDANCE_WEEKDAY_MARKERS


def _find_attendance_date_anchor(ws) -> tuple[int | None, int | None, int | None]:
    """
    在前10行中查找连续的考勤日期列，返回 (日期行, 首个数字日期列, 日期数字)。
    要求连续日期标记不少于5列，避免把汇总数值误判为考勤明细。
    """
    for row_idx in range(1, min(11, ws.max_row + 1)):
        run: list[tuple[int, int | None]] = []
        for col in range(1, ws.max_column + 2):
            value = ws.cell(row_idx, col).value if col <= ws.max_column else None
            if col <= ws.max_column and _is_attendance_date_marker(value):
                run.append((col, _coerce_day_number(value)))
                continue

            numeric_markers = [(c, d) for c, d in run if d is not None]
            if len(run) >= 5 and len(numeric_markers) >= 3:
                anchor_col, anchor_day = numeric_markers[0]
                return row_idx, anchor_col, anchor_day
            run = []
    return None, None, None


def looks_like_attendance_summary(filepath: str) -> bool:
    """判断工作簿是否包含可供补贴模块复用的日期级考勤结果列。"""
    if not filepath or not os.path.exists(filepath):
        return False

    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        ws = wb.active
        if not (_find_col(ws, "姓名") and _find_col(ws, "工号", "员工工号", "员工编号")):
            return False

        date_row, anchor_col, _ = _find_attendance_date_anchor(ws)
        if not date_row or not anchor_col:
            return False

        context_values = []
        for row_idx in range(1, min(date_row, 6) + 1):
            for col_idx in range(max(1, anchor_col - 4), min(ws.max_column, anchor_col + 4) + 1):
                context_values.append(str(ws.cell(row_idx, col_idx).value or ""))
        title_values = [ws.title]
        for row_idx in range(1, min(4, ws.max_row + 1)):
            for col_idx in range(1, min(8, ws.max_column) + 1):
                title_values.append(str(ws.cell(row_idx, col_idx).value or ""))

        context_text = " ".join(context_values)
        title_text = " ".join(title_values)
        return "考勤" in context_text or "月度汇总" in title_text
    finally:
        wb.close()


def _blank_to_none(value):
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return value


def _compact_number(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


_RE_CELL_REF = re.compile(r"(?<![\w'\]])\$?([A-Z]{1,3})\$?(\d+)")


def _numeric_value(value) -> float:
    value = _blank_to_none(value)
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("¥", "").replace(",", "")
    if not text or text == "-":
        return 0.0
    return float(text)


def _sum_formula_arg(arg_text: str, ws_values) -> float:
    total = 0.0
    for part in arg_text.split(","):
        part = part.strip()
        if not part:
            continue
        if "!" in part:
            raise ValueError("external or sheet-qualified refs are not evaluated")
        if ":" in part:
            min_col, min_row, max_col, max_row = range_boundaries(part.replace("$", ""))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    total += _numeric_value(ws_values.cell(row, col).value)
            continue
        match = re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", part)
        if match:
            min_col, min_row, _, _ = range_boundaries(part.replace("$", ""))
            total += _numeric_value(ws_values.cell(min_row, min_col).value)
            continue
        total += _numeric_value(part)
    return total


def _eval_numeric_ast(node):
    if isinstance(node, ast.Expression):
        return _eval_numeric_ast(node.body)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, bool) or not isinstance(node.value, (int, float)):
            raise ValueError("unsupported literal")
        return node.value
    if isinstance(node, ast.UnaryOp):
        value = _eval_numeric_ast(node.operand)
        if isinstance(node.op, ast.UAdd):
            return value
        if isinstance(node.op, ast.USub):
            return -value
        raise ValueError("unsupported unary operator")
    if isinstance(node, ast.BinOp):
        left = _eval_numeric_ast(node.left)
        right = _eval_numeric_ast(node.right)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return left / right
        raise ValueError("unsupported binary operator")
    raise ValueError("unsupported expression")


def _evaluate_simple_formula(formula, ws_values):
    """
    openpyxl 只能读取工作簿里保存的公式缓存；外链/公式缓存未更新时可能是旧值。
    这里仅兜底计算同表内的简单数值公式，避免把 =AF23*50 这类单元格误读成 0。
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return None

    expr = formula[1:].strip()
    if not expr or "!" in expr:
        return None

    try:
        expr = re.sub(
            r"(?i)\bSUM\s*\(([^()]*)\)",
            lambda match: str(_sum_formula_arg(match.group(1), ws_values)),
            expr,
        )

        def repl_cell(match):
            col, row = match.groups()
            min_col, min_row, _, _ = range_boundaries(f"{col}{row}")
            return str(_numeric_value(ws_values.cell(min_row, min_col).value))

        expr = _RE_CELL_REF.sub(repl_cell, expr)
        if not re.fullmatch(r"[\d\s+\-*/().]+", expr):
            return None
        return _compact_number(_eval_numeric_ast(ast.parse(expr, mode="eval")))
    except Exception:
        return None


def _read_source_cell_value(ws_values, ws_formulas, row_idx: int, col_idx: int):
    cached = _blank_to_none(ws_values.cell(row_idx, col_idx).value)
    formula = ws_formulas.cell(row_idx, col_idx).value
    evaluated = _evaluate_simple_formula(formula, ws_values)
    if evaluated is not None:
        return _blank_to_none(evaluated)
    return cached


def _normalize_keywords(keywords) -> list[str]:
    """关键字归一化：去空白、支持逗号分隔、去重并保序。"""
    if not keywords:
        return []
    if isinstance(keywords, str):
        items = [keywords]
    else:
        items = list(keywords)

    normalized: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item is None:
            continue
        for part in str(item).split(','):
            kw = part.strip()
            if not kw or kw in seen:
                continue
            normalized.append(kw)
            seen.add(kw)
    return normalized


def _is_rd_dept(dept1_value, keywords: list[str]) -> bool:
    dept1 = str(dept1_value or '')
    return any(kw in dept1 for kw in keywords)


def _clean_name(name: str) -> str:
    """去除姓名末尾的（已离职）/（离职）标记（兼容全角/半角括号）。"""
    return re.sub(r'[\uff08(](?:已)?离职[)\uff09]\s*$', '', name).strip()


def _normalize_names(names) -> list[str]:
    if not names:
        return []
    if isinstance(names, str):
        items = [names]
    else:
        items = list(names)

    normalized: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item is None:
            continue
        for part in str(item).replace("，", ",").split(","):
            name = _clean_name(part.strip())
            if not name or name in seen:
                continue
            normalized.append(name)
            seen.add(name)
    return normalized


def _is_intern_position(value) -> bool:
    return "实习" in str(value or "").strip()


def is_intern_record(record: dict) -> bool:
    return bool(record.get("is_intern")) or _is_intern_position(record.get("pos"))


def split_intern_records(records: list[dict]) -> tuple[list[dict], list[dict]]:
    non_intern_records: list[dict] = []
    intern_records: list[dict] = []
    for record in records:
        if is_intern_record(record):
            intern_records.append(record)
        else:
            non_intern_records.append(record)
    return non_intern_records, intern_records


def _normalize_emp_no(value):
    """工号归一化为字符串：去空白，数值型去掉小数 .0。"""
    value = _blank_to_none(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def _normalize_compare_value(value):
    value = _blank_to_none(value)
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    text = str(value).strip()
    if not text:
        return 0
    try:
        return round(float(text), 6)
    except ValueError:
        return text


def _values_match(existing, computed) -> bool:
    return _normalize_compare_value(existing) == _normalize_compare_value(computed)


def _empty_result(*keys: str) -> dict:
    return {key: None for key in keys}


def _should_exclude_late22_count(record: dict, included_names=None) -> bool:
    """
    判断员工是否应排除22点补贴。
    排除条件：
    1. 客服/售后部门（部门包含"客服"或"售后"）
    2. 不坐班业务人员：AI智慧文创事业部-销售组
    3. 排班到22点部门：运营管理中心-运营支撑部

    强制纳入名单优先于排除规则。
    """
    # 强制纳入名单优先
    name = str(record.get("name") or "").strip()
    default_names = LATE22_INCLUDED_NAMES if included_names is None else included_names
    if name in set(_normalize_names(default_names)):
        return False

    dept = str(record.get("dept1") or "").strip()
    if not dept:
        return False

    # 检查是否包含排除关键字
    for keyword in LATE22_EXCLUDED_DEPT_KEYWORDS:
        if keyword in dept:
            return True

    # 检查是否以排除部门开头（处理子部门）
    for excluded_dept in LATE22_EXCLUDED_DEPT_FULL:
        if dept.startswith(excluded_dept):
            return True

    return False


def parse_activity_checkin(filepath: str) -> dict[date, set]:
    """
    读取签到 Excel，提取活动签到表中「已参加」名单。
    返回 {date: {name, ...}}。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result: dict[date, set] = {}

    for sname in wb.sheetnames:
        if not any(keyword in sname for keyword in CHECKIN_SHEET_KEYWORDS):
            continue
        ws = wb[sname]
        date_row_idx = None
        date_cols: list[tuple[int, date]] = []
        name_cols: list[int] = []

        for row in ws.iter_rows():
            if not any(isinstance(cell.value, (datetime, date)) for cell in row):
                continue
            date_row_idx = row[0].row
            for cell in row:
                if isinstance(cell.value, datetime):
                    # datetime 是 date 的子类，需先判断
                    date_cols.append((cell.column, cell.value.date()))
                elif isinstance(cell.value, date):
                    # 纯日期格式（无时间部分），openpyxl 返回 date 而非 datetime
                    date_cols.append((cell.column, cell.value))
                elif isinstance(cell.value, str) and "姓名" in cell.value:
                    name_cols.append(cell.column)
            break

        if not date_row_idx or not date_cols or not name_cols:
            continue

        name_cols = sorted(set(name_cols))
        for idx, name_col in enumerate(name_cols):
            next_name_col = name_cols[idx + 1] if idx + 1 < len(name_cols) else ws.max_column + 1
            group_dates = [
                (col_idx, current_date)
                for col_idx, current_date in date_cols
                if name_col < col_idx < next_name_col
            ]
            if not group_dates:
                continue
            for r in range(date_row_idx + 1, ws.max_row + 1):
                name_val = ws.cell(r, name_col).value
                if not name_val or not isinstance(name_val, str):
                    continue
                name = name_val.strip()
                if not name:
                    continue
                for date_col, current_date in group_dates:
                    if ws.cell(r, date_col).value == "已参加":
                        result.setdefault(current_date, set()).add(name)

    return result


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  2. 解析现有补贴扣款表（钉钉导出）— 提取员工信息 + 扣款4字段             ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def parse_source_table(
    filepath: str,
    sheet: str | None = None,
    rd_dept_keywords: list[str] | None = None,
) -> list[dict]:
    """
    读取现有补贴扣款表，自动识别关键列位置（兼容不同格式/列序）。
    返回 [{row_idx, name, group, dept1, dept2, dept3, pos, is_rd, ...}]
    """
    wb_values = openpyxl.load_workbook(filepath, data_only=True)
    wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
    try:
        ws = wb_values[sheet] if sheet else wb_values.active
        ws_formula = wb_formulas[ws.title]
        rd_keywords = _normalize_keywords(rd_dept_keywords or RD_DEPT_KEYWORDS)

        col_name   = _find_col(ws, '姓名')   or 1
        col_emp_no = _find_col(ws, '工号', '员工工号', '员工编号')
        col_group  = _find_col(ws, '考勤组') or 2
        col_dept1  = _find_col(ws, '一级部门') or 3
        col_dept2  = _find_col(ws, '二级部门') or 4
        col_dept3  = _find_col(ws, '三级部门') or 5
        col_pos    = _find_col(ws, '职位')   or 6
        col_map = {
            'deduct_late': _find_source_detail_col(ws, '15-30分钟迟到扣款', '迟到扣款'),
            'deduct_early': _find_source_detail_col(ws, '15-30分钟早退扣款', '早退扣款'),
            'absent_days': _find_source_detail_col(ws, '旷工天数'),
            'missing_punch': _find_source_detail_col(ws, '缺卡次数'),
            'late22_days': _find_source_detail_col(ws, '晚于22点打卡天数', '晚于22点'),
            'taxi_subsidy': _find_source_detail_col(ws, '晚走补贴'),
            'rd_ot_gt4_days': _find_source_detail_col(ws, '产研休息日加班>4小时天数', '加班>4小时天数'),
            'rd_ot_subsidy': _find_source_detail_col(ws, '产研休息日加班补贴'),
        }

        if not all(col_map.values()):
            missing = [name for col, name in [
                (col_map['deduct_late'], '迟到扣款'),
                (col_map['deduct_early'], '早退扣款'),
                (col_map['absent_days'], '旷工天数'),
                (col_map['missing_punch'], '缺卡次数'),
                (col_map['late22_days'], '晚于22点打卡天数'),
                (col_map['taxi_subsidy'], '晚走补贴'),
                (col_map['rd_ot_gt4_days'], '产研休息日加班>4小时天数'),
                (col_map['rd_ot_subsidy'], '产研休息日加班补贴'),
            ] if not col]
            raise ValueError(f"源表中未找到以下列，请检查表头: {missing}")

        data_start = _find_data_start(ws, col_name)
        if data_start is None:
            return []

        result = []
        for r in range(data_start, ws.max_row + 1):
            name = ws.cell(r, col_name).value
            if not name or not isinstance(name, str):
                continue
            name = _clean_name(name.strip())
            if name in ('姓名', '合计', ''):
                continue
            emp_no = _normalize_emp_no(ws.cell(r, col_emp_no).value) if col_emp_no else None
            if not emp_no or not emp_no.startswith(VALID_EMP_NO_PREFIXES):
                continue
            dept1 = str(ws.cell(r, col_dept1).value or '')
            position = ws.cell(r, col_pos).value
            is_rd = _is_rd_dept(dept1, rd_keywords)
            row = {
                'row_idx':       r,
                'emp_no':        emp_no,
                'name':          name,
                'group':         ws.cell(r, col_group).value,
                'dept1':         ws.cell(r, col_dept1).value,
                'dept2':         ws.cell(r, col_dept2).value,
                'dept3':         ws.cell(r, col_dept3).value,
                'pos':           position,
                'is_intern':     _is_intern_position(position),
                'is_rd':         is_rd,
            }
            for field, col_idx in col_map.items():
                row[field] = _read_source_cell_value(ws, ws_formula, r, col_idx)
            result.append(row)
        return result
    finally:
        wb_values.close()
        wb_formulas.close()


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  3. 解析单个考勤单元格（仅用于获取打卡时间和加班时长，供补贴计算用）      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

_RE_LATE   = re.compile(r'(?:上班|严重)?(?:旷工)?迟到\s*(\d+)\s*分钟')
_RE_EARLY  = re.compile(r'(?:下班)?早退\s*(\d+)\s*分钟')
_RE_TIMES  = re.compile(r'\((\d{1,2}:\d{2}|-)(?:,(\d{1,2}:\d{2}|-))\)')
_RE_OT_HRS = re.compile(r'加班\d{2}-\d{2} \d{2}:\d{2}到\d{2}-\d{2} \d{2}:\d{2}\s*([\d.]+)\s*小时')


def _parse_time(ts: str):
    """'HH:MM' → (h, m)，'-' 或 None → None"""
    if not ts or ts == '-':
        return None
    parts = ts.split(':')
    try:
        return int(parts[0]), int(parts[1])
    except Exception:
        return None


def parse_cell(text) -> dict:
    """
    解析一个考勤单元格文字，返回:
      late_min        : int   – 迟到分钟数（系统注明或出差时从打卡时间推算）
      early_min       : int   – 早退分钟数
      is_absent       : bool  – 完全旷工（无打卡）
      is_roost_late   : bool  – 旷工级迟到（含 '旷工迟到'）
      missing_punch   : bool  – 含缺卡标记
      checkin         : tuple or None  – (h, m) 上班打卡时间
      checkout        : tuple or None  – (h, m) 下班打卡时间
      is_rest         : bool  – 休息日（无需出勤）
      is_weekend_work : bool  – 休息日但来了（休息并打卡）
      is_biztrip_work : bool  – 出差中实际到岗（有打卡记录）
      is_standard_shift : bool – 标准时间体系（含有下班打卡的出差工作日），用于 22:00 补贴判断
      ot_hours        : float – 休息日批准加班小时数（来自"加班X小时"文字）
    """
    r = dict(
        late_min=0, early_min=0,
        is_absent=False, is_roost_late=False,
        missing_punch=False, checkin=None, checkout=None,
        is_rest=False, is_weekend_work=False, is_biztrip_work=False,
        is_standard_shift=False,
        ot_hours=0.0,
    )
    if not text:
        r['is_rest'] = True
        return r

    text = str(text).strip()
    first_line = text.split('\n')[0]

    # 休息日判定
    if first_line.startswith('休息'):
        if '休息并打卡' in first_line:
            r['is_weekend_work'] = True
        else:
            r['is_rest'] = True
            return r
    elif first_line.startswith('休') and '休息' not in first_line:
        r['is_rest'] = True
        return r

    # 打卡时间（最后一行括号内）—— 在 is_leave 判断之前提取
    last_line = text.split('\n')[-1] if '\n' in text else text
    mt = _RE_TIMES.search(last_line)
    if mt:
        r['checkin']  = _parse_time(mt.group(1))
        r['checkout'] = _parse_time(mt.group(2))

    # 完全旷工
    if '旷工\n(-)' in text or text.strip() == '标准:旷工\n(-)' or \
       (first_line.startswith('标准:旷工') and '迟到' not in first_line):
        r['is_absent'] = True

    # 旷工迟到
    if '旷工迟到' in text:
        r['is_roost_late'] = True

    # 缺卡
    if '缺卡' in text:
        r['missing_punch'] = True

    # 出差工作日（出差中实际到岗）
    is_biztrip = first_line.startswith('出差')
    if is_biztrip and (r['checkin'] is not None or r['checkout'] is not None):
        r['is_biztrip_work'] = True

    # 标准班次判定（用于 22:00 晚走补贴）：
    # - "标准:" 或 "年会:" 前缀 → 标准班次
    # - "外出..." 前缀（临时外勤，非全天外勤班）→ 仍归属标准时间体系
    # - "出差..." 且有下班打卡 → 仍归属标准时间体系，可计 22:00 晚走补贴
    # - "上班外勤" / "下班外勤" 整日外勤班 → 排除
    # - "中班:" → 排除（非标准作息）
    is_biztrip_standard = is_biztrip and r['checkout'] is not None
    is_std_prefix = (first_line.startswith('标准:') or
                     first_line.startswith('年会:') or
                     first_line.startswith('外出') or
                     is_biztrip_standard)
    is_waiqin = '上班外勤' in first_line or '下班外勤' in first_line
    r['is_standard_shift'] = is_std_prefix and not is_waiqin

    # 请假关键字（仅用于控制是否从打卡时间推算迟到）
    is_leave_keyword = any(kw in first_line for kw in [
        '事假', '年假', '调休', '病假', '产假', '婚假', '陪产假',
        '上班外勤', '下班外勤',
    ])

    # ── 迟到分钟 ────────────────────────────────────────────────────────────
    # 规则1：始终提取系统已注明的「迟到X分钟」（含哺乳假班次、含年假的混合行）
    if not r['is_absent']:
        m = _RE_LATE.search(text)
        if m:
            r['late_min'] = int(m.group(1))

    # 规则2：出差工作日且无显式「迟到」时，从打卡时间推算（相对 09:00）
    # 仅当同时有上班和下班打卡时才计算（缺少打卡一侧可能是临时签到，不可靠）
    if r['late_min'] == 0 and r['is_biztrip_work'] and not is_leave_keyword:
        ci = r['checkin']
        co = r['checkout']   # 需要下班打卡存在，说明确实在办公室全勤
        if ci and co:
            late_min = (ci[0] - 9) * 60 + ci[1]
            if late_min > 0:
                r['late_min'] = late_min

    # ── 早退分钟 ────────────────────────────────────────────────────────────
    # 始终提取系统已注明的「早退X分钟」（含哺乳假班次）
    m = _RE_EARLY.search(text)
    if m:
        r['early_min'] = int(m.group(1))

    # 休息日批准加班小时数（产研补贴用）
    ot_m = _RE_OT_HRS.search(text)
    if ot_m:
        try:
            r['ot_hours'] = float(ot_m.group(1))
        except ValueError:
            pass

    return r


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  4. 计算补贴（仅补贴4字段，扣款字段由源表提供）                          ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def parse_attendance_result(filepath: str, year: int, month: int) -> dict[str, set[date]]:
    """
    读取考勤结果表（含批注），提取"已补流程"的旷工记录。
    返回 {员工姓名: {日期, ...}}。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 日期行在第3行，数据从第4行开始
    date_row_idx = 3
    data_start_row = 4

    # 构建列号→日期的映射（处理跨月）
    date_map: dict[int, date] = {}
    current_year = year
    current_month = month

    # 找到第一个有效日期，判断是否从上月开始
    first_day = None
    for col in range(4, ws.max_column + 1):
        try:
            first_day = int(ws.cell(date_row_idx, col).value)
            if 1 <= first_day <= 31:
                break
        except (ValueError, TypeError):
            continue
        first_day = None

    if first_day is not None and first_day > 25:
        current_month -= 1
        if current_month < 1:
            current_month = 12
            current_year -= 1

    prev_day = first_day if first_day is not None else 99  # 用于检测月份切换

    for col in range(4, ws.max_column + 1):
        val = ws.cell(date_row_idx, col).value
        if val is None:
            continue
        try:
            d = int(val)
        except (ValueError, TypeError):
            continue  # 跳过"六""日"等文字
        if not (1 <= d <= 31):
            continue

        # 日期从大变小（如31→1）表示月份切换
        if d < prev_day and prev_day > 25:
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

        try:
            date_map[col] = date(current_year, current_month, d)
        except ValueError:
            pass
        prev_day = d

    original_date_count = len(date_map)
    date_map = _filter_date_map_to_month(date_map, year, month)
    skipped_date_count = original_date_count - len(date_map)
    if skipped_date_count:
        print(f"[考勤结果表] 已忽略 {skipped_date_count} 个非 {year}年{month}月的日期列")

    result: dict[str, set[date]] = {}

    for row_idx in range(data_start_row, ws.max_row + 1):
        name_val = ws.cell(row_idx, 1).value
        if not name_val or not isinstance(name_val, str):
            continue
        name = name_val.strip()

        for col, d in date_map.items():
            cell = ws.cell(row_idx, col)
            if cell.comment and "已补流程" in (cell.comment.text or ""):
                result.setdefault(name, set()).add(d)

    return result


def _collect_calculation_details(
    day_cells: list[tuple[date, dict]],
    activity_days: dict[date, set],
    emp_name: str,
    is_rd: bool,
    approved_dates: set[date] | None,
    legal_holidays: set[date] | None,
    exclude_late22_count: bool = False,
) -> dict[str, str]:
    """
    收集各字段的计算明细，用于差异报告。
    返回字典：字段名 -> 明细说明文本
    """
    details = {}

    # 22:00 补贴明细
    late22_dates = []
    late22_excluded = []

    # 产研加班明细
    rd_ot_dates = []
    rd_ot_excluded = []

    # 扣款明细
    late_details = []
    early_details = []
    absent_details = []
    missing_punch_dates = []

    approved = approved_dates or set()
    holiday_dates = legal_holidays or set()
    remaining_exempt = MONTHLY_EXEMPT_MINUTES

    for d, cell in day_cells:
        if cell['is_rest']:
            continue

        # 休息日加班
        if cell['is_weekend_work']:
            if d in holiday_dates:
                rd_ot_excluded.append(f"{d.strftime('%m-%d')}(法定节假日)")
            elif is_rd:
                if cell['ot_hours'] >= RD_OT_THRESHOLD:
                    rd_ot_dates.append(f"{d.strftime('%m-%d')}({cell['ot_hours']:.1f}h)")
                elif cell['ot_hours'] == 0:
                    ci = cell.get('checkin')
                    co = cell.get('checkout')
                    if ci and co:
                        # 修复：先转换为总分钟数再计算小时数，避免负数
                        # 处理跨午夜情况
                        ci_minutes = ci[0] * 60 + ci[1]
                        co_minutes = co[0] * 60 + co[1]
                        if co_minutes < ci_minutes:
                            # 跨午夜：下班时间在第二天
                            co_minutes += 24 * 60
                        total_minutes = co_minutes - ci_minutes
                        actual_hours = total_minutes / 60.0
                        if actual_hours >= RD_OT_THRESHOLD:
                            rd_ot_dates.append(f"{d.strftime('%m-%d')}({actual_hours:.1f}h实际)")

            # 休息日加班仍需检查缺卡
            if cell['missing_punch']:
                missing_punch_dates.append(d.strftime('%m-%d'))

            continue

        if cell['is_absent']:
            if approved_dates is not None and d in approved:
                # 已补流程的旷工不计入明细（因为不计入总小时数）
                pass
            else:
                absent_details.append(f"{d.strftime('%m-%d')}({WORKDAY_HOURS}h)")
            continue

        # 22:00 补贴
        checkout = cell['checkout']
        if (cell['is_standard_shift'] and checkout
                and (checkout[0] >= LATE_CHECKIN_HOUR or checkout[0] < 6)):
            if exclude_late22_count:
                late22_excluded.append(f"{d.strftime('%m-%d')}(排除计数)")
            elif emp_name in activity_days.get(d, set()):
                late22_excluded.append(f"{d.strftime('%m-%d')}(活动签到)")
            else:
                late22_dates.append(f"{d.strftime('%m-%d')}({checkout[0]:02d}:{checkout[1]:02d})")

        # 缺卡
        if cell['missing_punch']:
            missing_punch_dates.append(d.strftime('%m-%d'))

        # 迟到早退
        total_minutes = cell['late_min'] + cell['early_min']
        if total_minutes > 0:
            if total_minutes <= 15:
                # 豁免额度处理：与 compute_deduction() 保持一致
                exempt_used = min(remaining_exempt, total_minutes)
                remaining_exempt = max(0, remaining_exempt - total_minutes)

                if cell['late_min'] > 0:
                    late_details.append(f"{d.strftime('%m-%d')}({cell['late_min']}分,豁免)")
                if cell['early_min'] > 0:
                    early_details.append(f"{d.strftime('%m-%d')}({cell['early_min']}分,豁免)")
            elif total_minutes <= 30:
                if cell['late_min'] > 0:
                    deduct = cell['late_min'] * DONATION_PER_MINUTE
                    late_details.append(f"{d.strftime('%m-%d')}({cell['late_min']}分,¥{deduct})")
                if cell['early_min'] > 0:
                    deduct = cell['early_min'] * DONATION_PER_MINUTE
                    early_details.append(f"{d.strftime('%m-%d')}({cell['early_min']}分,¥{deduct})")
            else:
                # 旷工：根据实际情况优化显示格式
                hours = max(1, math.ceil(total_minutes / 60))
                if cell['late_min'] > 0 and cell['early_min'] > 0:
                    detail_text = f"{d.strftime('%m-%d')}(迟{cell['late_min']}分+早{cell['early_min']}分,计{hours}h旷工)"
                elif cell['late_min'] > 0:
                    detail_text = f"{d.strftime('%m-%d')}(迟{cell['late_min']}分,计{hours}h旷工)"
                else:
                    detail_text = f"{d.strftime('%m-%d')}(早{cell['early_min']}分,计{hours}h旷工)"

                # 添加到迟到或早退明细中（优先迟到）
                if cell['late_min'] > 0:
                    late_details.append(detail_text)
                else:
                    early_details.append(detail_text)

    # 组装明细文本
    if late22_dates:
        details['late22_days'] = f"计入: {', '.join(late22_dates)}"
        if late22_excluded:
            details['late22_days'] += f" | 排除: {', '.join(late22_excluded)}"
    elif late22_excluded:
        details['late22_days'] = f"排除: {', '.join(late22_excluded)}"

    if late22_dates:
        details['taxi_subsidy'] = f"¥{len(late22_dates) * TAXI_SUBSIDY} ({len(late22_dates)}天×¥{TAXI_SUBSIDY})"

    if rd_ot_dates:
        details['rd_ot_gt4_days'] = f"计入: {', '.join(rd_ot_dates)}"
        if rd_ot_excluded:
            details['rd_ot_gt4_days'] += f" | 排除: {', '.join(rd_ot_excluded)}"
    elif rd_ot_excluded:
        details['rd_ot_gt4_days'] = f"排除: {', '.join(rd_ot_excluded)}"

    if rd_ot_dates:
        details['rd_ot_subsidy'] = f"¥{len(rd_ot_dates) * RD_OT_SUBSIDY} ({len(rd_ot_dates)}天×¥{RD_OT_SUBSIDY})"

    if late_details:
        details['deduct_late'] = '; '.join(late_details)

    if early_details:
        details['deduct_early'] = '; '.join(early_details)

    # 修复：即使没有完全旷工的天，也要显示迟到早退导致的旷工
    total_absent_hours = 0
    for d, cell in day_cells:
        if cell['is_rest'] or cell['is_weekend_work']:
            continue
        if cell['is_absent']:
            if approved_dates is None or d not in approved:
                total_absent_hours += WORKDAY_HOURS
        else:
            total_minutes = cell['late_min'] + cell['early_min']
            if total_minutes > 30:
                total_absent_hours += max(1, math.ceil(total_minutes / 60))

    if total_absent_hours > 0:
        if absent_details:
            details['absent_days'] = f"{', '.join(absent_details)} | 合计: {total_absent_hours / WORKDAY_HOURS:.3f}天"
        else:
            # 只有迟到早退导致的旷工，没有完全旷工的天
            details['absent_days'] = f"合计: {total_absent_hours / WORKDAY_HOURS:.3f}天 (由迟到早退累计)"

    if missing_punch_dates:
        details['missing_punch'] = ', '.join(missing_punch_dates)

    return details


def compute_subsidy(
    day_cells: list[tuple[date, dict]],
    activity_days: dict[date, set],
    emp_name: str,
    is_rd: bool,
    exclude_late22_count: bool = False,
    approved_dates: set[date] | None = None,
    legal_holidays: set[date] | None = None,
) -> dict:
    """
    从考勤明细中计算 22:00 补贴和产研加班补贴，并剔除活动签到当日打卡人员。
    普通休息日加班仅计入休息日加班补贴，不计入"晚于22点打卡天数/晚走补贴"。
    法定节假日优先按加班工资口径处理，不计入"产研休息日加班补贴"。
    day_cells 在解析阶段已按本次作息表月份过滤，跨月补回不在本次补贴中计算。
    扣款字段（迟到/早退扣款、旷工、缺卡）不在此计算，直接用源表数据。
    """
    late22_days    = 0
    taxi_subsidy   = 0
    rd_ot_gt4_days = 0
    rd_ot_subsidy  = 0

    approved = approved_dates or set()
    holiday_dates = legal_holidays or set()

    for d, cell in day_cells:
        if cell['is_rest']:
            continue

        # ── 休息日加班 ───────────────────────────────────────────────────────
        if cell['is_weekend_work']:
            if d in holiday_dates:
                continue
            if is_rd:
                if cell['ot_hours'] >= RD_OT_THRESHOLD:
                    rd_ot_gt4_days += 1
                    rd_ot_subsidy  += RD_OT_SUBSIDY
                elif cell['ot_hours'] == 0:
                    # 无"加班"注释的休息并打卡：按打卡时间差计算实际工时
                    ci = cell.get('checkin')
                    co = cell.get('checkout')
                    if ci and co:
                        # 修复：先转换为总分钟数再计算小时数，避免负数
                        # 处理跨午夜情况
                        ci_minutes = ci[0] * 60 + ci[1]
                        co_minutes = co[0] * 60 + co[1]
                        if co_minutes < ci_minutes:
                            # 跨午夜：下班时间在第二天
                            co_minutes += 24 * 60
                        total_minutes = co_minutes - ci_minutes
                        actual_hours = total_minutes / 60.0
                        if actual_hours >= RD_OT_THRESHOLD:
                            rd_ot_gt4_days += 1
                            rd_ot_subsidy  += RD_OT_SUBSIDY
            continue

        if cell['is_absent']:
            continue

        # ── 22:00 补贴（仅标准班次，剔除当日羽毛球打卡人员）─────────────────
        if exclude_late22_count:
            continue
        checkout = cell['checkout']
        if (cell['is_standard_shift'] and checkout
                and (checkout[0] >= LATE_CHECKIN_HOUR or checkout[0] < 6)):
            if emp_name not in activity_days.get(d, set()):
                late22_days  += 1
                taxi_subsidy += TAXI_SUBSIDY

    return dict(
        late22_days    = late22_days    or None,
        taxi_subsidy   = taxi_subsidy   or None,
        rd_ot_gt4_days = rd_ot_gt4_days or None,
        rd_ot_subsidy  = rd_ot_subsidy  or None,
    )


def compute_deduction(day_cells: list[tuple[date, dict]], approved_dates: set[date] | None = None) -> dict:
    """
    从考勤明细中计算扣款字段。
    """
    remaining_exempt = MONTHLY_EXEMPT_MINUTES
    deduct_late = 0
    deduct_early = 0
    absent_hours = 0
    missing_punch = 0
    approved = approved_dates or set()

    for d, cell in day_cells:
        if cell['is_rest']:
            continue

        # 休息日加班仍需检查缺卡
        if cell['is_weekend_work']:
            if cell['missing_punch']:
                missing_punch += 1
            continue

        if cell['missing_punch']:
            missing_punch += 1

        if cell['is_absent']:
            # 排除已补流程的旷工记录
            if approved_dates is not None and d in approved:
                continue
            absent_hours += WORKDAY_HOURS
            continue

        total_minutes = cell['late_min'] + cell['early_min']
        if total_minutes <= 0:
            continue

        if total_minutes <= 15:
            remaining_exempt = max(0, remaining_exempt - total_minutes)
            continue

        if total_minutes <= 30:
            deduct_late += cell['late_min'] * DONATION_PER_MINUTE
            deduct_early += cell['early_min'] * DONATION_PER_MINUTE
            continue

        absent_hours += max(1, math.ceil(total_minutes / 60))

    absent_days = absent_hours / WORKDAY_HOURS if absent_hours else None
    absent_days = round(absent_days, 3) if absent_days else None

    return {
        'deduct_late': deduct_late or None,
        'deduct_early': deduct_early or None,
        'absent_days': _compact_number(absent_days),
        'missing_punch': missing_punch or None,
    }


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  5. 解析考勤表（只提取打卡时间 / 休息日加班时长，供补贴计算用）          ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def parse_attendance(
    filepath: str,
    rd_dept_keywords: list[str] | None = None,
    *,
    year: int | None = None,
    month: int | None = None,
    period_source: str | None = None,
    source_name: str | None = None,
    context_names: list[str] | tuple[str, ...] | None = None,
) -> list[dict]:
    """
    读取考勤月度汇总 Excel，返回每个员工的完整记录列表。
    每条记录包含 info_cols + stats。
    """
    if year is not None or month is not None:
        year, month = _validate_year_month(year, month)
        period_source = period_source or "手动指定"
    else:
        year, month, period_source = resolve_attendance_period(
            filepath,
            source_name=source_name,
            context_names=context_names,
        )
    print(f"[考勤年月] 使用 {year}年{month}月（{period_source}）")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active  # 月度汇总
    col_emp_no = _find_col(ws, '工号', '员工工号', '员工编号')
    rd_keywords = _normalize_keywords(rd_dept_keywords or RD_DEPT_KEYWORDS)

    # 自动检测日期行：在前10行中查找连续的日期/周末标记列。
    date_row, anchor_col, anchor_day = _find_attendance_date_anchor(ws)

    # 构建日期映射（支持跨月）
    date_map: dict[int, date] = {}
    if anchor_col is not None and date_row is not None and anchor_day is not None:
        current_year = year
        current_month = month

        # 如果锚点日大于25，说明日期从上个月开始（如30, 31, 1, 2...）
        if anchor_day > 25:
            current_month -= 1
            if current_month < 1:
                current_month = 12
                current_year -= 1

        prev_day = anchor_day - 1  # 用于检测月份切换
        prev_col = None

        # 处理数字日期列，并顺推"六"/"日"等周末标记列。
        for col in range(anchor_col, ws.max_column + 1):
            day_val = ws.cell(date_row, col).value
            if day_val is None:
                continue

            d = _coerce_day_number(day_val)
            if d is not None:
                # 检测月份切换：当日期从大变小（如31→1）时，月份+1
                if d < prev_day and prev_day > 25:
                    current_month += 1
                    if current_month > 12:
                        current_month = 1
                        current_year += 1
                # 尝试创建日期
                try:
                    date_map[col] = date(current_year, current_month, d)
                    prev_day = d
                    prev_col = col
                except ValueError:
                    # 无效日期（如4月31日）→ 月份+1，日期=1
                    current_month += 1
                    if current_month > 12:
                        current_month = 1
                        current_year += 1
                    try:
                        date_map[col] = date(current_year, current_month, 1)
                        prev_day = 1
                        prev_col = col
                    except ValueError:
                        pass
                continue

            if str(day_val).strip() in _ATTENDANCE_WEEKDAY_MARKERS and prev_col is not None:
                inferred = date_map[prev_col] + timedelta(days=col - prev_col)
                date_map[col] = inferred
                prev_day = inferred.day
                current_year = inferred.year
                current_month = inferred.month
                prev_col = col

        # 如果月份开头是周末标记（例如"六"、"日"后才出现数字日期），从锚点向前补齐。
        if anchor_col in date_map:
            for col in range(anchor_col - 1, 0, -1):
                marker = str(ws.cell(date_row, col).value or "").strip()
                if marker not in _ATTENDANCE_WEEKDAY_MARKERS:
                    break
                date_map[col] = date_map[anchor_col] - timedelta(days=anchor_col - col)

    original_date_count = len(date_map)
    date_map = _filter_date_map_to_month(date_map, year, month)
    skipped_date_count = original_date_count - len(date_map)
    if skipped_date_count:
        print(f"[考勤年月] 已忽略 {skipped_date_count} 个非 {year}年{month}月的日期列")
    if original_date_count and not date_map:
        raise ValueError(f"考勤表未找到 {year}年{month}月的日期列，请检查作息表月份和考勤表是否匹配。")

    # 确定员工数据起始行（日期行的下一行）
    data_start_row = (date_row + 1) if date_row else 4

    return_list = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        emp_name = ws.cell(row_idx, 1).value
        if not emp_name or not isinstance(emp_name, str):
            continue
        emp_name = _clean_name(emp_name.strip())
        emp_no = _normalize_emp_no(ws.cell(row_idx, col_emp_no).value) if col_emp_no else None
        if not emp_no or not emp_no.startswith(VALID_EMP_NO_PREFIXES):
            continue

        # 从源表获取部门信息（考勤表可能没有部门列）
        # 先尝试从列3读取
        dept_raw = ws.cell(row_idx, 3).value
        dept1 = str(dept_raw or '')
        is_rd = _is_rd_dept(dept1, rd_keywords)

        info = {
            'emp_no': emp_no,
            'name':  emp_name,
            'group': ws.cell(row_idx, 2).value,
            'dept1': dept_raw,
            'dept2': ws.cell(row_idx, 4).value if ws.max_column > 3 else None,
            'dept3': ws.cell(row_idx, 5).value if ws.max_column > 4 else None,
            'pos':   ws.cell(row_idx, 6).value if ws.max_column > 5 else None,
        }

        # 按日期读取考勤格
        day_cells: list[tuple[date, dict]] = []
        for col, d in date_map.items():
            cell_val = ws.cell(row_idx, col).value
            parsed   = parse_cell(cell_val)
            day_cells.append((d, parsed))

        info['day_cells'] = day_cells
        info['is_rd']     = is_rd
        return_list.append(info)

    return return_list


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  5. 写出 Excel                                                           ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def write_output(
    source_path: str,
    source_records: list[dict],
    employees: list[dict],
    activity_days: dict[date, set],
    filepath: str,
    approved_dates_by_name: dict[str, set[date]] | None = None,
    legal_holidays: set[date] | None = None,
    late22_included_names: list[str] | tuple[str, ...] | None = None,
):
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    col_name = _find_col(ws, '姓名') or 1
    col_map = {
        'deduct_late': _find_source_detail_col(ws, '15-30分钟迟到扣款', '迟到扣款'),
        'deduct_early': _find_source_detail_col(ws, '15-30分钟早退扣款', '早退扣款'),
        'absent_days': _find_source_detail_col(ws, '旷工天数'),
        'missing_punch': _find_source_detail_col(ws, '缺卡次数'),
        'late22_days': _find_source_detail_col(ws, '晚于22点打卡天数', '晚于22点'),
        'taxi_subsidy': _find_source_detail_col(ws, '晚走补贴'),
        'rd_ot_gt4_days': _find_source_detail_col(ws, '产研休息日加班>4小时天数', '加班>4小时天数'),
        'rd_ot_subsidy': _find_source_detail_col(ws, '产研休息日加班补贴'),
    }
    late22_name_source = (
        late22_included_names
        if late22_included_names is not None
        else LATE22_INCLUDED_NAMES
    )
    late22_names = _normalize_names(late22_name_source)

    att_index_by_name: dict[str, list] = {e['name']: e['day_cells'] for e in employees}
    att_index_by_emp_no: dict[str, list] = {
        e['emp_no']: e['day_cells'] for e in employees if _normalize_emp_no(e.get('emp_no'))
    }
    mismatches: list[dict] = []
    missing_attendance: list[str] = []
    payable_records, intern_records = split_intern_records(source_records)

    for src in payable_records:
        name = src['name']
        src_emp_no = _normalize_emp_no(src.get('emp_no'))
        day_cells = []
        if src_emp_no:
            day_cells = att_index_by_emp_no.get(src_emp_no, [])
        if not day_cells:
            day_cells = att_index_by_name.get(name, [])
        if not day_cells:
            missing_attendance.append(name)
            computed = _empty_result(*TARGET_FIELDS)
            details = {}
        else:
            emp_approved = (approved_dates_by_name or {}).get(name, set())
            computed = compute_deduction(day_cells, approved_dates=emp_approved)
            exclude_late22_count = _should_exclude_late22_count(
                src,
                included_names=late22_names,
            )
            subsidy_result = compute_subsidy(
                day_cells,
                activity_days,
                name,
                src['is_rd'],
                exclude_late22_count=exclude_late22_count,
                approved_dates=emp_approved,
                legal_holidays=legal_holidays,
            )
            computed.update(subsidy_result)
            # 收集计算明细
            details = _collect_calculation_details(
                day_cells, activity_days, name, src['is_rd'],
                emp_approved, legal_holidays,
                exclude_late22_count,
            )

        row_idx = src['row_idx']
        # 写回去除（已离职）后的姓名
        name_cell = ws.cell(row_idx, col_name)
        if not isinstance(name_cell, openpyxl.cell.cell.MergedCell):
            name_cell.value = name

        for field in TARGET_FIELDS:
            cell = ws.cell(row_idx, col_map[field])
            # 跳过合并单元格的非主单元格
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue

            # The writable workbook preserves formulas, so read the already parsed
            # data_only source value for comparisons and diff explanations.
            existing = _blank_to_none(src.get(field))
            expected = computed[field]
            if not _values_match(existing, expected):
                change_type = "回填" if existing is None else "核对差异"
                mismatches.append({
                    'name': name,
                    'row_idx': row_idx,
                    'field': field,
                    'existing': existing,
                    'expected': expected,
                    'change_type': change_type,
                    'details': details.get(field, ''),
                    'explanation': _explain_mismatch(
                        src,
                        field,
                        existing,
                        expected,
                        details.get(field, ''),
                        late22_included_names=late22_names,
                    ),
                })
            # 补贴字段、旷工天数、缺卡次数以重新计算值为准；其他扣款字段仅回填空格。
            if field in OVERWRITE_FIELDS or existing is None:
                cell.value = expected

    # 实习生没有补贴，最终核对表直接删除对应行；倒序删除避免行号漂移。
    for row_idx in sorted({src['row_idx'] for src in intern_records}, reverse=True):
        ws.delete_rows(row_idx, 1)

    wb.save(filepath)
    return mismatches, missing_attendance


def _display_value(value) -> str:
    if value is None:
        return "空"
    return str(value)


def _same_amount_days(amount, unit=50) -> str | None:
    try:
        return str(int(float(amount) / unit))
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _explain_mismatch(
    record: dict,
    field: str,
    existing,
    expected,
    details: str = "",
    late22_included_names: list[str] | tuple[str, ...] | None = None,
) -> str:
    dept = str(record.get("dept1") or "").strip()
    source_text = _display_value(existing)
    expected_text = _display_value(expected)
    detail_text = str(details or "").strip()
    is_late22_excluded = _should_exclude_late22_count(
        record,
        included_names=late22_included_names,
    )

    if field == "missing_punch":
        if expected is None:
            return f"按日期级考勤重算，未发现需要计入的有效缺卡；源表为{source_text}次，不采用。"
        suffix = f"有效缺卡日期：{detail_text}；" if detail_text else ""
        return f"按日期级考勤重算，{suffix}应为{expected_text}次，源表为{source_text}次。"

    if field == "absent_days":
        if existing is None:
            return f"源表未填旷工；按日期级考勤/迟到早退累计应回填为{expected_text}天。{detail_text}"
        if detail_text:
            return f"按日期级旷工明细累计为{expected_text}天，源表为{source_text}天；以明细累计为准。{detail_text}"
        return f"源表旷工天数为{source_text}，重新计算为{expected_text}；以日期级考勤计算值为准。"

    if field == "late22_days":
        if expected is None:
            if is_late22_excluded:
                return f"该员工所在部门「{dept}」属于晚走补贴排除范围（不坐班业务/客服售后/22点排班等），即使有22点后打卡也不自动计晚走补贴。"
            if detail_text:
                return f"22点后打卡均被规则排除：{detail_text}；所以不计晚走天数。"
            return "源表有22点后打卡统计，但日期级考勤未找到符合“标准坐班且非活动签到/非排除人员”的有效晚走记录，所以不计。"
        if detail_text:
            return f"按有效22点后打卡重算，只计入/保留：{detail_text}；所以应为{expected_text}天。"
        return f"晚走天数按有效22点后打卡重算，应为{expected_text}天，源表为{source_text}天。"

    if field == "taxi_subsidy":
        days = _same_amount_days(expected)
        if expected is None:
            if is_late22_excluded:
                return f"晚走补贴按有效晚走天数×50计算；该员工所在部门「{dept}」属于晚走补贴排除范围，所以补贴为0。"
            return "晚走补贴按有效晚走天数×50计算；有效晚走天数为0，所以补贴为0。"
        if days:
            return f"晚走补贴按有效晚走天数×50计算，应为{expected_text}元（{days}天×50），源表为{source_text}元。"
        return f"晚走补贴重新计算为{expected_text}元，源表为{source_text}元。"

    if field == "rd_ot_gt4_days":
        if expected is None:
            if "法定节假日" in detail_text:
                return f"产研休息日加班补贴排除法定节假日；{detail_text}，所以不计入。"
            if not record.get("is_rd"):
                return f"该员工所在部门「{dept}」不属于产研补贴范围，源表计入的产研休息日加班不采用。"
            return "日期级考勤未找到符合“普通休息日且加班超过4小时”的有效记录，所以不计入产研休息日加班天数。"
        if detail_text:
            return f"按有效普通休息日加班>4小时重算：{detail_text}；所以应为{expected_text}天。"
        return f"产研休息日加班天数重新计算为{expected_text}天，源表为{source_text}天。"

    if field == "rd_ot_subsidy":
        days = _same_amount_days(expected)
        if expected is None:
            if not record.get("is_rd"):
                return f"产研休息日加班补贴只适用于产研补贴范围；该员工所在部门「{dept}」不适用，所以补贴为0。"
            return "产研休息日加班补贴按有效天数×50计算；有效产研休息日加班天数为0，所以补贴为0。"
        if days:
            return f"产研休息日加班补贴按有效天数×50计算，应为{expected_text}元（{days}天×50），源表为{source_text}元。"
        return f"产研休息日加班补贴重新计算为{expected_text}元，源表为{source_text}元。"

    return f"源表值为{source_text}，重新计算值为{expected_text}；以重新计算结果为准。"


def write_mismatch_report(mismatches: list[dict], filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异清单"
    ws.append([
        "序号",
        "姓名",
        "行号",
        "类型",
        "字段",
        "源表值",
        "计算值",
        "计算明细",
        "差异说明",
    ])
    for idx, item in enumerate(mismatches, start=1):
        ws.append([
            idx,
            item["name"],
            item["row_idx"],
            item.get("change_type"),
            FIELD_LABELS[item["field"]],
            item["existing"],
            item["expected"],
            item.get("details", ""),
            item.get("explanation", ""),
        ])

    # 调整列宽以便查看明细
    ws.column_dimensions['H'].width = 80
    ws.column_dimensions['I'].width = 80

    wb.save(filepath)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  6. 主流程                                                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def main():
    parser = argparse.ArgumentParser(
        description="读取补贴扣款表（钉钉导出）+ 签到表，核对后输出剔除羽毛球补贴后的结果"
    )
    parser.add_argument("--source", default=DEFAULT_SOURCE,
                        help=f"现有补贴扣款表 Excel（默认: {DEFAULT_SOURCE}）")
    parser.add_argument("--checkin", default=DEFAULT_CHECKIN,
                        help=f"签到表 Excel（默认: {DEFAULT_CHECKIN}）")
    parser.add_argument("--attendance", default=DEFAULT_ATTENDANCE,
                        help=f"考勤表 Excel（默认: {DEFAULT_ATTENDANCE}，用于计算补贴）")
    parser.add_argument("--schedule", default="",
                        help="作息表 Excel（可选，用于识别法定节假日，避免误算产研休息日加班补贴）")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help=f"输出文件路径（默认: {DEFAULT_OUTPUT}）")
    parser.add_argument("--diff-output", default=DEFAULT_DIFF_OUTPUT,
                        help=f"差异清单输出路径（默认: {DEFAULT_DIFF_OUTPUT}）")
    parser.add_argument("--attendance-result", default=None,
                        help="考勤结果表路径（含批注，用于排除已补流程的旷工）")
    parser.add_argument("--year", type=int, default=None,
                        help="手动指定考勤年份；需与 --month 同时使用")
    parser.add_argument("--month", type=int, default=None,
                        help="手动指定考勤月份；需与 --year 同时使用")
    parser.add_argument(
        "--rd-dept-keyword",
        action="append",
        dest="rd_dept_keywords",
        help="产研部门匹配关键字，可重复传入或使用逗号分隔（默认: 研发）",
    )
    parser.add_argument(
        "--late22-included-name",
        "--late-walk-subsidy-name",
        action="append",
        dest="late22_included_names",
        help="晚走补贴强制纳入人员姓名，可重复传入或使用逗号分隔（默认: 既有晚走补贴人员名单）",
    )
    args = parser.parse_args()
    rd_keywords = _normalize_keywords(args.rd_dept_keywords or RD_DEPT_KEYWORDS)
    late22_included_names = _normalize_names(
        LATE22_INCLUDED_NAMES
        if args.late22_included_names is None
        else args.late22_included_names
    )

    # 1. 读取现有补贴扣款表（扣款字段来源）
    print(f"[1/4] 读取源补贴扣款表: {args.source}")
    print(f"      产研部门关键字: {rd_keywords}")
    print(f"      晚走补贴人员: {late22_included_names or '无'}")
    source_records = parse_source_table(
        args.source,
        rd_dept_keywords=rd_keywords,
    )
    payable_records, intern_records = split_intern_records(source_records)
    print(f"      读取员工数: {len(source_records)}")
    if intern_records:
        intern_names = "、".join(src["name"] for src in intern_records)
        print(f"      实习生剔除: {len(intern_records)} 人（{intern_names}）")

    # 2. 读取签到表（活动名单）
    print(f"[2/4] 解析签到表: {args.checkin}")
    activity_days = parse_activity_checkin(args.checkin)
    for d, names in sorted(activity_days.items()):
        print(f"      {d} 活动参与人数: {len(names)} 人")

    # 3. 读取考勤表（用于计算补贴）
    print(f"[3/4] 解析考勤表: {args.attendance}")
    period_context = [
        args.source,
        args.checkin,
        args.attendance_result,
        args.schedule,
    ]
    if args.schedule:
        period_year, period_month, period_source = resolve_schedule_period(args.schedule)
        if args.year is not None or args.month is not None:
            manual_year, manual_month = _validate_year_month(args.year, args.month)
            if (manual_year, manual_month) != (period_year, period_month):
                raise ValueError(
                    f"手动指定年月为 {manual_year}年{manual_month}月，"
                    f"但作息表为 {period_year}年{period_month}月；请保持一致。"
                )
        print(f"[作息月份] 使用 {period_year}年{period_month}月（{period_source}）")
    else:
        period_year, period_month, period_source = resolve_attendance_period(
            args.attendance,
            source_name=args.attendance,
            context_names=period_context,
            year=args.year,
            month=args.month,
        )
    employees = parse_attendance(
        args.attendance,
        rd_dept_keywords=rd_keywords,
        year=period_year,
        month=period_month,
        period_source=period_source,
    )
    print(f"      读取员工数: {len(employees)}")

    legal_holidays = set()
    if args.schedule:
        print(f"[3.2] 解析作息表法定节假日: {args.schedule}")
        legal_holidays = load_statutory_holidays_from_schedule(
            args.schedule,
            target_year=period_year,
            target_month=period_month,
        )
        print(f"      法定节假日天数: {len(legal_holidays)}")
    else:
        print("[3.2] 未提供作息表，无法额外识别法定节假日；休息并打卡将按普通休息日口径计算")

    # 3.5 读取考勤结果表（含批注，用于排除已补流程的旷工）
    approved_dates_by_name = None
    if args.attendance_result:
        print(f"[3.5] 解析考勤结果表: {args.attendance_result}")
        approved_dates_by_name = parse_attendance_result(
            args.attendance_result, year=period_year, month=period_month,
        )
        total_excluded = sum(len(v) for v in approved_dates_by_name.values())
        print(f"      已补流程旷工记录: {total_excluded} 条，涉及 {len(approved_dates_by_name)} 人")

    # 4. 输出核对后的表
    print(f"[4/4] 生成核对表: {args.output}")
    mismatches, missing_attendance = write_output(
        args.source,
        source_records,
        employees,
        activity_days,
        args.output,
        approved_dates_by_name=approved_dates_by_name,
        legal_holidays=legal_holidays,
        late22_included_names=late22_included_names,
    )
    print(f"[OK] 已生成: {args.output}")
    print(f"     共处理 {len(payable_records)} 条员工记录")
    if intern_records:
        print(f"     已从核对表删除实习生: {len(intern_records)} 人")
    print(f"     核对差异: {len(mismatches)} 条")
    write_mismatch_report(mismatches, args.diff_output)
    print(f"     差异清单: {args.diff_output}")
    if missing_attendance:
        print(f"     考勤表缺失员工: {len(missing_attendance)} 人")
    for item in mismatches:
        label = FIELD_LABELS[item['field']]
        print(
            f"     [{item['change_type']}] {item['name']} / {label}: "
            f"源表={item['existing']}，计算={item['expected']}"
        )


if __name__ == "__main__":
    main()
