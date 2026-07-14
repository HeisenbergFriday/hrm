from __future__ import annotations

import argparse
import calendar
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_LEAVE_PATH = os.path.join(_BASE, "leave")
if _LEAVE_PATH not in sys.path:
    sys.path.insert(0, _LEAVE_PATH)

import calc_leave  # noqa: E402


STATIC_HEADERS = ["姓名", "所属公司"]
ROW_KEY_SEPARATOR = "\x1f"

OUTPUT_TITLE_PREFIX = "兼职&实习生"
OUTPUT_TITLE_SUFFIX = "月份出勤数据"
OUTPUT_SHEET_SUFFIX = "月兼职出勤汇总"

SOURCE_ATTENDANCE_DETAIL = "考勤明细"
SOURCE_MONTHLY_SUMMARY = "月度汇总"
SOURCE_DEFAULT_WEEKDAYS = "默认工作日"
SOURCE_SPECIAL_DEFAULT_WEEKDAYS = "特殊兼职工作日"

SCHEDULE_BEIJING_UNIVERSAL = "北京环球"
SCHEDULE_NANJING_SHOP = "南京地铁商铺"
SCHEDULE_NANJING_STATION = "南京地铁驻场"
SCHEDULE_CHANGSHA = "长沙安检"
SCHEDULE_MATRIX = "排班表"

SCHEDULE_SOURCE_ORDER = (
    SCHEDULE_BEIJING_UNIVERSAL,
    SCHEDULE_NANJING_SHOP,
    SCHEDULE_NANJING_STATION,
    SCHEDULE_CHANGSHA,
    SCHEDULE_MATRIX,
)

SPLIT_SCHEDULE_SOURCES = {
    SCHEDULE_NANJING_SHOP,
    SCHEDULE_NANJING_STATION,
}

ROW_HEIGHT_HEADER = 24
ROW_HEIGHT_DATA = 20
WORKDAY_MINUTES = 480
MONTHLY_EXEMPT_MINUTES = 120

HEADER_FILL = PatternFill("solid", fgColor="FFF2F2F2")
TITLE_FILL = PatternFill("solid", fgColor="FFD9EAF7")
ATTENDANCE_ALERT_FILL = PatternFill("solid", fgColor="FFFFE599")
ATTENDANCE_DEDUCT_FILL = PatternFill("solid", fgColor="FFF4CCCC")
ATTENDANCE_OFFSITE_FILL = PatternFill("solid", fgColor="FFC6EFCE")
ATTENDANCE_UNSCHEDULED_PUNCH_FILL = PatternFill("solid", fgColor="FF9DC3E6")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)

LATE_RE = re.compile(r"(?:上班|严重)?(?:旷工)?迟到\s*(\d+)\s*分钟")
EARLY_RE = re.compile(r"(?:下班)?早退\s*(\d+)\s*分钟")
PUNCH_TIME_RE = re.compile(r"\((\d{1,2}:\d{2}|-)(?:,(\d{1,2}:\d{2}|-))\)")
OVERTIME_HOURS_RE = re.compile(r"加班[^\n\r]*?(\d+(?:\.\d+)?)\s*小时")
MUTENG_KEYWORDS = ("沐腾",)

DailyEntry = dict[str, Any]
DailyMap = dict[int, DailyEntry]
SourceMap = dict[str, DailyMap]

# 已知源数据里的名字变体。
VARIANT_TO_CANONICAL = {
    "陈伟铭泉": "陈铭泉",
    "葛广州": "葛广洲",
}

LEAVE_KEYWORDS = (
    "年假",
    "调休",
    "婚假",
    "陪产假",
    "丧假",
    "产检假",
    "工伤假",
    "产假",
    "病假",
    "事假",
)

PRESENT_KEYWORDS = (
    "正常",
    "迟到",
    "早退",
    "缺卡",
    "外勤",
    "补卡",
)

SITE_ROLE_KEYWORDS = (
    "地铁",
    "环球",
    "商铺",
    "驻场",
    "东方明珠",
    "机场",
    "景区",
    "轻松游",
    "宁波站",
    "北京站",
    "杭州",
    "上海",
    "福州",
    "厦门",
    "新疆",
    "长沙",
)

ATTENDANCE_ROLE_KEYWORDS = (
    "实习",
    "兼职",
    "安检",
    "保洁",
    "清洁",
)

ATTENDANCE_OUTSOURCE_ROLE_KEYWORDS = (
    "寄存",
    "客服",
    "店长",
)

ATTENDANCE_OUTSOURCE_CODE_PREFIXES = ("WB",)

MONTHLY_SCOPE_KEYWORDS = (
    "兼职",
    "外包",
    "实习",
    "保洁",
    "安检",
)

DEFAULT_SPECIAL_DEFAULT_NAMES = tuple(
    dict.fromkeys(
        (
            "王心英",
            "刘芮",
            "汤颖",
            "周代林",
            "常雨凡",
            "王心英",
            "陈富庆",
        )
    )
)
DEFAULT_SPECIAL_CHENGDU_NAMES = {"王心英"}

SCHEDULE_PLACEHOLDER_NAMES = {"张三", "李四", "王五"}


def _normalize_name(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"（离职）$", "", text)
    text = re.sub(r"\s+", "", text)
    return VARIANT_TO_CANONICAL.get(text, text)


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _field_key(value: Any) -> str:
    return re.sub(r"[\s\u00a0\u2000-\u200f\u2028-\u202f\u205f\u3000\ufeff]+", "", str(value or "")).strip()


def _normalize_code(value: Any) -> str:
    return _normalize_text(value).upper()


def _make_person_key(name: Any, company: Any = "") -> str:
    normalized_name = _normalize_name(name)
    normalized_company = _normalize_text(company)
    if not normalized_name:
        return ""
    if normalized_company:
        return f"{normalized_name}{ROW_KEY_SEPARATOR}{normalized_company}"
    return normalized_name


def _split_person_key(key: Any) -> tuple[str, str]:
    text = _normalize_text(key)
    if ROW_KEY_SEPARATOR not in text:
        return _normalize_name(text), ""
    name, company = text.split(ROW_KEY_SEPARATOR, 1)
    return _normalize_name(name), _normalize_text(company)


def _contains_keyword(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def normalize_name(value: Any) -> str:
    return _normalize_name(value)


def _normalize_name_collection(values: list[str] | tuple[str, ...] | set[str] | None) -> tuple[str, ...]:
    if not values:
        return ()

    normalized_values: list[str] = []
    seen: set[str] = set()
    for value in values:
        name = _normalize_name(value)
        if not name or name in seen:
            continue
        normalized_values.append(name)
        seen.add(name)
    return tuple(normalized_values)


def _round2(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _round3(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))


def _format_hours(value: float) -> str:
    return f"{_round2(value):g}"


def _make_daily_entry(
    value: float,
    note: str = "",
    raw_text: str = "",
    alert: bool = False,
    deducted: bool = False,
    offsite: bool = False,
    unscheduled_punch: bool = False,
    comment: str = "",
    overtime_hours: float | None = None,
) -> DailyEntry:
    entry = {
        "value": _round3(max(value, 0.0)),
        "note": note,
        "raw_text": raw_text,
        "alert": alert,
        "deducted": deducted,
        "offsite": offsite,
        "unscheduled_punch": unscheduled_punch,
        "comment": comment,
    }
    if overtime_hours is not None:
        entry["overtime_hours"] = _round2(max(overtime_hours, 0.0))
    return entry


def _entry_value(entry: DailyEntry | float | int | None) -> float | None:
    if entry is None:
        return None
    if isinstance(entry, dict):
        value = entry.get("value")
    else:
        value = entry
    if value is None:
        return None
    return float(value)


def _entry_note(entry: DailyEntry | None) -> str:
    if not isinstance(entry, dict):
        return ""
    return _normalize_text(entry.get("note"))


def _entry_comment(entry: DailyEntry | None) -> str:
    if not isinstance(entry, dict):
        return ""
    return _normalize_text(entry.get("comment"))


def _merge_note(*notes: str) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for note in notes:
        for part in re.split(r"[；;]\s*", _normalize_text(note)):
            clean = part.strip()
            if clean and clean not in seen:
                result.append(clean)
                seen.add(clean)
    return "；".join(result)


def _is_alert_entry(entry: DailyEntry | None) -> bool:
    return bool(isinstance(entry, dict) and entry.get("alert"))


def _is_deducted_entry(entry: DailyEntry | None) -> bool:
    return bool(isinstance(entry, dict) and entry.get("deducted"))


def _is_offsite_entry(entry: DailyEntry | None) -> bool:
    return bool(isinstance(entry, dict) and entry.get("offsite"))


def _is_unscheduled_punch_entry(entry: DailyEntry | None) -> bool:
    return bool(isinstance(entry, dict) and entry.get("unscheduled_punch"))


def _entry_overtime_hours(entry: DailyEntry | None) -> float:
    if not isinstance(entry, dict):
        return 0.0
    try:
        return float(entry.get("overtime_hours") or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _is_overtime_entry(entry: DailyEntry | None) -> bool:
    return _entry_overtime_hours(entry) > 0


def _find_header_row(ws, required_keywords: tuple[str, ...], max_rows: int = 10) -> tuple[int | None, list[Any]]:
    required_keys = tuple(_field_key(keyword) for keyword in required_keywords)
    for row_idx in range(1, min(max_rows, ws.max_row) + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        texts = [_field_key(value) for value in values]
        if all(any(keyword and keyword in text for text in texts) for keyword in required_keys):
            return row_idx, values
    return None, []


def _find_col(header_values: list[Any], *keywords: str) -> int | None:
    keyword_keys = tuple(_field_key(keyword) for keyword in keywords)
    for idx, value in enumerate(header_values):
        text = _field_key(value)
        if any(keyword and keyword in text for keyword in keyword_keys):
            return idx + 1
    return None


def _find_company_col(header_values: list[Any]) -> int | None:
    return _find_col(header_values, "所属公司", "合同主体", "主体", "公司")


def _extract_year_month_from_title(text: str) -> tuple[int | None, int | None]:
    clean_text = _normalize_text(text)
    if not clean_text:
        return None, None

    match = re.search(r"(\d{4})[-年/.](\d{1,2})", clean_text)
    if match:
        return int(match.group(1)), int(match.group(2))

    match = re.search(r"(\d{1,2})月份", clean_text)
    if match:
        return None, int(match.group(1))

    return None, None


def _infer_year_month(
    attendance_detail_path: str | None,
    monthly_summary_paths: list[str],
    schedule_paths: list[str],
    default_schedule_path: str | None = None,
) -> tuple[int, int]:
    candidates: list[tuple[int | None, int | None]] = []

    def _collect_from_workbook(path: str, cell_ref: tuple[int, int] = (1, 1)) -> None:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        row_idx, col_idx = cell_ref
        value = ws.cell(row_idx, col_idx).value
        candidates.append(_extract_year_month_from_title(_normalize_text(value)))
        wb.close()

    if attendance_detail_path:
        _collect_from_workbook(attendance_detail_path)
    for ms_path in monthly_summary_paths:
        _collect_from_workbook(ms_path)

    for path in schedule_paths:
        filename = os.path.basename(path)
        candidates.append(_extract_year_month_from_title(filename))
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        candidates.append(_extract_year_month_from_title(_normalize_text(sheet.cell(1, 1).value)))
        candidates.append(_extract_year_month_from_title(_normalize_text(sheet.cell(5, 4).value)))
        wb.close()

    if default_schedule_path:
        candidates.append(_extract_year_month_from_title(os.path.basename(default_schedule_path)))
        wb = openpyxl.load_workbook(default_schedule_path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            candidates.append(_extract_year_month_from_title(_normalize_text(sheet.cell(1, 1).value)))
        wb.close()

    year: int | None = None
    month: int | None = None
    for candidate_year, candidate_month in candidates:
        if candidate_year and year is None:
            year = candidate_year
        if candidate_month and month is None:
            month = candidate_month

    if year is None:
        year = datetime.today().year
    if month is None:
        raise ValueError("无法从输入文件中识别月份，请至少提供一份带月份信息的考勤表。")
    return year, month


def _build_weekday_day_map(year: int, month: int) -> dict[int, float]:
    _, days_in_month = calendar.monthrange(year, month)
    result: dict[int, float] = {}
    for day in range(1, days_in_month + 1):
        current = date(year, month, day)
        if current.weekday() < 5:
            result[day] = 1.0
    return result


def _build_default_day_maps(
    year: int,
    month: int,
    default_schedule_path: str | None,
) -> dict[str, DailyMap]:
    if not default_schedule_path:
        weekday_day_map = _build_weekday_day_map(year, month)
        return {
            "main": {day: _make_daily_entry(value) for day, value in weekday_day_map.items()},
            "chengdu": {day: _make_daily_entry(value) for day, value in weekday_day_map.items()},
        }

    ctx = calc_leave.load_schedule_context(default_schedule_path)
    if ctx["year"] != year or ctx["month"] != month:
        raise ValueError(
            f"默认作息表月份与兼职汇总不一致：作息表 {ctx['year']}-{ctx['month']:02d}，"
            f"目标月份 {year}-{month:02d}"
        )

    main_day_map = {current.day: _make_daily_entry(1.0) for current in sorted(ctx["main_working_days"])}
    chengdu_day_map = {current.day: _make_daily_entry(1.0) for current in sorted(ctx["chengdu_working_days"])}
    return {
        "main": main_day_map,
        "chengdu": chengdu_day_map or dict(main_day_map),
    }


def _has_real_punch(text: str) -> bool:
    for group in re.findall(r"\(([^()]*)\)", text):
        if re.search(r"\d{1,2}:\d{2}", group):
            return True
    return False


def _extract_leave_hours(text: str) -> float | None:
    matches = re.findall(r"(\d+(?:\.\d+)?)小时", text)
    if not matches:
        return None
    values = [float(item) for item in matches]
    return max(values) if values else None


def _extract_overtime_hours(text: str) -> float | None:
    values: list[float] = []
    for match in OVERTIME_HOURS_RE.finditer(text):
        try:
            values.append(float(match.group(1)))
        except ValueError:
            continue
    if not values:
        return None
    return sum(values)


def _extract_attendance_anomalies(text: str) -> tuple[int, int, bool]:
    late_minutes = sum(int(match.group(1)) for match in LATE_RE.finditer(text))
    early_minutes = sum(int(match.group(1)) for match in EARLY_RE.finditer(text))
    is_absent = False
    first_line = text.split("\n")[0].strip()
    if "旷工" in text and "旷工迟到" not in text and "迟到" not in first_line:
        is_absent = True
    if "旷工\n(-)" in text or text.strip() == "标准:旷工\n(-)":
        is_absent = True
    return late_minutes, early_minutes, is_absent


def _format_minutes_note(late_minutes: int, early_minutes: int, is_absent: bool, is_missing_punch: bool = False) -> str:
    parts: list[str] = []
    if is_absent:
        parts.append("旷工")
    if late_minutes:
        parts.append(f"迟到{late_minutes}分钟")
    if early_minutes:
        parts.append(f"早退{early_minutes}分钟")
    if is_missing_punch:
        parts.append("缺卡仅提醒")
    return "、".join(parts)


def _has_intern_monthly_exemption(employee_role: str) -> bool:
    return "实习" in _normalize_text(employee_role)


def _apply_exception_minutes(
    base_value: float,
    exception_minutes: int,
    note_prefix: str,
    raw_text: str,
    monthly_exemption_remaining: list[int] | None,
) -> DailyEntry:
    if exception_minutes <= 0:
        return _make_daily_entry(base_value, raw_text=raw_text)

    deducted_minutes = exception_minutes
    exempted_minutes = 0
    if monthly_exemption_remaining is not None and monthly_exemption_remaining[0] > 0:
        exempted_minutes = min(exception_minutes, monthly_exemption_remaining[0])
        monthly_exemption_remaining[0] -= exempted_minutes
        deducted_minutes -= exempted_minutes

    if deducted_minutes <= 0:
        return _make_daily_entry(
            base_value,
            note=f"{note_prefix}，已用月累计豁免{exempted_minutes}分钟",
            raw_text=raw_text,
            alert=True,
        )

    deducted_days = deducted_minutes / WORKDAY_MINUTES
    note = f"{note_prefix}，扣减{_round3(deducted_days)}天"
    if exempted_minutes:
        note = f"{note_prefix}，豁免{exempted_minutes}分钟，扣减{_round3(deducted_days)}天"
    return _make_daily_entry(
        base_value - deducted_days,
        note=note,
        raw_text=raw_text,
        alert=True,
        deducted=True,
    )


def _parse_daily_text_value(
    value: Any,
    *,
    monthly_exemption_remaining: list[int] | None = None,
    count_outing_as_present: bool = False,
) -> DailyEntry | None:
    text = _normalize_text(value)
    if not text:
        return None

    has_real_punch = _has_real_punch(text)
    overtime_hours = _extract_overtime_hours(text)
    is_rest_or_unscheduled = "休息" in text or "未排班" in text
    if overtime_hours is not None and overtime_hours > 0 and is_rest_or_unscheduled:
        note_prefix = "未排班加班" if "未排班" in text else "休息日加班"
        hours_text = _format_hours(overtime_hours)
        return _make_daily_entry(
            overtime_hours * 60 / WORKDAY_MINUTES,
            note=f"{note_prefix}{hours_text}小时",
            raw_text=text,
            alert=True,
            unscheduled_punch="未排班" in text and has_real_punch,
            comment=f"排班为休息/未排班，按加班时长折算出勤\n加班时长：{hours_text}小时\n源记录：{text}",
            overtime_hours=overtime_hours,
        )

    if "未排班" in text and has_real_punch:
        return _make_daily_entry(
            1.0,
            note="未排班并打卡",
            raw_text=text,
            alert=True,
            unscheduled_punch=True,
            comment=f"考勤记录显示未排班但已打卡\n源记录：{text}",
        )

    late_minutes, early_minutes, is_absent = _extract_attendance_anomalies(text)
    exception_minutes = late_minutes + early_minutes
    is_missing_punch = "缺卡" in text
    is_outing = "外出" in text
    is_business_trip = "出差" in text

    if (is_outing or is_business_trip) and count_outing_as_present and not is_absent:
        offsite_label = "外出" if is_outing else "出差"
        return _make_daily_entry(
            1.0,
            raw_text=text,
            offsite=True,
            comment=f"{offsite_label}已计出勤\n源记录：{text}",
        )

    leave_hours = _extract_leave_hours(text)
    if leave_hours is not None:
        if leave_hours >= 7.5 and not has_real_punch:
            return None
        if leave_hours >= 3.5 and has_real_punch:
            base_value = 0.5
            if exception_minutes:
                return _apply_exception_minutes(
                    base_value,
                    exception_minutes,
                    _format_minutes_note(late_minutes, early_minutes, False),
                    text,
                    monthly_exemption_remaining,
                )
            return _make_daily_entry(base_value, raw_text=text)

    if "休息" in text and has_real_punch:
        return _make_daily_entry(
            1.0,
            note="休息并打卡",
            raw_text=text,
            alert=True,
            unscheduled_punch=True,
            comment=f"考勤记录显示休息日已打卡\n源记录：{text}",
        )

    if "休息" in text and not any(keyword in text for keyword in PRESENT_KEYWORDS):
        return None

    if is_outing and not has_real_punch:
        return None

    if is_business_trip and not has_real_punch:
        return None

    if is_absent and not has_real_punch:
        return _make_daily_entry(0.0, note="旷工，未计出勤", raw_text=text, alert=True, deducted=True)

    if is_absent and has_real_punch:
        note = _format_minutes_note(late_minutes, early_minutes, True, is_missing_punch) or "旷工异常"
        return _make_daily_entry(0.0, note=f"{note}，未计出勤", raw_text=text, alert=True, deducted=True)

    if is_missing_punch:
        note = _format_minutes_note(late_minutes, early_minutes, False, True) or "缺卡仅提醒"
        return _make_daily_entry(1.0, note=note, raw_text=text, alert=True)

    if exception_minutes:
        return _apply_exception_minutes(
            1.0,
            exception_minutes,
            _format_minutes_note(late_minutes, early_minutes, False),
            text,
            monthly_exemption_remaining,
        )

    if any(keyword in text for keyword in PRESENT_KEYWORDS):
        return _make_daily_entry(1.0, raw_text=text)

    if has_real_punch:
        return _make_daily_entry(1.0, raw_text=text)

    return None


def _merge_daily_maps(target: SourceMap, name: str, day_map: DailyMap) -> None:
    if not name:
        return
    merged = target.setdefault(name, {})
    for day, value in day_map.items():
        if value is None:
            continue
        current = merged.get(day)
        current_value = _entry_value(current)
        incoming_value = _entry_value(value)
        if incoming_value is None:
            continue
        if current_value is None or incoming_value > current_value:
            merged[day] = value
        elif incoming_value == current_value and isinstance(current, dict) and isinstance(value, dict):
            current["note"] = _merge_note(_entry_note(current), _entry_note(value))
            current["comment"] = _merge_note(_entry_comment(current), _entry_comment(value))
            current["alert"] = bool(current.get("alert") or value.get("alert"))
            current["deducted"] = bool(current.get("deducted") or value.get("deducted"))
            current["offsite"] = bool(current.get("offsite") or value.get("offsite"))
            current["unscheduled_punch"] = bool(
                current.get("unscheduled_punch") or value.get("unscheduled_punch")
            )


def _is_attendance_row_in_scope(
    attendance_group: Any,
    department: Any,
    employee_code: Any,
    position: Any,
) -> bool:
    group_text = _normalize_text(attendance_group)
    department_text = _normalize_text(department)
    position_text = _normalize_text(position)
    employee_code_text = _normalize_code(employee_code)
    combined_text = " ".join(
        text for text in (group_text, department_text, position_text, employee_code_text) if text
    )

    if _contains_keyword(position_text, ATTENDANCE_ROLE_KEYWORDS):
        return True
    if _contains_keyword(combined_text, ATTENDANCE_ROLE_KEYWORDS):
        return True
    if _contains_keyword(position_text, ATTENDANCE_OUTSOURCE_ROLE_KEYWORDS) and any(
        employee_code_text.startswith(prefix) for prefix in ATTENDANCE_OUTSOURCE_CODE_PREFIXES
    ):
        return True
    if "异地外勤" in group_text and _contains_keyword(position_text, ATTENDANCE_ROLE_KEYWORDS):
        return True
    return False


def _is_intern_attendance_row(
    attendance_group: Any,
    department: Any,
    position: Any,
) -> bool:
    combined_text = " ".join(
        text
        for text in (
            _normalize_text(attendance_group),
            _normalize_text(department),
            _normalize_text(position),
        )
        if text
    )
    return "实习" in combined_text


def _is_monthly_row_in_scope(
    attendance_group: Any,
    department: Any,
    employee_code: Any,
    position: Any,
) -> bool:
    group_text = _normalize_text(attendance_group)
    department_text = _normalize_text(department)
    employee_code_text = _normalize_code(employee_code)
    position_text = _normalize_text(position)
    combined_text = " ".join(
        text for text in (group_text, department_text, employee_code_text, position_text) if text
    )

    if employee_code_text.startswith("TXB"):
        return True
    if _contains_keyword(combined_text, MONTHLY_SCOPE_KEYWORDS):
        return True
    return False


def _source_exempts_late_early(path: str, ws) -> bool:
    filename = os.path.basename(path)
    if any(keyword in filename for keyword in MUTENG_KEYWORDS):
        return True
    for row_idx in range(1, min(ws.max_row, 5) + 1):
        row_text = " ".join(
            _normalize_text(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, min(ws.max_column, 8) + 1)
        )
        if any(keyword in row_text for keyword in MUTENG_KEYWORDS):
            return True
    return False


def _build_schedule_name_sources(
    schedules: dict[str, SourceMap],
) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for source_name in SCHEDULE_SOURCE_ORDER:
        for key in schedules.get(source_name, {}):
            name, _ = _split_person_key(key)
            mapping.setdefault(name, []).append(source_name)
    return mapping


def _has_company_schedule_row(schedules: dict[str, SourceMap], name: str) -> bool:
    for source_map in schedules.values():
        for key in source_map:
            key_name, company = _split_person_key(key)
            if key_name == name and company:
                return True
    return False


def _should_split_schedule_rows(sources: list[str]) -> bool:
    if len(sources) > 1:
        return True
    return any(source in SPLIT_SCHEDULE_SOURCES for source in sources) and len(sources) > 1


def _build_row_specs(
    attendance_detail: SourceMap,
    monthly_summary: SourceMap,
    schedules: dict[str, SourceMap],
    include_empty_monthly: bool = False,
) -> list[str]:
    names: list[str] = []
    existing_names: set[str] = set()
    schedule_name_sources = _build_schedule_name_sources(schedules)
    monthly_names = set(monthly_summary)
    if not include_empty_monthly:
        monthly_names = {
            name
            for name, day_map in monthly_summary.items()
            if day_map or name in schedule_name_sources
        }

    def _append(name: str) -> None:
        if name and name not in existing_names:
            names.append(name)
            existing_names.add(name)

    for name in attendance_detail:
        if _has_company_schedule_row(schedules, name):
            continue
        _append(name)

    for name in monthly_summary:
        if name not in monthly_names:
            continue
        if _has_company_schedule_row(schedules, name):
            continue
        _append(name)

    for source_name in SCHEDULE_SOURCE_ORDER:
        for key in schedules.get(source_name, {}):
            _append(key)

    return names


def _build_output_rows(
    attendance_detail: SourceMap,
    monthly_summary: SourceMap,
    schedules: dict[str, SourceMap],
    special_default_names: tuple[str, ...],
    special_chengdu_names: set[str],
) -> list[dict[str, Any]]:
    output_rows: list[dict[str, Any]] = []
    row_names = _build_row_specs(attendance_detail, monthly_summary, schedules)
    existing_names = {name for name in row_names}
    existing_plain_names = {_split_person_key(name)[0] for name in row_names}
    special_default_name_set = set(special_default_names)
    for name in special_default_names:
        if name not in existing_plain_names:
            row_names.append(name)
            existing_names.add(name)
            existing_plain_names.add(name)

    for key in row_names:
        name, company = _split_person_key(key)
        output_rows.append(
            {
                "姓名": name,
                "所属公司": company,
                "_allow_default_fill": name in special_default_name_set,
                "_default_schedule_key": "chengdu" if name in special_chengdu_names else "main",
            }
        )
    return output_rows


def parse_attendance_detail(path: str) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx, header_values = _find_header_row(ws, ("姓名", "工号"))
    if header_row_idx is None:
        wb.close()
        raise ValueError(f"未找到考勤明细表头：{path}")

    day_header_row_idx = header_row_idx + 1
    col_name = _find_col(header_values, "姓名")
    col_group = _find_col(header_values, "考勤组")
    col_department = _find_col(header_values, "部门")
    col_code = _find_col(header_values, "工号")
    col_position = _find_col(header_values, "职位")
    if col_name is None:
        wb.close()
        raise ValueError(f"考勤明细缺少姓名列：{path}")

    day_columns = list(range(7, ws.max_column + 1))
    result: SourceMap = {}
    intern_exemption_remaining: dict[str, list[int]] = {}
    for row_idx in range(day_header_row_idx + 1, ws.max_row + 1):
        name = _normalize_name(ws.cell(row_idx, col_name).value)
        if not name:
            continue
        attendance_group = ws.cell(row_idx, col_group).value if col_group else None
        department = ws.cell(row_idx, col_department).value if col_department else None
        employee_code = ws.cell(row_idx, col_code).value if col_code else None
        position = ws.cell(row_idx, col_position).value if col_position else None
        is_intern_row = _is_intern_attendance_row(attendance_group, department, position)
        has_intern_exemption = _has_intern_monthly_exemption(_normalize_text(position))
        remaining = intern_exemption_remaining.setdefault(name, [MONTHLY_EXEMPT_MINUTES]) if has_intern_exemption else None
        if not _is_attendance_row_in_scope(
            attendance_group=attendance_group,
            department=department,
            employee_code=employee_code,
            position=position,
        ):
            continue
        day_map: DailyMap = {}
        for offset, col_idx in enumerate(day_columns, start=1):
            daily_value = _parse_daily_text_value(
                ws.cell(row_idx, col_idx).value,
                monthly_exemption_remaining=remaining,
                count_outing_as_present=is_intern_row,
            )
            if daily_value is not None:
                day_map[offset] = daily_value
        _merge_daily_maps(result, name, day_map)

    wb.close()
    print(f"[{SOURCE_ATTENDANCE_DETAIL}] 解析 {len(result)} 人")
    return result


def parse_monthly_summary(path: str) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx, header_values = _find_header_row(ws, ("姓名", "出勤天数"))
    if header_row_idx is None:
        header_row_idx, header_values = _find_header_row(ws, ("姓名", "考勤结果"))
    if header_row_idx is None:
        wb.close()
        raise ValueError(f"未找到月度汇总表头：{path}")

    day_header_row_idx = header_row_idx + 1
    col_name = _find_col(header_values, "姓名")
    col_attendance_result = _find_col(header_values, "考勤结果")
    col_group = _find_col(header_values, "考勤组")
    col_department = _find_col(header_values, "部门")
    col_code = _find_col(header_values, "工号")
    col_position = _find_col(header_values, "职位")
    if col_name is None:
        wb.close()
        raise ValueError(f"月度汇总缺少姓名列：{path}")
    if col_attendance_result is None:
        wb.close()
        raise ValueError(f"月度汇总缺少考勤结果列：{path}")

    day_columns = list(range(col_attendance_result, ws.max_column + 1))
    has_scope_columns = any([col_group, col_department, col_code, col_position])
    result: SourceMap = {}
    intern_exemption_remaining: dict[str, list[int]] = {}
    for row_idx in range(day_header_row_idx + 1, ws.max_row + 1):
        name = _normalize_name(ws.cell(row_idx, col_name).value)
        if not name:
            continue
        attendance_group = ws.cell(row_idx, col_group).value if col_group else None
        department = ws.cell(row_idx, col_department).value if col_department else None
        employee_code = ws.cell(row_idx, col_code).value if col_code else None
        position = ws.cell(row_idx, col_position).value if col_position else None
        is_intern_row = _is_intern_attendance_row(attendance_group, department, position)
        has_intern_exemption = _has_intern_monthly_exemption(_normalize_text(position))
        remaining = intern_exemption_remaining.setdefault(name, [MONTHLY_EXEMPT_MINUTES]) if has_intern_exemption else None
        if has_scope_columns and not _is_monthly_row_in_scope(
            attendance_group=attendance_group,
            department=department,
            employee_code=employee_code,
            position=position,
        ):
            continue
        day_map: DailyMap = {}
        for offset, col_idx in enumerate(day_columns, start=1):
            daily_value = _parse_daily_text_value(
                ws.cell(row_idx, col_idx).value,
                monthly_exemption_remaining=remaining,
                count_outing_as_present=is_intern_row,
            )
            if daily_value is not None:
                day_map[offset] = daily_value
        _merge_daily_maps(result, name, day_map)

    wb.close()
    print(f"[{SOURCE_MONTHLY_SUMMARY}] 解析 {len(result)} 人")
    return result


def _split_role_chunks(text: str) -> list[str]:
    normalized = re.sub(r"\s+", " ", text.replace("\r", "\n"))
    pattern = r"(?=(?:大：|小：|大行宫：|驻场：))"
    return [chunk.strip() for chunk in re.split(pattern, normalized) if chunk.strip()]


def parse_beijing_universal_schedule(path: str) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    target_sheet = None
    for sheet in wb.worksheets:
        if _normalize_text(sheet.cell(7, 3).value) == "顺序" and _normalize_text(sheet.cell(7, 4).value) == "姓名":
            target_sheet = sheet
            break
    if target_sheet is None:
        wb.close()
        return {}

    result: SourceMap = {}
    for row_idx in range(9, target_sheet.max_row + 1):
        name = _normalize_name(target_sheet.cell(row_idx, 4).value)
        if not name:
            continue
        day_map: DailyMap = {}
        for day in range(1, 32):
            value = _normalize_text(target_sheet.cell(row_idx, day + 4).value)
            if value == "班":
                day_map[day] = _make_daily_entry(1.0)
        _merge_daily_maps(result, name, day_map)

    wb.close()
    print(f"[{SCHEDULE_BEIJING_UNIVERSAL}] 解析 {len(result)} 人")
    return result


def parse_nanjing_shop_schedule(path: str) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result: SourceMap = {}

    def _store(day: int, text: Any) -> None:
        if not isinstance(text, str):
            return
        for chunk in _split_role_chunks(text):
            if "：" not in chunk:
                continue
            _, raw_name = chunk.split("：", 1)
            clean_name = _normalize_name(re.sub(r"[（(].*?[）)]", "", raw_name))
            if not clean_name:
                continue
            value = 0.33 if "临时有事情" in chunk else 1.0
            _merge_daily_maps(result, clean_name, {day: _make_daily_entry(value)})

    for day in range(1, 16):
        _store(day, ws.cell(3, day).value)

    for offset, day in enumerate(range(16, 32), start=1):
        for row_idx in (6, 7, 8):
            _store(day, ws.cell(row_idx, offset).value)

    wb.close()
    print(f"[{SCHEDULE_NANJING_SHOP}] 解析 {len(result)} 人")
    return result


def parse_nanjing_station_schedule(path: str) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result: SourceMap = {}

    def _store(day: int, text: Any) -> None:
        if not isinstance(text, str) or "：" not in text:
            return
        _, raw_name = text.split("：", 1)
        clean_name = _normalize_name(raw_name)
        if not clean_name:
            return
        _merge_daily_maps(result, clean_name, {day: _make_daily_entry(1.0)})

    for day in range(1, 16):
        for row_idx in (3, 4):
            _store(day, ws.cell(row_idx, day).value)

    for offset, day in enumerate(range(16, 32), start=1):
        for row_idx in (7, 8, 9):
            _store(day, ws.cell(row_idx, offset).value)

    wb.close()
    print(f"[{SCHEDULE_NANJING_STATION}] 解析 {len(result)} 人")
    return result


def parse_changsha_schedule(path: str, month: int | None = None) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result: SourceMap = {}

    date_pattern = re.compile(r"(\d{1,2})月(\d{1,2})[日号]?")
    header_row_idx = 8
    duty_columns = [2, 4]
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        first_cell = _normalize_text(ws.cell(row_idx, 1).value)
        if "日期" not in first_cell:
            continue
        detected_columns = [
            col_idx
            for col_idx in range(2, ws.max_column + 1)
            if "班" in _normalize_text(ws.cell(row_idx, col_idx).value)
        ]
        if detected_columns:
            header_row_idx = row_idx
            duty_columns = detected_columns
            break

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        date_text = _normalize_text(ws.cell(row_idx, 1).value)
        match = date_pattern.search(date_text)
        if not match:
            continue
        parsed_month = int(match.group(1))
        day = int(match.group(2))
        if not 1 <= parsed_month <= 12:
            print(f"[{SCHEDULE_CHANGSHA}] 第{row_idx}行月份无效：{parsed_month}月，已跳过")
            continue
        if month is not None and parsed_month != month:
            continue
        daily_names = {_normalize_name(ws.cell(row_idx, col_idx).value) for col_idx in duty_columns}
        for name in daily_names:
            if name:
                _merge_daily_maps(result, name, {day: _make_daily_entry(1.0)})

    wb.close()
    print(f"[{SCHEDULE_CHANGSHA}] 解析 {len(result)} 人")
    return result


def _coerce_day_number(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.day
    if isinstance(value, date):
        return value.day
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float):
        if value.is_integer() and 1 <= int(value) <= 31:
            return int(value)
        return None

    text = _normalize_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{1,2}", text):
        day = int(text)
        return day if 1 <= day <= 31 else None

    match = re.search(r"\d{4}[-/.]\d{1,2}[-/.](\d{1,2})", text)
    if match:
        day = int(match.group(1))
        return day if 1 <= day <= 31 else None

    match = re.search(r"\d{1,2}月(\d{1,2})[日号]?", text)
    if match:
        day = int(match.group(1))
        return day if 1 <= day <= 31 else None

    return None


def _parse_schedule_marker(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return 1.0 if float(value) > 0 else None

    text = _normalize_text(value)
    if not text or text in {"/", "-", "0", "休"}:
        return None
    if text.startswith("休"):
        return None

    numeric_match = re.fullmatch(r"\d+(?:\.\d+)?", text)
    if numeric_match:
        return 1.0 if float(text) > 0 else None

    if any(marker in text for marker in ("班", "加班", "√", "是")):
        return 1.0
    if re.search(r"\d", text):
        return 1.0
    return None


def _find_matrix_schedule_header(ws) -> tuple[int | None, int | None, list[tuple[int, int]]]:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        header_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        col_name = _find_col(header_values, "姓名")
        if col_name is None:
            continue

        day_columns: list[tuple[int, int]] = []
        for col_idx in range(1, ws.max_column + 1):
            day = _coerce_day_number(ws.cell(row_idx, col_idx).value)
            if day is not None:
                day_columns.append((day, col_idx))
        if len(day_columns) >= 3:
            return row_idx, col_name, day_columns
    return None, None, []


def _infer_matrix_schedule_source(path: str) -> str:
    filename = os.path.basename(path)
    if "北京环球" in filename:
        return SCHEDULE_BEIJING_UNIVERSAL
    if "南京" in filename:
        return SCHEDULE_NANJING_SHOP
    return SCHEDULE_MATRIX


def parse_matrix_schedule(path: str, source_name: str | None = None) -> SourceMap:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result: SourceMap = {}

    header_row_idx, col_name, day_columns = _find_matrix_schedule_header(ws)
    if header_row_idx is None or col_name is None:
        wb.close()
        return {}
    header_values = [ws.cell(header_row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
    col_company = _find_company_col(header_values)

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        name = _normalize_name(ws.cell(row_idx, col_name).value)
        if not name or name in SCHEDULE_PLACEHOLDER_NAMES:
            continue
        company = _normalize_text(ws.cell(row_idx, col_company).value) if col_company else ""
        person_key = _make_person_key(name, company)

        day_map: DailyMap = {}
        for day, col_idx in day_columns:
            marker_value = _parse_schedule_marker(ws.cell(row_idx, col_idx).value)
            if marker_value is not None:
                day_map[day] = _make_daily_entry(marker_value)
        _merge_daily_maps(result, person_key, day_map)

    wb.close()
    label = source_name or SCHEDULE_MATRIX
    print(f"[{label}] 解析 {len(result)} 人")
    return result


def _looks_like_changsha_schedule(ws) -> bool:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        first_cell = _normalize_text(ws.cell(row_idx, 1).value)
        if "日期" not in first_cell:
            continue
        if any("班" in _normalize_text(ws.cell(row_idx, col_idx).value) for col_idx in range(2, ws.max_column + 1)):
            return True
    return False


def parse_schedule_file(path: str, month: int | None = None) -> tuple[str | None, SourceMap]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    max_row = ws.max_row
    max_col = ws.max_column
    sheet_names = tuple(wb.sheetnames)
    matrix_header_row_idx, _, _ = _find_matrix_schedule_header(ws)
    is_changsha_schedule = _looks_like_changsha_schedule(ws)
    wb.close()

    if "设置" in sheet_names:
        return SCHEDULE_BEIJING_UNIVERSAL, parse_beijing_universal_schedule(path)
    if max_col == 16 and max_row == 8:
        return SCHEDULE_NANJING_SHOP, parse_nanjing_shop_schedule(path)
    if max_col == 16 and max_row == 9:
        return SCHEDULE_NANJING_STATION, parse_nanjing_station_schedule(path)
    if is_changsha_schedule:
        return SCHEDULE_CHANGSHA, parse_changsha_schedule(path, month=month)
    if matrix_header_row_idx is not None:
        source_name = _infer_matrix_schedule_source(path)
        return source_name, parse_matrix_schedule(path, source_name=source_name)
    return None, {}


def _lookup_source_day_map(source_map: SourceMap, name: str) -> DailyMap | None:
    day_map = source_map.get(name)
    if day_map:
        return dict(day_map)
    return None


def _lookup_schedule_day_map(source_map: SourceMap, name: str, company: str) -> DailyMap | None:
    if company:
        day_map = source_map.get(_make_person_key(name, company))
        if day_map:
            return dict(day_map)
    return _lookup_source_day_map(source_map, name)


def _merge_day_maps(*day_maps: DailyMap) -> DailyMap:
    merged: DailyMap = {}
    for day_map in day_maps:
        _merge_daily_maps({"_": merged}, "_", day_map)
    return merged


def _entry_has_real_punch(entry: DailyEntry | None) -> bool:
    if not isinstance(entry, dict):
        return False
    if _is_unscheduled_punch_entry(entry):
        return True
    return _has_real_punch(_normalize_text(entry.get("raw_text")))


def _mark_unscheduled_punch_entry(entry: DailyEntry, schedule_source_name: str) -> DailyEntry | None:
    value = _entry_value(entry)
    if value is None or value <= 0:
        return None

    marked = dict(entry)
    raw_text = _normalize_text(marked.get("raw_text"))
    has_real_punch = _entry_has_real_punch(entry)
    has_overtime = _is_overtime_entry(entry)
    if has_overtime and has_real_punch:
        note = "未排班加班并打卡"
        action_text = "已打卡且有加班记录"
    elif has_overtime:
        note = "未排班加班"
        action_text = "有加班记录"
    else:
        note = "未排班并打卡"
        action_text = "已打卡"

    comment = f"排班表当日无排班，但考勤记录显示{action_text}（排班来源：{schedule_source_name}）"
    if raw_text:
        comment = f"{comment}\n源记录：{raw_text}"

    marked["note"] = _merge_note(_entry_note(marked), note)
    marked["comment"] = _merge_note(_entry_comment(marked), comment)
    marked["alert"] = True
    marked["unscheduled_punch"] = True
    return marked


def _merge_unscheduled_punch_days(
    schedule_day_map: DailyMap,
    attendance_day_map: DailyMap,
    schedule_source_name: str,
) -> DailyMap:
    merged = dict(schedule_day_map)
    for day, entry in attendance_day_map.items():
        if day in merged or not (_entry_has_real_punch(entry) or _is_overtime_entry(entry)):
            continue
        marked = _mark_unscheduled_punch_entry(entry, schedule_source_name)
        if marked is not None:
            merged[day] = marked
    return merged


def _resolve_row_days(
    row: dict[str, Any],
    attendance_detail: SourceMap,
    monthly_summary: SourceMap,
    schedules: dict[str, SourceMap],
    default_day_maps: dict[str, DailyMap],
    special_default_name_set: set[str],
) -> tuple[DailyMap, str, list[str]]:
    name = _normalize_name(row.get("姓名"))
    company = _normalize_text(row.get("所属公司"))
    warnings: list[str] = []
    allow_default_fill = bool(row.get("_allow_default_fill"))
    default_schedule_key = _normalize_text(row.get("_default_schedule_key")) or "main"
    fallback_day_map = default_day_maps.get(default_schedule_key) or default_day_maps["main"]

    if name in special_default_name_set:
        return dict(fallback_day_map), SOURCE_SPECIAL_DEFAULT_WEEKDAYS, warnings

    schedule_sources = [
        source_name
        for source_name in SCHEDULE_SOURCE_ORDER
        if _lookup_schedule_day_map(schedules.get(source_name, {}), name, company)
    ]
    if schedule_sources:
        schedule_day_maps = [
            day_map
            for source_name in schedule_sources
            if (day_map := _lookup_schedule_day_map(schedules[source_name], name, company))
        ]
        merged_schedule = _merge_day_maps(*schedule_day_maps)
        if merged_schedule:
            schedule_source_name = "、".join(schedule_sources)
            attendance_day_map = _merge_day_maps(
                _lookup_source_day_map(monthly_summary, name) or {},
                _lookup_source_day_map(attendance_detail, name) or {},
            )
            return (
                _merge_unscheduled_punch_days(
                    merged_schedule,
                    attendance_day_map,
                    schedule_source_name,
                ),
                schedule_source_name,
                warnings,
            )

    day_map = _lookup_source_day_map(monthly_summary, name)
    if day_map:
        return day_map, SOURCE_MONTHLY_SUMMARY, warnings

    day_map = _lookup_source_day_map(attendance_detail, name)
    if day_map:
        return day_map, SOURCE_ATTENDANCE_DETAIL, warnings

    if allow_default_fill:
        warnings.append(f"{name} 未匹配到日明细，按工作日默认填充。")
        return dict(fallback_day_map), SOURCE_DEFAULT_WEEKDAYS, warnings

    warnings.append(f"{name} 未匹配到日明细，保留为空白。")
    return {}, SOURCE_DEFAULT_WEEKDAYS, warnings


def _write_title_and_headers(ws, month: int, days_in_month: int) -> None:
    total_cols = len(STATIC_HEADERS) + days_in_month + 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(1, 1)
    title_cell.value = f"{OUTPUT_TITLE_PREFIX}{month}{OUTPUT_TITLE_SUFFIX}"
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = TITLE_FILL
    title_cell.border = THIN_BORDER

    for col_idx, header in enumerate(STATIC_HEADERS, start=1):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    start_day_col = len(STATIC_HEADERS) + 1
    for day in range(1, days_in_month + 1):
        cell = ws.cell(2, start_day_col + day - 1)
        cell.value = day
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    total_cell = ws.cell(2, start_day_col + days_in_month)
    total_cell.value = "实际出勤天数"
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.fill = HEADER_FILL
    total_cell.border = THIN_BORDER

    reminder_cell = ws.cell(2, start_day_col + days_in_month + 1)
    reminder_cell.value = "提醒"
    reminder_cell.font = Font(bold=True)
    reminder_cell.alignment = Alignment(horizontal="center", vertical="center")
    reminder_cell.fill = HEADER_FILL
    reminder_cell.border = THIN_BORDER

    ws.row_dimensions[1].height = ROW_HEIGHT_HEADER
    ws.row_dimensions[2].height = ROW_HEIGHT_HEADER


def _apply_layout(ws, days_in_month: int) -> None:
    widths = {
        1: 14,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    start_day_col = len(STATIC_HEADERS) + 1
    for col_idx in range(start_day_col, start_day_col + days_in_month):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    ws.column_dimensions[get_column_letter(start_day_col + days_in_month)].width = 12
    ws.column_dimensions[get_column_letter(start_day_col + days_in_month + 1)].width = 42

    ws.freeze_panes = ws.cell(3, start_day_col)
    ws.sheet_view.showGridLines = True


def _write_data_rows(
    ws,
    output_rows: list[dict[str, Any]],
    days_in_month: int,
    attendance_detail: SourceMap,
    monthly_summary: SourceMap,
    schedules: dict[str, SourceMap],
    default_day_maps: dict[str, DailyMap],
    special_default_name_set: set[str],
) -> tuple[list[str], int]:
    warnings: list[str] = []
    start_day_col = len(STATIC_HEADERS) + 1
    total_col = start_day_col + days_in_month
    reminder_col = total_col + 1
    next_row_idx = 3

    for output_row in output_rows:
        daily_map, source_name, row_warnings = _resolve_row_days(
            output_row,
            attendance_detail=attendance_detail,
            monthly_summary=monthly_summary,
            schedules=schedules,
            default_day_maps=default_day_maps,
            special_default_name_set=special_default_name_set,
        )

        for warn in row_warnings:
            warnings.append(f"[{source_name}] {warn}")

        total_value = _round3(
            sum(_entry_value(value) or 0.0 for value in daily_map.values() if value is not None)
        )
        if total_value <= 0:
            print(f"[跳过] {output_row.get('姓名')} <- {source_name}（实际出勤天数为 0）")
            continue

        row_idx = next_row_idx
        row_reminders: list[str] = []
        for col_idx, header in enumerate(STATIC_HEADERS, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = output_row.get(header)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for day in range(1, days_in_month + 1):
            cell = ws.cell(row_idx, start_day_col + day - 1)
            entry = daily_map.get(day)
            value = _entry_value(entry)
            if value is not None:
                cell.value = _round3(value)
                note = _entry_note(entry)
                comment = _merge_note(note, _entry_comment(entry))
                if comment:
                    cell.comment = Comment(comment, "考勤汇总")
                if note:
                    row_reminders.append(f"{day}日{note}")
                if _is_unscheduled_punch_entry(entry):
                    cell.fill = ATTENDANCE_UNSCHEDULED_PUNCH_FILL
                    warnings.append(f"[{source_name}] {output_row.get('姓名')} {day}日 {note}")
                elif _is_deducted_entry(entry):
                    cell.fill = ATTENDANCE_DEDUCT_FILL
                    warnings.append(f"[{source_name}] {output_row.get('姓名')} {day}日 {note}")
                elif _is_alert_entry(entry):
                    cell.fill = ATTENDANCE_ALERT_FILL
                    warnings.append(f"[{source_name}] {output_row.get('姓名')} {day}日 {note}")
                elif _is_offsite_entry(entry):
                    cell.fill = ATTENDANCE_OFFSITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        total_cell = ws.cell(row_idx, total_col)
        start_ref = get_column_letter(start_day_col)
        end_ref = get_column_letter(total_col - 1)
        total_cell.value = f"=ROUND(SUM({start_ref}{row_idx}:{end_ref}{row_idx}),2)"
        total_cell.border = THIN_BORDER
        total_cell.alignment = Alignment(horizontal="center", vertical="center")

        reminder_cell = ws.cell(row_idx, reminder_col)
        if row_reminders:
            reminder_cell.value = "；".join(row_reminders)
            reminder_cell.fill = ATTENDANCE_ALERT_FILL
        reminder_cell.border = THIN_BORDER
        reminder_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.row_dimensions[row_idx].height = ROW_HEIGHT_DATA
        print(f"[生成] {output_row.get('姓名')} <- {source_name}")
        next_row_idx += 1

    return warnings, next_row_idx - 3


def generate_parttime_summary(
    output_path: str,
    attendance_detail_path: str | None = None,
    monthly_summary_paths: list[str] | None = None,
    schedule_paths: list[str] | None = None,
    default_schedule_path: str | None = None,
    special_default_names: list[str] | tuple[str, ...] | set[str] | None = None,
) -> dict[str, Any]:
    schedule_paths = schedule_paths or []
    monthly_summary_paths = monthly_summary_paths or []
    if not any([attendance_detail_path, monthly_summary_paths, schedule_paths]):
        raise ValueError("至少需要上传一份考勤来源文件。")

    resolved_special_default_names = _normalize_name_collection(
        DEFAULT_SPECIAL_DEFAULT_NAMES if special_default_names is None else special_default_names
    )
    special_default_name_set = set(resolved_special_default_names)
    special_chengdu_names = set(
        name
        for name in _normalize_name_collection(DEFAULT_SPECIAL_CHENGDU_NAMES)
        if name in special_default_name_set
    )

    year, month = _infer_year_month(
        attendance_detail_path=attendance_detail_path,
        monthly_summary_paths=monthly_summary_paths,
        schedule_paths=schedule_paths,
        default_schedule_path=default_schedule_path,
    )
    _, days_in_month = calendar.monthrange(year, month)

    attendance_detail = parse_attendance_detail(attendance_detail_path) if attendance_detail_path else {}

    monthly_summary: SourceMap = {}
    for ms_path in monthly_summary_paths:
        parsed = parse_monthly_summary(ms_path)
        for name, day_map in parsed.items():
            _merge_daily_maps(monthly_summary, name, day_map)

    schedules: dict[str, SourceMap] = {}
    for path in schedule_paths:
        source_name, parsed = parse_schedule_file(path, month=month)
        if source_name and parsed:
            target = schedules.setdefault(source_name, {})
            for name, day_map in parsed.items():
                _merge_daily_maps(target, name, day_map)
            print(f"[排班] 已识别 {source_name}: {os.path.basename(path)}")
        else:
            print(f"[排班] 未识别文件，跳过: {os.path.basename(path)}")

    output_rows = _build_output_rows(
        attendance_detail=attendance_detail,
        monthly_summary=monthly_summary,
        schedules=schedules,
        special_default_names=resolved_special_default_names,
        special_chengdu_names=special_chengdu_names,
    )
    default_day_maps = _build_default_day_maps(year, month, default_schedule_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month}{OUTPUT_SHEET_SUFFIX}"
    _write_title_and_headers(ws, month=month, days_in_month=days_in_month)
    warnings, written_rows = _write_data_rows(
        ws,
        output_rows=output_rows,
        days_in_month=days_in_month,
        attendance_detail=attendance_detail,
        monthly_summary=monthly_summary,
        schedules=schedules,
        default_day_maps=default_day_maps,
        special_default_name_set=special_default_name_set,
    )
    _apply_layout(ws, days_in_month=days_in_month)

    wb.save(output_path)
    wb.close()

    warning_count = len(warnings)
    if warnings:
        print("[警告] 以下记录使用了兜底或存在未匹配情况：")
        for warning in warnings:
            print(f"  - {warning}")

    print(f"[完成] 已输出：{output_path}")
    return {
        "year": year,
        "month": month,
        "rows": written_rows,
        "warnings": warning_count,
        "output_path": output_path,
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="生成兼职汇总表")
    parser.add_argument("--attendance-detail", help="考勤明细.xlsx")
    parser.add_argument("--monthly-summary", action="append", default=[], help="月度汇总.xlsx，可多次传入")
    parser.add_argument("--schedule", action="append", default=[], help="排班表，可多次传入")
    parser.add_argument("--default-schedule", help="默认作息表，用于免打卡/兜底员工")
    parser.add_argument("--output", required=True, help="输出文件路径")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    generate_parttime_summary(
        attendance_detail_path=args.attendance_detail,
        monthly_summary_paths=args.monthly_summary,
        schedule_paths=args.schedule,
        default_schedule_path=args.default_schedule,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
