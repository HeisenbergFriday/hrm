from datetime import date, datetime
import os
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from overtime import fill_overtime_fields as ot
import runner


def _write_workbook(rows, sheet_name: str = "Sheet1", extra_sheets=None):
    """Save to a temp xlsx and return the path. rows[0] is the header."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(list(row))
    for name, sheet_rows in (extra_sheets or []):
        extra = wb.create_sheet(title=name)
        for row in sheet_rows:
            extra.append(list(row))
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def _build_attendance_map_for_punch(employee_code: str, punch_dates: dict):
    """Build a real attendance clock hours map for testing bilateral punch.

    punch_dates: {date: clock_hours_float} — simulates bilateral punch records.
    """
    ot._reset_attendance_runtime_state()
    ot._ATTENDANCE_CLOCK_HOURS_BY_CODE[employee_code] = punch_dates


def _build_no_punch_map(employee_code: str, no_punch_dates: set):
    """Build attendance no-punch map for dates with (-) or single-side punch."""
    ot._ATTENDANCE_NO_PUNCH_BY_CODE[employee_code] = no_punch_dates


class OpsGroupUnaddedTest(unittest.TestCase):
    """运营支撑部运维组强制判定为「未加」的回归测试。"""

    def _build_overtime_ws(self, dept1, dept2, dept3):
        wb = Workbook()
        ws = wb.active
        ws.title = "加班明细"
        ws.append(list(ot.CLEAN_OUTPUT_COLS))
        ws.append([
            "MT9999", "测试运维", dept1, dept2, dept3,
            datetime(2026, 6, 15, 9, 0), datetime(2026, 6, 15, 18, 0),
            None, None, None, None, 8.0, None, None,
        ])
        header_row = ot.find_header_row(ws)
        header_map = ot.build_header_map(ws, header_row)
        header_map = ot.ensure_target_columns(ws, header_row, header_map)
        return ws, header_row, header_map

    def _system_col(self, header_map):
        return header_map.get(ot.normalize_header_name("系统操作"))

    def _remark_col(self, header_map):
        return header_map.get(ot.normalize_header_name("备注"))

    def _overtime_type_col(self, header_map):
        return header_map.get(ot.normalize_header_name("加班类型"))

    def _final_hours_col(self, header_map):
        return header_map.get(ot.normalize_header_name("最终加班时长（小时）"))

    def _two_x_hours_col(self, header_map):
        return header_map.get(ot.normalize_header_name("2倍加班小时"))

    def _three_x_hours_col(self, header_map):
        return header_map.get(ot.normalize_header_name("3倍加班小时"))

    def test_ops_group_with_bilateral_punch_is_unadded(self):
        """运营支撑部+运维组+正常双边打卡 → 未加、备注非空、计算列清空。"""
        ws, header_row, header_map = self._build_overtime_ws("总部", "运营支撑部", "运维组")
        # 构造真实双边打卡数据：(09:00, 18:30) → 9.5 小时跨度
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertEqual(result, "未加")
        self.assertEqual(ws.cell(header_row + 1, self._system_col(header_map)).value, "未加")
        remark = ws.cell(header_row + 1, self._remark_col(header_map)).value
        self.assertTrue(remark and "运维组" in remark, f"备注应包含'运维组'，实际: {remark}")
        self.assertIsNone(ws.cell(header_row + 1, self._overtime_type_col(header_map)).value)
        self.assertIsNone(ws.cell(header_row + 1, self._final_hours_col(header_map)).value)
        self.assertIsNone(ws.cell(header_row + 1, self._two_x_hours_col(header_map)).value)
        self.assertIsNone(ws.cell(header_row + 1, self._three_x_hours_col(header_map)).value)

    def test_smart_locker_ops_group_with_bilateral_punch_is_unadded(self):
        """运营支撑部+智慧寄存运维组+正常双边打卡 → 未加、备注非空、计算列清空。"""
        ws, header_row, header_map = self._build_overtime_ws("总部", "运营支撑部", "智慧寄存运维组")
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertEqual(result, "未加")
        self.assertEqual(ws.cell(header_row + 1, self._system_col(header_map)).value, "未加")
        remark = ws.cell(header_row + 1, self._remark_col(header_map)).value
        self.assertTrue(remark and "运维组" in remark)
        self.assertIsNone(ws.cell(header_row + 1, self._overtime_type_col(header_map)).value)
        self.assertIsNone(ws.cell(header_row + 1, self._final_hours_col(header_map)).value)

    def test_ops_group_without_punch_is_unadded(self):
        """运营支撑部+运维组+无打卡 → 未加、备注非空。"""
        ws, header_row, header_map = self._build_overtime_ws("总部", "运营支撑部", "运维组")
        _build_no_punch_map("MT9999", {date(2026, 6, 15)})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertEqual(result, "未加")
        remark = ws.cell(header_row + 1, self._remark_col(header_map)).value
        self.assertTrue(remark and "运维组" in remark)

    def test_roster_maps_to_ops_group_then_unadded(self):
        """花名册通过员工编号把员工映射到运营支撑部/智慧寄存运维组 → 最终输出未加。"""
        rows = [
            ["员工编号", "员工姓名", "部门名称"],
            ["MT9999", "测试运维", "总部-运营支撑部-智慧寄存运维组"],
        ]
        roster_path = _write_workbook(rows, "花名册")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(roster_path)
        finally:
            os.remove(roster_path)
        self.assertIn("MT9999", mapping)
        self.assertEqual(mapping["MT9999"]["dept2"], "运营支撑部")
        self.assertEqual(mapping["MT9999"]["dept3"], "智慧寄存运维组")
        self.assertIsNone(diagnostic)

        # 用映射后部门构建加班表
        ws, header_row, header_map = self._build_overtime_ws("其他部门", "其他部门", "其他部门")
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, mapping,
        )
        self.assertEqual(result, "未加")
        remark = ws.cell(header_row + 1, self._remark_col(header_map)).value
        self.assertTrue(remark and "运维组" in remark)

    def test_operations_support_non_ops_group_keeps_normal_rule(self):
        """运营支撑部+客服组 → 不命中运维组规则，走普通规则。"""
        ws, header_row, header_map = self._build_overtime_ws("总部", "运营支撑部", "客服组")
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertNotEqual(result, "未加")
        remark = ws.cell(header_row + 1, self._remark_col(header_map)).value
        self.assertNotEqual(remark, ot.OPS_GROUP_DEPT_REMARK)

    def test_leaf_relative_department_contract_is_exact(self):
        """与 Go rosterDepartmentLevels 镜像：只按连续 dept2/dept3 精确命中。"""
        cases = [
            ("三层路径", "运营管理中心", "运营支撑部", "智慧寄存运维组", True),
            ("含企业根的四层路径", "运营管理中心", "运营支撑部", "智慧寄存运维组", True),
            ("超过四层路径", "运营管理中心", "运营支撑部", "智慧寄存运维组", True),
            ("非目标部门", "研发中心", "平台部", "运维组", False),
            ("相似但非精确名称", "运营管理中心", "运营支撑部", "智慧寄存运维组一组", False),
        ]
        for name, dept1, dept2, dept3, expected in cases:
            with self.subTest(name=name):
                self.assertEqual(ot._is_ops_group_by_depts(dept1, dept2, dept3), expected)

    def test_other_dept_with_ops_group_name_not_matched(self):
        """产品技术部+运维组 → 不命中（二级部门不是运营支撑部）。"""
        ws, header_row, header_map = self._build_overtime_ws("总部", "产品技术部", "运维组")
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertNotEqual(result, "未加")
        remark = ws.cell(header_row + 1, self._remark_col(header_map)).value
        self.assertNotEqual(remark, ot.OPS_GROUP_DEPT_REMARK)

    def test_dept_with_surrounding_whitespace_still_matched(self):
        """部门字段包含首尾空格与不可见字符，清洗后仍应正确识别运维组。"""
        ws, header_row, header_map = self._build_overtime_ws(
            " 总部 ", " 运营支撑部\n ", " 运维组 ",
        )
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertEqual(result, "未加")

    def test_similar_name_not_in_allowlist_not_matched(self):
        """名称相似但不在允许名单中的三级部门 → 不命中。"""
        ws, header_row, header_map = self._build_overtime_ws("总部", "运营支撑部", "运维组（临时）")
        _build_attendance_map_for_punch("MT9999", {date(2026, 6, 15): 9.5})
        result = ot.fill_row(
            ws, header_row + 1, header_map, {}, date(2026, 6, 1),
            None, None, set(), None, None,
        )
        self.assertNotEqual(result, "未加")


class RosterDepartmentMapTest(unittest.TestCase):
    """花名册/钉钉自动同步花名册部门映射识别的回归测试。"""

    def test_standard_roster_maps_employee_and_department(self):
        rows = [
            ["工号", "姓名", "一级部门", "二级部门", "三级部门", "岗位"],
            ["MT0001", "张三", "总部", "运营支撑部", "运维组", "工程师"],
        ]
        path = _write_workbook(rows, "花名册")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIn("MT0001", mapping)
        self.assertEqual(mapping["MT0001"]["dept2"], "运营支撑部")
        self.assertEqual(mapping["MT0001"]["dept3"], "运维组")
        self.assertIsNone(diagnostic)

    def test_header_offset_by_title_row_is_recognized(self):
        rows = [
            ["钉钉自动同步花名册", None, None, None, None, None],
            [None, None, None, None, None, None],
            ["工号", "姓名", "一级部门", "二级部门", "三级部门", "岗位"],
            ["MT0002", "李四", "总部", "运营支撑部", "客服组", "客服"],
        ]
        path = _write_workbook(rows, "花名册")
        try:
            mapping, _ = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIn("MT0002", mapping)

    def test_dept_header_with_whitespace_or_newline_is_recognized(self):
        rows = [
            ["工号", "姓名", " 一级部门 ", "二级部门 ", "三级部门"],
            ["MT0003", "王五", "总部", "运营支撑部", "运维组"],
        ]
        path = _write_workbook(rows, "花名册")
        try:
            mapping, _ = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIn("MT0003", mapping)

    def test_dingtalk_employee_number_and_dept_name_is_recognized(self):
        rows = [
            ["员工编号", "员工姓名", "部门名称"],
            ["MT0004", "赵六", "总部-运营支撑部-运维组"],
        ]
        path = _write_workbook(rows, "钉钉花名册")
        try:
            mapping, _ = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIn("MT0004", mapping)
        self.assertEqual(mapping["MT0004"]["dept2"], "运营支撑部")
        self.assertEqual(mapping["MT0004"]["dept3"], "运维组")

    def test_smart_locker_ops_group_dept_path_is_recognized(self):
        """花名册部门路径包含'智慧寄存运维组'时正确解析。"""
        rows = [
            ["员工编号", "员工姓名", "部门名称"],
            ["MT0010", "测试员工", "运营管理中心-运营支撑部-智慧寄存运维组"],
        ]
        path = _write_workbook(rows, "花名册")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIn("MT0010", mapping)
        self.assertEqual(mapping["MT0010"]["dept2"], "运营支撑部")
        self.assertEqual(mapping["MT0010"]["dept3"], "智慧寄存运维组")
        self.assertIsNone(diagnostic)

    def test_multiple_worksheets_picks_correct_one(self):
        good = [
            ["工号", "姓名", "一级部门", "二级部门", "三级部门"],
            ["MT0005", "钱七", "总部", "运营支撑部", "运维组"],
        ]
        bad = [
            ["说明", None],
            ["这是说明页，不包含员工与部门字段。", None],
        ]
        path = _write_workbook(bad, "说明", extra_sheets=[("花名册", good)])
        try:
            mapping, _ = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIn("MT0005", mapping)

    def test_missing_department_column_returns_diagnostic(self):
        rows = [
            ["工号", "姓名", "岗位"],
            ["MT0006", "孙八", "工程师"],
        ]
        path = _write_workbook(rows, "花名册")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertEqual(mapping, {})
        self.assertIsNotNone(diagnostic)
        joined = str(diagnostic)
        self.assertIn("花名册", joined)
        self.assertIn("岗位", joined)

    def test_transfer_process_file_not_recognized_as_roster(self):
        """岗位异动流程表不应被识别为花名册。"""
        rows = [
            ["实际申请人工号", "实际申请人", "发起人工号", "发起人姓名",
             "异动日期", "异动类型", "生效日期"],
            ["MT0007", "张三", "MT0007", "张三", "2026-07-01", "转岗", "2026-07-01"],
        ]
        path = _write_workbook(rows, "岗位异动流程表")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        # 不应匹配到任何员工（缺少部门列）
        self.assertEqual(mapping, {})
        self.assertIsNotNone(diagnostic)

    def test_transfer_process_with_initiator_dept_not_recognized_as_roster(self):
        """含有'发起人部门'的异动流程表不应被识别为花名册。"""
        rows = [
            ["实际申请人工号", "实际申请人", "发起人部门", "异动类型"],
            ["MT0008", "李四", "总部-运营支撑部-运维组", "转岗"],
        ]
        path = _write_workbook(rows, "岗位异动流程表")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        # "发起人部门"已从 DEPARTMENT_PATH_HEADERS 中移除，不应匹配
        self.assertEqual(mapping, {})

    def test_duplicate_names_do_not_create_ambiguous_name_fallback(self):
        rows = [
            ["工号", "姓名", "一级部门", "二级部门", "三级部门"],
            ["MT2001", "同名员工", "总部", "运营支撑部", "智慧寄存运维组"],
            ["MT2002", "同名员工", "总部", "产品部", "研发组"],
        ]
        path = _write_workbook(rows, "花名册")
        try:
            mapping, diagnostic = ot.parse_employee_department_map(path)
        finally:
            os.remove(path)
        self.assertIsNone(diagnostic)
        self.assertEqual(mapping["MT2001"]["dept3"], "智慧寄存运维组")
        self.assertEqual(mapping["MT2002"]["dept3"], "研发组")
        self.assertNotIn("name:同名员工", mapping)
        self.assertEqual(ot._lookup_employee_department(mapping, "", "同名员工"), {})

    def test_name_only_roster_is_rejected_by_overtime_entry_with_clear_error(self):
        roster_path = _write_workbook([["姓名"], ["测试运维"]], "在职花名册")
        try:
            with tempfile.TemporaryDirectory() as workdir:
                with self.assertRaisesRegex(ValueError, "未识别到可用部门映射"):
                    runner.run_overtime(
                        {"overtime_src": roster_path, "overtime_roster": roster_path},
                        Path(workdir),
                    )
        finally:
            os.remove(roster_path)


class GeneratedRosterOvertimeEndToEndTest(unittest.TestCase):
    """生产花名册生成器 -> 部门解析 -> 加班完整入口回归。"""

    @staticmethod
    def _write_overtime_export(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "加班导出"
        sheet.append([
            "审批编号", "审批状态", "审批结果", "发起人工号", "发起人姓名", "发起人部门",
            "开始时间", "结束时间", "明细", "时长",
        ])
        sheet.append([
            "OT-001", "完成", "同意", "MT9999", "测试运维", "其他部门-其他部门-其他部门",
            datetime(2026, 6, 15, 9, 0), datetime(2026, 6, 15, 18, 0), date(2026, 6, 15), 8,
        ])
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _write_attendance(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "考勤明细"
        sheet.append(["考勤时间：2026-06-01 至 2026-06-30"])
        sheet.append(["姓名", "工号", "考勤组", "考勤结果"])
        sheet.append([None, None, None, 15])
        sheet.append(["测试运维", "MT9999", "默认考勤组", "正常 (09:00, 18:30)"])
        workbook.save(path)
        workbook.close()

    def test_generated_roster_maps_ops_group_and_full_overtime_entry_marks_unadded(self):
        with tempfile.TemporaryDirectory() as workdir:
            root = Path(workdir)
            roster_output = root / "roster-output"
            roster_output.mkdir()
            generated = runner.action_generate_roster({
                "org_name": "测试组织",
                "employees": [{
                    "emp_no": "MT9999",
                    "name": "测试运维",
                    "dept1": "运营管理中心",
                    "dept2": "运营支撑部",
                    "dept3": "智慧寄存运维组",
                    "position": "运维工程师",
                    "emp_type": "正式",
                }],
            }, roster_output)
            roster_path = Path(generated[0]["path"])

            workbook = load_workbook(roster_path, data_only=True)
            try:
                sheet = workbook["在职花名册"]
                headers = [cell.value for cell in sheet[1]]
                generated_row = [sheet.cell(2, column).value for column in range(1, len(headers) + 1)]
            finally:
                workbook.close()
            self.assertEqual(headers, runner._ROSTER_HEADERS)
            self.assertEqual(generated_row[0:6], [
                "MT9999", "测试运维", None, "运营管理中心", "运营支撑部", "智慧寄存运维组",
            ])

            mapping, diagnostic = ot.parse_employee_department_map(str(roster_path))
            self.assertIsNone(diagnostic)
            self.assertEqual(mapping["MT9999"]["dept2"], "运营支撑部")
            self.assertEqual(mapping["MT9999"]["dept3"], "智慧寄存运维组")

            overtime_path = root / "overtime.xlsx"
            attendance_path = root / "attendance.xlsx"
            output_dir = root / "overtime-output"
            output_dir.mkdir()
            self._write_overtime_export(overtime_path)
            self._write_attendance(attendance_path)

            outputs = runner.run_overtime({
                "overtime_src": str(overtime_path),
                "overtime_roster": str(roster_path),
                "overtime_attendance": str(attendance_path),
                "overtime_target_month": "2026-06",
            }, output_dir)
            result_path = next(Path(item["path"]) for item in outputs if item.get("path", "").endswith(".xlsx"))
            result_workbook = load_workbook(result_path, data_only=True)
            try:
                result_sheet = result_workbook.active
                header_row = ot.find_header_row(result_sheet)
                header_map = ot.build_header_map(result_sheet, header_row)
                values = {
                    name: result_sheet.cell(header_row + 1, column).value
                    for name, column in header_map.items()
                }
            finally:
                result_workbook.close()

            self.assertEqual(values["系统操作"], "未加")
            self.assertIn("运维组", values["备注"])
            self.assertIsNone(values["加班类型"])
            self.assertIsNone(values["最终加班时长（小时）"])
            self.assertIsNone(values["2倍加班小时"])
            self.assertIsNone(values["3倍加班小时"])


class RunnerRosterErrorTest(unittest.TestCase):
    """runner.format_roster_department_error 应给出明确提示而非笼统文案。"""

    def _import_runner(self):
        import sys
        for key in ("tools/attendance_toolbox/python",):
            if key not in sys.path:
                sys.path.insert(0, key)
        import runner
        return runner

    def test_missing_dept_error_contains_actual_headers(self):
        runner = self._import_runner()
        diagnostic = {
            "matched_sheets": [],
            "headers_preview": [
                {"sheet": "花名册", "headers": ["工号", "姓名", "岗位"]},
            ],
            "missing": "员工标识字段或部门字段缺失",
        }
        msg = runner.format_roster_department_error(diagnostic, "花名册_钉钉自动同步.xlsx")
        self.assertIn("花名册_钉钉自动同步.xlsx", msg)
        self.assertIn("工号", msg)
        self.assertIn("岗位", msg)


if __name__ == "__main__":
    unittest.main()
