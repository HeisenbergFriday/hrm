"""
rules_engine.py
──────────────────────────────────────────────────────────────────────────────
加班规则引擎：将硬编码的倍数规则、部门匹配、日期判定等抽成可配置的模块。

数据来源：
  - OvertimeConfig：内存中的完整配置
  - 加班规则配置.xlsx：Excel 配置文件（可选，不提供时使用默认配置）

核心函数：
  get_default_config()          — 返回与当前硬编码逻辑一致的默认配置
  load_config(path)             — 从 Excel 读取配置
  save_config(config, path)     — 写入 Excel
  classify_employee(...)        — 判断员工所属部门组
  classify_overtime_date(...)   — 判断日期类型（5 类加班日期类型）
  classify_date(...)            — classify_overtime_date 的兼容包装
  determine_action(...)         — 按优先级匹配倍数规则
  calculate_hours(...)          — 计算最终加班时长
  should_mark_no_punch(...)     — 判断是否标记"未加"
──────────────────────────────────────────────────────────────────────────────
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from functools import lru_cache
from typing import Iterable, Protocol

import openpyxl


# ── 日期类型常量 ─────────────────────────────────────────────────────────────

LEGAL_HOLIDAY = "LEGAL_HOLIDAY"
HOLIDAY_ADJUST_REST = "HOLIDAY_ADJUST_REST"
ORDINARY_WEEKEND = "ORDINARY_WEEKEND"
MAKEUP_WORKDAY = "MAKEUP_WORKDAY"
NORMAL_WORKDAY = "NORMAL_WORKDAY"

DATE_TYPE_LABELS = {
    LEGAL_HOLIDAY: "法定节假日当天",
    HOLIDAY_ADJUST_REST: "因法定节假日放假调休形成的休息日",
    ORDINARY_WEEKEND: "普通周六日",
    MAKEUP_WORKDAY: "调休上班日",
    NORMAL_WORKDAY: "普通工作日",
}

SPECIAL_DEPARTMENTS = {"运营支撑部"}


class HolidayDataSourceError(RuntimeError):
    """节假日数据源缺失、覆盖不足或无法可靠分类时抛出。"""


class HolidayDataSource(Protocol):
    """可替换的节假日数据源接口。"""

    def classify(self, overtime_date: date) -> str:
        """返回 LEGAL_HOLIDAY / HOLIDAY_ADJUST_REST / ORDINARY_WEEKEND / MAKEUP_WORKDAY / NORMAL_WORKDAY。"""

    def validate_year(self, year: int) -> None:
        """确认指定年份有可靠数据。"""

    def available_years(self) -> tuple[int, int] | None:
        """返回数据覆盖年份范围；未知时返回 None。"""


def _normalize_date(value: date | datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise TypeError(f"日期类型不支持：{type(value)!r}，需要 datetime.date")


def _dates_from_mapping_or_set(values) -> Iterable[date]:
    if values is None:
        return ()
    if hasattr(values, "keys"):
        return values.keys()
    return values


class StaticHolidayDataSource:
    """
    静态节假日数据源，主要用于测试或企业日历 API 已经返回清晰分类后的适配。
    生产默认不使用它，避免把手工表重新塞回主逻辑。
    """

    def __init__(
        self,
        *,
        legal_holidays: Iterable[date] = (),
        holiday_adjust_rest_days: Iterable[date] = (),
        makeup_workdays: Iterable[date] = (),
        supported_years: Iterable[int] | None = None,
    ) -> None:
        self.legal_holidays = set(legal_holidays)
        self.holiday_adjust_rest_days = set(holiday_adjust_rest_days)
        self.makeup_workdays = set(makeup_workdays)
        if supported_years is None:
            years = {
                d.year
                for d in (
                    self.legal_holidays
                    | self.holiday_adjust_rest_days
                    | self.makeup_workdays
                )
            }
        else:
            years = set(supported_years)
        self._years = years

    def available_years(self) -> tuple[int, int] | None:
        if not self._years:
            return None
        return min(self._years), max(self._years)

    def validate_year(self, year: int) -> None:
        if self._years and year not in self._years:
            raise HolidayDataSourceError(
                f"节假日数据源未覆盖 {year} 年，请升级节假日数据源或补充企业日历数据。"
            )

    def classify(self, overtime_date: date) -> str:
        d = _normalize_date(overtime_date)
        self.validate_year(d.year)
        if d in self.legal_holidays:
            return LEGAL_HOLIDAY
        if d in self.holiday_adjust_rest_days:
            return HOLIDAY_ADJUST_REST
        if d in self.makeup_workdays:
            return MAKEUP_WORKDAY
        if d.weekday() >= 5:
            return ORDINARY_WEEKEND
        return NORMAL_WORKDAY


class ChineseCalendarHolidayDataSource:
    """
    默认中国节假日数据源。

    chinese_calendar 能识别国务院放假安排中的休息日、调休上班日和普通周末，
    但它的 holidays 集合不区分“3 倍法定当天”和“连休中的其他休息日”。
    因此这里叠加《全国年节及纪念日放假办法》的法定日计算器来拆分 3 倍日。
    """

    def __init__(self) -> None:
        self._calendar = None

    def _load_calendar(self):
        if self._calendar is not None:
            return self._calendar
        try:
            import chinese_calendar as calendar
        except ImportError as exc:
            raise HolidayDataSourceError(
                "未安装 chinesecalendar，无法自动识别中国节假日。"
                "请执行：python -m pip install -U chinesecalendar lunardate"
            ) from exc

        required = ("holidays", "workdays", "is_workday", "get_solar_terms")
        missing = [name for name in required if not hasattr(calendar, name)]
        if missing:
            raise HolidayDataSourceError(
                "当前 chinesecalendar 缺少必要接口/常量："
                f"{', '.join(missing)}；请升级：python -m pip install -U chinesecalendar"
            )
        self._calendar = calendar
        return calendar

    def available_years(self) -> tuple[int, int] | None:
        calendar = self._load_calendar()
        years = {
            d.year
            for d in (
                list(_dates_from_mapping_or_set(calendar.holidays))
                + list(_dates_from_mapping_or_set(calendar.workdays))
            )
            if isinstance(d, date)
        }
        if not years:
            return None
        return min(years), max(years)

    def validate_year(self, year: int) -> None:
        years = self.available_years()
        if years and not (years[0] <= year <= years[1]):
            raise HolidayDataSourceError(
                f"节假日数据源未覆盖 {year} 年，当前 chinesecalendar 仅覆盖 "
                f"{years[0]}-{years[1]} 年；请升级依赖或接入更新的企业日历 API。"
            )
        try:
            self._statutory_legal_holidays(year)
        except HolidayDataSourceError:
            raise
        except Exception as exc:
            raise HolidayDataSourceError(
                f"{year} 年法定节假日日期计算失败，请升级节假日数据源。"
            ) from exc

    def classify(self, overtime_date: date) -> str:
        d = _normalize_date(overtime_date)
        self.validate_year(d.year)
        calendar = self._load_calendar()

        try:
            is_workday = bool(calendar.is_workday(d))
        except NotImplementedError as exc:
            raise HolidayDataSourceError(
                f"节假日数据源未覆盖 {d.year} 年：{exc}；请升级 chinesecalendar。"
            ) from exc

        statutory_days = self._statutory_legal_holidays(d.year)
        holiday_arrangement_days = set(_dates_from_mapping_or_set(calendar.holidays))
        makeup_workdays = set(_dates_from_mapping_or_set(calendar.workdays))

        if d in makeup_workdays:
            return MAKEUP_WORKDAY
        if d in statutory_days:
            return LEGAL_HOLIDAY
        if d in holiday_arrangement_days:
            return HOLIDAY_ADJUST_REST
        if is_workday:
            return NORMAL_WORKDAY
        if d.weekday() >= 5:
            return ORDINARY_WEEKEND

        raise HolidayDataSourceError(
            f"{d.isoformat()} 被识别为休息日，但无法判断是否属于节假日调休休息日；"
            "请升级节假日数据源或接入企业日历 API。"
        )

    @lru_cache(maxsize=None)
    def _statutory_legal_holidays(self, year: int) -> frozenset[date]:
        if year < 2014:
            raise HolidayDataSourceError(
                f"暂未内置 {year} 年的法定节假日当天规则；"
                "请接入能直接返回 LEGAL_HOLIDAY / HOLIDAY_ADJUST_REST 的数据源。"
            )

        try:
            from lunardate import LunarDate
        except ImportError as exc:
            raise HolidayDataSourceError(
                "未安装 lunardate，无法计算春节/端午/中秋的法定当天。"
                "请执行：python -m pip install -U lunardate"
            ) from exc

        calendar = self._load_calendar()
        lunar_new_year = LunarDate(year, 1, 1).toSolarDate()

        if year >= 2025:
            spring_festival = {lunar_new_year - timedelta(days=1)}
            spring_festival.update(lunar_new_year + timedelta(days=offset) for offset in range(3))
            labour_day = {date(year, 5, 1), date(year, 5, 2)}
        else:
            spring_festival = {lunar_new_year + timedelta(days=offset) for offset in range(3)}
            labour_day = {date(year, 5, 1)}

        qingming_terms = [
            term_date
            for term_date, _ in calendar.get_solar_terms(date(year, 4, 1), date(year, 4, 7))
            if term_date.month == 4 and term_date.day in (4, 5, 6)
        ]
        if not qingming_terms:
            raise HolidayDataSourceError(f"无法计算 {year} 年清明节日期，请升级节假日数据源。")

        legal_days = {
            date(year, 1, 1),
            qingming_terms[0],
            LunarDate(year, 5, 5).toSolarDate(),
            LunarDate(year, 8, 15).toSolarDate(),
            date(year, 10, 1),
            date(year, 10, 2),
            date(year, 10, 3),
        }
        legal_days.update(spring_festival)
        legal_days.update(labour_day)
        return frozenset(legal_days)


_DEFAULT_HOLIDAY_DATA_SOURCE: HolidayDataSource | None = None


def get_default_holiday_data_source() -> HolidayDataSource:
    global _DEFAULT_HOLIDAY_DATA_SOURCE
    if _DEFAULT_HOLIDAY_DATA_SOURCE is None:
        _DEFAULT_HOLIDAY_DATA_SOURCE = ChineseCalendarHolidayDataSource()
    return _DEFAULT_HOLIDAY_DATA_SOURCE


def classify_overtime_date(
    overtime_date: date,
    data_source: HolidayDataSource | None = None,
) -> str:
    """
    返回日期类型：
    - LEGAL_HOLIDAY：法定节假日当天，所有人 3 倍
    - HOLIDAY_ADJUST_REST：因法定节假日放假调休形成的休息日，运营支撑部 2 倍，其他人调休
    - ORDINARY_WEEKEND：普通周六日，所有人调休
    - MAKEUP_WORKDAY：调休上班日，不触发 2 倍 / 3 倍
    - NORMAL_WORKDAY：普通工作日，不触发 2 倍 / 3 倍
    """
    source = data_source or get_default_holiday_data_source()
    return source.classify(overtime_date)


def get_overtime_result(
    overtime_date: date,
    department: str,
    data_source: HolidayDataSource | None = None,
) -> str:
    date_type = classify_overtime_date(overtime_date, data_source=data_source)

    if date_type == LEGAL_HOLIDAY:
        return "3倍工资"

    if date_type == HOLIDAY_ADJUST_REST and department in SPECIAL_DEPARTMENTS:
        return "2倍工资"

    return "调休"


def validate_holiday_years_available(
    years: Iterable[int],
    data_source: HolidayDataSource | None = None,
) -> None:
    source = data_source or get_default_holiday_data_source()
    for year in sorted(set(years)):
        source.validate_year(year)


# ── 数据结构 ─────────────────────────────────────────────────────────────────

@dataclass
class DepartmentRule:
    """部门组定义：一个匹配条件。"""
    group_name: str
    match_field: str        # "一级部门" | "二级部门" | "三级部门" | "考勤组"
    match_method: str       # "包含" | "等于"
    match_value: str


@dataclass
class PremiumRule:
    """倍数规则：一行优先级匹配。"""
    priority: int
    date_type: str          # LEGAL_HOLIDAY / HOLIDAY_ADJUST_REST / ...
    department_group: str   # 部门组名称，或 "全部"
    action: str             # "加班工资" | "调休"
    multiplier: float


@dataclass
class OvertimeConfig:
    """完整加班规则配置。"""
    premium_rules: list[PremiumRule]
    department_rules: list[DepartmentRule]
    standard_hours_per_day: float = 8.0
    no_punch_mark: str = "未加"
    schedule_augment_holidays: bool = True
    schedule_augment_rest_dept_group: str = "运营支撑部"
    chengdu_use_separate_calendar: bool = True
    legal_holidays_override: set[date] | None = None
    holiday_data_source: HolidayDataSource | None = field(default=None, repr=False, compare=False)

    # 内部索引：group_name -> [DepartmentRule]
    _dept_group_index: dict[str, list[DepartmentRule]] = field(default_factory=dict, repr=False)

    def build_lookups(self) -> None:
        """构建内部索引。"""
        self._dept_group_index = {}
        for r in self.department_rules:
            self._dept_group_index.setdefault(r.group_name, []).append(r)


# ── 默认配置 ─────────────────────────────────────────────────────────────────

def get_default_config() -> OvertimeConfig:
    """返回与当前硬编码逻辑完全一致的默认配置。"""
    config = OvertimeConfig(
        premium_rules=[
            PremiumRule(1, LEGAL_HOLIDAY, "全部", "加班工资", 3.0),
            PremiumRule(2, HOLIDAY_ADJUST_REST, "运营支撑部", "加班工资", 2.0),
            PremiumRule(3, HOLIDAY_ADJUST_REST, "全部", "调休", 1.0),
            PremiumRule(4, ORDINARY_WEEKEND, "全部", "调休", 1.0),
            PremiumRule(5, MAKEUP_WORKDAY, "全部", "调休", 1.0),
            PremiumRule(6, NORMAL_WORKDAY, "全部", "调休", 1.0),
        ],
        department_rules=[
            DepartmentRule("运营支撑部", "一级部门", "包含", "运营支撑部"),
            DepartmentRule("运营支撑部", "二级部门", "包含", "运营支撑部"),
            DepartmentRule("运营支撑部", "三级部门", "包含", "运营支撑部"),
            DepartmentRule("运营支撑部", "考勤组",   "包含", "运维组"),
            DepartmentRule("运营支撑部", "考勤组",   "包含", "深圳客服"),
        ],
    )
    config.build_lookups()
    return config


# ── 配置加载 / 保存 ─────────────────────────────────────────────────────────

def _header_key(value) -> str:
    return re.sub(r"[\s\u00a0\u2000-\u200f\u2028-\u202f\u205f\u3000\ufeff]+", "", str(value or "")).strip()

def _read_sheet(ws) -> list[dict]:
    """读取一个 sheet 为 list[dict]，首行为表头。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_header_key(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({
            headers[i]: row[i] if i < len(row) else None
            for i in range(len(headers))
            if headers[i]
        })
    return result


def _norm_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_float(v, default=0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return default


def _norm_int(v, default=0) -> int:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except ValueError:
        return default


def _parse_date_cell(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def load_config(config_path: str) -> OvertimeConfig:
    """从 Excel 配置文件加载配置。"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"规则配置文件不存在：{config_path}")

    wb = openpyxl.load_workbook(config_path, data_only=True)

    # 倍数规则
    premium_rules = []
    if "倍数规则" in wb.sheetnames:
        for r in _read_sheet(wb["倍数规则"]):
            premium_rules.append(PremiumRule(
                priority=_norm_int(r.get("优先级"), 99),
                date_type=_norm_str(r.get("日期类型")),
                department_group=_norm_str(r.get("部门组")) or "全部",
                action=_norm_str(r.get("加班类型")),
                multiplier=_norm_float(r.get("倍数")),
            ))
    premium_rules.sort(key=lambda r: r.priority)

    # 部门组
    department_rules = []
    if "部门组" in wb.sheetnames:
        for r in _read_sheet(wb["部门组"]):
            department_rules.append(DepartmentRule(
                group_name=_norm_str(r.get("部门组名称")),
                match_field=_norm_str(r.get("匹配字段")),
                match_method=_norm_str(r.get("匹配方式")) or "包含",
                match_value=_norm_str(r.get("匹配值")),
            ))

    # 参数设置
    params = {}
    if "参数设置" in wb.sheetnames:
        for r in _read_sheet(wb["参数设置"]):
            k = _norm_str(r.get("参数名"))
            v = _norm_str(r.get("值"))
            if k:
                params[k] = v

    # 法定节假日
    holidays_override = None
    if "法定节假日" in wb.sheetnames:
        dates = set()
        for r in _read_sheet(wb["法定节假日"]):
            d = _parse_date_cell(r.get("日期"))
            if d:
                dates.add(d)
        if dates:
            holidays_override = dates

    wb.close()

    config = OvertimeConfig(
        premium_rules=premium_rules,
        department_rules=department_rules,
        standard_hours_per_day=_norm_float(params.get("标准每日工时"), 8.0),
        no_punch_mark=params.get("未打卡标记", "未加"),
        schedule_augment_holidays=params.get("排班补录-法定节假日", "是") == "是",
        schedule_augment_rest_dept_group=params.get("排班补录-休息日部门组", "运营支撑部"),
        chengdu_use_separate_calendar=params.get("成都使用独立日历", "是") == "是",
        legal_holidays_override=holidays_override,
    )
    config.build_lookups()
    return config


def load_config_or_default(config_path: str | None) -> OvertimeConfig:
    """安全加载：路径有效则读取，否则返回默认配置。"""
    if config_path and os.path.exists(config_path):
        try:
            return load_config(config_path)
        except Exception as exc:
            print(f"[规则引擎] 配置文件读取失败：{exc}，使用默认规则")
    return get_default_config()


def save_config(config: OvertimeConfig, config_path: str) -> None:
    """将配置写入 Excel 文件。"""
    wb = openpyxl.Workbook()

    # Sheet 1: 倍数规则
    ws = wb.active
    ws.title = "倍数规则"
    ws.append(["优先级", "日期类型", "部门组", "加班类型", "倍数"])
    for r in config.premium_rules:
        ws.append([r.priority, r.date_type, r.department_group, r.action, r.multiplier])

    # Sheet 2: 部门组
    ws2 = wb.create_sheet("部门组")
    ws2.append(["部门组名称", "匹配字段", "匹配方式", "匹配值"])
    for r in config.department_rules:
        ws2.append([r.group_name, r.match_field, r.match_method, r.match_value])

    # Sheet 3: 参数设置
    ws3 = wb.create_sheet("参数设置")
    ws3.append(["参数名", "值"])
    ws3.append(["标准每日工时", config.standard_hours_per_day])
    ws3.append(["未打卡标记", config.no_punch_mark])
    ws3.append(["排班补录-法定节假日", "是" if config.schedule_augment_holidays else "否"])
    ws3.append(["排班补录-休息日部门组", config.schedule_augment_rest_dept_group])
    ws3.append(["成都使用独立日历", "是" if config.chengdu_use_separate_calendar else "否"])

    # Sheet 4: 法定节假日
    ws4 = wb.create_sheet("法定节假日")
    ws4.append(["日期", "备注"])
    if config.legal_holidays_override:
        for d in sorted(config.legal_holidays_override):
            ws4.append([d, ""])

    # Sheet 5: 版本
    ws5 = wb.create_sheet("版本")
    ws5.append(["参数", "值"])
    ws5.append(["配置版本", "1.0"])
    ws5.append(["创建日期", date.today().isoformat()])

    out_dir = os.path.dirname(os.path.abspath(config_path))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(config_path)
    wb.close()


# ── 规则判定函数 ─────────────────────────────────────────────────────────────

def classify_employee(
    config: OvertimeConfig,
    dept_d1: str | None,
    dept_d2: str | None,
    dept_d3: str | None,
    attendance_group: str | None,
) -> set[str]:
    """返回员工所属的所有部门组名称集合。"""
    dept_values = {
        "一级部门": str(dept_d1 or ""),
        "二级部门": str(dept_d2 or ""),
        "三级部门": str(dept_d3 or ""),
        "考勤组":   str(attendance_group or ""),
    }
    matched: set[str] = set()
    for group_name, rules in config._dept_group_index.items():
        for rule in rules:
            field_val = dept_values.get(rule.match_field, "")
            if not field_val:
                continue
            if rule.match_method == "等于":
                if field_val == rule.match_value:
                    matched.add(group_name)
                    break
            else:  # 包含
                if rule.match_value in field_val:
                    matched.add(group_name)
                    break
    return matched


def classify_date(
    config: OvertimeConfig,
    d: date,
    legal_holidays: set[date],
    work_calendar: set[date] | None,
    calendar_loaded: bool,
) -> str:
    """兼容旧调用的日期分类包装；实际分类由节假日数据源完成。"""
    if legal_holidays and d in legal_holidays:
        return LEGAL_HOLIDAY
    return classify_overtime_date(d, data_source=config.holiday_data_source)


def determine_action(
    config: OvertimeConfig,
    date_type: str,
    employee_groups: set[str],
) -> tuple[str, float]:
    """按优先级匹配倍数规则，返回 (加班类型, 倍数)。"""
    for rule in config.premium_rules:
        if rule.date_type != date_type:
            continue
        if rule.department_group == "全部" or rule.department_group in employee_groups:
            return rule.action, rule.multiplier
    # 无匹配 → 默认调休
    return "调休", 1.0


def calculate_hours(
    config: OvertimeConfig,
    action: str,
    raw_hours: float | None,
    two_x_days: float,
    three_x_days: float,
    day_count: int,
) -> float | None:
    """计算最终加班时长。"""
    if action == "加班工资":
        return round((two_x_days + three_x_days) * config.standard_hours_per_day, 2)
    # 调休
    if raw_hours is None:
        return None
    max_hours = max(day_count, 1) * config.standard_hours_per_day
    return round(min(raw_hours, max_hours), 2)


def should_mark_no_punch(
    config: OvertimeConfig,
    attendance_missing_checkout_map: dict[str, set[date]],
    employee_code: str,
    overtime_date: date | None,
    action: str,
) -> bool:
    """仅调休记录检查是否标记'未加'。"""
    if action != "加班工资" and employee_code and overtime_date is not None:
        return overtime_date in attendance_missing_checkout_map.get(employee_code, set())
    return False
