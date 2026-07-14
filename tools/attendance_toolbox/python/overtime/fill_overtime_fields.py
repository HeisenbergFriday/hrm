"""
fill_overtime_fields.py
──────────────────────────────────────────────────────────────────────────────
功能：
  读取系统导出的加班表，按当前加班规则补充以下 8 个字段：
    - 最终加班时长（小时）
    - 加班类型
    - 2倍加班小时
    - 3倍加班小时
    - 2倍加班天数
    - 3倍加班天数
    - 备注
    - 系统操作

默认口径（按行政说明实现）：
  - 法定节假日：所有员工记为 3 倍加班工资
  - 节假日调休形成的休息日：运营支撑部员工记为 2 倍加班工资（陈星雨除外），其余员工记为调休
  - 普通周六日：所有员工记为调休
  - 调休上班日/普通工作日：所有员工记为调休
  - 当月数据：
      最终加班时长按现有规则回填具体时长
  - 非当月数据：
      最终加班时长（小时）列仅标明月份，不显示具体时长
  - 调休时：
      最终加班时长 = min(系统时长, 8 * 天数)
      若考勤明细匹配到当日打卡时间为”(-)”（未打卡），系统操作 = 未加
      其余情况系统操作留空
  - 加班工资时：
      最终加班时长按 2倍/3倍加班小时回填；天数由小时 / 标准每日工时派生；
      已审批的加班工资日若完全无打卡，按 1 天计并保留核实备注
      系统操作留空
  - 跨天记录：
      逐日独立判定类型，分别累加 2 倍/3 倍天数

说明：
  - 排班表是可选的：
      用于补录审批导出中缺失的加班记录，不参与 2 倍 / 3 倍主规则判定。
      法定节假日全员补录（3倍）；节假日调休休息日仅在考勤明细可识别特殊部门时补录（2倍）。
  - 作息表是可选的：
      仅用于辅助识别处理月份；2 倍 / 3 倍主规则由节假日数据源判断。
  - 备注列会标出无打卡核实项，以及最终回填时长和源系统时长不一致的记录。
  - 若历史表里存在“法定按2倍计算”“本人要求转成调休时长”等人工例外，
    仍需人工补充或在脚本中新增规则。

用法：
  python fill_overtime_fields.py
  python fill_overtime_fields.py --export 加班系统导出.xlsx --schedule 春节排班.xlsx
  python fill_overtime_fields.py --attendance 考勤明细.xlsx
  python fill_overtime_fields.py --output 加班明细_回填.xlsx
  python fill_overtime_fields.py --schedule ""
──────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime, timedelta

import openpyxl


_HERE = os.path.dirname(os.path.abspath(__file__))
_BASE = os.path.dirname(_HERE)
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)
_leave_path = os.path.join(_BASE, "leave")
if _leave_path not in sys.path:
    sys.path.insert(0, _leave_path)

import calc_leave  # noqa: E402  # 复用特殊名单解析和默认成都名单
from excel_compat import load_workbook_compat  # noqa: E402

DEFAULT_EXPORT_FILE = os.path.join(_HERE, "加班系统导出.xlsx")
DEFAULT_SCHEDULE = os.path.join(_HERE, "排班表.xlsx")
DEFAULT_ATTENDANCE = os.path.join(_HERE, "考勤明细.xlsx")
DEFAULT_WORK_CALENDAR = os.path.join(_HERE, "作息表.xlsx")
DEFAULT_OUTPUT = os.path.join(_HERE, "加班明细_回填.xlsx")
DEFAULT_CHENGDU_WORK_LOCATION_NAMES = calc_leave.DEFAULT_CHENGDU_WORK_LOCATION_NAMES

# ── 清洗：允许通过的审批状态 ────────────────────────────────────────────────
KEEP_STATUSES = {"完成", "审批中", "已修改"}

# ── 清洗后输出的目标列（与加班明细.xlsx 前13列对齐）──────────────────────────
CLEAN_OUTPUT_COLS = (
    "发起人工号", "发起人姓名", "一级部门", "二级部门", "三级部门",
    "开始时间", "结束时间", "明细", "加班时间", "加班时长", "时长",
    "是否包含法定节假日期", "2026法定节假日如下：",
)

TARGET_COLUMNS = [
    "最终加班时长（小时）",
    "加班类型",
    "2倍加班小时",
    "3倍加班小时",
    "2倍加班天数",
    "3倍加班天数",
    "备注",
    "系统操作",
]

STANDARD_HOURS_PER_DAY = 8.0
PREMIUM_HOUR_VALUE_PRECISION = 6
PREMIUM_DAY_VALUE_PRECISION = 6
PREMIUM_HOUR_NUMBER_FORMAT = "0.00"
PREMIUM_DAY_NUMBER_FORMAT = "0.00"
REST_PREMIUM_EXCLUDED_NAMES = {"陈星雨"}
REST_PREMIUM_EXCLUDED_CODES = {"MT0019"}
NO_ATTENDANCE_REMARK = "当天无打卡记录需核实"
NO_DOUBLE_ATTENDANCE_REMARK = "无双边打卡，按系统时长兜底，需核实"
NO_SYSTEM_HOURS_REMARK = "无双边打卡且系统时长为空，暂按0小时，需人工核实"
SYSTEM_REMAINDER_EXHAUSTED_REMARK = "系统剩余时长不足，无双边打卡日期暂按0小时，需人工核实"
APPROVAL_TOTAL_CAP_EXHAUSTED_REMARK = "审批总时长上限已用完，本日暂按0小时，需核实"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据当前加班规则回填加班表中的人工字段。"
    )
    parser.add_argument(
        "--export",
        default=DEFAULT_EXPORT_FILE,
        help=f"加班系统导出文件，默认: {DEFAULT_EXPORT_FILE}",
    )
    parser.add_argument(
        "--schedule",
        default=DEFAULT_SCHEDULE,
        help=f"排班表；仅用于补录缺失记录，可不提供或文件不存在，默认: {DEFAULT_SCHEDULE}",
    )
    parser.add_argument(
        "--attendance",
        default="",
        help=f"考勤明细表；可不提供或文件不存在，示例: {DEFAULT_ATTENDANCE}",
    )
    parser.add_argument(
        "--roster",
        default="",
        help="花名册/员工信息表；可选，用于按工号修正部门归属后再判断部门组",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"输出文件路径，默认: {DEFAULT_OUTPUT}",
    )
    parser.add_argument(
        "--schedule-sheet",
        default=None,
        help="排班表 sheet 名；不填时默认第一个 sheet",
    )
    parser.add_argument(
        "--work-calendar",
        default=DEFAULT_WORK_CALENDAR,
        help=f"作息表文件路径；用于判断工作日/休息日，默认: {DEFAULT_WORK_CALENDAR}",
    )
    parser.add_argument(
        "--config",
        default=None,
        help="加班规则配置 Excel 文件路径；不提供时使用默认规则",
    )
    parser.add_argument(
        "--special-chengdu-names",
        default=None,
        help="成都作息名单姓名，支持换行、逗号、顿号、空格分隔；不填时使用默认名单",
    )
    return parser.parse_args()


def normalize_name(value) -> str:
    return str(value or "").strip()


def parse_special_chengdu_names(text: str | None = None) -> tuple[str, ...]:
    if text:
        return calc_leave.parse_special_employee_names(text)
    return DEFAULT_CHENGDU_WORK_LOCATION_NAMES


def normalize_header_name(value) -> str:
    text = str(value or "")
    return re.sub(r"[\s\u00a0\u2000-\u200f\u2028-\u202f\u205f\u3000\ufeff]+", "", text).strip()


def normalize_employee_code(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().upper()


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def compact_number(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def round2(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value + 1e-9, 2)


def _format_hours_for_remark(value: float | int | None) -> str:
    rounded = round2(float(value)) if value is not None else None
    return str(compact_number(rounded))


def _join_remarks(*parts: str | None) -> str | None:
    remarks: list[str] = []
    for part in parts:
        text = normalize_name(part)
        if text and text not in remarks:
            remarks.append(text)
    return "；".join(remarks) if remarks else None


def _system_hours_mismatch_remark(
    raw_hours: float | None,
    final_hours: float | None,
) -> str | None:
    """标出源系统时长和回填时长不一致的记录，便于人工复核。"""
    if raw_hours is None or final_hours is None:
        return None
    if abs(float(final_hours) - float(raw_hours)) <= 0.01:
        return None
    return (
        "系统时长"
        f"{_format_hours_for_remark(raw_hours)}小时，"
        f"回填{_format_hours_for_remark(final_hours)}小时，需核实"
    )


def _double_attendance_remark(final_hours: float | int | None) -> str | None:
    if final_hours is None:
        return None
    return f"当天有双边打卡，按打卡回填{_format_hours_for_remark(final_hours)}小时"


def _recognized_hours_mismatch_remark(
    system_hours: float | None,
    final_hours: float | None,
    basis: str,
) -> str | None:
    if system_hours is None or final_hours is None:
        return None
    if abs(float(final_hours) - float(system_hours)) <= 0.01:
        return None
    basis_text = "按打卡认定" if basis == "clock" else "最终认定"
    return (
        f"系统时长{_format_hours_for_remark(system_hours)}小时，"
        f"{basis_text}{_format_hours_for_remark(final_hours)}小时，需核实"
    )


def _system_total_mismatch_remark(
    system_hours: float | None,
    final_hours: float | None,
) -> str | None:
    if system_hours is None or final_hours is None:
        return None
    diff = round2(float(final_hours) - float(system_hours))
    if diff is None or abs(diff) <= 0.01:
        return None
    direction = "超出" if diff > 0 else "少于"
    return (
        f"系统总时长{_format_hours_for_remark(system_hours)}小时，"
        f"逐日认定{_format_hours_for_remark(final_hours)}小时，"
        f"{direction}{_format_hours_for_remark(abs(diff))}小时，需核实"
    )


def parse_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_datetime(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d",
        ):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def parse_date_cell(value) -> date | None:
    dt = parse_datetime(value)
    return dt.date() if dt else None


def parse_day_number(value) -> int | None:
    if isinstance(value, (int, float)):
        day_num = int(value)
        if float(value) == day_num and 1 <= day_num <= 31:
            return day_num
        return None
    text = normalize_name(value)
    if not text:
        return None
    try:
        day_num = int(text)
    except ValueError:
        return None
    return day_num if 1 <= day_num <= 31 else None


def get_current_month_anchor() -> date:
    return datetime.now().date().replace(day=1)


def infer_target_month_anchor(ws, header_map: dict[str, int], header_row: int) -> date:
    """从加班数据里推断本次处理月份，避免用运行当天月份误判历史数据。"""
    month_counts: dict[tuple[int, int], int] = {}
    month_order: list[tuple[int, int]] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if all(is_blank(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)):
            continue
        start_dt = parse_datetime(ws.cell(row_idx, header_map["开始时间"]).value)
        end_dt = parse_datetime(ws.cell(row_idx, header_map["结束时间"]).value)
        overtime_date = resolve_overtime_date(ws, row_idx, header_map, start_dt, end_dt)
        if overtime_date is None:
            continue
        month_key = (overtime_date.year, overtime_date.month)
        if month_key not in month_counts:
            month_counts[month_key] = 0
            month_order.append(month_key)
        month_counts[month_key] += 1

    if month_counts:
        year, month = max(month_order, key=lambda key: month_counts[key])
        print(f"[月份] 按加班数据推断处理月份：{year}年{month}月")
        return date(year, month, 1)

    fallback = get_current_month_anchor()
    print(f"[月份] 未能从加班数据推断月份，临时按运行月份处理：{fallback.year}年{fallback.month}月")
    return fallback


def resolve_overtime_date(
    ws,
    row_idx: int,
    header_map: dict[str, int],
    start_dt: datetime | None,
    end_dt: datetime | None,
) -> date | None:
    # 跨天记录拆成逐日行后，当前行的明细日期才是本行应参与判定的加班日期。
    if "明细" in header_map:
        detail_date = parse_date_cell(ws.cell(row_idx, header_map["明细"]).value)
        if detail_date:
            if start_dt and end_dt:
                s_date = start_dt.date()
                e_date = end_dt.date()
                if e_date < s_date:
                    s_date, e_date = e_date, s_date
                if s_date <= detail_date <= e_date:
                    return detail_date
            else:
                return detail_date
    if "加班时间" in header_map:
        overtime_date = parse_date_cell(ws.cell(row_idx, header_map["加班时间"]).value)
        if overtime_date:
            return overtime_date
    if start_dt:
        return start_dt.date()
    if end_dt:
        return end_dt.date()
    return None


def is_same_month(left: date | None, right: date | None) -> bool:
    if not left or not right:
        return False
    return left.year == right.year and left.month == right.month


def format_overtime_month_label(overtime_date: date | None, current_month_anchor: date) -> str | None:
    if not overtime_date:
        return None
    if overtime_date.year == current_month_anchor.year:
        return f"{overtime_date.month}月"
    return f"{overtime_date.year}年{overtime_date.month}月"


def should_count_in_target_month(
    overtime_date: date | None,
    current_month_anchor: date,
) -> bool:
    """Only the target month should write numeric overtime values."""
    if overtime_date is None:
        return True
    return is_same_month(overtime_date, current_month_anchor)


def has_no_punch_record(cell_text: str) -> bool:
    """判断考勤单元格是否属于"未正常打卡"，即系统操作需标记为"未加"。

    触发条件（满足任一）：
      1. 打卡时间为 (-) —— 完全无打卡记录
      2. 括号内只有一个时间值，如 (18:39) —— 仅打了单边卡

    正常打卡示例：(09:00,22:22)（两个时间，逗号分隔）
    """
    # 条件1：(-)
    if re.search(r'\(\s*-\s*\)', cell_text):
        return True
    # 条件2：括号内仅一个时间，格式 (HH:MM) 或 (H:MM)，无逗号
    if re.search(r'\(\d{1,2}:\d{2}\)', cell_text):
        return True
    return False


def has_empty_punch_record(cell_text: str) -> bool:
    """判断考勤单元格是否为完全无打卡记录。"""
    return bool(re.search(r'\(\s*-\s*\)', cell_text or ""))


def extract_attendance_clock_hours(cell_text: str) -> float | None:
    """从考勤单元格括号内提取当天最早到最晚打卡跨度（小时）。"""
    text = normalize_name(cell_text)
    if not text or has_empty_punch_record(text):
        return None

    # 单元格正文可能包含流程申请时间，实际打卡时间稳定放在括号内。
    groups = re.findall(r"\(([^()]*)\)", text)
    candidates = groups if groups else [text]
    for group in reversed(candidates):
        times = re.findall(r"\b(\d{1,2}):(\d{2})\b", group)
        if len(times) < 2:
            continue
        minutes = [int(h) * 60 + int(m) for h, m in times]
        first = min(minutes)
        last = max(minutes)
        if last < first:
            last += 24 * 60
        return round2((last - first) / 60)
    return None


def find_attendance_header_row(ws) -> int:
    name_key = normalize_header_name("姓名")
    code_key = normalize_header_name("工号")
    for row_idx in range(1, min(6, ws.max_row) + 1):
        row_values = [normalize_header_name(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if name_key in row_values and code_key in row_values:
            return row_idx
    raise ValueError("未找到考勤明细表表头行（需包含“姓名”“工号”）。")


def parse_attendance_year_month(ws) -> tuple[int, int]:
    # 格式1：文字描述，如 "2026-04-01 至 2026-04-30"
    pattern = re.compile(r"(\d{4})-(\d{1,2})-\d{1,2}\s*至\s*(\d{4})-(\d{1,2})")
    for row_idx in range(1, min(4, ws.max_row) + 1):
        for col_idx in range(1, min(4, ws.max_column) + 1):
            text = normalize_name(ws.cell(row_idx, col_idx).value)
            if not text:
                continue
            matched = pattern.search(text)
            if matched:
                return int(matched.group(1)), int(matched.group(2))

    # 格式2：日期数字行（如 30,31,1,2,3...），找到 31→1 分界推断月份
    for row_idx in range(1, min(5, ws.max_row) + 1):
        day_nums: list[tuple[int, int]] = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row_idx, col_idx).value
            n = None
            if isinstance(val, (int, float)) and float(val) == int(val):
                n = int(val)
            elif isinstance(val, str) and val.strip().isdigit():
                n = int(val.strip())
            if n is not None and 1 <= n <= 31:
                day_nums.append((col_idx, n))
        if len(day_nums) < 5:
            continue
        for i in range(1, len(day_nums)):
            if day_nums[i - 1][1] >= 28 and day_nums[i][1] == 1:
                now = datetime.now()
                target_month = now.month - 1 or 12
                target_year = now.year if now.month > 1 else now.year - 1
                return target_year, target_month
        return datetime.now().year, datetime.now().month
    raise ValueError("考勤明细表未识别到统计月份。")

def build_attendance_date_map(
    ws,
    header_row: int,
    header_map: dict[str, int],
    year: int,
    month: int,
) -> dict[int, date]:
    date_row = header_row + 1
    if date_row > ws.max_row:
        raise ValueError("考勤明细表缺少日期行。")

    anchor_col = None
    anchor_day = None
    for col in range(1, ws.max_column + 1):
        day_num = parse_day_number(ws.cell(date_row, col).value)
        if day_num is not None:
            anchor_col = col
            anchor_day = day_num
            break

    if anchor_col is None or anchor_day is None:
        raise ValueError("考勤明细表未识别到日期列。")

    attendance_result_col = header_map.get("考勤结果")
    if attendance_result_col is not None and 0 <= anchor_col - attendance_result_col <= 1:
        start_col = attendance_result_col
    else:
        start_col = anchor_col
    date_map: dict[int, date] = {}
    for col in range(start_col, ws.max_column + 1):
        day_num = anchor_day + (col - anchor_col)
        try:
            date_map[col] = date(year, month, day_num)
        except ValueError:
            continue

    if not date_map:
        raise ValueError("考勤明细表日期范围解析失败。")
    return date_map


def _load_attendance_sheet(attendance_path: str):
    """加载考勤明细表并返回 (wb, ws, header_row, header_map, year, month, date_map)。"""
    wb = openpyxl.load_workbook(attendance_path, data_only=True)
    ws = wb.active
    header_row = find_attendance_header_row(ws)
    header_map = build_header_map(ws, header_row)
    if "工号" not in header_map:
        wb.close()
        raise ValueError("考勤明细表缺少必要字段：工号")
    year, month = parse_attendance_year_month(ws)
    date_map = build_attendance_date_map(ws, header_row, header_map, year, month)
    return wb, ws, header_row, header_map, year, month, date_map


def _reset_attendance_runtime_state() -> None:
    global _ATTENDANCE_CLOCK_HOURS_BY_CODE, _ATTENDANCE_CLOCK_HOURS_BY_NAME
    global _ATTENDANCE_NO_PUNCH_BY_CODE, _ATTENDANCE_NO_PUNCH_BY_NAME
    _ATTENDANCE_CLOCK_HOURS_BY_CODE = {}
    _ATTENDANCE_CLOCK_HOURS_BY_NAME = {}
    _ATTENDANCE_NO_PUNCH_BY_CODE = {}
    _ATTENDANCE_NO_PUNCH_BY_NAME = {}


def parse_attendance_missing_checkout_map(attendance_path: str) -> dict[str, set[date]]:
    global _ATTENDANCE_CLOCK_HOURS_BY_CODE, _ATTENDANCE_CLOCK_HOURS_BY_NAME
    global _ATTENDANCE_NO_PUNCH_BY_CODE, _ATTENDANCE_NO_PUNCH_BY_NAME
    _reset_attendance_runtime_state()
    wb, ws, header_row, header_map, year, month, date_map = _load_attendance_sheet(attendance_path)
    try:
        no_punch_result: dict[str, set[date]] = {}
        clock_by_code: dict[str, dict[date, float]] = {}
        clock_by_name: dict[str, dict[date, float]] = {}
        empty_by_code: dict[str, set[date]] = {}
        empty_by_name: dict[str, set[date]] = {}
        name_col = header_map.get("姓名")

        for row_idx in range(header_row + 2, ws.max_row + 1):
            employee_code = normalize_employee_code(ws.cell(row_idx, header_map["工号"]).value)
            employee_name = normalize_name(ws.cell(row_idx, name_col).value) if name_col else ""
            if not employee_code and not employee_name:
                continue

            for col_idx, current_date in date_map.items():
                cell_text = normalize_name(ws.cell(row_idx, col_idx).value)
                if has_no_punch_record(cell_text):
                    if employee_code:
                        no_punch_result.setdefault(employee_code, set()).add(current_date)
                    if has_empty_punch_record(cell_text):
                        if employee_code:
                            empty_by_code.setdefault(employee_code, set()).add(current_date)
                        if employee_name:
                            empty_by_name.setdefault(employee_name, set()).add(current_date)

                clock_hours = extract_attendance_clock_hours(cell_text)
                if clock_hours is None:
                    continue
                if employee_code:
                    clock_by_code.setdefault(employee_code, {})[current_date] = clock_hours
                if employee_name:
                    clock_by_name.setdefault(employee_name, {})[current_date] = clock_hours
    finally:
        wb.close()

    _ATTENDANCE_CLOCK_HOURS_BY_CODE = clock_by_code
    _ATTENDANCE_CLOCK_HOURS_BY_NAME = clock_by_name
    _ATTENDANCE_NO_PUNCH_BY_CODE = empty_by_code
    _ATTENDANCE_NO_PUNCH_BY_NAME = empty_by_name
    no_punch_count = sum(len(days) for days in no_punch_result.values())
    clock_count = sum(len(days) for days in clock_by_code.values())
    empty_count = sum(len(days) for days in empty_by_code.values())
    print(f"[考勤明细] 未打卡人员 {len(no_punch_result)} 人，{no_punch_count} 条记录（打卡时间为 (-) 或单边卡）")
    print(f"[考勤明细] 解析到 {clock_count} 条双边打卡跨度，{empty_count} 条完全无打卡记录")
    return no_punch_result


def load_attendance_if_available(attendance_path: str | None) -> tuple[dict[str, set[date]], bool]:
    path = normalize_name(attendance_path)
    if not path:
        _reset_attendance_runtime_state()
        print("未提供考勤明细表，调休记录的系统操作列默认留空。")
        return {}, False
    if not os.path.exists(path):
        _reset_attendance_runtime_state()
        print(f"未找到考勤明细表文件：{path}，调休记录的系统操作列默认留空。")
        return {}, False
    try:
        return parse_attendance_missing_checkout_map(path), True
    except Exception as exc:
        _reset_attendance_runtime_state()
        print(f"考勤明细表解析失败：{exc}；调休记录的系统操作列默认留空。")
        return {}, False


def build_attendance_name_group_map(attendance_path: str) -> dict[str, str]:
    """从考勤明细表构建 员工姓名→考勤组 的映射，用于补录时判断部门归属。"""
    wb, ws, header_row, header_map, year, month, date_map = _load_attendance_sheet(attendance_path)
    try:
        name_col = header_map.get("姓名")
        if not name_col:
            return {}
        group_col = header_map.get("考勤组")
        result: dict[str, str] = {}
        for row_idx in range(header_row + 2, ws.max_row + 1):
            name = normalize_name(ws.cell(row_idx, name_col).value)
            if not name:
                continue
            group = ""
            if group_col:
                group = normalize_name(ws.cell(row_idx, group_col).value)
            result[name] = group
        return result
    finally:
        wb.close()


def load_attendance_name_group_map_if_available(
    attendance_path: str | None,
) -> dict[str, str]:
    path = normalize_name(attendance_path)
    if not path or not os.path.exists(path):
        return {}
    try:
        return build_attendance_name_group_map(path)
    except Exception:
        return {}


def should_mark_system_unadded(
    attendance_missing_checkout_map: dict[str, set[date]],
    employee_code: str,
    overtime_date: date | None,
) -> bool:
    if not employee_code or overtime_date is None:
        return False
    return overtime_date in attendance_missing_checkout_map.get(employee_code, set())


def daterange_inclusive(start_date: date, end_date: date) -> list[date]:
    if end_date < start_date:
        return []
    result = []
    current = start_date
    while current <= end_date:
        result.append(current)
        current = current.fromordinal(current.toordinal() + 1)
    return result


def split_dept(dept_str) -> tuple:
    """将 '一级-二级-三级' 形式的部门字符串拆分为三元组，不足三级用 None 补齐。"""
    if not dept_str:
        return None, None, None
    parts = str(dept_str).split("-")
    d1 = parts[0] or None
    d2 = parts[1] if len(parts) > 1 else None
    d3 = "-".join(parts[2:]) if len(parts) > 2 else None
    return d1, d2, (d3 or None)


def _first_text(value) -> str | None:
    text = normalize_name(value)
    return text or None


def _find_col_in_headers(header_vals: tuple | list, *names: str) -> int | None:
    normalized = [normalize_header_name(value) for value in header_vals]
    for name in names:
        key = normalize_header_name(name)
        if key in normalized:
            return normalized.index(key)
    return None


def _split_roster_department_text(value, contract_entity: str | None = None) -> tuple[str | None, str | None, str | None]:
    text = normalize_name(value)
    if not text:
        return None, None, None
    parts = [part.strip() for part in re.split(r"[-/／>＞]+", text) if part and part.strip()]
    if not parts:
        return None, None, None

    contract = normalize_name(contract_entity)
    first = parts[0]
    if contract and (first == contract or first in contract or contract in first):
        parts = parts[1:]
    elif len(parts) > 1 and ("公司" in first or "有限公司" in first):
        parts = parts[1:]

    d1 = parts[0] if len(parts) > 0 else None
    d2 = parts[1] if len(parts) > 1 else None
    d3 = "-".join(parts[2:]) if len(parts) > 2 else None
    return d1, d2, d3


def _department_name_key(name: str | None) -> str:
    return f"name:{normalize_name(name)}"


def parse_employee_department_map(roster_path: str | None) -> dict[str, dict[str, str]]:
    """从花名册/最终表类文件中提取工号到部门的映射，用于加班部门组判定。"""
    path = normalize_name(roster_path)
    if not path:
        return {}
    if not os.path.exists(path):
        print(f"[员工部门] 未找到花名册/员工信息表：{path}，沿用加班导出部门")
        return {}

    wb = load_workbook_compat(path, data_only=True)
    result: dict[str, dict[str, str]] = {}
    matched_sheets = 0
    matched_rows = 0

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_idx = None
        header_vals = None
        for idx, row in enumerate(rows[:20]):
            normalized = {normalize_header_name(value) for value in row}
            has_emp = bool(normalized & {normalize_header_name("工号"), normalize_header_name("员工工号"), normalize_header_name("发起人工号")})
            has_name = bool(normalized & {normalize_header_name("姓名"), normalize_header_name("员工姓名"), normalize_header_name("发起人姓名")})
            has_dept = bool(normalized & {
                normalize_header_name("一级部门"), normalize_header_name("1级部门"),
                normalize_header_name("二级部门"), normalize_header_name("2级部门"),
                normalize_header_name("三级部门"), normalize_header_name("3级部门"),
                normalize_header_name("部门路径"), normalize_header_name("所属部门"),
                normalize_header_name("发起人部门"), normalize_header_name("部门"),
            })
            if (has_emp or has_name) and has_dept:
                header_row_idx = idx
                header_vals = row
                break

        if header_row_idx is None or header_vals is None:
            continue

        matched_sheets += 1
        col_emp = _find_col_in_headers(header_vals, "工号", "员工工号", "发起人工号")
        col_name = _find_col_in_headers(header_vals, "姓名", "员工姓名", "发起人姓名")
        col_contract = _find_col_in_headers(header_vals, "合同主体", "所属公司", "公司主体", "主体")
        col_dept1 = _find_col_in_headers(header_vals, "一级部门", "1级部门", "一级组织", "1级组织")
        col_dept2 = _find_col_in_headers(header_vals, "二级部门", "2级部门", "二级组织", "2级组织")
        col_dept3 = _find_col_in_headers(header_vals, "三级部门", "3级部门", "三级组织", "3级组织")
        col_dept_path = _find_col_in_headers(header_vals, "部门路径", "完整部门", "主部门", "所属部门", "发起人部门", "部门")
        if col_dept_path in {col_dept1, col_dept2, col_dept3}:
            col_dept_path = None
        col_att_group = _find_col_in_headers(header_vals, "考勤组", "考勤组名称")

        for row in rows[header_row_idx + 1:]:
            emp_no = normalize_employee_code(row[col_emp]) if col_emp is not None and col_emp < len(row) else ""
            name = normalize_name(row[col_name]) if col_name is not None and col_name < len(row) else ""
            if not emp_no and not name:
                continue

            def cell(col_idx: int | None):
                if col_idx is None or col_idx >= len(row):
                    return None
                return row[col_idx]

            contract = _first_text(cell(col_contract))
            dept1 = _first_text(cell(col_dept1))
            dept2 = _first_text(cell(col_dept2))
            dept3 = _first_text(cell(col_dept3))
            if not (dept1 or dept2 or dept3) and col_dept_path is not None:
                dept1, dept2, dept3 = _split_roster_department_text(cell(col_dept_path), contract)

            attendance_group = _first_text(cell(col_att_group))
            if not (dept1 or dept2 or dept3 or attendance_group):
                continue

            info = {
                "emp_no": emp_no,
                "name": name,
                "dept1": dept1 or "",
                "dept2": dept2 or "",
                "dept3": dept3 or "",
                "attendance_group": attendance_group or "",
            }
            if emp_no:
                result[emp_no] = info
            if name and _department_name_key(name) not in result:
                result[_department_name_key(name)] = info
            matched_rows += 1

    wb.close()
    if matched_sheets:
        unique_codes = sum(1 for key in result if not key.startswith("name:"))
        print(f"[员工部门] 已从花名册/员工信息表读取 {unique_codes} 个工号部门映射（{matched_sheets} 个 sheet）")
    else:
        print("[员工部门] 花名册/员工信息表未识别到可用部门列，沿用加班导出部门")
    return result


def _lookup_employee_department(
    employee_department_map: dict[str, dict[str, str]] | None,
    employee_code: str | None,
    emp_name: str | None,
) -> dict[str, str]:
    if not employee_department_map:
        return {}
    code = normalize_employee_code(employee_code)
    if code and code in employee_department_map:
        return employee_department_map[code]
    name_key = _department_name_key(emp_name)
    return employee_department_map.get(name_key, {})


def _resolve_employee_department_fields(
    ws,
    row_idx: int,
    header_map: dict[str, int],
    employee_code: str,
    emp_name: str,
    name_group_map: dict[str, str] | None,
    employee_department_map: dict[str, dict[str, str]] | None,
) -> tuple[str, str, str, str]:
    dept_d1 = str(ws.cell(row_idx, header_map["一级部门"]).value or "") if "一级部门" in header_map else ""
    dept_d2 = str(ws.cell(row_idx, header_map["二级部门"]).value or "") if "二级部门" in header_map else ""
    dept_d3 = str(ws.cell(row_idx, header_map["三级部门"]).value or "") if "三级部门" in header_map else ""
    att_group = name_group_map.get(emp_name, "") if name_group_map else ""

    roster_info = _lookup_employee_department(employee_department_map, employee_code, emp_name)
    if roster_info:
        dept_d1 = roster_info.get("dept1") or dept_d1
        dept_d2 = roster_info.get("dept2") or dept_d2
        dept_d3 = roster_info.get("dept3") or dept_d3
        att_group = roster_info.get("attendance_group") or att_group

    return dept_d1, dept_d2, dept_d3, att_group


def clean_export_overtime(export_file: str) -> list[tuple]:
    """
    读取加班系统导出文件（多 sheet），合并、过滤并整理为标准列格式，返回行列表。
    第 0 行为表头元组，后续为数据行元组。

    保留条件：
      审批状态 in {'完成', '审批中', '已修改'}
      且 NOT (审批状态 == '完成' AND 审批结果 == '拒绝')

    返回列顺序（CLEAN_OUTPUT_COLS）：
      发起人工号 / 发起人姓名 / 一级部门 / 二级部门 / 三级部门
      / 开始时间 / 结束时间 / 明细 / 加班时间 / 加班时长 / 时长
      / 是否包含法定节假日期 / 2026法定节假日如下：
    """
    if not os.path.exists(export_file):
        raise FileNotFoundError(f"未找到加班系统导出文件：{export_file}")

    wb_src = openpyxl.load_workbook(export_file, data_only=True)
    out_rows: list[tuple] = [CLEAN_OUTPUT_COLS]
    kept = skipped = 0

    for ws in wb_src.worksheets:
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue

        header = sheet_rows[0]
        col: dict[str, int] = {}
        for idx, value in enumerate(header):
            header_key = normalize_header_name(value)
            if header_key and header_key not in col:
                col[header_key] = idx

        required = {"审批状态", "审批结果", "发起人工号", "发起人姓名",
                    "发起人部门", "开始时间", "结束时间", "时长"}
        required_keys = {normalize_header_name(name) for name in required}
        if required_keys - col.keys():
            continue   # 缺必要字段的 sheet 跳过

        approved_ids: set[str] = set()
        for row in sheet_rows[1:]:
            if not any(row):
                continue
            status = row[col[normalize_header_name("审批状态")]]
            result = row[col[normalize_header_name("审批结果")]]

            if status not in KEEP_STATUSES:
                skipped += 1
                continue
            if status == "完成" and result == "拒绝":
                skipped += 1
                continue

            # 按审批编号+明细日期去重（同一审批不同天的明细保留多行）
            approval_key = normalize_header_name("审批编号")
            detail_col_key = normalize_header_name("明细")
            approval_id = str(row[col[approval_key]] if approval_key in col else "").strip()
            detail_raw = row[col.get(detail_col_key, -1)] if detail_col_key in col else None
            detail_key = str(detail_raw).strip() if detail_raw else ""
            dedup_key = (approval_id, detail_key)
            if dedup_key in approved_ids:
                skipped += 1
                continue
            if approval_id:
                approved_ids.add(dedup_key)

            d1, d2, d3 = split_dept(row[col[normalize_header_name("发起人部门")]])
            raw_start = row[col[normalize_header_name("开始时间")]]
            raw_end = row[col[normalize_header_name("结束时间")]]
            raw_duration = row[col[normalize_header_name("时长")]]

            # 跨天记录拆分为每天一行（匹配手动核算格式）
            dt_start = parse_datetime(raw_start)
            dt_end = parse_datetime(raw_end)
            detail_date = parse_date_cell(detail_raw)

            if dt_start and dt_end:
                s_date = dt_start.date()
                e_date = dt_end.date()
                if e_date < s_date:
                    s_date, e_date = e_date, s_date
                span_dates = daterange_inclusive(s_date, e_date)
            else:
                span_dates = []

            # 仅当明细为空或不在跨天范围内时才拆分（导出已有逐日明细则保留）
            need_expand = len(span_dates) > 1 and (not detail_date or detail_date not in span_dates)
            if need_expand:
                for d in span_dates:
                    day_key = (approval_id, str(d))
                    if day_key in approved_ids:
                        continue
                    approved_ids.add(day_key)
                    out_rows.append((
                        row[col["发起人工号"]],
                        row[col["发起人姓名"]],
                        d1, d2, d3,
                        raw_start,
                        raw_end,
                        d,
                        row[col["加班时间"]] if "加班时间" in col else None,
                        row[col["加班时长"]] if "加班时长" in col else None,
                        raw_duration,
                        row[col["是否包含法定节假日期"]] if "是否包含法定节假日期" in col else None,
                        row[col["2026法定节假日如下："]] if "2026法定节假日如下：" in col else None,
                    ))
            else:
                # 系统时长是钉钉审批申请时长，逐日明细也保留原值，不做跨天均分。
                out_rows.append((
                    row[col["发起人工号"]],
                    row[col["发起人姓名"]],
                    d1, d2, d3,
                    raw_start,
                    raw_end,
                    detail_date if detail_date else detail_raw,
                    row[col["加班时间"]] if "加班时间" in col else None,
                    row[col["加班时长"]] if "加班时长" in col else None,
                    raw_duration,
                    row[col["是否包含法定节假日期"]] if "是否包含法定节假日期" in col else None,
                    row[col["2026法定节假日如下："]] if "2026法定节假日如下：" in col else None,
                ))
            kept += 1

    wb_src.close()
    print(f'[清洗] 合并 {len(wb_src.worksheets)} 个 sheet，'
          f'保留 {kept} 行，剔除 {skipped} 行（含不符合审批状态及"完成+拒绝"数据）')
    return out_rows


# ── 旧导出字段状态：保留列结构，主规则不再依赖导出表节假日清单 ─────────────
_LEGAL_HOLIDAYS: set[date] = set()
_HOLIDAY_REST_DAYS: set[date] = set()
_ALL_SCHEDULE_PREMIUM_DATES: set[date] = set()
_SCHEDULE_ONLY_REST_PREMIUM_ROWS: set[tuple[str, date]] = set()

# ── 作息表（工作日历）──────────────────────────────────────────────
_WORK_CALENDAR_MAIN: set[date] = set()
_WORK_CALENDAR_CHENGDU: set[date] = set()
_WORK_CALENDAR_LOADED: bool = False
_WORK_CALENDAR_MONTH_ANCHOR: date | None = None
_WORK_CALENDAR_LEGAL_HOLIDAYS: set[date] = set()
_WORK_CALENDAR_HOLIDAY_ADJUST_REST_DAYS: set[date] = set()
_ATTENDANCE_CLOCK_HOURS_BY_CODE: dict[str, dict[date, float]] = {}
_ATTENDANCE_CLOCK_HOURS_BY_NAME: dict[str, dict[date, float]] = {}
_ATTENDANCE_NO_PUNCH_BY_CODE: dict[str, set[date]] = {}
_ATTENDANCE_NO_PUNCH_BY_NAME: dict[str, set[date]] = {}

_WORK_CAL_COLOR = "FFFFFF00"  # 作息表中工作日单元格背景色
_WORK_CAL_STATUTORY_HOLIDAY_COLOR = getattr(calc_leave, "STATUTORY_HOLIDAY_COLOR", "FFFF0000")
_WORK_CAL_COMPANY_WELFARE_COLOR = getattr(calc_leave, "COMPANY_WELFARE_COLOR", "FF0070C0")


def _reset_schedule_runtime_state() -> None:
    global _ALL_SCHEDULE_PREMIUM_DATES, _SCHEDULE_ONLY_REST_PREMIUM_ROWS
    _ALL_SCHEDULE_PREMIUM_DATES = set()
    _SCHEDULE_ONLY_REST_PREMIUM_ROWS = set()


def _reset_work_calendar_runtime_state() -> None:
    global _WORK_CALENDAR_MAIN, _WORK_CALENDAR_CHENGDU, _WORK_CALENDAR_LOADED, _WORK_CALENDAR_MONTH_ANCHOR
    global _WORK_CALENDAR_LEGAL_HOLIDAYS, _WORK_CALENDAR_HOLIDAY_ADJUST_REST_DAYS
    _WORK_CALENDAR_MAIN = set()
    _WORK_CALENDAR_CHENGDU = set()
    _WORK_CALENDAR_LOADED = False
    _WORK_CALENDAR_MONTH_ANCHOR = None
    _WORK_CALENDAR_LEGAL_HOLIDAYS = set()
    _WORK_CALENDAR_HOLIDAY_ADJUST_REST_DAYS = set()


def _cell_fg_rgb(cell) -> str:
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb":
            normalizer = getattr(calc_leave, "normalize_schedule_color", None)
            return normalizer(fg.rgb) if normalizer else fg.rgb
    except Exception:
        pass
    return ""


def _parse_cal_title(title_str: str):
    match = re.search(r"(\d{4})\D*(\d{1,2})\D*月", title_str or "")
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _fix_cal_date(cell_val, target_year: int, target_month: int):
    if isinstance(cell_val, datetime):
        try:
            return date(target_year, target_month, cell_val.day)
        except ValueError:
            return None
    if isinstance(cell_val, (int, float)):
        try:
            return date(target_year, target_month, int(cell_val))
        except ValueError:
            return None
    return None


def _parse_cal_block(ws, title_row_idx: int) -> dict | None:
    title = str(ws.cell(title_row_idx, 1).value or "").strip()
    year, month = _parse_cal_title(title)
    if not year or not month:
        return None

    header_row_idx = None
    week_key = normalize_header_name("周数")
    for row_idx in range(title_row_idx + 1, min(ws.max_row, title_row_idx + 8) + 1):
        if normalize_header_name(ws.cell(row_idx, 1).value) == week_key:
            header_row_idx = row_idx
            break
    if header_row_idx is None:
        return None

    working_days = set()
    statutory_holiday_days = set()
    company_welfare_days = set()
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        first_val = ws.cell(row_idx, 1).value
        if isinstance(first_val, (int, float)):
            for col_idx in range(2, min(8, ws.max_column) + 1):
                cell = ws.cell(row_idx, col_idx)
                current = _fix_cal_date(cell.value, year, month)
                if not current:
                    continue
                color = _cell_fg_rgb(cell)
                if color == _WORK_CAL_COLOR:
                    working_days.add(current)
                elif color == _WORK_CAL_STATUTORY_HOLIDAY_COLOR:
                    statutory_holiday_days.add(current)
                elif color == _WORK_CAL_COMPANY_WELFARE_COLOR:
                    company_welfare_days.add(current)
            continue
        if first_val is not None:
            break

    return {
        "title": title,
        "year": year,
        "month": month,
        "is_chengdu": "成都" in title or "成都" in ws.title,
        "working_days": working_days,
        "statutory_holiday_days": statutory_holiday_days,
        "company_welfare_days": company_welfare_days,
    }


def load_work_calendar(calendar_file: str, rules_config=None) -> None:
    """加载作息表，设置全局 _WORK_CALENDAR_MAIN / _WORK_CALENDAR_CHENGDU。"""
    global _WORK_CALENDAR_MAIN, _WORK_CALENDAR_CHENGDU, _WORK_CALENDAR_LOADED, _WORK_CALENDAR_MONTH_ANCHOR
    global _WORK_CALENDAR_LEGAL_HOLIDAYS, _WORK_CALENDAR_HOLIDAY_ADJUST_REST_DAYS

    _reset_work_calendar_runtime_state()

    if not calendar_file or not os.path.exists(calendar_file):
        return

    wb = openpyxl.load_workbook(calendar_file)
    try:
        blocks = []
        for ws in wb.worksheets:
            for row_idx in range(1, ws.max_row + 1):
                title = ws.cell(row_idx, 1).value
                if isinstance(title, str) and "作息时间表" in title:
                    block = _parse_cal_block(ws, row_idx)
                    if block:
                        blocks.append(block)
    finally:
        wb.close()

    if not blocks:
        return

    main_days = None
    chengdu_days = None
    month_anchor = None
    all_working_days: set[date] = set()
    all_statutory_raw_days: set[date] = set()
    all_company_welfare_days: set[date] = set()
    for block in blocks:
        if month_anchor is None:
            month_anchor = date(block["year"], block["month"], 1)
        all_working_days.update(block["working_days"])
        all_statutory_raw_days.update(block["statutory_holiday_days"])
        all_company_welfare_days.update(block["company_welfare_days"])
        if block["is_chengdu"]:
            chengdu_days = block["working_days"]
        elif main_days is None:
            main_days = block["working_days"]

    if main_days is None:
        main_days = blocks[0]["working_days"]
    if chengdu_days is None:
        chengdu_days = set(main_days)

    _WORK_CALENDAR_MAIN = main_days
    _WORK_CALENDAR_CHENGDU = chengdu_days
    _WORK_CALENDAR_LOADED = True
    _WORK_CALENDAR_MONTH_ANCHOR = month_anchor

    legal_days = calc_leave._normalize_paid_statutory_holidays(all_statutory_raw_days)
    holiday_adjust_rest_days = all_statutory_raw_days - legal_days - all_company_welfare_days
    _WORK_CALENDAR_LEGAL_HOLIDAYS = set(legal_days)
    _WORK_CALENDAR_HOLIDAY_ADJUST_REST_DAYS = set(holiday_adjust_rest_days)

    makeup_workdays = {d for d in all_working_days if d.weekday() >= 5}

    if rules_config is not None and (legal_days or holiday_adjust_rest_days or makeup_workdays):
        from rules_engine import StaticHolidayDataSource

        supported_years = {
            d.year
            for d in (set(legal_days) | set(holiday_adjust_rest_days) | set(makeup_workdays))
        }
        rules_config.holiday_data_source = StaticHolidayDataSource(
            legal_holidays=legal_days,
            holiday_adjust_rest_days=holiday_adjust_rest_days,
            makeup_workdays=makeup_workdays,
            supported_years=supported_years,
        )

    print(f"[作息表] 已加载 {len(main_days)} 个工作日（主）/ {len(chengdu_days)} 个（成都）")
    if legal_days or holiday_adjust_rest_days:
        legal_text = "、".join(d.strftime("%m/%d") for d in sorted(legal_days)) or "-"
        rest_text = "、".join(d.strftime("%m/%d") for d in sorted(holiday_adjust_rest_days)) or "-"
        print(f"[作息表] 加班倍数日期：法定3倍 {legal_text}；调休休息日2倍候选 {rest_text}")


def get_work_calendar_month_key() -> tuple[int, int] | None:
    if _WORK_CALENDAR_MONTH_ANCHOR is None:
        return None
    return (_WORK_CALENDAR_MONTH_ANCHOR.year, _WORK_CALENDAR_MONTH_ANCHOR.month)


def resolve_target_month_anchor(ws, header_map: dict[str, int], header_row: int) -> date:
    if _WORK_CALENDAR_MONTH_ANCHOR is not None:
        print(
            f"[月份] 优先使用作息表月份："
            f"{_WORK_CALENDAR_MONTH_ANCHOR.year}年{_WORK_CALENDAR_MONTH_ANCHOR.month}月"
        )
        return _WORK_CALENDAR_MONTH_ANCHOR
    return infer_target_month_anchor(ws, header_map, header_row)


def resolve_target_month_anchor_from_rows(
    src_rows: list[tuple],
    target_month_anchor: date | None = None,
) -> date:
    if target_month_anchor is not None:
        print(
            f"[月份] 手动指定处理月份："
            f"{target_month_anchor.year}年{target_month_anchor.month}月"
        )
        return target_month_anchor
    if _WORK_CALENDAR_MONTH_ANCHOR is not None:
        print(
            f"[月份] 优先使用作息表月份："
            f"{_WORK_CALENDAR_MONTH_ANCHOR.year}年{_WORK_CALENDAR_MONTH_ANCHOR.month}月"
        )
        return _WORK_CALENDAR_MONTH_ANCHOR
    target_key = _infer_target_month_key(src_rows)
    if target_key:
        year, month = target_key
        print(f"[月份] 按加班数据推断处理月份：{year}年{month}月")
        return date(year, month, 1)
    fallback = get_current_month_anchor()
    print(f"[月份] 未能识别月份，临时按运行月份处理：{fallback.year}年{fallback.month}月")
    return fallback


def resolve_overtime_date_from_src_row(row: tuple) -> date | None:
    if len(row) < len(CLEAN_OUTPUT_COLS):
        return None
    start_dt = parse_datetime(row[5])
    end_dt = parse_datetime(row[6])
    detail_date = parse_date_cell(row[7])
    if detail_date:
        if start_dt and end_dt:
            s_date = start_dt.date()
            e_date = end_dt.date()
            if e_date < s_date:
                s_date, e_date = e_date, s_date
            if s_date <= detail_date <= e_date:
                return detail_date
        else:
            return detail_date
    overtime_date = parse_date_cell(row[8])
    if overtime_date:
        return overtime_date
    if start_dt:
        return start_dt.date()
    if end_dt:
        return end_dt.date()
    return None


def filter_rows_by_target_month(
    src_rows: list[tuple],
    current_month_anchor: date,
) -> tuple[list[tuple], int]:
    if not src_rows:
        return src_rows, 0
    filtered_rows = [src_rows[0]]
    skipped_rows = 0
    for row in src_rows[1:]:
        overtime_date = resolve_overtime_date_from_src_row(row)
        if overtime_date is not None and not should_count_in_target_month(
            overtime_date, current_month_anchor
        ):
            skipped_rows += 1
            continue
        filtered_rows.append(row)
    return filtered_rows, skipped_rows


def _extract_legal_holidays(rows: list[tuple], year: int) -> set[date]:
    """兼容旧导出字段的提取函数；主规则不再使用该清单判断 2 倍 / 3 倍。"""
    if len(rows) <= 1:
        return set()
    holiday_col = 12  # CLEAN_OUTPUT_COLS 中 "2026法定节假日如下：" 的位置
    vals: set[str] = set()
    for row in rows[1:]:
        v = row[holiday_col] if len(row) > holiday_col else None
        if v:
            for part in str(v).split(";"):
                part = part.strip()
                if part:
                    vals.add(part)
    result: set[date] = set()
    for val in vals:
        m = re.match(r"(\d{1,2})\.(\d{1,2})", val)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            try:
                result.add(date(year, month, day))
            except ValueError:
                pass
    return result


def _set_holidays(legal: set[date], rest_days: set[date], schedule_premium: set[date] | None = None) -> None:
    global _LEGAL_HOLIDAYS, _HOLIDAY_REST_DAYS, _ALL_SCHEDULE_PREMIUM_DATES
    _LEGAL_HOLIDAYS = legal
    _HOLIDAY_REST_DAYS = rest_days
    if schedule_premium is not None:
        _ALL_SCHEDULE_PREMIUM_DATES = schedule_premium


def _collect_holiday_years(
    src_rows: list[tuple],
    schedule_map: dict[str, dict[str, set[date]]] | None = None,
) -> set[int]:
    years: set[int] = set()
    for row in src_rows[1:]:
        overtime_date = resolve_overtime_date_from_src_row(row)
        if overtime_date is not None:
            years.add(overtime_date.year)
    if schedule_map:
        for dates in schedule_map.values():
            for date_set in dates.values():
                years.update(d.year for d in date_set)
    return years


def find_header_row(ws) -> int:
    emp_no_key = normalize_header_name("发起人工号")
    emp_name_key = normalize_header_name("发起人姓名")
    for row_idx in range(1, min(5, ws.max_row) + 1):
        row_values = [normalize_header_name(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if emp_no_key in row_values and emp_name_key in row_values:
            return row_idx
    raise ValueError("未找到加班表表头行（需包含“发起人工号”“发起人姓名”）。")


def build_header_map(ws, header_row: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(header_row, col).value
        header_key = normalize_header_name(header)
        if header_key and header_key not in header_map:
            header_map[header_key] = col
    return header_map


def ensure_target_columns(ws, header_row: int, header_map: dict[str, int]) -> dict[str, int]:
    next_col = ws.max_column + 1
    for column_name in TARGET_COLUMNS:
        column_key = normalize_header_name(column_name)
        if column_key not in header_map:
            ws.cell(header_row, next_col).value = column_name
            header_map[column_key] = next_col
            next_col += 1
    return header_map


def _parse_premium_dates_from_notes(notes: str, year: int, month: int) -> set[date]:
    """从排班表备注列提取加班工资日期。

    支持格式：
      "4月4号-6号计算加班"  → 4/4, 4/5, 4/6
      "4月4号,4月6号计算加班" → 4/4, 4/6
      "4月4号-5号计算加班"   → 4/4, 4/5
      "4号-6号计算加班"    → 目标月份 4/4, 4/5, 4/6
      "5.1-5.3计算加班"     → 5/1, 5/2, 5/3
      "5.1-5.5计算加班"     → 5/1..5/5
    """
    result: set[date] = set()

    # "X月A号-B号" 跨天范围（同月）
    for m in re.finditer(r'(\d{1,2})\s*月\s*(\d{1,2})\s*[号日]?\s*[-—至]\s*(\d{1,2})\s*[号日]?', notes):
        m_val, d_start, d_end = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= m_val <= 12 and 1 <= d_start <= 31 and 1 <= d_end <= 31:
            for d in range(d_start, d_end + 1):
                try:
                    result.add(date(year, m_val, d))
                except ValueError:
                    pass

    # "A号-B号" 跨天范围（默认目标月份）
    for m in re.finditer(r'(?<!月)(\d{1,2})\s*[号日]\s*[-—至]\s*(\d{1,2})\s*[号日]?', notes):
        d_start, d_end = int(m.group(1)), int(m.group(2))
        if 1 <= d_start <= 31 and 1 <= d_end <= 31:
            for d in range(d_start, d_end + 1):
                try:
                    result.add(date(year, month, d))
                except ValueError:
                    pass

    # "A月B号,C月D号" 逗号分隔的多个日期
    for m in re.finditer(r'(\d{1,2})\s*月\s*(\d{1,2})\s*[号日]', notes):
        m_val, d_val = int(m.group(1)), int(m.group(2))
        # 排除已被范围匹配覆盖的
        if not re.search(rf'{m_val}\s*月\s*{d_val}\s*[号日]?\s*[-—至]', notes):
            if 1 <= m_val <= 12 and 1 <= d_val <= 31:
                try:
                    result.add(date(year, m_val, d_val))
                except ValueError:
                    pass

    # "B号" / "B日" 单日（默认目标月份）
    for m in re.finditer(r'(?<!月)(\d{1,2})\s*[号日]', notes):
        d_val = int(m.group(1))
        if not re.search(rf'{d_val}\s*[号日]\s*[-—至]', notes) and 1 <= d_val <= 31:
            try:
                result.add(date(year, month, d_val))
            except ValueError:
                pass

    # "M.D-M.D" 格式（如 5.1-5.3）
    for m in re.finditer(r'(\d{1,2})\.(\d{1,2})\s*[-—至]\s*(\d{1,2})\.(\d{1,2})', notes):
        m1, d1, m2, d2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        try:
            start = date(year, m1, d1)
            end = date(year, m2, d2)
            cur = start
            while cur <= end:
                result.add(cur)
                cur = date.fromordinal(cur.toordinal() + 1)
        except ValueError:
            pass

    # "M.D" 单日
    for m in re.finditer(r'(\d{1,2})\.(\d{1,2})', notes):
        m_val, d_val = int(m.group(1)), int(m.group(2))
        if not re.search(rf'{m_val}\.{d_val}\s*[-—至]', notes):
            try:
                result.add(date(year, m_val, d_val))
            except ValueError:
                pass

    return result


def _infer_year_month_from_text(value) -> tuple[int, int] | None:
    text = normalize_name(value)
    if not text:
        return None
    for pattern in (
        r"(\d{4})\s*年\s*(\d{1,2})\s*月",
        r"(\d{4})[-_/\.](\d{1,2})",
        r"(?<!\d)(\d{4})(\d{2})(?!\d)",
    ):
        match = re.search(pattern, text)
        if not match:
            continue
        year, month = int(match.group(1)), int(match.group(2))
        if 1900 <= year <= 2100 and 1 <= month <= 12:
            return year, month
    return None


def _resolve_schedule_year_month(
    ws,
    schedule_path: str,
    target_year_month: tuple[int, int] | None = None,
) -> tuple[int | None, int | None]:
    if target_year_month:
        return target_year_month
    for value in (ws.title, os.path.basename(schedule_path)):
        parsed = _infer_year_month_from_text(value)
        if parsed:
            return parsed
    for row in range(1, min(ws.max_row, 8) + 1):
        for col in range(1, min(ws.max_column, 8) + 1):
            parsed = _infer_year_month_from_text(ws.cell(row, col).value)
            if parsed:
                return parsed
    return None, None


def _parse_schedule_date_cell(value, target_year: int | None, target_month: int | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and float(value).is_integer() and 1 <= int(value) <= 31:
        if target_year and target_month:
            try:
                return date(target_year, target_month, int(value))
            except ValueError:
                return None
        return None
    parsed_date = parse_date_cell(value)
    if parsed_date and parsed_date.year >= 2000:
        return parsed_date
    text = normalize_name(value)
    if not text:
        return None
    if target_year and target_month:
        match = re.fullmatch(r"(\d{1,2})\s*[号日]?", text)
        if match:
            try:
                return date(target_year, target_month, int(match.group(1)))
            except ValueError:
                return None
    match = re.fullmatch(r"(\d{1,2})\s*月\s*(\d{1,2})\s*[号日]?", text)
    if match and target_year:
        try:
            return date(target_year, int(match.group(1)), int(match.group(2)))
        except ValueError:
            return None
    match = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})", text)
    if match and target_year:
        try:
            return date(target_year, int(match.group(1)), int(match.group(2)))
        except ValueError:
            return None
    return None


def _add_schedule_premium_date(
    result: dict[str, dict[str, set[date]]],
    name: str,
    current_date: date,
    data_source,
) -> None:
    from rules_engine import HOLIDAY_ADJUST_REST, LEGAL_HOLIDAY, classify_overtime_date

    result.setdefault(name, {"2x": set(), "3x": set()})
    date_type = classify_overtime_date(current_date, data_source=data_source)
    if date_type == LEGAL_HOLIDAY:
        result[name]["3x"].add(current_date)
        _ALL_SCHEDULE_PREMIUM_DATES.add(current_date)
    elif date_type == HOLIDAY_ADJUST_REST:
        result[name]["2x"].add(current_date)
        _ALL_SCHEDULE_PREMIUM_DATES.add(current_date)


def _merge_schedule_maps(
    target: dict[str, dict[str, set[date]]],
    source: dict[str, dict[str, set[date]]],
) -> None:
    for name, dates in (source or {}).items():
        target_dates = target.setdefault(name, {"2x": set(), "3x": set()})
        for bucket_name, values in dates.items():
            target_dates.setdefault(bucket_name, set()).update(values or set())


def parse_schedule(
    schedule_path: str,
    sheet_name: str | None,
    legal_holidays: set[date] | None = None,
    rules_config=None,
    target_year_month: tuple[int, int] | None = None,
) -> dict[str, dict[str, set[date]]]:
    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    target_year, target_month = _resolve_schedule_year_month(ws, schedule_path, target_year_month)
    data_source = getattr(rules_config, "holiday_data_source", None)

    # 尝试旧格式：第2行有"2倍"/"3倍"标记列
    two_x_label_col = None
    three_x_label_col = None
    for col in range(1, ws.max_column + 1):
        marker = ws.cell(2, col).value
        if marker == "2倍":
            two_x_label_col = col
        elif marker == "3倍":
            three_x_label_col = col

    if two_x_label_col is not None and three_x_label_col is not None:
        # ── 旧格式 ──
        two_x_date_cols: list[tuple[int, date]] = []
        three_x_date_cols: list[tuple[int, date]] = []

        for col in range(2, two_x_label_col):
            current_date = _parse_schedule_date_cell(ws.cell(1, col).value, target_year, target_month)
            if current_date:
                two_x_date_cols.append((col, current_date))

        for col in range(two_x_label_col + 1, three_x_label_col):
            current_date = _parse_schedule_date_cell(ws.cell(1, col).value, target_year, target_month)
            if current_date:
                three_x_date_cols.append((col, current_date))

        if not two_x_date_cols and not three_x_date_cols:
            raise ValueError("排班表未解析到日期列。")

        result: dict[str, dict[str, set[date]]] = {}
        for row in range(3, ws.max_row + 1):
            name = normalize_name(ws.cell(row, 1).value)
            if not name:
                continue
            result[name] = {"2x": set(), "3x": set()}
            for col, current_date in two_x_date_cols:
                value = ws.cell(row, col).value
                if not is_blank(value):
                    result[name]["2x"].add(current_date)
            for col, current_date in three_x_date_cols:
                value = ws.cell(row, col).value
                if not is_blank(value):
                    result[name]["3x"].add(current_date)
        return result

    # ── 新格式：支持周块或整月横排；按节假日数据源逐日分类。 ──
    date_rows: list[tuple[int, list[tuple[int, date]]]] = []
    for row in range(1, ws.max_row + 1):
        first_cell = normalize_header_name(ws.cell(row, 1).value)
        if first_cell and "日期" not in first_cell:
            continue
        block_date_cols: list[tuple[int, date]] = []
        for col in range(2, ws.max_column + 1):
            current_date = _parse_schedule_date_cell(ws.cell(row, col).value, target_year, target_month)
            if current_date:
                block_date_cols.append((col, current_date))
        if len(block_date_cols) >= 2:
            date_rows.append((row, block_date_cols))

    if not date_rows:
        raise ValueError("排班表未找到日期行。")

    result = {}

    for i, (header_row, block_date_cols) in enumerate(date_rows):
        end_row = date_rows[i + 1][0] if i + 1 < len(date_rows) else ws.max_row + 1

        for row in range(header_row + 2, end_row):
            raw_name = ws.cell(row, 1).value
            if not raw_name:
                continue
            name = normalize_name(raw_name)
            if not name:
                continue
            if "排班" in name or "作息" in name:
                continue

            # 根据实际排班判定加班。
            has_premium_date = False
            for col, current_date in block_date_cols:
                if target_year and target_month:
                    if current_date.year != target_year or current_date.month != target_month:
                        continue
                cell_val = ws.cell(row, col).value
                if cell_val is None or str(cell_val).strip().upper() == "OFF":
                    continue  # 无排班 or OFF → 不算加班
                before_count = sum(len(values) for values in result.get(name, {}).values())
                _add_schedule_premium_date(result, name, current_date, data_source)
                after_count = sum(len(values) for values in result.get(name, {}).values())
                has_premium_date = has_premium_date or after_count > before_count

            # 兼容客服排班等备注口径：行内备注明确写“计算加班”时，按备注日期补录。
            if target_year and target_month:
                for col in range(1, ws.max_column + 1):
                    note_text = normalize_name(ws.cell(row, col).value)
                    if "计算加班" not in note_text:
                        continue
                    for current_date in _parse_premium_dates_from_notes(note_text, target_year, target_month):
                        if current_date.year == target_year and current_date.month == target_month:
                            _add_schedule_premium_date(result, name, current_date, data_source)
                            has_premium_date = True

            if not has_premium_date and name in result and not any(result[name].values()):
                result.pop(name, None)

    count = sum(len(v["2x"]) + len(v["3x"]) for v in result.values())
    print(f'[排班表] 解析到 {len(result)} 名员工，共 {count} 条加班工资日期（从排班数据判定）')
    return result


def load_schedule_if_available(
    schedule_path: str | None,
    sheet_name: str | None = None,
    target_year_month: tuple[int, int] | None = None,
    legal_holidays: set[date] | None = None,
    rules_config=None,
) -> tuple[dict[str, dict[str, set[date]]], bool]:
    _reset_schedule_runtime_state()
    path = normalize_name(schedule_path)
    if not path:
        print("未提供排班表，按无排班表处理。")
        return {}, False
    if not os.path.exists(path):
        print(f"未找到排班表文件：{path}，按无排班表处理。")
        return {}, False
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
    except Exception as exc:
        print(f"排班表打开失败：{exc}；按无排班表处理。")
        return {}, False

    if sheet_name:
        candidate_sheets = [sheet_name]
    else:
        candidate_sheets = []
        if target_year_month:
            target_tag = f"{target_year_month[0]}{target_year_month[1]:02d}"
            for sn in sheet_names:
                clean_sheet_name = sn.replace(" ", "")
                if clean_sheet_name == target_tag or target_tag in clean_sheet_name:
                    candidate_sheets.append(sn)
                    print(f'[排班表] 自动匹配 sheet: "{sn}"')
        if not candidate_sheets:
            candidate_sheets.extend(sheet_names)

    merged: dict[str, dict[str, set[date]]] = {}
    parse_errors: list[str] = []
    loaded_sheets = 0
    for candidate_sheet in candidate_sheets:
        try:
            parsed = parse_schedule(
                path,
                candidate_sheet,
                legal_holidays or set(),
                rules_config,
                target_year_month,
            )
        except Exception as exc:
            parse_errors.append(f"{candidate_sheet}: {exc}")
            continue
        count = sum(len(v["2x"]) + len(v["3x"]) for v in parsed.values())
        if count <= 0:
            continue
        _merge_schedule_maps(merged, parsed)
        loaded_sheets += 1
        if sheet_name:
            break

    if loaded_sheets:
        print(f"[排班表] {os.path.basename(path)} 成功载入 {loaded_sheets} 个 sheet")
        return merged, True

    detail = "；".join(parse_errors[:3]) if parse_errors else "未识别到加班日期"
    print(f"排班表解析失败：{detail}；按无排班表处理。")
    return {}, False


def _infer_target_month_key(src_rows: list[tuple]) -> tuple[int, int] | None:
    """从清洗后的导出数据推断 (year, month)。"""
    month_counts: dict[tuple[int, int], int] = {}
    for row in src_rows[1:]:
        start = parse_datetime(row[5]) if len(row) > 5 else None
        if start:
            key = (start.year, start.month)
            month_counts[key] = month_counts.get(key, 0) + 1
    if not month_counts:
        return None
    return max(month_counts, key=month_counts.get)


def _parse_group_names(value: str | None) -> set[str]:
    text = normalize_name(value)
    if not text:
        return set()
    return {part.strip() for part in re.split(r"[，,;；/|]+", text) if part.strip()}


def _is_rest_premium_employee_group(
    employee_groups: set[str],
    rules_config,
) -> bool:
    """判断员工是否属于假期安排日可按 2 倍补录的部门组。"""
    allowed_groups = _parse_group_names(rules_config.schedule_augment_rest_dept_group)
    return bool(employee_groups & allowed_groups)


def _is_rest_premium_excluded_employee(emp_name: str | None, employee_code: str | None = None) -> bool:
    """节假日调休休息日 2 倍加班工资的人员例外名单。"""
    return (
        normalize_name(emp_name) in REST_PREMIUM_EXCLUDED_NAMES
        or normalize_employee_code(employee_code) in REST_PREMIUM_EXCLUDED_CODES
    )


def _lookup_attendance_clock_hours(
    employee_code: str | None,
    emp_name: str | None,
    current_date: date,
) -> float | None:
    code = normalize_employee_code(employee_code)
    name = normalize_name(emp_name)
    if code:
        value = _ATTENDANCE_CLOCK_HOURS_BY_CODE.get(code, {}).get(current_date)
        if value is not None:
            return value
    if name:
        return _ATTENDANCE_CLOCK_HOURS_BY_NAME.get(name, {}).get(current_date)
    return None


def _has_effective_clock_hours(clock_hours: float | int | None) -> bool:
    return clock_hours is not None and float(clock_hours) > 0.01


def _has_empty_attendance_punch(
    employee_code: str | None,
    emp_name: str | None,
    current_date: date,
) -> bool:
    code = normalize_employee_code(employee_code)
    name = normalize_name(emp_name)
    return (
        (bool(code) and current_date in _ATTENDANCE_NO_PUNCH_BY_CODE.get(code, set()))
        or (bool(name) and current_date in _ATTENDANCE_NO_PUNCH_BY_NAME.get(name, set()))
    )


def _calc_premium_day_value(
    raw_hours: float | None,
    clock_hours: float | None,
    standard_hours_per_day: float,
) -> float:
    premium_hours = _calc_premium_hour_value(
        raw_hours,
        clock_hours,
        standard_hours_per_day,
    )
    return _calc_premium_days_from_hours(premium_hours, standard_hours_per_day)


def _calc_premium_hour_value(
    raw_hours: float | None,
    clock_hours: float | None,
    standard_hours_per_day: float,
) -> float:
    standard_hours = _normalize_standard_hours(standard_hours_per_day)
    if _has_effective_clock_hours(clock_hours):
        return round(
            min(float(clock_hours), standard_hours),
            PREMIUM_HOUR_VALUE_PRECISION,
        ) or 0.0
    if raw_hours is None:
        return standard_hours
    return round(
        min(raw_hours, standard_hours),
        PREMIUM_HOUR_VALUE_PRECISION,
    ) or 0.0


def _calc_premium_days_from_hours(
    premium_hours: float,
    standard_hours_per_day: float,
) -> float:
    standard_hours = _normalize_standard_hours(standard_hours_per_day)
    return round(
        premium_hours / standard_hours,
        PREMIUM_DAY_VALUE_PRECISION,
    ) or 0.0


def _normalize_standard_hours(standard_hours_per_day: float) -> float:
    if standard_hours_per_day <= 0:
        return STANDARD_HOURS_PER_DAY
    return float(standard_hours_per_day)


def augment_with_schedule_only(
    src_rows: list[tuple],
    schedule_map: dict[str, dict[str, set[date]]],
    legal_holidays: set[date] | None = None,
    name_group_map: dict[str, str] | None = None,
    rules_config=None,
    special_chengdu_names: set[str] | tuple[str, ...] | None = None,
    employee_department_map: dict[str, dict[str, str]] | None = None,
    target_month_anchor: date | None = None,
) -> list[tuple]:
    """为排班表中有"计算加班"标注、但审批导出中无记录的人员生成合成行。"""
    if rules_config is None:
        from rules_engine import get_default_config
        rules_config = get_default_config()

    from rules_engine import (
        HOLIDAY_ADJUST_REST,
        LEGAL_HOLIDAY,
        classify_employee,
        classify_overtime_date,
    )

    data_source = getattr(rules_config, "holiday_data_source", None)
    if not schedule_map or len(src_rows) <= 1:
        return src_rows

    target_key = (
        (target_month_anchor.year, target_month_anchor.month)
        if target_month_anchor is not None
        else (get_work_calendar_month_key() or _infer_target_month_key(src_rows))
    )
    if not target_key:
        return src_rows
    target_year, target_month = target_key

    export_dates_by_name: dict[str, set[date]] = {}
    for row in src_rows[1:]:
        name = normalize_name(row[1]) if len(row) > 1 else ""
        if name:
            overtime_date = resolve_overtime_date_from_src_row(row)
            if overtime_date:
                export_dates_by_name.setdefault(name, set()).add(overtime_date)

    extra_rows: list[tuple] = []
    skipped_existing_rows = 0
    skipped_rest_rows = 0
    skipped_rest_people: set[str] = set()
    for name, dates in schedule_map.items():
        all_premium = dates.get("2x", set()) | dates.get("3x", set())
        month_dates = sorted(
            d for d in all_premium
            if d.year == target_year and d.month == target_month
            and d not in export_dates_by_name.get(name, set())
        )
        skipped_existing_rows += sum(
            1 for d in all_premium
            if d.year == target_year and d.month == target_month
            and d in export_dates_by_name.get(name, set())
        )
        if not month_dates:
            continue
        roster_info = _lookup_employee_department(employee_department_map, "", name)
        att_group = roster_info.get("attendance_group") or (name_group_map.get(name, "") if name_group_map else "")
        dept_d1 = roster_info.get("dept1", "")
        dept_d2 = roster_info.get("dept2", "")
        dept_d3 = roster_info.get("dept3", "") or att_group
        for d in month_dates:
            date_type = classify_overtime_date(d, data_source=data_source)
            if date_type == LEGAL_HOLIDAY and rules_config.schedule_augment_holidays:
                start_dt = datetime(d.year, d.month, d.day, 9, 0)
                end_dt = datetime(d.year, d.month, d.day, 18, 30)
                extra_rows.append((
                    "", name, dept_d1 or None, dept_d2 or None, dept_d3 or None,
                    start_dt, end_dt, d,
                    None, None, 8,
                    "含法定日期", None,
                ))
            elif date_type == HOLIDAY_ADJUST_REST:
                # 节假日调休休息日 → 用规则引擎判断部门组
                groups = classify_employee(rules_config, dept_d1, dept_d2, dept_d3, att_group)
                # 客服排班这类场景里，部分补录人员不在考勤导入表中，拿不到考勤组。
                # 只要排班备注已明确标注"计算加班"日期，就按排班口径补录，避免整行漏掉。
                allow_schedule_fallback = not any((dept_d1, dept_d2, dept_d3, att_group))
                if (
                    not _is_rest_premium_excluded_employee(name)
                    and (_is_rest_premium_employee_group(groups, rules_config) or allow_schedule_fallback)
                ):
                    if allow_schedule_fallback:
                        _SCHEDULE_ONLY_REST_PREMIUM_ROWS.add((name, d))
                    start_dt = datetime(d.year, d.month, d.day, 9, 0)
                    end_dt = datetime(d.year, d.month, d.day, 18, 30)
                    extra_rows.append((
                        "", name, dept_d1 or None, dept_d2 or None, dept_d3 or None,
                        start_dt, end_dt, d,
                        None, None, 8,
                        None, None,
                    ))
                else:
                    skipped_rest_rows += 1
                    skipped_rest_people.add(name)

    holiday_count = sum(1 for r in extra_rows if len(r) > 11 and r[11] == "含法定日期")
    rest_count = len(extra_rows) - holiday_count
    if extra_rows:
        parts = []
        if holiday_count:
            parts.append(f"法定节假日 {holiday_count} 行")
        if rest_count:
            parts.append(f"休息日 {rest_count} 行")
        print(f"[排班表补录] 新增 {len(extra_rows)} 行排班表合成记录（{'，'.join(parts)}），涉及 "
              f"{len({r[1] for r in extra_rows})} 人")
    if skipped_rest_rows:
        allowed_groups = "、".join(sorted(_parse_group_names(rules_config.schedule_augment_rest_dept_group)))
        print(
            f"[排班表补录] 跳过 {skipped_rest_rows} 行休息日补录记录，涉及 "
            f"{len(skipped_rest_people)} 人（缺少部门信息或不属于{allowed_groups}，无法补录）"
        )
    if skipped_existing_rows:
        print(f"[排班表补录] 已跳过 {skipped_existing_rows} 行导出表中已存在的同人同日记录")
    return src_rows + extra_rows


def calc_final_hours_for_time_off(raw_hours: float | None, day_count: int = 1) -> float | None:
    if raw_hours is None:
        return None
    max_hours = max(day_count, 1) * STANDARD_HOURS_PER_DAY
    return round2(min(raw_hours, max_hours))


def _is_operations_support_dept(ws, row_idx: int, header_map: dict[str, int]) -> bool:
    """判断员工是否属于运营支撑部（任一级别部门包含即可）。"""
    for col_name in ("一级部门", "二级部门", "三级部门"):
        if col_name in header_map:
            val = ws.cell(row_idx, header_map[col_name]).value
            if val and "运营支撑部" in str(val):
                return True
    return False


def _is_chengdu_dept(ws, row_idx: int, header_map: dict[str, int]) -> bool:
    """判断员工是否属于成都部门（任一级别部门包含'成都'即可）。"""
    for col_name in ("一级部门", "二级部门", "三级部门"):
        if col_name in header_map:
            val = ws.cell(row_idx, header_map[col_name]).value
            if val and "成都" in str(val):
                return True
    return False


def _row_span_dates(start_dt: datetime | None, end_dt: datetime | None, overtime_date: date | None) -> list[date]:
    if start_dt and end_dt:
        s_date = start_dt.date()
        e_date = end_dt.date()
        if e_date < s_date:
            s_date, e_date = e_date, s_date
        return daterange_inclusive(s_date, e_date)
    return [overtime_date] if overtime_date else []


def _is_special_overtime_rule_row(
    employee_groups: set[str],
    emp_name: str | None,
    employee_code: str | None,
    row_dates: list[date],
    rules_config,
) -> bool:
    """特殊部门/特殊人员继续走旧口径，不进入普通逐日认定规则。"""
    if _is_rest_premium_employee_group(employee_groups, rules_config):
        return True
    if _is_rest_premium_excluded_employee(emp_name, employee_code):
        return True
    name = normalize_name(emp_name)
    return any((name, d) in _SCHEDULE_ONLY_REST_PREMIUM_ROWS for d in row_dates)


def _general_context_group_key(item: dict) -> tuple:
    if item["detail_hours"] is not None:
        return ("row", item["row_idx"])
    if len(item["span_dates"]) > 1 and item["raw_hours"] is not None:
        start_dt = item["start_dt"]
        end_dt = item["end_dt"]
        return (
            "span",
            item["employee_code"] or "",
            item["emp_name"] or "",
            start_dt.isoformat() if start_dt else "",
            end_dt.isoformat() if end_dt else "",
            _format_hours_for_remark(item["raw_hours"]),
        )
    return ("row", item["row_idx"])


def _build_general_overtime_contexts(
    ws,
    header_map: dict[str, int],
    row_indices: list[int],
    name_group_map: dict[str, str] | None,
    rules_config,
    employee_department_map: dict[str, dict[str, str]] | None = None,
) -> dict[int, dict]:
    """为普通人员/部门预先计算逐日认定时长。

    单日特殊部门仍走原有 fill_row 逻辑；跨多日审批统一在这里按审批总时长封顶。
    """
    from rules_engine import classify_employee

    groups: dict[tuple, list[dict]] = {}
    standard_hours = _normalize_standard_hours(rules_config.standard_hours_per_day)

    for row_idx in row_indices:
        employee_code = normalize_employee_code(ws.cell(row_idx, header_map["发起人工号"]).value)
        emp_name = normalize_name(ws.cell(row_idx, header_map["发起人姓名"]).value)
        start_dt = parse_datetime(ws.cell(row_idx, header_map["开始时间"]).value)
        end_dt = parse_datetime(ws.cell(row_idx, header_map["结束时间"]).value)
        overtime_date = resolve_overtime_date(ws, row_idx, header_map, start_dt, end_dt)
        if overtime_date is None:
            continue

        dept_d1, dept_d2, dept_d3, att_group = _resolve_employee_department_fields(
            ws,
            row_idx,
            header_map,
            employee_code,
            emp_name,
            name_group_map,
            employee_department_map,
        )
        employee_groups = classify_employee(rules_config, dept_d1, dept_d2, dept_d3, att_group)
        span_dates = _row_span_dates(start_dt, end_dt, overtime_date)
        raw_hours = parse_float(ws.cell(row_idx, header_map["时长"]).value) if "时长" in header_map else None
        detail_hours = parse_float(ws.cell(row_idx, header_map["加班时长"]).value) if "加班时长" in header_map else None
        is_span_total_row = len(span_dates) > 1 and raw_hours is not None and detail_hours is None
        if (
            _is_special_overtime_rule_row(employee_groups, emp_name, employee_code, span_dates, rules_config)
            and not is_span_total_row
        ):
            continue

        clock_hours = _lookup_attendance_clock_hours(employee_code, emp_name, overtime_date)
        item = {
            "row_idx": row_idx,
            "employee_code": employee_code,
            "emp_name": emp_name,
            "start_dt": start_dt,
            "end_dt": end_dt,
            "overtime_date": overtime_date,
            "span_dates": span_dates,
            "raw_hours": raw_hours,
            "detail_hours": detail_hours,
            "system_hours": detail_hours if detail_hours is not None else raw_hours,
            "clock_hours": clock_hours,
            "employee_groups": set(employee_groups),
            "no_attendance": _has_empty_attendance_punch(employee_code, emp_name, overtime_date),
            "rest_premium_excluded": _is_rest_premium_excluded_employee(emp_name, employee_code),
        }
        groups.setdefault(_general_context_group_key(item), []).append(item)

    contexts: dict[int, dict] = {}

    def _item_priority(item: dict) -> tuple[int, date, int]:
        has_effective_clock = _has_effective_clock_hours(item.get("clock_hours"))
        no_attendance = bool(item.get("no_attendance"))
        attendance_rank = 0 if has_effective_clock else (2 if no_attendance else 1)
        return attendance_rank, item["overtime_date"], item["row_idx"]

    def _item_candidate_hours(item: dict) -> tuple[float, str, list[str]]:
        remarks: list[str] = []
        clock_hours = item["clock_hours"]
        if _has_effective_clock_hours(clock_hours):
            return (
                round(min(float(clock_hours), standard_hours), PREMIUM_HOUR_VALUE_PRECISION) or 0.0,
                "clock",
                remarks,
            )
        system_hours = item["system_hours"]
        if system_hours is None:
            remarks.append(NO_SYSTEM_HOURS_REMARK)
            return 0.0, "system", remarks
        remarks.append(NO_DOUBLE_ATTENDANCE_REMARK)
        return (
            round(min(float(system_hours), standard_hours), PREMIUM_HOUR_VALUE_PRECISION) or 0.0,
            "system",
            remarks,
        )

    for key, items in groups.items():
        ordered_items = sorted(items, key=lambda item: (item["overtime_date"], item["row_idx"]))
        is_total_group = key[0] == "span" and len(ordered_items) > 1

        if is_total_group:
            total_system_hours = ordered_items[0]["raw_hours"]
            assignments: dict[int, tuple[float, str, list[str]]] = {}
            final_total = 0.0
            remaining_hours = float(total_system_hours) if total_system_hours is not None else None

            for item in sorted(ordered_items, key=_item_priority):
                candidate_hours, basis, remarks = _item_candidate_hours(item)
                if item.get("no_attendance"):
                    remarks.append(NO_ATTENDANCE_REMARK)

                if remaining_hours is None:
                    final_hours = candidate_hours
                elif remaining_hours > 0.01 and candidate_hours > 0.01:
                    final_hours = round(
                        min(candidate_hours, remaining_hours),
                        PREMIUM_HOUR_VALUE_PRECISION,
                    ) or 0.0
                    remaining_hours -= final_hours
                    if final_hours + 0.01 < candidate_hours:
                        remarks.append(
                            f"受审批总时长上限影响，本日按{_format_hours_for_remark(final_hours)}小时回填，需核实"
                        )
                else:
                    final_hours = 0.0
                    remarks.append(APPROVAL_TOTAL_CAP_EXHAUSTED_REMARK)

                assignments[item["row_idx"]] = (final_hours, basis, remarks)
                final_total += final_hours

            group_remark = _system_total_mismatch_remark(total_system_hours, final_total)
            for item in ordered_items:
                final_hours, basis, remarks = assignments[item["row_idx"]]
                contexts[item["row_idx"]] = {
                    "final_hours": round2(final_hours) or 0.0,
                    "basis": basis,
                    "overtime_date": item["overtime_date"],
                    "remarks": _join_remarks(*(remarks + [group_remark])),
                    "employee_groups": set(item.get("employee_groups") or set()),
                    "rest_premium_excluded": item.get("rest_premium_excluded", False),
                }
            continue

        for item in ordered_items:
            system_hours = item["system_hours"]
            clock_hours = item["clock_hours"]
            remarks: list[str] = []
            if _has_effective_clock_hours(clock_hours):
                final_hours = round(min(clock_hours, standard_hours), PREMIUM_HOUR_VALUE_PRECISION) or 0.0
                basis = "clock"
                remarks.append(_recognized_hours_mismatch_remark(system_hours, final_hours, basis))
            elif system_hours is None:
                final_hours = 0.0
                basis = "system"
                remarks.append(NO_SYSTEM_HOURS_REMARK)
            else:
                final_hours = round(min(system_hours, standard_hours), PREMIUM_HOUR_VALUE_PRECISION) or 0.0
                basis = "system"
                remarks.append(NO_DOUBLE_ATTENDANCE_REMARK)
                remarks.append(_recognized_hours_mismatch_remark(system_hours, final_hours, basis))

            contexts[item["row_idx"]] = {
                "final_hours": round2(final_hours) or 0.0,
                "basis": basis,
                "overtime_date": item["overtime_date"],
                "remarks": _join_remarks(*remarks),
                "employee_groups": set(item.get("employee_groups") or set()),
                "rest_premium_excluded": item.get("rest_premium_excluded", False),
            }

    return contexts


def _fill_general_overtime_row(
    ws,
    row_idx: int,
    header_map: dict[str, int],
    attendance_missing_checkout_map: dict[str, set[date]],
    current_month_anchor: date,
    employee_code: str,
    context: dict,
    rules_config,
) -> str | None:
    from rules_engine import HOLIDAY_ADJUST_REST, classify_date, determine_action, should_mark_no_punch

    overtime_date = context["overtime_date"]
    final_hours = float(context.get("final_hours") or 0.0)
    employee_groups = set(context.get("employee_groups") or set())
    count_in_target_month = should_count_in_target_month(overtime_date, current_month_anchor)
    date_type = classify_date(rules_config, overtime_date, set(), None, False)
    action, multiplier = determine_action(rules_config, date_type, employee_groups)
    if date_type == HOLIDAY_ADJUST_REST and context.get("rest_premium_excluded"):
        action, multiplier = "调休", 1.0

    two_x_hours = 0.0
    three_x_hours = 0.0
    if action == "加班工资":
        if abs(float(multiplier) - 2.0) <= 0.01:
            two_x_hours = final_hours
        elif abs(float(multiplier) - 3.0) <= 0.01:
            three_x_hours = final_hours

    two_x_days = _calc_premium_days_from_hours(
        two_x_hours,
        rules_config.standard_hours_per_day,
    )
    three_x_days = _calc_premium_days_from_hours(
        three_x_hours,
        rules_config.standard_hours_per_day,
    )

    final_hours_col = header_map["最终加班时长（小时）"]
    overtime_type_col = header_map["加班类型"]
    two_x_hours_col = header_map["2倍加班小时"]
    three_x_hours_col = header_map["3倍加班小时"]
    two_x_col = header_map["2倍加班天数"]
    three_x_col = header_map["3倍加班天数"]
    remark_col = header_map["备注"]
    system_col = header_map["系统操作"]

    if count_in_target_month:
        ws.cell(row_idx, final_hours_col).value = compact_number(round2(final_hours))
    else:
        ws.cell(row_idx, final_hours_col).value = format_overtime_month_label(
            overtime_date,
            current_month_anchor,
        )

    ws.cell(row_idx, overtime_type_col).value = action
    two_x_hours_cell = ws.cell(row_idx, two_x_hours_col)
    three_x_hours_cell = ws.cell(row_idx, three_x_hours_col)
    two_x_cell = ws.cell(row_idx, two_x_col)
    three_x_cell = ws.cell(row_idx, three_x_col)
    two_x_hours_cell.value = compact_number(two_x_hours) if count_in_target_month and two_x_hours else None
    three_x_hours_cell.value = compact_number(three_x_hours) if count_in_target_month and three_x_hours else None
    two_x_cell.value = compact_number(two_x_days) if count_in_target_month and two_x_days else None
    three_x_cell.value = compact_number(three_x_days) if count_in_target_month and three_x_days else None
    two_x_hours_cell.number_format = PREMIUM_HOUR_NUMBER_FORMAT
    three_x_hours_cell.number_format = PREMIUM_HOUR_NUMBER_FORMAT
    two_x_cell.number_format = PREMIUM_DAY_NUMBER_FORMAT
    three_x_cell.number_format = PREMIUM_DAY_NUMBER_FORMAT
    ws.cell(row_idx, remark_col).value = context.get("remarks")

    system_operation = (
        rules_config.no_punch_mark
        if should_mark_no_punch(
            rules_config,
            attendance_missing_checkout_map,
            employee_code,
            overtime_date,
            action,
        )
        else None
    )
    ws.cell(row_idx, system_col).value = system_operation
    return system_operation


def fill_row(
    ws,
    row_idx: int,
    header_map: dict[str, int],
    attendance_missing_checkout_map: dict[str, set[date]],
    current_month_anchor: date,
    name_group_map: dict[str, str] | None = None,
    rules_config=None,
    special_chengdu_names: set[str] | tuple[str, ...] | None = None,
    general_overtime_contexts: dict[int, dict] | None = None,
    employee_department_map: dict[str, dict[str, str]] | None = None,
) -> str | None:
    # 延迟导入避免循环依赖
    if rules_config is None:
        from rules_engine import get_default_config
        rules_config = get_default_config()

    from rules_engine import (
        HOLIDAY_ADJUST_REST,
        LEGAL_HOLIDAY,
        calculate_hours,
        classify_date,
        classify_employee,
        should_mark_no_punch,
    )

    employee_code = normalize_employee_code(ws.cell(row_idx, header_map["发起人工号"]).value)
    start_dt = parse_datetime(ws.cell(row_idx, header_map["开始时间"]).value)
    end_dt = parse_datetime(ws.cell(row_idx, header_map["结束时间"]).value)
    overtime_date = resolve_overtime_date(ws, row_idx, header_map, start_dt, end_dt)
    count_in_target_month = should_count_in_target_month(overtime_date, current_month_anchor)
    raw_hours = parse_float(ws.cell(row_idx, header_map["时长"]).value) if "时长" in header_map else None

    # 跨天记录已拆分为每天一行时，用明细日期而非起止范围计算
    if overtime_date and start_dt and end_dt:
        s_date = start_dt.date()
        e_date = end_dt.date()
        if e_date < s_date:
            s_date, e_date = e_date, s_date
        if s_date != e_date:
            single_dt = datetime.combine(overtime_date, datetime.min.time())
            start_dt = single_dt
            end_dt = single_dt

    # ── 用规则引擎判断员工部门组 ───────────────────────────────
    emp_name = normalize_name(ws.cell(row_idx, header_map["发起人姓名"]).value)
    dept_d1, dept_d2, dept_d3, att_group = _resolve_employee_department_fields(
        ws,
        row_idx,
        header_map,
        employee_code,
        emp_name,
        name_group_map,
        employee_department_map,
    )
    employee_groups = classify_employee(rules_config, dept_d1, dept_d2, dept_d3, att_group)

    general_context = (general_overtime_contexts or {}).get(row_idx)
    if general_context is not None:
        return _fill_general_overtime_row(
            ws,
            row_idx,
            header_map,
            attendance_missing_checkout_map,
            current_month_anchor,
            employee_code,
            general_context,
            rules_config,
        )

    # ── 逐日分类 + 按规则引擎确定倍数 ─────────────────────────
    rest_hours = 0.0
    legal_hours = 0.0
    no_attendance_premium_dates: list[date] = []
    double_attendance_premium_dates: list[date] = []

    if start_dt and end_dt:
        s_date = start_dt.date()
        e_date = end_dt.date()
        if e_date < s_date:
            s_date, e_date = e_date, s_date
        all_dates = daterange_inclusive(s_date, e_date)
        use_attendance_for_premium = _is_rest_premium_employee_group(employee_groups, rules_config)

        for d in all_dates:
            date_type = classify_date(rules_config, d, set(), None, False)
            clock_hours = (
                _lookup_attendance_clock_hours(employee_code, emp_name, d)
                if use_attendance_for_premium
                else None
            )
            no_attendance = _has_empty_attendance_punch(employee_code, emp_name, d)
            # 已审批的加班工资日如果完全无打卡，按整天封顶计薪，备注仍提示核实。
            premium_hours = (
                _normalize_standard_hours(rules_config.standard_hours_per_day)
                if no_attendance
                else _calc_premium_hour_value(
                    raw_hours,
                    clock_hours,
                    rules_config.standard_hours_per_day,
                )
            )

            # ── 逐日独立判定倍数 ────────────────────────────────
            if date_type == LEGAL_HOLIDAY:
                # 法定节假日：全员 3 倍加班工资
                legal_hours += premium_hours
                if _has_effective_clock_hours(clock_hours):
                    double_attendance_premium_dates.append(d)
                if no_attendance:
                    no_attendance_premium_dates.append(d)
            elif date_type == HOLIDAY_ADJUST_REST:
                # 节假日调休休息日：
                # 运营支撑部 → 2 倍加班工资；其余 → 调休
                if (
                    not _is_rest_premium_excluded_employee(emp_name, employee_code)
                    and (
                        _is_rest_premium_employee_group(employee_groups, rules_config)
                        or (emp_name, d) in _SCHEDULE_ONLY_REST_PREMIUM_ROWS
                    )
                ):
                    rest_hours += premium_hours
                    if _has_effective_clock_hours(clock_hours):
                        double_attendance_premium_dates.append(d)
                    if no_attendance:
                        no_attendance_premium_dates.append(d)
            # 普通休息日 / 工作日 → 调休，不计入

    two_x_hours = round(rest_hours, PREMIUM_HOUR_VALUE_PRECISION)
    three_x_hours = round(legal_hours, PREMIUM_HOUR_VALUE_PRECISION)
    two_x_days = _calc_premium_days_from_hours(
        two_x_hours,
        rules_config.standard_hours_per_day,
    )
    three_x_days = _calc_premium_days_from_hours(
        three_x_hours,
        rules_config.standard_hours_per_day,
    )

    final_hours_col = header_map["最终加班时长（小时）"]
    overtime_type_col = header_map["加班类型"]
    two_x_hours_col = header_map["2倍加班小时"]
    three_x_hours_col = header_map["3倍加班小时"]
    two_x_col = header_map["2倍加班天数"]
    three_x_col = header_map["3倍加班天数"]
    remark_col = header_map["备注"]
    system_col = header_map["系统操作"]

    if two_x_hours or three_x_hours:
        final_hours = round2(two_x_hours + three_x_hours)
        if count_in_target_month:
            ws.cell(row_idx, final_hours_col).value = compact_number(final_hours)
        else:
            ws.cell(row_idx, final_hours_col).value = format_overtime_month_label(
                overtime_date,
                current_month_anchor,
            )
        ws.cell(row_idx, overtime_type_col).value = "加班工资"
        two_x_hours_cell = ws.cell(row_idx, two_x_hours_col)
        three_x_hours_cell = ws.cell(row_idx, three_x_hours_col)
        two_x_cell = ws.cell(row_idx, two_x_col)
        three_x_cell = ws.cell(row_idx, three_x_col)
        two_x_hours_cell.value = (
            compact_number(two_x_hours) if count_in_target_month and two_x_hours else None
        )
        three_x_hours_cell.value = (
            compact_number(three_x_hours) if count_in_target_month and three_x_hours else None
        )
        two_x_cell.value = (
            compact_number(two_x_days) if count_in_target_month and two_x_days else None
        )
        three_x_cell.value = (
            compact_number(three_x_days) if count_in_target_month and three_x_days else None
        )
        two_x_hours_cell.number_format = PREMIUM_HOUR_NUMBER_FORMAT
        three_x_hours_cell.number_format = PREMIUM_HOUR_NUMBER_FORMAT
        two_x_cell.number_format = PREMIUM_DAY_NUMBER_FORMAT
        three_x_cell.number_format = PREMIUM_DAY_NUMBER_FORMAT
        ws.cell(row_idx, remark_col).value = _join_remarks(
            _double_attendance_remark(final_hours) if double_attendance_premium_dates else None,
            NO_ATTENDANCE_REMARK if no_attendance_premium_dates else None,
            _system_hours_mismatch_remark(raw_hours, final_hours)
            if count_in_target_month
            else None,
        )
        ws.cell(row_idx, system_col).value = None
        return None

    # 调休
    day_count = 1
    if start_dt and end_dt:
        s = start_dt.date()
        e = end_dt.date()
        if e < s:
            s, e = e, s
        day_count = (e - s).days + 1
    final_hours = calculate_hours(rules_config, "调休", raw_hours, 0, 0, day_count)
    if count_in_target_month:
        ws.cell(row_idx, final_hours_col).value = compact_number(final_hours)
    else:
        ws.cell(row_idx, final_hours_col).value = format_overtime_month_label(
            overtime_date,
            current_month_anchor,
        )
    ws.cell(row_idx, overtime_type_col).value = "调休"
    ws.cell(row_idx, two_x_hours_col).value = None
    ws.cell(row_idx, three_x_hours_col).value = None
    ws.cell(row_idx, two_x_col).value = None
    ws.cell(row_idx, three_x_col).value = None
    ws.cell(row_idx, remark_col).value = (
        _system_hours_mismatch_remark(raw_hours, final_hours)
        if count_in_target_month
        else None
    )
    system_operation = (
        rules_config.no_punch_mark
        if should_mark_no_punch(
            rules_config,
            attendance_missing_checkout_map,
            employee_code,
            overtime_date,
            "调休",
        )
        else None
    )
    ws.cell(row_idx, system_col).value = system_operation
    return system_operation


def process_overtime(
    src_rows: list[tuple],
    out_file: str,
    schedule_map: dict,
    attendance_missing_checkout_map: dict[str, set[date]],
    name_group_map: dict[str, str] | None = None,
    rules_config=None,
    target_month_anchor: date | None = None,
    special_chengdu_names: tuple[str, ...] | set[str] | None = None,
    employee_department_map: dict[str, dict[str, str]] | None = None,
) -> None:
    """
    将清洗后的内存行列表写入工作簿，执行回填逻辑，保存到 out_file。
    src_rows[0] 为表头，后续为数据行。
    """
    if rules_config is None:
        from rules_engine import get_default_config
        rules_config = get_default_config()

    current_month_anchor = resolve_target_month_anchor_from_rows(
        src_rows,
        target_month_anchor,
    )
    special_chengdu_name_set = {
        normalize_name(name) for name in (special_chengdu_names or ()) if normalize_name(name)
    }
    from rules_engine import validate_holiday_years_available

    rows_for_holiday_validation, _ = filter_rows_by_target_month(
        src_rows,
        current_month_anchor,
    )
    holiday_years = _collect_holiday_years(rows_for_holiday_validation, schedule_map)
    if not holiday_years:
        holiday_years = {current_month_anchor.year}
    validate_holiday_years_available(
        holiday_years,
        getattr(rules_config, "holiday_data_source", None),
    )
    print(f"[节假日] 已校验年份数据：{', '.join(str(y) for y in sorted(holiday_years))}")
    if employee_department_map:
        employee_department_codes = sum(1 for key in employee_department_map if not key.startswith("name:"))
        print(f"[员工部门] 本次回填使用 {employee_department_codes} 个工号部门映射修正部门组判定")
    else:
        print("[员工部门] 未提供可用部门映射，沿用加班导出部门和考勤组判定部门组")
    _set_holidays(set(), set())

    # 为排班表有"计算加班"标注但审批导出无记录的人员生成合成行
    src_rows = augment_with_schedule_only(
        src_rows,
        schedule_map,
        set(),
        name_group_map,
        rules_config,
        special_chengdu_name_set,
        employee_department_map,
        current_month_anchor,
    )
    src_rows, skipped_cross_month_rows = filter_rows_by_target_month(
        src_rows,
        current_month_anchor,
    )

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "加班明细"

    for row in src_rows:
        ws.append(list(row))

    header_row = find_header_row(ws)
    header_map = build_header_map(ws, header_row)

    required_columns = ["发起人姓名", "开始时间", "结束时间", "时长"]
    missing = [c for c in required_columns if c not in header_map]
    if missing:
        raise ValueError(f"加班导出数据缺少必要字段：{', '.join(missing)}")

    header_map = ensure_target_columns(ws, header_row, header_map)

    if not _WORK_CALENDAR_LOADED:
        print("[作息表] 未加载作息表，处理月份将按加班数据自动推断")

    marked_unadded_count = 0
    data_row_indices = [
        row_idx
        for row_idx in range(header_row + 1, ws.max_row + 1)
        if not all(is_blank(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1))
    ]
    general_overtime_contexts = _build_general_overtime_contexts(
        ws,
        header_map,
        data_row_indices,
        name_group_map,
        rules_config,
        employee_department_map,
    )

    for row_idx in data_row_indices:
        system_operation = fill_row(
            ws,
            row_idx,
            header_map,
            attendance_missing_checkout_map,
            current_month_anchor,
            name_group_map,
            rules_config,
            special_chengdu_name_set,
            general_overtime_contexts,
            employee_department_map,
        )
        if system_operation == "未加":
            marked_unadded_count += 1

    output_dir = os.path.dirname(os.path.abspath(out_file))
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb_out.save(out_file)
    print(f"[回填] 系统操作列标记“未加” {marked_unadded_count} 行")
    if special_chengdu_names is not None:
        print(f"[成都作息名单] 设置 {len(special_chengdu_name_set)} 个姓名（不参与 2 倍加班判断）")
    if skipped_cross_month_rows:
        print(
            f"[跨月] 已过滤 {skipped_cross_month_rows} 行非目标月份记录，"
            "输出文件仅保留目标月份数据"
        )
    print(f"[完成] 已写出 → {out_file}，共 {ws.max_row - 1} 行")


def main() -> None:
    args = parse_args()

    print("── 加班字段自动计算工具 ─────────────────────────────────────────\n")

    # 加载规则配置
    from rules_engine import load_config_or_default
    rules_config = load_config_or_default(args.config)
    if args.config and os.path.exists(args.config):
        print(f"[规则] 已加载配置文件：{args.config}")
    else:
        print("[规则] 使用默认规则配置")

    print("── 步骤一：清洗加班系统导出表 ────────────────────────────────────")
    src_rows = clean_export_overtime(args.export)
    print()

    print("── 步骤二：回填加班字段 ──────────────────────────────────────────")
    load_work_calendar(args.work_calendar, rules_config)
    # 与回填阶段保持一致：优先使用作息表月份匹配排班表sheet。
    target_ym = get_work_calendar_month_key() or _infer_target_month_key(src_rows)
    schedule_map, _ = load_schedule_if_available(
        args.schedule,
        args.schedule_sheet,
        target_year_month=target_ym,
        legal_holidays=set(),
        rules_config=rules_config,
    )
    attendance_missing_checkout_map, _ = load_attendance_if_available(args.attendance)
    name_group_map = load_attendance_name_group_map_if_available(args.attendance)
    employee_department_map = parse_employee_department_map(args.roster)
    special_chengdu_names = parse_special_chengdu_names(args.special_chengdu_names)

    process_overtime(
        src_rows,
        args.output,
        schedule_map,
        attendance_missing_checkout_map,
        name_group_map,
        rules_config,
        special_chengdu_names=special_chengdu_names,
        employee_department_map=employee_department_map,
    )


if __name__ == "__main__":
    main()
