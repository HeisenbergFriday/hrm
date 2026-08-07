# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

import openpyxl

from tools.attendance_toolbox.python.templates import build_subsidy_source_template

MODULE_PATH = Path(__file__).resolve().parent / "subsidy" / "calc_subsidy_deduction.py"


def _load_module():
    spec = importlib.util.spec_from_file_location("calc_subsidy_deduction_all_people_test", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


class SubsidyAllPeopleMonthlySummaryTests(unittest.TestCase):
    DINGTALK_HEADERS = [
        "姓名", "考勤组", "部门", "工号", "职位", "UserId",
        "15-30分钟迟到扣款", "15-30分钟早退扣款",
        "旷工天数", "缺卡次数", "晚于22点打卡天数", "晚走补贴",
        "产研休息日加班>4小时天数", "产研休息日加班补贴",
    ]

    @classmethod
    def setUpClass(cls) -> None:
        cls.module = _load_module()

    def _write_workbook(self, path: Path, *, all_people: bool) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "月度汇总"
        ws["A1"] = "月度汇总表（补贴及扣款） 统计日期：2026-07-01 至 2026-07-31" if all_people else "旧格式月度汇总"
        headers = [
            "姓名", "考勤组", "部门" if all_people else "一级部门", "工号", "职位",
            "UserId" if all_people else "备注", "入职时间", "休息日&节假日加班打卡时长",
            "本月加班审批流时长", "本月加班时长（工作日加班）", "15分钟内迟到次数",
            "15分钟内迟到分钟数", "15-30分钟迟到次数", "15-30分钟迟到分钟数",
            "超30分钟迟到次数", "超30分钟迟到分钟数", "迟到总次数", "迟到总时长(分钟)",
            "15分钟内早退次数", "15分钟内早退分钟数", "15-30分钟早退次数",
            "15-30分钟早退分钟数", "超30分钟早退次数", "超30分钟早退分钟数",
            "早退总次数", "早退总时长(分钟)", "15-30分钟迟到扣款", "15-30分钟早退扣款",
            "旷工天数", "缺卡次数", "晚于22点打卡天数", "晚走补贴",
            "产研休息日加班>4小时天数", "产研休息日加班补贴", "1", "2", "3", "六", "日",
        ]
        for col, value in enumerate(headers, 1):
            ws.cell(4, col, value)
        rows = [
            ["空工号", "未加入考勤组", "研发部", None, "工程师", "ding-empty", "2026-01-01"],
            ["数字工号", "全员", "支持部", 7, "专员", "ding-seven", "2026-01-01"],
            ["标准工号", "全员", "研发部", "MT0001", "工程师", "ding-mt", "2026-01-01"],
        ]
        for row_index, values in enumerate(rows, 5):
            for col, value in enumerate(values, 1):
                ws.cell(row_index, col, value)
            for col in range(35, 40):
                ws.cell(row_index, col, "标准:正常\n(09:00,18:30)")
        wb.save(path)
        wb.close()

    def _write_dingtalk_summary(self, path: Path, a1_text: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "月度汇总"
        ws["A1"] = a1_text
        for col, value in enumerate(self.DINGTALK_HEADERS, 1):
            ws.cell(4, col, value)
        ws.cell(5, 1, "张三")
        ws.cell(5, 4, "MT0001")
        wb.save(path)
        wb.close()

    def _assert_full_month_error(
        self,
        error: str,
        *,
        actual_start: str,
        actual_end: str,
        year: int,
        month: int,
        expected_start: str,
        expected_end: str,
    ) -> None:
        self.assertIn(f"报表统计范围为{actual_start}至{actual_end}", error)
        self.assertIn(f"不是{year}年{month}月的完整自然月", error)
        self.assertIn(f"当前处理月份为{year}年{month}月", error)
        self.assertIn(f"系统要求完整范围为{expected_start}至{expected_end}", error)
        self.assertIn(f"重新导出{expected_start}至{expected_end}", error)
        self.assertIn("月度汇总表（补贴及扣款）", error)

    def _assert_unparseable_date_error(
        self,
        error: str,
        *,
        year: int,
        month: int,
        expected_start: str,
        expected_end: str,
    ) -> None:
        self.assertIn("报表A1单元格未包含可识别的统计日期范围", error)
        self.assertIn("统计开始日期和统计结束日期", error)
        self.assertIn(f"当前处理月份为{year}年{month}月", error)
        self.assertIn(f"系统要求完整范围为{expected_start}至{expected_end}", error)
        self.assertIn(f"重新导出{expected_start}至{expected_end}", error)

    def test_all_people_table_keeps_empty_and_special_employee_numbers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "all_people.xlsx"
            self._write_workbook(path, all_people=True)
            source = self.module.parse_source_table(str(path))
            attendance = self.module.parse_attendance(str(path), year=2026, month=7)
            self.assertEqual([row["name"] for row in source], ["空工号", "数字工号", "标准工号"])
            self.assertEqual([row["name"] for row in attendance], ["空工号", "数字工号", "标准工号"])
            self.assertEqual(source[0]["dept1"], "研发部")
            self.assertIsNone(source[0]["dept2"])
            self.assertEqual(attendance[0]["pos"], "工程师")

    def test_old_table_still_filters_nonstandard_employee_numbers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "legacy.xlsx"
            self._write_workbook(path, all_people=False)
            source = self.module.parse_source_table(str(path), year=2026, month=7)
            attendance = self.module.parse_attendance(str(path), year=2026, month=7)
            self.assertEqual([row["name"] for row in source], ["标准工号"])
            self.assertEqual([row["name"] for row in attendance], ["标准工号"])

    def test_system_template_without_a1_dates_accepts_year_and_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "subsidy_template.xlsx"
            path.write_bytes(build_subsidy_source_template())
            source = self.module.parse_source_table(str(path), year=2026, month=7)
            self.assertEqual([row["name"] for row in source], ["张三"])
            self.assertEqual(source[0]["emp_no"], "MT0001")

    def test_old_format_table_without_a1_date_accepts_year_and_month(self) -> None:
        """旧格式补贴表 A1 不含统计日期时，传入 year/month 仍应正常解析（不触发日期校验）。"""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "old_format.xlsx"
            self._write_workbook(path, all_people=False)
            # A1 = "旧格式月度汇总"，不含"月度汇总表（补贴及扣款）"，不应触发日期校验
            source = self.module.parse_source_table(str(path), year=2026, month=7)
            self.assertEqual([row["name"] for row in source], ["标准工号"])
            self.assertEqual(source[0]["emp_no"], "MT0001")

    def test_missing_deduction_columns_reported_not_misidentified(self) -> None:
        """当报表只有迟到次数/分钟数而没有迟到扣款列时，必须报告缺少扣款列，禁止误将次数/分钟数当作金额。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "月度汇总"
        ws["A1"] = "月度汇总表（补贴及扣款） 统计日期：2026-07-01 至 2026-07-31"
        headers = [
            "姓名", "考勤组", "部门", "工号", "职位", "UserId",
            "入职时间", "15-30分钟迟到次数", "15-30分钟迟到分钟数",
            "15-30分钟早退次数", "15-30分钟早退分钟数",
            "旷工天数", "缺卡次数", "晚于22点打卡天数", "晚走补贴",
            "产研休息日加班>4小时天数", "产研休息日加班补贴",
        ]
        for col, value in enumerate(headers, 1):
            ws.cell(4, col, value)
        ws.cell(5, 1, "张三")
        ws.cell(5, 4, "MT0001")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "no_deduction.xlsx"
            wb.save(path)
            wb.close()
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            error_msg = str(ctx.exception)
            self.assertIn("迟到扣款", error_msg,
                          "缺少迟到扣款列时错误信息应明确提及'迟到扣款'，"
                          "不应将'15-30分钟迟到次数'误当作扣款列。")
            self.assertIn("早退扣款", error_msg,
                          "缺少早退扣款列时错误信息应明确提及'早退扣款'。")

    def test_a1_date_mismatch_stops_processing(self) -> None:
        """报表A1统计日期与处理月份不一致时应停止并报错。"""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "wrong_month.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-08-01 至 2026-08-31",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_full_month_error(
                str(ctx.exception),
                actual_start="2026-08-01",
                actual_end="2026-08-31",
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )

    def test_full_31_day_month_allows_processing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "correct_month.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-07-01 至 2026-07-31",
            )
            source = self.module.parse_source_table(str(path), year=2026, month=7)
            self.assertEqual(len(source), 1)
            self.assertEqual(source[0]["name"], "张三")

    def test_full_30_day_month_allows_processing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "april.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-04-01 至 2026-04-30",
            )
            source = self.module.parse_source_table(str(path), year=2026, month=4)
            self.assertEqual([row["name"] for row in source], ["张三"])

    def test_full_non_leap_february_allows_processing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "february_2026.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-02-01 至 2026-02-28",
            )
            source = self.module.parse_source_table(str(path), year=2026, month=2)
            self.assertEqual([row["name"] for row in source], ["张三"])

    def test_full_leap_february_allows_processing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "february_2028.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2028-02-01 至 2028-02-29",
            )
            source = self.module.parse_source_table(str(path), year=2028, month=2)
            self.assertEqual([row["name"] for row in source], ["张三"])

    def test_leap_february_ending_on_28_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "incomplete_february_2028.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2028-02-01 至 2028-02-28",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2028, month=2)
            self._assert_full_month_error(
                str(ctx.exception),
                actual_start="2028-02-01",
                actual_end="2028-02-28",
                year=2028,
                month=2,
                expected_start="2028-02-01",
                expected_end="2028-02-29",
            )

    def test_start_date_after_month_start_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "late_start.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-07-02 至 2026-07-31",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_full_month_error(
                str(ctx.exception),
                actual_start="2026-07-02",
                actual_end="2026-07-31",
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )

    def test_end_date_before_month_end_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "early_end.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-07-01 至 2026-07-30",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_full_month_error(
                str(ctx.exception),
                actual_start="2026-07-01",
                actual_end="2026-07-30",
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )

    def test_partial_range_within_same_month_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "partial_july.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-07-10 至 2026-07-20",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_full_month_error(
                str(ctx.exception),
                actual_start="2026-07-10",
                actual_end="2026-07-20",
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )

    def test_a1_no_date_fails_closed(self) -> None:
        """A1 不含任何日期时应报错（fail-closed）。"""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "no_date.xlsx"
            self._write_dingtalk_summary(path, "月度汇总表（补贴及扣款）")
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            error = str(ctx.exception)
            self._assert_unparseable_date_error(
                error,
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )
            self.assertIn("A1内容：月度汇总表（补贴及扣款）", error)

    def test_a1_unparseable_date_fails_closed(self) -> None:
        """A1 含统计日期文字但日期格式无法解析时应报错（fail-closed）。"""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "unparseable_date.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026年7月 至 2026年7月",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_unparseable_date_error(
                str(ctx.exception),
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )

    def test_a1_with_only_one_date_fails_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "one_date.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-07-01",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_unparseable_date_error(
                str(ctx.exception),
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )

    def test_a1_cross_month_range_fails(self) -> None:
        """A1 统计日期跨月时应报错。"""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "cross_month.xlsx"
            self._write_dingtalk_summary(
                path,
                "月度汇总表（补贴及扣款） 统计日期：2026-07-01 至 2026-08-01",
            )
            with self.assertRaises(ValueError) as ctx:
                self.module.parse_source_table(str(path), year=2026, month=7)
            self._assert_full_month_error(
                str(ctx.exception),
                actual_start="2026-07-01",
                actual_end="2026-08-01",
                year=2026,
                month=7,
                expected_start="2026-07-01",
                expected_end="2026-07-31",
            )


if __name__ == "__main__":
    unittest.main()
